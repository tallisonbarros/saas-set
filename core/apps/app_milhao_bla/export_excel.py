from collections import defaultdict
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import pixels_to_EMU


BALANCE_ORDER = ("LIMBL01", "CLABL01", "CLABL02", "SECBL01", "SECBL02")
BALANCE_LABELS = {
    "LIMBL01": "MILHO",
    "SECBL01": "GERMEN",
    "SECBL02": "RESIDUO",
    "CLABL01": "MIUDO",
    "CLABL02": "GRAUDO",
}
HEADER_FILL = PatternFill(fill_type="solid", fgColor="FF7A1A")
HEADER_FONT = Font(color="FFFFFF", bold=True)
BLACK_FILL = PatternFill(fill_type="solid", fgColor="000000")


def _format_footer(export_dt):
    return f"exportado dia {export_dt.strftime('%d/%m/%Y %H:%M')} de setbrasil.club"


def _style_header(ws, row_idx=1):
    for cell in ws[row_idx]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autosize_columns(ws, min_width=12, max_width=42):
    for column_cells in ws.columns:
        first_cell = column_cells[0]
        column_letter = getattr(first_cell, "column_letter", None)
        if not column_letter:
            continue
        content_width = 0
        for cell in column_cells:
            value = cell.value
            if value is None:
                continue
            content_width = max(content_width, len(str(value)))
        ws.column_dimensions[column_letter].width = max(min_width, min(max_width, content_width + 2))


def _append_footer(ws, footer_text, spacer_rows=2):
    footer_row = ws.max_row + max(1, int(spacer_rows))
    max_col = max(1, ws.max_column)
    ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=max_col)
    footer_cell = ws.cell(row=footer_row, column=1, value=footer_text)
    footer_cell.font = Font(italic=True, size=10, color="475569")
    footer_cell.alignment = Alignment(horizontal="left", vertical="center")


def _column_width_to_pixels(width):
    return int((float(width) * 7) + 5)


def _add_logo(ws, logo_path, anchor, width=None, height=None):
    if not logo_path:
        return
    path = Path(logo_path)
    if not path.exists():
        return
    try:
        image = XLImage(str(path))
    except Exception:
        return
    if width and image.width:
        ratio = float(width) / float(image.width)
        image.width = int(width)
        image.height = int(image.height * ratio)
    elif height and image.height:
        ratio = float(height) / float(image.height)
        image.height = int(height)
        image.width = int(image.width * ratio)
    ws.add_image(image, anchor)


def _paint_merged_range_black(ws, *, row_idx, start_col=1, end_col=2):
    ws.merge_cells(start_row=row_idx, start_column=start_col, end_row=row_idx, end_column=end_col)
    for col_idx in range(start_col, end_col + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.fill = BLACK_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _add_logo_centered_in_range(
    ws,
    *,
    logo_path,
    row_idx,
    start_col=1,
    end_col=2,
    width=None,
    height=None,
):
    if not logo_path:
        return
    path = Path(logo_path)
    if not path.exists():
        return
    try:
        image = XLImage(str(path))
    except Exception:
        return

    if width and image.width:
        ratio = float(width) / float(image.width)
        image.width = int(width)
        image.height = int(image.height * ratio)
    elif height and image.height:
        ratio = float(height) / float(image.height)
        image.height = int(height)
        image.width = int(image.width * ratio)

    area_width_px = 0
    for col_idx in range(start_col, end_col + 1):
        col_letter = get_column_letter(col_idx)
        col_width = ws.column_dimensions[col_letter].width
        if col_width is None:
            col_width = 8.43
        area_width_px += _column_width_to_pixels(col_width)

    row_height_points = ws.row_dimensions[row_idx].height or 15
    area_height_px = int(float(row_height_points) * (96.0 / 72.0))

    offset_x = max(0, int((area_width_px - image.width) / 2))
    offset_y = max(0, int((area_height_px - image.height) / 2))
    marker = AnchorMarker(
        col=start_col - 1,
        row=row_idx - 1,
        colOff=pixels_to_EMU(offset_x),
        rowOff=pixels_to_EMU(offset_y),
    )
    image.anchor = OneCellAnchor(
        _from=marker,
        ext=XDRPositiveSize2D(
            cx=pixels_to_EMU(image.width),
            cy=pixels_to_EMU(image.height),
        ),
    )
    ws.add_image(image)


def _build_summary_sheet(wb, *, filename, start_date, end_date, export_dt, logo_set_path, logo_milhao_path, footer_text):
    ws = wb.active
    ws.title = "Resumo"

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 26
    ws.row_dimensions[1].height = 58

    _paint_merged_range_black(ws, row_idx=1, start_col=1, end_col=2)
    _add_logo_centered_in_range(
        ws,
        logo_path=logo_milhao_path,
        row_idx=1,
        start_col=1,
        end_col=2,
        height=50,
    )

    ws["A3"] = f"Arquivo: {filename}"
    ws["A3"].font = Font(size=12, bold=True)
    ws["A4"] = "Periodo inicial"
    ws["B4"] = start_date.strftime("%d/%m/%Y")
    ws["A5"] = "Periodo final"
    ws["B5"] = end_date.strftime("%d/%m/%Y")
    ws["A6"] = "Exportado em"
    ws["B6"] = export_dt.strftime("%d/%m/%Y %H:%M")

    set_logo_row = 8
    ws.row_dimensions[set_logo_row].height = 58
    _paint_merged_range_black(ws, row_idx=set_logo_row, start_col=1, end_col=2)
    _add_logo_centered_in_range(
        ws,
        logo_path=logo_set_path,
        row_idx=set_logo_row,
        start_col=1,
        end_col=2,
        height=19,
    )

    _append_footer(ws, footer_text, spacer_rows=1)


def _build_totals_by_balance_sheet(wb, *, entries, footer_text):
    ws = wb.create_sheet("Totais por balanca")
    ws.append(["Balanca", "Total_kg", "% do total (sem LIMBL01)"])
    _style_header(ws, 1)

    totals_map = defaultdict(float)
    for item in entries:
        balance = item.get("balance")
        totals_map[balance] += item.get("value") or 0.0

    total_without_milho = sum(value for balance, value in totals_map.items() if balance != "LIMBL01")
    for balance in BALANCE_ORDER:
        total_value = totals_map.get(balance, 0.0)
        percent = None
        if balance != "LIMBL01" and total_without_milho > 0:
            percent = (total_value / total_without_milho) * 100.0
        ws.append(
            [
                balance,
                round(total_value, 2),
                round(percent, 2) if percent is not None else None,
            ]
        )

    ws.auto_filter.ref = f"A1:C{ws.max_row}"
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=2).number_format = "#,##0.00"
        ws.cell(row=row_idx, column=3).number_format = "0.00\\%"
    _autosize_columns(ws)
    _append_footer(ws, footer_text)


def _build_hourly_readings_sheet(wb, *, entries, footer_text):
    ws = wb.create_sheet("Leituras por hora")
    ws.append(["Data", "Hora", "Balanca", "Valor_kg"])
    _style_header(ws, 1)

    if not entries:
        ws.append(["", "", "Sem dados no periodo selecionado.", ""])
        _autosize_columns(ws)
        _append_footer(ws, footer_text)
        return

    for item in entries:
        item_date = item.get("date")
        balance = item.get("balance")
        ws.append(
            [
                item_date.strftime("%d/%m/%Y") if item_date else "",
                item.get("hour") or "",
                balance,
                item.get("value"),
            ]
        )

    ws.auto_filter.ref = f"A1:D{ws.max_row}"
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=4).number_format = "#,##0.00"
    _autosize_columns(ws)
    _append_footer(ws, footer_text)


def _build_daily_totals_sheet(wb, *, entries, footer_text):
    ws = wb.create_sheet("Totais por dia")
    ws.append(
        [
            "Data",
            "LIMBL01_kg",
            "CLABL01_kg",
            "CLABL01_%",
            "CLABL02_kg",
            "CLABL02_%",
            "SECBL01_kg",
            "SECBL01_%",
            "SECBL02_kg",
            "SECBL02_%",
            "TOTAL_sem_milho_kg",
        ]
    )
    _style_header(ws, 1)

    if not entries:
        ws.append(["Sem dados no periodo selecionado.", 0, 0, 0, 0, 0, 0, 0, 0, 0, 0])
        _autosize_columns(ws)
        _append_footer(ws, footer_text)
        return

    daily_totals = defaultdict(lambda: defaultdict(float))
    for item in entries:
        day = item.get("date")
        balance = item.get("balance")
        daily_totals[day][balance] += item.get("value") or 0.0

    for day in sorted(daily_totals.keys()):
        day_values = daily_totals[day]
        composition_balances = ("CLABL01", "CLABL02", "SECBL01", "SECBL02")
        total_without_milho = (
            day_values.get("CLABL01", 0.0)
            + day_values.get("CLABL02", 0.0)
            + day_values.get("SECBL01", 0.0)
            + day_values.get("SECBL02", 0.0)
        )
        composition_percent = {}
        if total_without_milho > 0:
            for balance in composition_balances:
                composition_percent[balance] = (day_values.get(balance, 0.0) / total_without_milho) * 100.0
        else:
            for balance in composition_balances:
                composition_percent[balance] = 0.0

        ws.append(
            [
                day.strftime("%d/%m/%Y"),
                round(day_values.get("LIMBL01", 0.0), 2),
                round(day_values.get("CLABL01", 0.0), 2),
                round(composition_percent["CLABL01"], 2),
                round(day_values.get("CLABL02", 0.0), 2),
                round(composition_percent["CLABL02"], 2),
                round(day_values.get("SECBL01", 0.0), 2),
                round(composition_percent["SECBL01"], 2),
                round(day_values.get("SECBL02", 0.0), 2),
                round(composition_percent["SECBL02"], 2),
                round(total_without_milho, 2),
            ]
        )

    ws.auto_filter.ref = f"A1:K{ws.max_row}"
    for row_idx in range(2, ws.max_row + 1):
        for column_idx in (2, 3, 5, 7, 9, 11):
            ws.cell(row=row_idx, column=column_idx).number_format = "#,##0.00"
        for column_idx in (4, 6, 8, 10):
            ws.cell(row=row_idx, column=column_idx).number_format = "0.00\\%"
    _autosize_columns(ws)
    _append_footer(ws, footer_text)


def build_milhao_excel_export(
    *,
    filename,
    start_date,
    end_date,
    export_dt,
    entries,
    logo_set_path=None,
    logo_milhao_path=None,
):
    workbook = Workbook()
    footer_text = _format_footer(export_dt)

    _build_summary_sheet(
        workbook,
        filename=filename,
        start_date=start_date,
        end_date=end_date,
        export_dt=export_dt,
        logo_set_path=logo_set_path,
        logo_milhao_path=logo_milhao_path,
        footer_text=footer_text,
    )
    _build_totals_by_balance_sheet(workbook, entries=entries, footer_text=footer_text)
    _build_hourly_readings_sheet(workbook, entries=entries, footer_text=footer_text)
    _build_daily_totals_sheet(workbook, entries=entries, footer_text=footer_text)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
