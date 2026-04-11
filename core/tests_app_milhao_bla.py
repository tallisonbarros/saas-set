from datetime import date, datetime
from io import BytesIO

from django.contrib.auth.models import User
from django.db.models import Q
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from core.apps.app_milhao_bla.views import EXPORT_ACCESS_AUDIT_MODULE, MURAL_ACCESS_AUDIT_MODULE, MURAL_INTRO_SEEN_AUDIT_MODULE
from core.models import AdminAccessLog, App, AppMilhaoBlaMuralDia, AppMilhaoBlaMuralDiaLeitura, IngestRecord, PerfilUsuario, TipoPerfil


class AppMilhaoBlaIngestConfigTests(TestCase):
    def setUp(self):
        self.app = App.objects.create(
            slug="appmilhaobla",
            nome="App Milhao Bla",
            ativo=True,
            ingest_client_id="UBS3-UN1",
            ingest_agent_id="VMSCADA",
            ingest_source="balanca_acumulado_hora",
        )
        self.user = User.objects.create_user(username="bla_user", password="123456", email="bla@example.com")
        self.perfil = PerfilUsuario.objects.create(
            nome="Bla User",
            email="bla@example.com",
            usuario=self.user,
            ativo=True,
        )
        self.perfil.apps.add(self.app)
        self.other_user = User.objects.create_user(username="bla_peer", password="123456", email="peer@example.com")
        self.other_perfil = PerfilUsuario.objects.create(
            nome="Peer User",
            email="peer@example.com",
            usuario=self.other_user,
            ativo=True,
        )
        self.other_perfil.apps.add(self.app)

    def test_dashboard_uses_ingest_config_from_app(self):
        now_iso = timezone.now().isoformat()
        IngestRecord.objects.create(
            source_id="bla-1",
            client_id="UBS3-UN1",
            agent_id="VMSCADA",
            source="balanca_acumulado_hora",
            payload={"TagName": "LIMBL01", "Hora": now_iso, "ProducaoHora": "10"},
        )
        IngestRecord.objects.create(
            source_id="bla-2",
            client_id="clienteA",
            agent_id="agente01",
            source="balanca_acumulado_hora",
            payload={"TagName": "LIMBL01", "Hora": now_iso, "ProducaoHora": "99"},
        )

        self.client.force_login(self.user)
        response = self.client.get(reverse("app_milhao_bla_dashboard"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Cliente: UBS3-UN1")
        entries = response.context["entries"]
        self.assertEqual(len(entries), 1)
        self.assertEqual(entries[0]["value"], 10.0)

    def test_cards_data_endpoint_returns_live_payload(self):
        now_iso = timezone.now().isoformat()
        IngestRecord.objects.create(
            source_id="bla-3",
            client_id="UBS3-UN1",
            agent_id="VMSCADA",
            source="balanca_acumulado_hora",
            payload={"TagName": "LIMBL01", "Hora": now_iso, "ProducaoHora": "10"},
        )
        IngestRecord.objects.create(
            source_id="bla-4",
            client_id="UBS3-UN1",
            agent_id="VMSCADA",
            source="balanca_acumulado_hora",
            payload={"TagName": "SECBL01", "Hora": now_iso, "ProducaoHora": "5"},
        )

        self.client.force_login(self.user)
        response = self.client.get(reverse("app_milhao_bla_cards_data"))
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["ok"])
        self.assertEqual(payload["total_value_display"], "10")
        total_card = next(item for item in payload["totals_by_balance"] if item["balance"] == "TOTAL")
        self.assertEqual(total_card["total_display"], "5")
        self.assertGreaterEqual(len(payload["composition"]), 1)

    def test_dashboard_mural_exibe_publicas_e_privadas_do_autor_no_dia(self):
        public_note = AppMilhaoBlaMuralDia.objects.create(
            data_referencia=date(2026, 3, 5),
            texto="Nota publica do dia",
            visibilidade=AppMilhaoBlaMuralDia.Visibilidade.PUBLICA,
            autor=self.other_user,
        )
        own_private_note = AppMilhaoBlaMuralDia.objects.create(
            data_referencia=date(2026, 3, 5),
            texto="Nota privada do autor logado",
            visibilidade=AppMilhaoBlaMuralDia.Visibilidade.PRIVADA,
            autor=self.user,
        )
        AppMilhaoBlaMuralDia.objects.create(
            data_referencia=date(2026, 3, 5),
            texto="Nota privada de outro usuario",
            visibilidade=AppMilhaoBlaMuralDia.Visibilidade.PRIVADA,
            autor=self.other_user,
        )
        AppMilhaoBlaMuralDia.objects.filter(pk=public_note.pk).update(
            criado_em=timezone.make_aware(datetime(2026, 3, 5, 8, 0))
        )
        AppMilhaoBlaMuralDia.objects.filter(pk=own_private_note.pk).update(
            criado_em=timezone.make_aware(datetime(2026, 3, 5, 9, 0))
        )

        self.client.force_login(self.user)
        response = self.client.get(reverse("app_milhao_bla_dashboard"), {"date": "2026-03-05"})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["selected_date"].isoformat(), "2026-03-05")
        self.assertIn(date(2026, 3, 5), response.context["dates"])
        self.assertEqual(response.context["mural_notes_count"], 2)
        self.assertTrue(response.context["mural_has_unread"])
        notes_texts = [item["text"] for item in response.context["mural_notes"]]
        self.assertEqual(notes_texts, ["Nota publica do dia", "Nota privada do autor logado"])
        self.assertIn("Nota publica do dia", notes_texts)
        self.assertIn("Nota privada do autor logado", notes_texts)
        self.assertNotIn("Nota privada de outro usuario", notes_texts)

    def test_dashboard_mural_sem_nota_nova_apos_visualizacao_do_dia(self):
        note = AppMilhaoBlaMuralDia.objects.create(
            data_referencia=date(2026, 3, 5),
            texto="Nota ja visualizada",
            visibilidade=AppMilhaoBlaMuralDia.Visibilidade.PUBLICA,
            autor=self.other_user,
        )
        AppMilhaoBlaMuralDiaLeitura.objects.create(
            usuario=self.user,
            data_referencia=date(2026, 3, 5),
            visualizado_em=note.criado_em,
        )

        self.client.force_login(self.user)
        response = self.client.get(reverse("app_milhao_bla_dashboard"), {"date": "2026-03-05"})

        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.context["mural_has_unread"])

    def test_dashboard_sinaliza_intro_do_mural_para_usuario_que_ainda_nao_viu(self):
        self.client.force_login(self.user)

        first_response = self.client.get(reverse("app_milhao_bla_dashboard"))
        self.assertEqual(first_response.status_code, 200)
        self.assertTrue(first_response.context["mural_intro_should_open"])

        seen_response = self.client.post(
            reverse("app_milhao_bla_mural_day_intro_seen"),
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )
        self.assertEqual(seen_response.status_code, 200)
        self.assertTrue(
            AdminAccessLog.objects.filter(user=self.user, module=MURAL_INTRO_SEEN_AUDIT_MODULE).exists()
        )

        second_response = self.client.get(reverse("app_milhao_bla_dashboard"))
        self.assertEqual(second_response.status_code, 200)
        self.assertFalse(second_response.context["mural_intro_should_open"])

    def test_mural_create_endpoint_persiste_nota_e_retorna_html_atualizado(self):
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("app_milhao_bla_mural_day_create"),
            {
                "data_referencia": "2026-03-06",
                "texto": "Primeira nota do mural",
                "visibilidade": AppMilhaoBlaMuralDia.Visibilidade.PRIVADA,
            },
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["ok"])
        self.assertEqual(payload["notes_count"], 1)
        self.assertFalse(payload["has_unread"])
        note = AppMilhaoBlaMuralDia.objects.get(pk=payload["note_id"])
        self.assertEqual(note.data_referencia.isoformat(), "2026-03-06")
        self.assertEqual(note.texto, "Primeira nota do mural")
        self.assertEqual(note.visibilidade, AppMilhaoBlaMuralDia.Visibilidade.PRIVADA)
        self.assertEqual(note.autor, self.user)
        self.assertIn("Primeira nota do mural", payload["list_html"])
        self.assertIn("so eu", payload["list_html"])
        self.assertTrue(
            AppMilhaoBlaMuralDiaLeitura.objects.filter(
                usuario=self.user,
                data_referencia=date(2026, 3, 6),
            ).exists()
        )

    def test_mural_mark_viewed_endpoint_registra_visualizacao(self):
        AppMilhaoBlaMuralDia.objects.create(
            data_referencia=date(2026, 3, 8),
            texto="Nota nova para leitura",
            visibilidade=AppMilhaoBlaMuralDia.Visibilidade.PUBLICA,
            autor=self.other_user,
        )

        self.client.force_login(self.user)
        response = self.client.post(
            reverse("app_milhao_bla_mural_day_mark_viewed"),
            {"data_referencia": "2026-03-08"},
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["ok"])
        self.assertFalse(payload["has_unread"])
        self.assertTrue(
            AppMilhaoBlaMuralDiaLeitura.objects.filter(
                usuario=self.user,
                data_referencia=date(2026, 3, 8),
            ).exists()
        )

    def test_mural_live_endpoint_retorna_html_quando_existe_nota_nova(self):
        note = AppMilhaoBlaMuralDia.objects.create(
            data_referencia=date(2026, 3, 9),
            texto="Nova nota de outro usuario",
            visibilidade=AppMilhaoBlaMuralDia.Visibilidade.PUBLICA,
            autor=self.other_user,
        )

        self.client.force_login(self.user)
        response = self.client.get(
            reverse("app_milhao_bla_mural_day_live"),
            {"date": "2026-03-09", "latest_note_id": ""},
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["ok"])
        self.assertTrue(payload["has_changed"])
        self.assertTrue(payload["has_unread"])
        self.assertEqual(payload["latest_note_id"], note.id)
        self.assertIn("Nova nota de outro usuario", payload["list_html"])

    def test_mural_live_endpoint_nao_retorna_html_sem_mudanca(self):
        note = AppMilhaoBlaMuralDia.objects.create(
            data_referencia=date(2026, 3, 10),
            texto="Nota estavel",
            visibilidade=AppMilhaoBlaMuralDia.Visibilidade.PUBLICA,
            autor=self.other_user,
        )

        self.client.force_login(self.user)
        response = self.client.get(
            reverse("app_milhao_bla_mural_day_live"),
            {"date": "2026-03-10", "latest_note_id": str(note.id)},
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["ok"])
        self.assertFalse(payload["has_changed"])
        self.assertNotIn("list_html", payload)

    def test_autor_pode_excluir_nota_do_mural(self):
        note = AppMilhaoBlaMuralDia.objects.create(
            data_referencia=date(2026, 3, 7),
            texto="Nota que sera removida",
            visibilidade=AppMilhaoBlaMuralDia.Visibilidade.PUBLICA,
            autor=self.user,
        )

        self.client.force_login(self.user)
        response = self.client.post(
            reverse("app_milhao_bla_mural_day_delete", args=[note.id]),
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["ok"])
        self.assertEqual(payload["notes_count"], 0)
        self.assertFalse(AppMilhaoBlaMuralDia.objects.filter(pk=note.id).exists())
        self.assertIn("Nenhuma nota registrada", payload["list_html"])

    def test_outro_usuario_nao_pode_excluir_nota_do_mural(self):
        note = AppMilhaoBlaMuralDia.objects.create(
            data_referencia=date(2026, 3, 7),
            texto="Nota protegida",
            visibilidade=AppMilhaoBlaMuralDia.Visibilidade.PUBLICA,
            autor=self.user,
        )

        self.client.force_login(self.other_user)
        response = self.client.post(
            reverse("app_milhao_bla_mural_day_delete", args=[note.id]),
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 403)
        self.assertTrue(AppMilhaoBlaMuralDia.objects.filter(pk=note.id).exists())

    def test_usuario_dev_enxerga_notas_privadas_de_outros(self):
        tipo_dev = TipoPerfil.objects.filter(Q(nome__iexact="DEV") | Q(codigo__iexact="DEV")).first()
        if not tipo_dev:
            tipo_dev = TipoPerfil.objects.create(nome="DEV", codigo="DEV")
        self.perfil.tipos.add(tipo_dev)
        AppMilhaoBlaMuralDia.objects.create(
            data_referencia=date(2026, 3, 11),
            texto="Privada de outro usuario",
            visibilidade=AppMilhaoBlaMuralDia.Visibilidade.PRIVADA,
            autor=self.other_user,
        )

        self.client.force_login(self.user)
        response = self.client.get(reverse("app_milhao_bla_dashboard"), {"date": "2026-03-11"})

        self.assertEqual(response.status_code, 200)
        notes_texts = [item["text"] for item in response.context["mural_notes"]]
        self.assertIn("Privada de outro usuario", notes_texts)

    def test_mural_access_endpoint_registra_auditoria_a_cada_abertura(self):
        self.client.force_login(self.user)

        first_response = self.client.post(
            reverse("app_milhao_bla_mural_day_access"),
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )
        second_response = self.client.post(
            reverse("app_milhao_bla_mural_day_access"),
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(first_response.status_code, 200)
        self.assertEqual(second_response.status_code, 200)
        self.assertEqual(
            AdminAccessLog.objects.filter(user=self.user, module=MURAL_ACCESS_AUDIT_MODULE).count(),
            2,
        )

    def test_export_access_endpoint_registra_auditoria(self):
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("app_milhao_bla_export_excel_access"),
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(
            AdminAccessLog.objects.filter(user=self.user, module=EXPORT_ACCESS_AUDIT_MODULE).exists()
        )

    def test_export_excel_endpoint_returns_workbook_with_expected_structure(self):
        IngestRecord.objects.create(
            source_id="bla-export-1",
            client_id="UBS3-UN1",
            agent_id="VMSCADA",
            source="balanca_acumulado_hora",
            payload={"TagName": "LIMBL01", "Hora": "2026-03-01T08:00:00", "ProducaoHora": "10"},
        )
        IngestRecord.objects.create(
            source_id="bla-export-2",
            client_id="UBS3-UN1",
            agent_id="VMSCADA",
            source="balanca_acumulado_hora",
            payload={"TagName": "SECBL01", "Hora": "2026-03-01T09:00:00", "ProducaoHora": "5"},
        )
        IngestRecord.objects.create(
            source_id="bla-export-3",
            client_id="UBS3-UN1",
            agent_id="VMSCADA",
            source="balanca_acumulado_hora",
            payload={"TagName": "CLABL01", "Hora": "2026-03-02T10:00:00", "ProducaoHora": "7"},
        )

        self.client.force_login(self.user)
        response = self.client.post(
            reverse("app_milhao_bla_export_excel"),
            {"start_date": "2026-03-01", "end_date": "2026-03-02"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            response["Content-Type"],
        )
        self.assertIn(
            "milhao_bla_20260301_a_20260302.xlsx",
            response["Content-Disposition"],
        )

        workbook = load_workbook(filename=BytesIO(response.content))
        self.assertListEqual(
            workbook.sheetnames,
            ["Resumo", "Totais por balanca", "Leituras por hora", "Totais por dia"],
        )
        resumo = workbook["Resumo"]
        self.assertEqual(resumo["A3"].value, "Arquivo: milhao_bla_20260301_a_20260302.xlsx")
        merged_ranges = {str(rng) for rng in resumo.merged_cells.ranges}
        self.assertIn("A1:B1", merged_ranges)
        self.assertIn("A8:B8", merged_ranges)
        self.assertNotIn("Cliente", {str(resumo[f"A{line}"].value) for line in range(1, 9)})

        totals_balance = workbook["Totais por balanca"]
        self.assertEqual(totals_balance["A1"].value, "Balanca")
        self.assertEqual(totals_balance["B1"].value, "Total_kg")
        self.assertEqual(totals_balance["C1"].value, "% do total (sem LIMBL01)")
        self.assertIsNone(totals_balance["D1"].value)

        hourly = workbook["Leituras por hora"]
        self.assertEqual(hourly["A1"].value, "Data")
        self.assertEqual(hourly["B1"].value, "Hora")
        self.assertEqual(hourly["C1"].value, "Balanca")
        self.assertEqual(hourly["D1"].value, "Valor_kg")
        self.assertIsNone(hourly["E1"].value)

        daily = workbook["Totais por dia"]
        self.assertEqual(daily["A1"].value, "Data")
        self.assertEqual(daily["B1"].value, "LIMBL01_kg")
        self.assertEqual(daily["C1"].value, "CLABL01_kg")
        self.assertEqual(daily["D1"].value, "CLABL01_%")
        self.assertEqual(daily["E1"].value, "CLABL02_kg")
        self.assertEqual(daily["F1"].value, "CLABL02_%")
        self.assertEqual(daily["G1"].value, "SECBL01_kg")
        self.assertEqual(daily["H1"].value, "SECBL01_%")
        self.assertEqual(daily["I1"].value, "SECBL02_kg")
        self.assertEqual(daily["J1"].value, "SECBL02_%")
        self.assertEqual(daily["K1"].value, "TOTAL_sem_milho_kg")
        self.assertIsNone(daily["L1"].value)

        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            footer_text = str(ws.cell(row=ws.max_row, column=1).value or "").lower()
            self.assertIn("setbrasil.club", footer_text)

    def test_export_excel_rejects_interval_over_limit(self):
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("app_milhao_bla_export_excel"),
            {"start_date": "2026-01-01", "end_date": "2026-05-01"},
        )
        self.assertEqual(response.status_code, 400)
        payload = response.json()
        self.assertFalse(payload["ok"])
        self.assertIn("Intervalo maximo", payload["error"])

    def test_export_excel_requires_app_permission(self):
        outsider = User.objects.create_user(username="outsider", password="123456", email="out@example.com")
        PerfilUsuario.objects.create(
            nome="Out",
            email="out@example.com",
            usuario=outsider,
            ativo=True,
        )
        self.client.force_login(outsider)
        response = self.client.post(
            reverse("app_milhao_bla_export_excel"),
            {"start_date": "2026-03-01", "end_date": "2026-03-02"},
        )
        self.assertEqual(response.status_code, 403)

    def test_dev_can_access_dashboard_even_when_app_is_inactive(self):
        tipo_dev = TipoPerfil.objects.filter(Q(nome__iexact="DEV") | Q(codigo__iexact="DEV")).first()
        if not tipo_dev:
            tipo_dev = TipoPerfil.objects.create(nome="DEV", codigo="DEV")
        self.perfil.tipos.add(tipo_dev)
        self.app.ativo = False
        self.app.save(update_fields=["ativo"])

        self.client.force_login(self.user)
        response = self.client.get(reverse("app_milhao_bla_dashboard"))
        self.assertEqual(response.status_code, 200)
