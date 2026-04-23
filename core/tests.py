import json
from io import BytesIO
from pathlib import Path
from unittest.mock import patch
from datetime import datetime, timedelta
from decimal import Decimal

from django.contrib.auth.models import User
from django.core.management import call_command
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import Client, SimpleTestCase, TestCase
from django.utils import timezone
from openpyxl import Workbook

from core.apps.app_rotas.views import _global_point_visual_flags, _route_point_visual_flags
from core.access_control import has_tipo_code, normalize_access_code
from core.models import (
    AcessoProdutoUsuario,
    AdminAccessLog,
    App,
    AssinaturaUsuario,
    ConsumoImportacaoDiaria,
    ConfiguracaoPagamento,
    GrupoRackIO,
    IPImportJob,
    IPImportSettings,
    IOImportAICache,
    IOImportJob,
    IOImportSettings,
    ListaIP,
    ListaIPItem,
    PerfilUsuario,
    PlanoComercial,
    IngestRecord,
    CanalRackIO,
    Caderno,
    Compra,
    LocalRackIO,
    ModuloAcesso,
    ModuloIO,
    ModuloRackIO,
    PlantaIO,
    Proposta,
    Radar,
    RadarAtividade,
    RadarAtividadeDiaExecucao,
    RadarColaborador,
    RadarClassificacao,
    RadarContrato,
    RadarID,
    RackIO,
    RadarTrabalho,
    RadarTrabalhoColaborador,
    RadarTrabalhoObservacao,
    RackSlotIO,
    SystemConfiguration,
    TipoCanalIO,
    TipoPerfil,
)
from core.services.billing import (
    DOCUMENTATION_PRODUCT_CODE,
    activate_starter_plan,
    activate_trial,
    ensure_billing_catalog,
    register_successful_import_usage,
    resolve_import_quota,
)
from core.services.io_import import (
    _call_openai_responses,
    ParsedSpreadsheet,
    apply_import_job,
    build_file_sha256,
    build_import_proposal,
    normalize_rows,
    parse_workbook,
    reprocess_import_job,
    run_ai_analysis,
    serialize_module_catalog,
)
from core.services.ip_import import (
    _call_openai_responses as call_ip_openai_responses,
    apply_import_job as apply_ip_import_job,
    reprocess_import_job as reprocess_ip_import_job,
)
from core.views import (
    DOCUMENTACAO_TECNICA_LANDING_AUDIT_MODULE,
    _build_proposta_pdf_context,
    _build_radar_relatorio_pdf_context,
    _reprocess_ip_import_job,
    _sanitize_proposta_descricao,
    _user_role,
)


class AccessControlFoundationTests(TestCase):
    def test_system_types_are_seeded_with_stable_codes(self):
        self.assertTrue(TipoPerfil.objects.filter(codigo="DEV", sistema=True, ativo=True).exists())
        self.assertTrue(TipoPerfil.objects.filter(codigo="MASTER", sistema=True, ativo=True).exists())

    def test_access_modules_are_seeded_with_expected_base_data(self):
        modulo = ModuloAcesso.objects.get(codigo="APP_MILHAO_BLA")
        self.assertEqual(modulo.tipo, ModuloAcesso.Tipo.APP)
        self.assertTrue(modulo.ativo)

    def test_tipo_perfil_generates_code_from_name_when_missing(self):
        tipo = TipoPerfil.objects.create(nome="Radar Operacional")
        self.assertEqual(tipo.codigo, "RADAR_OPERACIONAL")

    def test_modulo_acesso_normalizes_code_on_save(self):
        modulo = ModuloAcesso.objects.create(codigo="", nome="Modulo Piloto")
        self.assertEqual(modulo.codigo, "MODULO_PILOTO")

    def test_tipo_code_helper_uses_stable_codes(self):
        user = User.objects.create_user(username="acl@set.local", email="acl@set.local", password="123456")
        perfil = PerfilUsuario.objects.create(nome="ACL", email="acl@set.local", usuario=user)
        tipo = TipoPerfil.objects.create(nome="Teste ACL", codigo="ACL_TESTE")
        perfil.tipos.add(tipo)
        self.assertEqual(normalize_access_code("acl teste"), "ACL_TESTE")
        self.assertTrue(has_tipo_code(user, "acl teste"))


class AccessControlAdminAndVisibilityTests(TestCase):
    def setUp(self):
        self.client_http = Client()
        self.tipo_dev = TipoPerfil.objects.get(codigo="DEV")
        self.tipo_financeiro = TipoPerfil.objects.get(codigo="FINANCEIRO")
        self.dev_user = User.objects.create_user(username="shadow-dev@set.local", email="shadow-dev@set.local", password="123456")
        self.dev_perfil = PerfilUsuario.objects.create(
            nome="Shadow Dev",
            email="shadow-dev@set.local",
            usuario=self.dev_user,
        )
        self.dev_perfil.tipos.add(self.tipo_dev)

        self.user = User.objects.create_user(username="cliente-shadow@set.local", email="cliente-shadow@set.local", password="123456")
        self.perfil = PerfilUsuario.objects.create(
            nome="Cliente Shadow",
            email="cliente-shadow@set.local",
            usuario=self.user,
        )

    def test_dev_can_open_module_access_management(self):
        self.client_http.force_login(self.dev_user)
        response = self.client_http.get("/modulos-acesso/")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Modulos de acesso")
        self.assertNotContains(response, "Novo modulo")
        self.assertNotContains(response, "App vinculado")
        self.assertContains(response, "APP ID em `perfil.apps`", html=False)
        self.assertContains(response, "Os IDs atuais continuam apenas como escopo e compartilhamento interno entre usuarios.", html=False)
        self.assertNotContains(response, "Modos de autorizacao")

    def test_dev_can_update_module_access_management(self):
        modulo = ModuloAcesso.objects.get(codigo="FINANCEIRO")
        self.client_http.force_login(self.dev_user)
        response = self.client_http.post(
            "/modulos-acesso/",
            {
                "action": "update_module",
                "module_id": modulo.id,
                "tipos": [self.tipo_financeiro.id],
                "ativo": "on",
            },
        )
        self.assertEqual(response.status_code, 302)
        modulo.refresh_from_db()
        self.assertSetEqual(set(modulo.tipos.values_list("codigo", flat=True)), {"FINANCEIRO"})

    def test_internal_module_route_is_blocked_without_matching_type(self):
        modulo = ModuloAcesso.objects.get(codigo="FINANCEIRO")
        modulo.tipos.clear()

        self.client_http.force_login(self.user)
        response = self.client_http.get("/financeiro/")
        self.assertEqual(response.status_code, 403)

    def test_internal_module_card_stays_hidden_without_matching_type(self):
        modulo = ModuloAcesso.objects.get(codigo="FINANCEIRO")
        modulo.tipos.set([self.tipo_financeiro])
        self.client_http.force_login(self.user)
        response = self.client_http.get("/painel/")
        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, 'href="/financeiro/"')

    def test_internal_module_card_and_route_are_enabled_by_tipo(self):
        modulo = ModuloAcesso.objects.get(codigo="FINANCEIRO")
        modulo.tipos.set([self.tipo_financeiro])
        self.perfil.tipos.add(self.tipo_financeiro)
        self.client_http.force_login(self.user)
        painel_response = self.client_http.get("/painel/")
        self.assertEqual(painel_response.status_code, 200)
        self.assertContains(painel_response, 'href="/financeiro/"')
        route_response = self.client_http.get("/financeiro/")
        self.assertEqual(route_response.status_code, 200)

    def test_dedicated_apps_stay_reference_only_in_module_access_screen(self):
        self.client_http.force_login(self.dev_user)
        response = self.client_http.get("/modulos-acesso/")
        self.assertContains(response, "Apps dedicados")
        self.assertContains(response, "o acesso real continua sendo decidido exclusivamente pelo APP ID do usuario", html=False)

    def test_dedicated_apps_remain_visible_by_app_id_only(self):
        app = App.objects.create(nome="BLA", slug="bla", ativo=True)
        self.perfil.apps.add(app)
        self.client_http.force_login(self.user)
        response = self.client_http.get("/painel/")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "BLA")
        self.assertNotContains(response, 'href="/financeiro/"')

    def test_dev_keeps_inactive_dedicated_app_visible_in_painel(self):
        app = App.objects.create(nome="BLA Oculto", slug="bla_oculto", ativo=False)
        self.dev_perfil.apps.add(app)
        self.client_http.force_login(self.dev_user)
        response = self.client_http.get("/painel/")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "BLA Oculto")

    def test_dev_can_open_inactive_dedicated_app_home_when_assigned(self):
        app = App.objects.create(nome="BLA Oculto", slug="bla_oculto", ativo=False)
        self.dev_perfil.apps.add(app)
        self.client_http.force_login(self.dev_user)
        response = self.client_http.get("/apps/bla_oculto/")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "BLA Oculto")


class BillingAndPlanFlowTests(TestCase):
    def setUp(self):
        ensure_billing_catalog()
        self.client_http = Client()
        self.user = User.objects.create_user(
            username="billing-user@set.local",
            email="billing-user@set.local",
            password="123456",
        )
        self.perfil = PerfilUsuario.objects.create(
            nome="Billing User",
            email="billing-user@set.local",
            usuario=self.user,
        )
        self.client_http.force_login(self.user)

    def _create_rack(self, name):
        rack = RackIO.objects.create(
            cliente=self.perfil,
            nome=name,
            descricao="Teste",
            slots_total=4,
        )
        RackSlotIO.objects.bulk_create([RackSlotIO(rack=rack, posicao=index) for index in range(1, 5)])
        return rack

    def test_trial_activation_enables_module_access(self):
        response = self.client_http.post("/produtos/documentacao-tecnica/ativar/", {"next": "/ios/"})
        self.assertEqual(response.status_code, 302)
        access = AcessoProdutoUsuario.objects.get(usuario=self.user, produto__codigo=DOCUMENTATION_PRODUCT_CODE)
        self.assertEqual(access.status, AcessoProdutoUsuario.Status.TRIAL_ATIVO)
        response = self.client_http.get("/ios/")
        self.assertEqual(response.status_code, 200)

    def test_expired_trial_redirects_to_plan_selection(self):
        product = ensure_billing_catalog()
        AcessoProdutoUsuario.objects.create(
            usuario=self.user,
            produto=product,
            origem=AcessoProdutoUsuario.Origem.TRIAL,
            status=AcessoProdutoUsuario.Status.EXPIRADO,
        )
        response = self.client_http.get("/ios/")
        self.assertEqual(response.status_code, 302)
        self.assertIn("/produtos/planos/", response["Location"])

    def test_starter_plan_activation_is_available_with_three_or_fewer_racks(self):
        self._create_rack("Rack 01")
        self._create_rack("Rack 02")
        subscription, error = activate_starter_plan(self.user, DOCUMENTATION_PRODUCT_CODE)
        self.assertIsNotNone(subscription)
        self.assertEqual(error, "")
        self.assertEqual(subscription.plano.codigo, PlanoComercial.Codigo.STARTER)
        self.assertEqual(subscription.status, AssinaturaUsuario.Status.ACTIVE)

    def test_starter_plan_is_blocked_with_more_than_three_racks(self):
        for index in range(1, 5):
            self._create_rack(f"Rack {index:02d}")
        subscription, error = activate_starter_plan(self.user, DOCUMENTATION_PRODUCT_CODE)
        self.assertIsNone(subscription)
        self.assertIn("3 racks", error)
        response = self.client_http.get("/produtos/planos/?state=starter_bloqueado")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Plano Iniciante")

    def test_starter_over_limit_redirects_module_to_plan_page(self):
        for index in range(1, 5):
            self._create_rack(f"Rack {index:02d}")
        starter = PlanoComercial.objects.get(produto__codigo=DOCUMENTATION_PRODUCT_CODE, codigo=PlanoComercial.Codigo.STARTER)
        product = ensure_billing_catalog()
        AssinaturaUsuario.objects.create(
            usuario=self.user,
            produto=product,
            plano=starter,
            provider=AssinaturaUsuario.Provider.INTERNAL,
            status=AssinaturaUsuario.Status.ACTIVE,
            billing_interval=AssinaturaUsuario.BillingInterval.MONTHLY,
            auto_renew=True,
        )
        AcessoProdutoUsuario.objects.create(
            usuario=self.user,
            produto=product,
            origem=AcessoProdutoUsuario.Origem.INTERNO,
            status=AcessoProdutoUsuario.Status.ATIVO,
        )
        response = self.client_http.get("/ios/")
        self.assertEqual(response.status_code, 302)
        self.assertIn("/produtos/planos/", response["Location"])

    def test_trial_daily_import_quota_is_tracked_separately_for_io_and_ip(self):
        activate_trial(self.user, DOCUMENTATION_PRODUCT_CODE)
        io_quota = resolve_import_quota(self.user, "IO")
        ip_quota = resolve_import_quota(self.user, "IP")

        self.assertTrue(io_quota["enforced"])
        self.assertEqual(io_quota["limit"], 3)
        self.assertEqual(io_quota["used"], 0)
        self.assertEqual(ip_quota["limit"], 3)
        self.assertEqual(ip_quota["used"], 0)

        for _ in range(3):
            register_successful_import_usage(self.user, "IO")
        refreshed_io_quota = resolve_import_quota(self.user, "IO")
        refreshed_ip_quota = resolve_import_quota(self.user, "IP")

        self.assertEqual(refreshed_io_quota["used"], 3)
        self.assertEqual(refreshed_io_quota["remaining"], 0)
        self.assertEqual(refreshed_ip_quota["used"], 0)
        self.assertEqual(
            ConsumoImportacaoDiaria.objects.get(
                usuario=self.user,
                produto__codigo=DOCUMENTATION_PRODUCT_CODE,
                modulo=ConsumoImportacaoDiaria.Modulo.IO,
            ).importacoes_bem_sucedidas,
            3,
        )

        with self.assertRaisesMessage(ValueError, "3 importacoes concluidas de planilhas de IO por dia"):
            register_successful_import_usage(self.user, "IO")

    def test_billing_admin_page_is_available_for_dev(self):
        dev_type = TipoPerfil.objects.get(codigo="DEV")
        admin_user = User.objects.create_user(username="billing-dev@set.local", email="billing-dev@set.local", password="123456")
        admin_profile = PerfilUsuario.objects.create(nome="Billing Dev", email="billing-dev@set.local", usuario=admin_user)
        admin_profile.tipos.add(dev_type)
        self.client_http.force_login(admin_user)
        response = self.client_http.get("/pagamentos-planos/")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Pagamentos e Planos")
        self.assertContains(response, "/pagamentos/checkout/sucesso/")
        self.assertContains(response, "/pagamentos/checkout/falha/")
        self.assertContains(response, "/pagamentos/checkout/pendente/")

    def test_checkout_status_pages_are_available(self):
        for url, expected in (
            ("/pagamentos/checkout/sucesso/", "Pagamento confirmado"),
            ("/pagamentos/checkout/falha/", "Pagamento nao concluido"),
            ("/pagamentos/checkout/pendente/", "Pagamento em analise"),
        ):
            response = self.client_http.get(url)
            self.assertEqual(response.status_code, 200)
            self.assertContains(response, expected)

    def test_plans_page_is_separate_from_landing(self):
        response = self.client_http.get("/produtos/planos/?next=/ios/")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Ver página do produto")
        self.assertContains(response, "Ative um plano da plataforma para continuar usando os produtos liberados na sua conta.")
        self.assertContains(response, "Plano Iniciante")


class DevAdminPrivilegesTests(TestCase):
    def setUp(self):
        self.client_http = Client()
        self.tipo_dev, _ = TipoPerfil.objects.get_or_create(nome="Dev")
        self.tipo_cliente, _ = TipoPerfil.objects.get_or_create(nome="Cliente")

        self.dev_user = User.objects.create_user(username="dev@set.local", email="dev@set.local", password="123456")
        self.dev_perfil = PerfilUsuario.objects.create(
            nome="Dev User",
            email="dev@set.local",
            usuario=self.dev_user,
        )
        self.dev_perfil.tipos.add(self.tipo_dev)

        self.user = User.objects.create_user(username="user@set.local", email="user@set.local", password="123456")
        self.perfil = PerfilUsuario.objects.create(
            nome="Regular User",
            email="user@set.local",
            usuario=self.user,
        )
        self.perfil.tipos.add(self.tipo_cliente)

    def test_dev_role_is_admin(self):
        self.assertEqual(_user_role(self.dev_user), "ADMIN")

    def test_dev_can_access_admin_logs(self):
        self.client_http.force_login(self.dev_user)
        response = self.client_http.get("/admin-logs/")
        self.assertEqual(response.status_code, 200)

    def test_documentacao_landing_records_anonymous_access(self):
        response = self.client_http.get("/produtos/documentacao-tecnica/")
        self.assertEqual(response.status_code, 200)
        self.assertTrue(
            AdminAccessLog.objects.filter(
                user__isnull=True,
                module=DOCUMENTACAO_TECNICA_LANDING_AUDIT_MODULE,
            ).exists()
        )

    def test_admin_logs_exposes_documentacao_landing_metrics(self):
        AdminAccessLog.objects.create(user=None, module=DOCUMENTACAO_TECNICA_LANDING_AUDIT_MODULE)
        AdminAccessLog.objects.create(user=self.user, module=DOCUMENTACAO_TECNICA_LANDING_AUDIT_MODULE)

        self.client_http.force_login(self.dev_user)
        response = self.client_http.get("/admin-logs/")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Landing de documentacao tecnica")
        self.assertEqual(response.context["landing_metrics"]["total"], 2)
        self.assertEqual(response.context["landing_metrics"]["anonymous"], 1)
        self.assertEqual(response.context["landing_metrics"]["authenticated"], 1)

    def test_dev_is_promoted_to_staff_by_middleware(self):
        self.assertFalse(User.objects.get(pk=self.dev_user.pk).is_staff)
        self.client_http.force_login(self.dev_user)
        response = self.client_http.get("/painel/")
        self.assertEqual(response.status_code, 200)
        self.dev_user.refresh_from_db()
        self.assertTrue(self.dev_user.is_staff)

    def test_non_dev_cannot_access_admin_logs(self):
        self.client_http.force_login(self.user)
        response = self.client_http.get("/admin-logs/")
        self.assertEqual(response.status_code, 403)


class ForbiddenPagePresentationTests(TestCase):
    def setUp(self):
        self.client_http = Client()
        self.user = User.objects.create_user(
            username="forbidden-user@set.local",
            email="forbidden-user@set.local",
            password="123456",
        )
        self.perfil = PerfilUsuario.objects.create(
            nome="Forbidden User",
            email="forbidden-user@set.local",
            usuario=self.user,
        )

    def test_html_forbidden_uses_discreet_custom_page(self):
        self.client_http.force_login(self.user)
        response = self.client_http.get("/admin-logs/")
        self.assertEqual(response.status_code, 403)
        self.assertContains(response, "contate a set", status_code=403)
        self.assertContains(response, "Acesso restrito", status_code=403)

    def test_html_not_found_uses_discreet_custom_page(self):
        App.objects.create(
            slug="appmilhaobla",
            nome="App Milhao Bla",
            ativo=False,
        )
        self.perfil.apps.add(App.objects.get(slug="appmilhaobla"))
        self.client_http.force_login(self.user)
        response = self.client_http.get("/apps/appmilhaobla/")
        self.assertEqual(response.status_code, 404)
        self.assertContains(response, "contate a set", status_code=404)
        self.assertContains(response, "Pagina indisponivel", status_code=404)

    def test_ajax_json_forbidden_keeps_json_payload(self):
        App.objects.create(
            slug="appmilhaobla",
            nome="App Milhao Bla",
            ativo=True,
        )
        self.client_http.force_login(self.user)
        response = self.client_http.get(
            "/apps/appmilhaobla/cards-data/",
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )
        self.assertEqual(response.status_code, 403)
        self.assertIn("application/json", response["Content-Type"])
        self.assertEqual(response.json()["error"], "forbidden")


class MaintenanceModeTests(TestCase):
    def setUp(self):
        self.client_http = Client()
        self.tipo_dev = TipoPerfil.objects.get(codigo="DEV")
        self.tipo_cliente = TipoPerfil.objects.get(codigo="CLIENTE")

        self.dev_user = User.objects.create_user(username="maintenance-dev@set.local", email="maintenance-dev@set.local", password="123456")
        self.dev_perfil = PerfilUsuario.objects.create(
            nome="Maintenance Dev",
            email="maintenance-dev@set.local",
            usuario=self.dev_user,
        )
        self.dev_perfil.tipos.add(self.tipo_dev)

        self.user = User.objects.create_user(username="maintenance-user@set.local", email="maintenance-user@set.local", password="123456")
        self.perfil = PerfilUsuario.objects.create(
            nome="Maintenance User",
            email="maintenance-user@set.local",
            usuario=self.user,
        )
        self.perfil.tipos.add(self.tipo_cliente)

        self.config = SystemConfiguration.load()
        self.config.maintenance_mode_enabled = True
        self.config.maintenance_message = "Sistema temporariamente indisponivel."
        self.config.save()

    def test_non_dev_user_is_redirected_to_maintenance(self):
        self.client_http.force_login(self.user)
        response = self.client_http.get("/painel/")
        self.assertRedirects(response, "/manutencao/", fetch_redirect_response=False)

    def test_dev_user_bypasses_maintenance_mode(self):
        self.client_http.force_login(self.dev_user)
        response = self.client_http.get("/painel/")
        self.assertEqual(response.status_code, 200)

    def test_maintenance_page_renders_configured_message(self):
        self.client_http.force_login(self.user)
        response = self.client_http.get("/manutencao/")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Sistema temporariamente indisponivel.")

    def test_logout_remains_accessible_during_maintenance(self):
        self.client_http.force_login(self.user)
        response = self.client_http.post("/logout/")
        self.assertEqual(response.status_code, 302)


class RouteTimelineStateTests(SimpleTestCase):
    def _timeline_points(self, start, count, step_minutes=5):
        points = []
        for idx in range(count):
            ts = start + timedelta(minutes=step_minutes * idx)
            points.append({"idx": idx, "timestamp": ts, "iso": ts.isoformat(), "label": ts.isoformat()})
        return points

    def test_route_visual_flags_follow_same_rule_as_status(self):
        tz = timezone.get_current_timezone()
        start = timezone.make_aware(datetime(2026, 1, 10, 8, 0, 0), tz)
        timeline = self._timeline_points(start, 4)
        events = [
            # Linha ainda nao deve ficar verde sem LIGADA.
            {"timestamp": timeline[1]["timestamp"], "atributo": "LIGAR", "valor": 1},
            # Linha ligada: deve ficar verde.
            {"timestamp": timeline[1]["timestamp"], "atributo": "LIGADA", "valor": 1},
            # Linha desligando: deve sair do verde mesmo que LIGADA ainda esteja 1.
            {"timestamp": timeline[2]["timestamp"], "atributo": "DESLIGAR", "valor": 1},
        ]
        flags = _route_point_visual_flags(
            day_events=events,
            timeline=timeline,
            available_until=timeline[-1]["timestamp"],
            baseline_attrs=None,
        )
        self.assertEqual(flags, [False, True, False, False])

    def test_global_visual_flags_use_same_play_on_semantics(self):
        tz = timezone.get_current_timezone()
        start = timezone.make_aware(datetime(2026, 1, 10, 9, 0, 0), tz)
        timeline = self._timeline_points(start, 3)
        seed_states = {
            "ENS01": {
                "attrs": {"LIGAR": 1, "DESLIGAR": 0, "LIGADA": 1, "ORIGEM": None, "DESTINO": None},
            }
        }
        events = [
            # Mesmo com LIGADA ainda 1, DESLIGAR deve apagar o verde global.
            {"prefixo": "ENS01", "timestamp": timeline[1]["timestamp"], "atributo": "DESLIGAR", "valor": 1},
        ]
        flags = _global_point_visual_flags(
            day_events=events,
            timeline=timeline,
            available_until=timeline[-1]["timestamp"],
            seed_states=seed_states,
        )
        self.assertEqual(flags, [True, False, False])


class IngestCleanupByDateTests(TestCase):
    def setUp(self):
        self.client_http = Client()
        self.tipo_dev, _ = TipoPerfil.objects.get_or_create(nome="Dev")
        self.dev_user = User.objects.create_user(username="ingest-dev@set.local", email="ingest-dev@set.local", password="123456")
        self.dev_perfil = PerfilUsuario.objects.create(
            nome="Ingest Dev",
            email="ingest-dev@set.local",
            usuario=self.dev_user,
        )
        self.dev_perfil.tipos.add(self.tipo_dev)

    def _create_ingest_record(self, source_id, created_at, client_id="CLIENTE-A", agent_id="AGENTE-A", source="SOURCE-A"):
        record = IngestRecord.objects.create(
            source_id=source_id,
            client_id=client_id,
            agent_id=agent_id,
            source=source,
            payload={"source_id": source_id},
        )
        IngestRecord.objects.filter(pk=record.pk).update(created_at=created_at, updated_at=created_at)
        record.refresh_from_db()
        return record

    def test_admin_can_delete_only_records_within_selected_created_at_range(self):
        tz = timezone.get_current_timezone()
        keep_before = self._create_ingest_record(
            "before-range",
            timezone.make_aware(datetime(2026, 3, 10, 9, 0, 0), tz),
        )
        remove_first = self._create_ingest_record(
            "inside-range-1",
            timezone.make_aware(datetime(2026, 3, 11, 10, 0, 0), tz),
        )
        remove_second = self._create_ingest_record(
            "inside-range-2",
            timezone.make_aware(datetime(2026, 3, 12, 11, 0, 0), tz),
        )
        keep_other_source = self._create_ingest_record(
            "other-source",
            timezone.make_aware(datetime(2026, 3, 11, 12, 0, 0), tz),
            source="SOURCE-B",
        )

        self.client_http.force_login(self.dev_user)
        response = self.client_http.post(
            "/ingest-gerenciar/limpar/",
            {
                "action": "delete_filtered_ingest",
                "client_id": "CLIENTE-A",
                "agent_id": "AGENTE-A",
                "source": "SOURCE-A",
                "data_inicial": "2026-03-11",
                "data_final": "2026-03-12",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "2 registro(s) removido(s)")
        self.assertTrue(IngestRecord.objects.filter(pk=keep_before.pk).exists())
        self.assertFalse(IngestRecord.objects.filter(pk=remove_first.pk).exists())
        self.assertFalse(IngestRecord.objects.filter(pk=remove_second.pk).exists())
        self.assertTrue(IngestRecord.objects.filter(pk=keep_other_source.pk).exists())

    def test_preview_can_show_only_records_that_match_selected_filters(self):
        tz = timezone.get_current_timezone()
        matching = self._create_ingest_record(
            "preview-match",
            timezone.make_aware(datetime(2026, 3, 11, 9, 0, 0), tz),
        )
        self._create_ingest_record(
            "preview-outside-date",
            timezone.make_aware(datetime(2026, 3, 15, 9, 0, 0), tz),
        )
        self._create_ingest_record(
            "preview-other-source",
            timezone.make_aware(datetime(2026, 3, 11, 11, 0, 0), tz),
            source="SOURCE-B",
        )

        self.client_http.force_login(self.dev_user)
        response = self.client_http.get(
            "/ingest-gerenciar/limpar/",
            {
                "client_id": "CLIENTE-A",
                "agent_id": "AGENTE-A",
                "source": "SOURCE-A",
                "data_inicial": "2026-03-11",
                "data_final": "2026-03-12",
                "preview": "1",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Pre-visualizacao da limpeza")
        self.assertContains(response, matching.source_id)
        self.assertNotContains(response, "preview-outside-date")
        self.assertNotContains(response, "preview-other-source")


class PropostaTrabalhoVinculoTests(TestCase):
    def setUp(self):
        self.client_http = Client()
        self.vendedor = User.objects.create_user(username="vend", email="vend@set.local", password="123456")
        self.destinatario_user = User.objects.create_user(
            username="cliente",
            email="cliente@set.local",
            password="123456",
        )
        self.destinatario = PerfilUsuario.objects.create(
            nome="Cliente",
            email="cliente@set.local",
            usuario=self.destinatario_user,
        )
        self.radar = Radar.objects.create(cliente=self.destinatario, nome="Radar Norte")
        self.classificacao = RadarClassificacao.objects.create(nome="Critica")
        self.contrato = RadarContrato.objects.create(nome="Contrato A")
        self.trabalho = RadarTrabalho.objects.create(
            radar=self.radar,
            classificacao=self.classificacao,
            contrato=self.contrato,
            nome="Trabalho TEC",
            descricao="Descricao tecnica inicial",
            setor="Manutencao",
            solicitante="Engenharia",
            responsavel="Paulo",
            criado_por=self.vendedor,
        )
        RadarTrabalhoColaborador.objects.create(trabalho=self.trabalho, nome="Ana")
        RadarTrabalhoColaborador.objects.create(trabalho=self.trabalho, nome="Bruno")
        RadarAtividade.objects.create(trabalho=self.trabalho, nome="Inspecao", descricao="Linha 1")
        self.proposta = Proposta.objects.create(
            cliente=self.destinatario,
            criada_por=self.vendedor,
            nome="Proposta 1",
            descricao="Escopo comercial.",
            trabalho=self.trabalho,
        )

    def test_proposta_vinculada_reflete_alteracoes_do_trabalho(self):
        self.trabalho.descricao = "Descricao tecnica atualizada"
        self.trabalho.save(update_fields=["descricao"])
        RadarAtividade.objects.create(trabalho=self.trabalho, nome="Comissionamento", descricao="Linha 2")
        proposta = Proposta.objects.select_related("trabalho", "trabalho__radar").prefetch_related("trabalho__atividades").get(
            pk=self.proposta.pk
        )
        context = _build_proposta_pdf_context(proposta, "Pendente", include_origem=True, trabalho=proposta.trabalho)
        labels = {row["label"]: row["value"] for row in context["origem_rows"]}
        self.assertEqual(labels.get("Descricao do trabalho"), "Descricao tecnica atualizada")
        atividades_nomes = [atividade["nome"] for atividade in context["atividades"]]
        self.assertIn("Comissionamento", atividades_nomes)

    def test_descricao_comercial_remove_bloco_tecnico_duplicado(self):
        raw = (
            "Origem tecnica\n"
            "Radar: Radar Norte\n"
            "Trabalho: Trabalho TEC\n"
            "Descricao do trabalho: xpto\n"
            "Resumo das atividades\n"
            "- item legado\n\n"
            "Somente condicoes comerciais."
        )
        clean = _sanitize_proposta_descricao(raw)
        self.assertNotIn("Origem tecnica", clean)
        self.assertNotIn("Radar:", clean)
        self.assertIn("Somente condicoes comerciais.", clean)

    def test_pdf_context_mostra_origem_tecnica_e_atividades_corretas(self):
        proposta = Proposta.objects.select_related(
            "trabalho",
            "trabalho__radar",
            "trabalho__classificacao",
            "trabalho__contrato",
        ).prefetch_related("trabalho__atividades").get(pk=self.proposta.pk)
        context = _build_proposta_pdf_context(proposta, "Pendente", include_origem=True, trabalho=proposta.trabalho)
        self.assertTrue(context["has_trabalho_vinculado"])
        self.assertFalse(context["trabalho_indisponivel"])
        labels = {row["label"]: row["value"] for row in context["origem_rows"]}
        self.assertEqual(labels.get("Radar"), "Radar Norte")
        self.assertEqual(labels.get("Trabalho"), "Trabalho TEC")
        self.assertEqual(labels.get("Colaboradores"), "Ana, Bruno")
        self.assertEqual(context["atividades"][0]["nome"], "Inspecao")

    def test_proposta_sem_vinculo_permanece_funcional(self):
        proposta = Proposta.objects.create(
            cliente=self.destinatario,
            criada_por=self.vendedor,
            nome="Proposta Avulsa",
            descricao="Conteudo comercial avulso.",
        )
        context = _build_proposta_pdf_context(proposta, "Pendente", include_origem=True)
        self.assertFalse(context["has_trabalho_vinculado"])
        self.assertEqual(context["origem_rows"], [])
        self.assertEqual(context["descricao_blocks"][0]["text"], "Conteudo comercial avulso.")

    def test_nao_autorizado_nao_cria_vinculo(self):
        outro = User.objects.create_user(username="outro", email="outro@set.local", password="123456")
        self.client_http.force_login(outro)
        response = self.client_http.post(
            "/propostas/nova/",
            {
                "email": self.destinatario.email,
                "nome": "Proposta sem permissao",
                "descricao": "Comercial",
                "trabalho_id": str(self.trabalho.id),
            },
        )
        self.assertEqual(response.status_code, 200)
        proposta = Proposta.objects.filter(nome="Proposta sem permissao").first()
        self.assertIsNone(proposta)


class RadarCreatorPermissionTests(TestCase):
    def setUp(self):
        self.client_http = Client()
        self.owner_user = User.objects.create_user(username="owner", email="owner@set.local", password="123456")
        self.viewer_user = User.objects.create_user(username="viewer", email="viewer@set.local", password="123456")
        self.owner = PerfilUsuario.objects.create(nome="Owner", email="owner@set.local", usuario=self.owner_user)
        self.viewer = PerfilUsuario.objects.create(nome="Viewer", email="viewer@set.local", usuario=self.viewer_user)
        self.radar_id = RadarID.objects.create(codigo="R-001")
        self.viewer.radares.add(self.radar_id)
        self.radar = Radar.objects.create(
            cliente=self.owner,
            id_radar=self.radar_id,
            nome="Radar Permissao",
            criador=self.owner_user,
        )
        self.trabalho = RadarTrabalho.objects.create(
            radar=self.radar,
            nome="Trabalho Permissao",
            criado_por=self.owner_user,
        )

    def test_usuario_com_id_radar_nao_altera_radar_trabalho_atividade(self):
        self.client_http.force_login(self.viewer_user)
        resp_radar = self.client_http.post(
            f"/radar-atividades/{self.radar.id}/",
            {"action": "create_trabalho", "nome": "Nao deve criar"},
        )
        self.assertEqual(resp_radar.status_code, 403)
        resp_trabalho = self.client_http.post(
            f"/radar-atividades/{self.radar.id}/trabalhos/{self.trabalho.id}/",
            {"action": "create_atividade", "nome": "Nao deve criar"},
        )
        self.assertEqual(resp_trabalho.status_code, 403)

    def test_criador_do_radar_pode_alterar(self):
        self.client_http.force_login(self.owner_user)
        resp = self.client_http.post(
            f"/radar-atividades/{self.radar.id}/",
            {"action": "create_trabalho", "nome": "Pode criar"},
        )
        self.assertEqual(resp.status_code, 302)
        self.assertTrue(RadarTrabalho.objects.filter(radar=self.radar, nome="Pode criar").exists())

    def test_create_trabalho_ajax_persiste_colaboradores(self):
        self.client_http.force_login(self.owner_user)
        response = self.client_http.post(
            f"/radar-atividades/{self.radar.id}/",
            {
                "action": "create_trabalho",
                "nome": "Trabalho Equipe",
                "colaboradores": "Ana, Bruno, ana",
            },
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["ok"])
        self.assertEqual(payload["row"]["total_colaboradores"], 2)
        self.assertEqual(payload["row"]["total_horas"], "0.00")
        trabalho = RadarTrabalho.objects.get(radar=self.radar, nome="Trabalho Equipe")
        nomes = list(trabalho.colaboradores.order_by("nome").values_list("nome", flat=True))
        self.assertEqual(nomes, ["Ana", "Bruno"])

    def test_radar_detail_exibe_total_horas_por_trabalho(self):
        self.client_http.force_login(self.owner_user)
        atividade = RadarAtividade.objects.create(trabalho=self.trabalho, nome="Ativ Horas Lista")
        self.client_http.post(
            f"/radar-atividades/{self.radar.id}/trabalhos/{self.trabalho.id}/",
            {
                "action": "set_agenda_atividade",
                "atividade_id": str(atividade.id),
                "dias_execucao": json.dumps(["2026-03-01", "2026-03-02"]),
            },
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        response = self.client_http.get(f"/radar-atividades/{self.radar.id}/")
        self.assertEqual(response.status_code, 200)
        rows = response.context["trabalhos_table_data"]
        row = next(item for item in rows if item["id"] == self.trabalho.id)
        self.assertEqual(row["total_horas"], "16.00")

    def test_agenda_exibe_observacoes_do_dia_e_marcador_no_calendario(self):
        self.client_http.force_login(self.owner_user)
        atividade = RadarAtividade.objects.create(trabalho=self.trabalho, nome="Ativ Agenda Obs")
        RadarAtividadeDiaExecucao.objects.create(atividade=atividade, data_execucao="2026-03-10")
        RadarTrabalhoObservacao.objects.create(
            trabalho=self.trabalho,
            texto="Observacao no mesmo dia",
            data_observacao=datetime(2026, 3, 10).date(),
        )
        RadarTrabalhoObservacao.objects.create(
            trabalho=self.trabalho,
            texto="Observacao em outro dia",
            data_observacao=datetime(2026, 3, 11).date(),
        )

        response = self.client_http.get(f"/radar-atividades/{self.radar.id}/agenda/?dia=2026-03-10")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["daily_total_atividades"], 1)
        self.assertEqual(response.context["daily_total_observacoes"], 1)
        self.assertEqual(len(response.context["daily_groups"]), 1)
        group = response.context["daily_groups"][0]
        self.assertEqual(group["trabalho_id"], self.trabalho.id)
        self.assertEqual(group["total_atividades"], 1)
        self.assertEqual(group["total_observacoes"], 1)
        self.assertEqual(group["observacoes"][0]["texto"], "Observacao no mesmo dia")
        calendar_cells = [
            cell
            for week in response.context["calendar_weeks"]
            for cell in week
            if cell["iso"] == "2026-03-10"
        ]
        self.assertEqual(len(calendar_cells), 1)
        self.assertTrue(calendar_cells[0]["has_observation"])
        self.assertEqual(calendar_cells[0]["observation_count"], 1)

    def test_agenda_exibe_trabalho_com_apenas_observacao_no_dia(self):
        self.client_http.force_login(self.owner_user)
        RadarTrabalhoObservacao.objects.create(
            trabalho=self.trabalho,
            texto="Apenas observacao",
            data_observacao=datetime(2026, 3, 12).date(),
        )

        response = self.client_http.get(f"/radar-atividades/{self.radar.id}/agenda/?dia=2026-03-12")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["daily_total_atividades"], 0)
        self.assertEqual(response.context["daily_total_observacoes"], 1)
        self.assertEqual(len(response.context["daily_groups"]), 1)
        group = response.context["daily_groups"][0]
        self.assertEqual(group["trabalho_id"], self.trabalho.id)
        self.assertEqual(group["total_atividades"], 0)
        self.assertEqual(group["total_observacoes"], 1)
        self.assertEqual(group["observacoes"][0]["texto"], "Apenas observacao")

    def test_update_trabalho_sincroniza_colaboradores(self):
        self.client_http.force_login(self.owner_user)
        response = self.client_http.post(
            f"/radar-atividades/{self.radar.id}/trabalhos/{self.trabalho.id}/",
            {
                "action": "update_trabalho",
                "nome": self.trabalho.nome,
                "descricao": self.trabalho.descricao or "",
                "setor": self.trabalho.setor or "",
                "solicitante": self.trabalho.solicitante or "",
                "responsavel": self.trabalho.responsavel or "",
                "colaboradores": "Carlos, Diana, Carlos",
            },
        )
        self.assertEqual(response.status_code, 302)
        nomes = list(self.trabalho.colaboradores.order_by("nome").values_list("nome", flat=True))
        self.assertEqual(nomes, ["Carlos", "Diana"])

    def test_agenda_define_datas_e_status_permanece_manual(self):
        self.client_http.force_login(self.owner_user)
        create_resp = self.client_http.post(
            f"/radar-atividades/{self.radar.id}/trabalhos/{self.trabalho.id}/",
            {"action": "create_atividade", "nome": "Ativ X", "status": "EXECUTANDO"},
        )
        self.assertEqual(create_resp.status_code, 302)
        atividade = RadarAtividade.objects.get(trabalho=self.trabalho, nome="Ativ X")
        self.assertIsNone(atividade.inicio_execucao_em)
        self.assertIsNone(atividade.finalizada_em)

        agenda_resp = self.client_http.post(
            f"/radar-atividades/{self.radar.id}/trabalhos/{self.trabalho.id}/",
            {
                "action": "set_agenda_atividade",
                "atividade_id": str(atividade.id),
                "dias_execucao": json.dumps(["2026-03-05", "2026-03-07", "2026-03-06"]),
            },
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )
        self.assertEqual(agenda_resp.status_code, 200)
        self.assertEqual(agenda_resp.json()["agenda_total_dias"], 3)
        atividade.refresh_from_db()
        self.assertIsNotNone(atividade.inicio_execucao_em)
        self.assertIsNotNone(atividade.finalizada_em)
        self.assertEqual(atividade.inicio_execucao_em.date().isoformat(), "2026-03-05")
        self.assertEqual(atividade.finalizada_em.date().isoformat(), "2026-03-07")

        status_resp = self.client_http.post(
            f"/radar-atividades/{self.radar.id}/trabalhos/{self.trabalho.id}/",
            {
                "action": "quick_status_atividade",
                "atividade_id": str(atividade.id),
                "status": "FINALIZADA",
            },
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )
        self.assertEqual(status_resp.status_code, 200)
        atividade.refresh_from_db()
        self.assertEqual(atividade.inicio_execucao_em.date().isoformat(), "2026-03-05")
        self.assertEqual(atividade.finalizada_em.date().isoformat(), "2026-03-07")

        clear_resp = self.client_http.post(
            f"/radar-atividades/{self.radar.id}/trabalhos/{self.trabalho.id}/",
            {
                "action": "set_agenda_atividade",
                "atividade_id": str(atividade.id),
                "dias_execucao": "[]",
            },
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )
        self.assertEqual(clear_resp.status_code, 200)
        atividade.refresh_from_db()
        self.assertIsNone(atividade.inicio_execucao_em)
        self.assertIsNone(atividade.finalizada_em)
        self.assertFalse(RadarAtividadeDiaExecucao.objects.filter(atividade=atividade).exists())

    def test_set_agenda_atividade_retorna_erro_para_data_invalida(self):
        self.client_http.force_login(self.owner_user)
        atividade = RadarAtividade.objects.create(trabalho=self.trabalho, nome="Ativ Y")
        response = self.client_http.post(
            f"/radar-atividades/{self.radar.id}/trabalhos/{self.trabalho.id}/",
            {
                "action": "set_agenda_atividade",
                "atividade_id": str(atividade.id),
                "dias_execucao": json.dumps(["2026-02-30"]),
            },
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )
        self.assertEqual(response.status_code, 400)
        payload = response.json()
        self.assertFalse(payload["ok"])

    def test_horas_trabalho_e_calculada_por_agenda_x_horas_dia(self):
        self.client_http.force_login(self.owner_user)
        atividade = RadarAtividade.objects.create(trabalho=self.trabalho, nome="Ativ Horas")

        response = self.client_http.post(
            f"/radar-atividades/{self.radar.id}/trabalhos/{self.trabalho.id}/",
            {
                "action": "set_agenda_atividade",
                "atividade_id": str(atividade.id),
                "dias_execucao": json.dumps(["2026-03-01", "2026-03-02", "2026-03-03"]),
            },
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["horas_trabalho"], "24.00")

        atividade.refresh_from_db()
        self.assertEqual(atividade.horas_trabalho, Decimal("24.00"))

    def test_horas_trabalho_multiplica_quantidade_colaboradores(self):
        self.client_http.force_login(self.owner_user)
        RadarTrabalhoColaborador.objects.create(trabalho=self.trabalho, nome="Ana")
        RadarTrabalhoColaborador.objects.create(trabalho=self.trabalho, nome="Bruno")
        atividade = RadarAtividade.objects.create(trabalho=self.trabalho, nome="Ativ Equipe")

        response = self.client_http.post(
            f"/radar-atividades/{self.radar.id}/trabalhos/{self.trabalho.id}/",
            {
                "action": "set_agenda_atividade",
                "atividade_id": str(atividade.id),
                "dias_execucao": json.dumps(["2026-03-01", "2026-03-02", "2026-03-03"]),
            },
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["horas_trabalho"], "48.00")

        atividade.refresh_from_db()
        self.assertEqual(atividade.horas_trabalho, Decimal("48.00"))

    def test_create_atividade_herda_colaboradores_do_trabalho(self):
        self.client_http.force_login(self.owner_user)
        RadarTrabalhoColaborador.objects.create(trabalho=self.trabalho, nome="Ana")
        RadarTrabalhoColaborador.objects.create(trabalho=self.trabalho, nome="Bruno")

        response = self.client_http.post(
            f"/radar-atividades/{self.radar.id}/trabalhos/{self.trabalho.id}/",
            {"action": "create_atividade", "nome": "Ativ Herdada"},
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )
        self.assertEqual(response.status_code, 200)
        atividade = RadarAtividade.objects.get(trabalho=self.trabalho, nome="Ativ Herdada")
        nomes = list(atividade.colaboradores.order_by("nome").values_list("nome", flat=True))
        self.assertEqual(nomes, ["Ana", "Bruno"])

    def test_atualizar_horas_dia_recalcula_horas_de_todas_atividades(self):
        self.client_http.force_login(self.owner_user)
        atividade = RadarAtividade.objects.create(trabalho=self.trabalho, nome="Ativ Recalc")

        agenda_resp = self.client_http.post(
            f"/radar-atividades/{self.radar.id}/trabalhos/{self.trabalho.id}/",
            {
                "action": "set_agenda_atividade",
                "atividade_id": str(atividade.id),
                "dias_execucao": json.dumps(["2026-03-01", "2026-03-02"]),
            },
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )
        self.assertEqual(agenda_resp.status_code, 200)
        atividade.refresh_from_db()
        self.assertEqual(atividade.horas_trabalho, Decimal("16.00"))

        update_resp = self.client_http.post(
            f"/radar-atividades/{self.radar.id}/trabalhos/{self.trabalho.id}/",
            {
                "action": "update_trabalho",
                "nome": self.trabalho.nome,
                "descricao": self.trabalho.descricao or "",
                "setor": self.trabalho.setor or "",
                "solicitante": self.trabalho.solicitante or "",
                "responsavel": self.trabalho.responsavel or "",
                "horas_dia": "6",
                "colaboradores": "",
            },
        )
        self.assertEqual(update_resp.status_code, 302)
        atividade.refresh_from_db()
        self.assertEqual(atividade.horas_trabalho, Decimal("12.00"))

    def test_atualizar_colaboradores_do_trabalho_nao_muda_equipe_ja_herdada(self):
        self.client_http.force_login(self.owner_user)
        atividade = RadarAtividade.objects.create(trabalho=self.trabalho, nome="Ativ Recalc Equipe")

        agenda_resp = self.client_http.post(
            f"/radar-atividades/{self.radar.id}/trabalhos/{self.trabalho.id}/",
            {
                "action": "set_agenda_atividade",
                "atividade_id": str(atividade.id),
                "dias_execucao": json.dumps(["2026-03-01", "2026-03-02"]),
            },
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )
        self.assertEqual(agenda_resp.status_code, 200)
        atividade.refresh_from_db()
        self.assertEqual(atividade.horas_trabalho, Decimal("16.00"))

        update_resp = self.client_http.post(
            f"/radar-atividades/{self.radar.id}/trabalhos/{self.trabalho.id}/",
            {
                "action": "update_trabalho",
                "nome": self.trabalho.nome,
                "descricao": self.trabalho.descricao or "",
                "setor": self.trabalho.setor or "",
                "solicitante": self.trabalho.solicitante or "",
                "responsavel": self.trabalho.responsavel or "",
                "horas_dia": "8",
                "colaboradores": "Ana, Bruno",
            },
        )
        self.assertEqual(update_resp.status_code, 302)
        atividade.refresh_from_db()
        self.assertEqual(atividade.horas_trabalho, Decimal("16.00"))

    def test_update_atividade_recalcula_horas_com_colaboradores_da_atividade(self):
        self.client_http.force_login(self.owner_user)
        ana = RadarColaborador.objects.create(perfil=self.owner, nome="Ana")
        bruno = RadarColaborador.objects.create(perfil=self.owner, nome="Bruno")
        RadarTrabalhoColaborador.objects.create(trabalho=self.trabalho, nome="Ana", colaborador=ana)
        RadarTrabalhoColaborador.objects.create(trabalho=self.trabalho, nome="Bruno", colaborador=bruno)
        atividade = RadarAtividade.objects.create(trabalho=self.trabalho, nome="Ativ Eq Editavel")

        agenda_resp = self.client_http.post(
            f"/radar-atividades/{self.radar.id}/trabalhos/{self.trabalho.id}/",
            {
                "action": "set_agenda_atividade",
                "atividade_id": str(atividade.id),
                "dias_execucao": json.dumps(["2026-03-01", "2026-03-02"]),
            },
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )
        self.assertEqual(agenda_resp.status_code, 200)

        update_resp = self.client_http.post(
            f"/radar-atividades/{self.radar.id}/trabalhos/{self.trabalho.id}/",
            {
                "action": "update_atividade",
                "atividade_id": str(atividade.id),
                "nome": atividade.nome,
                "descricao": "",
                "status": "PENDENTE",
                "colaborador_ids": [str(ana.id), str(bruno.id)],
            },
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )
        self.assertEqual(update_resp.status_code, 200)
        atividade.refresh_from_db()
        self.assertEqual(atividade.horas_trabalho, Decimal("32.00"))
        self.assertEqual(
            list(atividade.colaboradores.order_by("nome").values_list("nome", flat=True)),
            ["Ana", "Bruno"],
        )

    def test_update_atividade_mantem_colaborador_ja_vinculado_mesmo_apos_remocao_do_trabalho(self):
        self.client_http.force_login(self.owner_user)
        ana = RadarColaborador.objects.create(perfil=self.owner, nome="Ana")
        RadarTrabalhoColaborador.objects.create(trabalho=self.trabalho, nome="Ana", colaborador=ana)
        atividade = RadarAtividade.objects.create(trabalho=self.trabalho, nome="Ativ Preserva Equipe")

        remove_resp = self.client_http.post(
            f"/radar-atividades/{self.radar.id}/trabalhos/{self.trabalho.id}/",
            {
                "action": "update_trabalho",
                "nome": self.trabalho.nome,
                "descricao": self.trabalho.descricao or "",
                "setor": self.trabalho.setor or "",
                "solicitante": self.trabalho.solicitante or "",
                "responsavel": self.trabalho.responsavel or "",
                "horas_dia": "8",
                "colaborador_ids": [],
                "colaboradores": "",
            },
        )
        self.assertEqual(remove_resp.status_code, 302)

        update_resp = self.client_http.post(
            f"/radar-atividades/{self.radar.id}/trabalhos/{self.trabalho.id}/",
            {
                "action": "update_atividade",
                "atividade_id": str(atividade.id),
                "nome": atividade.nome,
                "descricao": "",
                "status": "PENDENTE",
                "colaborador_ids": [str(ana.id)],
            },
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )
        self.assertEqual(update_resp.status_code, 200)
        self.assertEqual(
            list(atividade.colaboradores.order_by("nome").values_list("nome", flat=True)),
            ["Ana"],
        )

    def test_status_trabalho_nao_aceita_update_manual(self):
        self.client_http.force_login(self.owner_user)
        response = self.client_http.post(
            f"/radar-atividades/{self.radar.id}/",
            {
                "action": "quick_status_trabalho",
                "trabalho_id": str(self.trabalho.id),
                "status": "FINALIZADA",
            },
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )
        self.assertEqual(response.status_code, 400)
        payload = response.json()
        self.assertFalse(payload["ok"])
        self.assertEqual(payload["status"], RadarTrabalho.Status.PENDENTE)
        self.assertIn("automatico", payload["message"].lower())

    def test_radar_trabalho_detail_recalcula_horas_stale_ao_carregar(self):
        self.client_http.force_login(self.owner_user)
        RadarTrabalhoColaborador.objects.create(trabalho=self.trabalho, nome="Ana")
        RadarTrabalhoColaborador.objects.create(trabalho=self.trabalho, nome="Bruno")
        atividade = RadarAtividade.objects.create(
            trabalho=self.trabalho,
            nome="Ativ stale",
            horas_trabalho=Decimal("0.00"),
        )
        RadarAtividadeDiaExecucao.objects.create(atividade=atividade, data_execucao="2026-03-01")
        RadarAtividadeDiaExecucao.objects.create(atividade=atividade, data_execucao="2026-03-02")

        response = self.client_http.get(
            f"/radar-atividades/{self.radar.id}/trabalhos/{self.trabalho.id}/"
        )
        self.assertEqual(response.status_code, 200)
        atividade.refresh_from_db()
        self.assertEqual(atividade.horas_trabalho, Decimal("32.00"))

    def test_export_relatorio_pdf_permite_apenas_criador(self):
        self.client_http.force_login(self.viewer_user)
        response = self.client_http.get(
            f"/radar-atividades/{self.radar.id}/relatorio/pdf/?mes=2026-03",
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )
        self.assertEqual(response.status_code, 403)

    def test_export_relatorio_pdf_bloqueia_sem_dados_no_mes(self):
        self.client_http.force_login(self.owner_user)
        response = self.client_http.get(
            f"/radar-atividades/{self.radar.id}/relatorio/pdf/?mes=2026-03",
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )
        self.assertEqual(response.status_code, 400)
        payload = response.json()
        self.assertFalse(payload["ok"])
        self.assertIn("Sem atividades executadas", payload["message"])

    def test_export_relatorio_pdf_gera_arquivo_com_nome_padrao(self):
        self.client_http.force_login(self.owner_user)
        atividade = RadarAtividade.objects.create(trabalho=self.trabalho, nome="Ativ Export")
        RadarAtividadeDiaExecucao.objects.create(atividade=atividade, data_execucao="2026-03-10")

        with patch("core.views._render_radar_relatorio_pdf", return_value=BytesIO(b"%PDF-1.4\nfake")):
            response = self.client_http.get(
                f"/radar-atividades/{self.radar.id}/relatorio/pdf/?mes=2026-03",
                HTTP_X_REQUESTED_WITH="XMLHttpRequest",
            )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/pdf")
        self.assertIn('attachment; filename="relatorio_Radar_Permissao_2026-03.pdf"', response["Content-Disposition"])

    def test_criador_pode_criar_observacao_com_data_padrao(self):
        self.client_http.force_login(self.owner_user)
        response = self.client_http.post(
            f"/radar-atividades/{self.radar.id}/trabalhos/{self.trabalho.id}/",
            {
                "action": "create_observacao",
                "observacao_texto": "Primeira observacao",
            },
        )
        self.assertEqual(response.status_code, 302)
        observacao = RadarTrabalhoObservacao.objects.get(trabalho=self.trabalho)
        self.assertEqual(observacao.texto, "Primeira observacao")
        self.assertEqual(observacao.data_observacao, timezone.localdate())

    def test_viewer_nao_cria_observacao(self):
        self.client_http.force_login(self.viewer_user)
        response = self.client_http.post(
            f"/radar-atividades/{self.radar.id}/trabalhos/{self.trabalho.id}/",
            {
                "action": "create_observacao",
                "observacao_texto": "Nao pode",
            },
        )
        self.assertEqual(response.status_code, 403)
        self.assertFalse(RadarTrabalhoObservacao.objects.filter(trabalho=self.trabalho).exists())

    def test_criador_pode_editar_e_excluir_observacao(self):
        self.client_http.force_login(self.owner_user)
        observacao = RadarTrabalhoObservacao.objects.create(
            trabalho=self.trabalho,
            texto="Obs inicial",
            data_observacao=datetime(2026, 3, 1).date(),
        )
        update_response = self.client_http.post(
            f"/radar-atividades/{self.radar.id}/trabalhos/{self.trabalho.id}/",
            {
                "action": "update_observacao",
                "observacao_id": str(observacao.id),
                "observacao_texto": "Obs atualizada",
                "observacao_data": "2026-03-05",
            },
        )
        self.assertEqual(update_response.status_code, 302)
        observacao.refresh_from_db()
        self.assertEqual(observacao.texto, "Obs atualizada")
        self.assertEqual(observacao.data_observacao.isoformat(), "2026-03-05")

        delete_response = self.client_http.post(
            f"/radar-atividades/{self.radar.id}/trabalhos/{self.trabalho.id}/",
            {
                "action": "delete_observacao",
                "observacao_id": str(observacao.id),
            },
        )
        self.assertEqual(delete_response.status_code, 302)
        self.assertFalse(RadarTrabalhoObservacao.objects.filter(pk=observacao.id).exists())

    def test_contexto_pdf_do_radar_inclui_todas_observacoes_do_trabalho(self):
        self.client_http.force_login(self.owner_user)
        atividade = RadarAtividade.objects.create(trabalho=self.trabalho, nome="Ativ PDF Obs")
        RadarAtividadeDiaExecucao.objects.create(atividade=atividade, data_execucao="2026-03-10")
        RadarTrabalhoObservacao.objects.create(
            trabalho=self.trabalho,
            texto="Observacao antiga",
            data_observacao=datetime(2026, 1, 15).date(),
        )
        RadarTrabalhoObservacao.objects.create(
            trabalho=self.trabalho,
            texto="Observacao mais recente",
            data_observacao=datetime(2026, 4, 2).date(),
        )

        context = _build_radar_relatorio_pdf_context(
            self.radar,
            datetime(2026, 3, 1).date(),
            datetime(2026, 3, 31).date(),
        )
        trabalho_ctx = next(item for item in context["trabalho_pages"] if item["id"] == self.trabalho.id)
        self.assertEqual(len(trabalho_ctx["observacoes_resumo"]), 2)
        self.assertEqual(trabalho_ctx["observacoes_resumo"][0]["texto"], "Observacao mais recente")
        self.assertEqual(trabalho_ctx["observacoes_resumo"][1]["texto"], "Observacao antiga")


class IOImportPipelineTests(TestCase):
    FIXTURES_DIR = Path(__file__).resolve().parent.parent / "io_import_test_files"

    def setUp(self):
        self.client_http = Client()
        self.tipo_dev = TipoPerfil.objects.get(codigo="DEV")
        self.user = User.objects.create_user(
            username="io-import@set.local",
            email="io-import@set.local",
            password="123456",
        )
        self.perfil = PerfilUsuario.objects.create(
            nome="IO Import",
            email="io-import@set.local",
            usuario=self.user,
        )
        self.perfil.tipos.add(self.tipo_dev)
        self.client_http.force_login(self.user)

        self.local = LocalRackIO.objects.create(cliente=self.perfil, nome="Fabrica 01")
        self.grupo = GrupoRackIO.objects.create(cliente=self.perfil, nome="Recebimento")
        self.planta = PlantaIO.objects.create(codigo="PLANTA-001")
        self.perfil.plantas.add(self.planta)

        self.tipo_di = TipoCanalIO.objects.get(nome="DI")
        self.tipo_do = TipoCanalIO.objects.get(nome="DO")
        self.modulo_di = ModuloIO.objects.filter(is_default=True, modelo="DI-04").first()
        if not self.modulo_di:
            self.modulo_di = ModuloIO.objects.create(
                cliente=None,
                nome="DI-04",
                modelo="DI-04",
                marca="SET",
                quantidade_canais=4,
                tipo_base=self.tipo_di,
                is_default=True,
            )
        self.modulo_di16 = ModuloIO.objects.filter(is_default=True, modelo="DI-16").first()
        if not self.modulo_di16:
            self.modulo_di16 = ModuloIO.objects.create(
                cliente=None,
                nome="DI-16",
                modelo="DI-16",
                marca="SET",
                quantidade_canais=16,
                tipo_base=self.tipo_di,
                is_default=True,
            )
        self.modulo_do16 = ModuloIO.objects.filter(is_default=True, modelo="DO-16").first()
        if not self.modulo_do16:
            self.modulo_do16 = ModuloIO.objects.create(
                cliente=None,
                nome="DO-16",
                modelo="DO-16",
                marca="SET",
                quantidade_canais=16,
                tipo_base=self.tipo_do,
                is_default=True,
            )

    def _build_slot_block_workbook(self):
        workbook = Workbook()
        sheet1 = workbook.active
        sheet1.title = "UBS3-ET200SP"
        for row in [
            ["", "", "LISTA DE IO", "", "", "", "", ""],
            ["", "", "SLOT 01 SIMATIC ET 200SP IM 155-6PN - 64 1/0", "", "", "", "", ""],
            ["", "", "SLOT 02 6ES7131-6BH01-0BA0: 16 DI", "", "", "", "", ""],
            ["", "", "IO", "Local", "Min", "Max", "TAG 's", "DESCRIÇÃO"],
            ["", "", "DI.0", "", "0", "1", "IGNORAR_A", "Primeira aba"],
        ]:
            sheet1.append(row)

        sheet2 = workbook.create_sheet("REM01 6ES7 155-6AU01-0BN0")
        for row in [
            ["", "", "Nutrien Unidade - Rio Verde", "", "", "", "", ""],
            ["", "", "LISTA DE IO", "", "", "", "", ""],
            ["", "", "REM-01 6ES7 155-6AU01-0BN0", "", "", "", "", ""],
            ["", "", "SLOT 01 6ES7131-6BH01-0BA0: 16 DI", "", "", "", "", ""],
            ["", "", "IO", "Local", "Min", "Max", "TAG 's", "DESCRIÇÃO"],
            ["", "", "DI.0", "", "0", "1", "FB_S2EX01", "FEEDBACK EXAUSTOR 01"],
            ["", "", "DI.1", "", "0", "1", "FB_S2EX02", "FEEDBACK EXAUSTOR 02"],
            ["", "", "SLOT 02 6ES7132-6BH01-0BA0: 16 DQ", "", "", "", "", ""],
            ["", "", "IO", "Local", "Min", "Max", "TAG 's", "DESCRIÇÃO"],
            ["", "", "DQ.0", "", "0", "1", "CMD_S2EX01", "COMANDO EXAUSTOR 01"],
            ["", "", "DQ.1", "", "0", "1", "CMD_S2EX02", "COMANDO EXAUSTOR 02"],
        ]:
            sheet2.append(row)

        payload = BytesIO()
        workbook.save(payload)
        return payload.getvalue()

    def _build_single_rack_slot_block_workbook(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "REM01 6ES7 155-6AU01-0BN0"
        for row in [
            ["", "", "Nutrien Unidade - Rio Verde", "", "", "", "", ""],
            ["", "", "LISTA DE IO", "", "", "", "", ""],
            ["", "", "REM-01 6ES7 155-6AU01-0BN0", "", "", "", "", ""],
            ["", "", "SLOT 01 6ES7131-6BH01-0BA0: 16 DI", "", "", "", "", ""],
            ["", "", "IO", "Local", "Min", "Max", "TAG 's", "DESCRIÃ‡ÃƒO"],
            ["", "", "DI.0", "", "0", "1", "FB_S2EX01", "FEEDBACK EXAUSTOR 01"],
            ["", "", "DI.1", "", "0", "1", "FB_S2EX02", "FEEDBACK EXAUSTOR 02"],
            ["", "", "SLOT 02 6ES7132-6BH01-0BA0: 16 DQ", "", "", "", "", ""],
            ["", "", "IO", "Local", "Min", "Max", "TAG 's", "DESCRIÃ‡ÃƒO"],
            ["", "", "DQ.0", "", "0", "1", "CMD_S2EX01", "COMANDO EXAUSTOR 01"],
            ["", "", "DQ.1", "", "0", "1", "CMD_S2EX02", "COMANDO EXAUSTOR 02"],
        ]:
            sheet.append(row)

        payload = BytesIO()
        workbook.save(payload)
        return payload.getvalue()

    def _build_single_sheet_multi_rack_workbook(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "IO Consolidado"
        for row in [
            ["", "", "REM01 6ES7 155-6AU01-0BN0", "", "", "", "", ""],
            ["", "", "SLOT 01 6ES7131-6BH01-0BA0: 16 DI", "", "", "", "", ""],
            ["", "", "IO", "Local", "Min", "Max", "TAG 's", "DESCRIÃ‡ÃƒO"],
            ["", "", "DI.0", "", "0", "1", "REM01_FB_01", "Feedback REM01 01"],
            ["", "", "DI.1", "", "0", "1", "REM01_FB_02", "Feedback REM01 02"],
            ["", "", "UBS3 ET200SP", "", "", "", "", ""],
            ["", "", "SLOT 01 6ES7131-6BH01-0BA0: 16 DI", "", "", "", "", ""],
            ["", "", "IO", "Local", "Min", "Max", "TAG 's", "DESCRIÃ‡ÃƒO"],
            ["", "", "DI.0", "", "0", "1", "UBS3_FB_01", "Feedback UBS3 01"],
            ["", "", "DI.1", "", "0", "1", "UBS3_FB_02", "Feedback UBS3 02"],
        ]:
            sheet.append(row)

        payload = BytesIO()
        workbook.save(payload)
        return payload.getvalue()

    def _build_slot_block_with_internal_point_titles_workbook(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "AENTR"
        for row in [
            ["", "", "Milhao Ingredients", "", "", "", "", ""],
            ["", "", "CPU CONTROLOGIX 1756-L73 192.100.3.1", "", "", "", "", ""],
            ["", "", "LISTA DE IO", "", "", "", "", ""],
            ["", "", "SLOT 01 1746-IB32", "", "", "", "", ""],
            ["", "", "IO", "Local", "Min", "Max", "TAG 's", "DESCRICAO"],
            ["", "", "DI.0", "", "0", "1", "M21VR59_FB", "Feedback 59"],
            ["", "", "DI.1", "", "0", "1", "M21VR30_FB", "Feedback 30"],
            ["", "", "AO2-M21RE5-CAMPO", "", "", "", "", ""],
            ["", "", "SLOT 02 1746-NO4I", "", "", "", "", ""],
            ["", "", "IO", "Local", "Min", "Max", "TAG 's", "DESCRICAO"],
            ["", "", "AO.0", "", "0", "100", "M21RE5", "Comando analogico"],
            ["", "", "AO5-V20MV3-OUT", "", "", "", "", ""],
            ["", "", "SLOT 03 1746-OB32", "", "", "", "", ""],
            ["", "", "IO", "Local", "Min", "Max", "TAG 's", "DESCRICAO"],
            ["", "", "DO.0", "", "0", "1", "V20MV3_CMD", "Comando valvula"],
        ]:
            sheet.append(row)
        payload = BytesIO()
        workbook.save(payload)
        return payload.getvalue()

    def _build_ai_first_obscure_workbook(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Raw Export"
        for row in [
            ["Projeto X", "", "", ""],
            ["REFX", "NARRATIVE", "HARDWAREPATH", "SIGCLASS"],
            ["P101_RUN_FB", "Feedback bomba 101", "PNL-PRC-01 / Rack 1 / Slot 02 / DI.00", "DI"],
            ["P101_START_CMD", "Comando partida bomba 101", "PNL-PRC-01 / Rack 1 / Slot 03 / DQ.01", "DO"],
        ]:
            sheet.append(row)

        payload = BytesIO()
        workbook.save(payload)
        return payload.getvalue()

    def test_import_csv_creates_review_job(self):
        upload = SimpleUploadedFile(
            "lista.csv",
            (
                b"TAG;DESCRICAO;TIPO;SLOT;CANAL\n"
                b"MTR_01;Motor principal;DI;1;1\n"
                b"MTR_02;Motor reserva;DI;1;2\n"
            ),
            content_type="text/csv",
        )
        response = self.client_http.post(
            "/ios/importacoes/nova/",
            {
                "arquivo": upload,
            },
        )
        self.assertEqual(response.status_code, 302)

        job = IOImportJob.objects.get()
        self.assertEqual(job.status, IOImportJob.Status.REVIEW)
        self.assertEqual(job.file_format, IOImportJob.FileFormat.CSV)
        self.assertEqual(job.rows_parsed, 2)
        self.assertEqual(job.mode, IOImportJob.Mode.CREATE_RACK)
        self.assertEqual(job.requested_rack_name, "")
        self.assertIsNone(job.requested_local)
        self.assertIsNone(job.requested_grupo)
        self.assertEqual(job.requested_planta_code, "")
        self.assertFalse(job.target_rack_id)
        self.assertTrue(job.proposal_payload["rack"]["name"])
        self.assertEqual(job.proposal_payload["summary"]["modules"], 1)

    def test_import_ajax_returns_json_redirect_url(self):
        upload = SimpleUploadedFile(
            "PLANILHA DE IO UBS3 NUTRIEN - REM01 REV03.xlsx",
            self._build_single_rack_slot_block_workbook(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        with patch("core.views._spawn_io_import_job_processor") as spawn_mock:
            with self.captureOnCommitCallbacks(execute=True):
                response = self.client_http.post(
                    "/ios/importacoes/nova/",
                    {
                        "arquivo": upload,
                    },
                    HTTP_X_REQUESTED_WITH="XMLHttpRequest",
                )

        self.assertEqual(response.status_code, 200)
        self.assertIn("application/json", response["Content-Type"])
        payload = response.json()
        self.assertTrue(payload["ok"])
        self.assertIn("/ios/importacoes/", payload["redirect_url"])
        self.assertIn("/status/", payload["status_url"])
        spawn_mock.assert_called_once()

        job = IOImportJob.objects.latest("id")
        self.assertEqual(job.status, IOImportJob.Status.UPLOADED)
        self.assertEqual((job.progress_payload or {}).get("stage"), "upload")
        self.assertGreater((job.progress_payload or {}).get("percent", 0), 0)

    def test_import_ajax_without_file_returns_json_error(self):
        response = self.client_http.post(
            "/ios/importacoes/nova/",
            {},
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("application/json", response["Content-Type"])
        payload = response.json()
        self.assertFalse(payload["ok"])
        self.assertIn("Selecione um arquivo", payload["message"])

    def test_import_ajax_status_endpoint_reports_processing(self):
        upload = SimpleUploadedFile(
            "PLANILHA DE IO UBS3 NUTRIEN - REM01 REV03.xlsx",
            self._build_single_rack_slot_block_workbook(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        with patch("core.views._spawn_io_import_job_processor"):
            with self.captureOnCommitCallbacks(execute=False):
                create_response = self.client_http.post(
                    "/ios/importacoes/nova/",
                    {
                        "arquivo": upload,
                    },
                    HTTP_X_REQUESTED_WITH="XMLHttpRequest",
                )

        payload = create_response.json()
        response = self.client_http.get(payload["status_url"], HTTP_X_REQUESTED_WITH="XMLHttpRequest")
        self.assertEqual(response.status_code, 200)
        self.assertIn("application/json", response["Content-Type"])
        payload = response.json()
        self.assertTrue(payload["ok"])
        job = IOImportJob.objects.latest("id")
        self.assertEqual(payload["job_id"], job.pk)
        self.assertTrue(payload["processing"])
        self.assertFalse(payload["complete"])
        self.assertEqual(payload["status"], IOImportJob.Status.UPLOADED)
        self.assertIn("progress", payload)
        self.assertEqual(payload["progress"]["stage"], "upload")
        self.assertIn("steps", payload["progress"])
        self.assertEqual(payload["progress"]["steps"]["upload"], "active")

    def test_background_command_invalid_xlsx_marks_job_failed(self):
        upload = SimpleUploadedFile(
            "bad.xlsx",
            b"this-is-not-a-real-xlsx",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        job = IOImportJob.objects.create(
            created_by=self.user,
            cliente=self.perfil,
            original_filename="bad.xlsx",
            source_file=upload,
            file_sha256=build_file_sha256(b"this-is-not-a-real-xlsx"),
            mode=IOImportJob.Mode.CREATE_RACK,
        )
        with self.assertRaises(Exception):
            call_command("process_io_import_job", str(job.pk))

        job.refresh_from_db()
        self.assertEqual(job.status, IOImportJob.Status.FAILED)
        self.assertTrue((job.progress_payload or {}).get("failed"))
        self.assertTrue(
            any(
                "planilha real" in warning.lower() or "segundo plano" in warning.lower()
                for warning in (job.warnings or [])
            )
        )

    def test_status_endpoint_returns_real_progress_snapshots(self):
        job = IOImportJob.objects.create(
            created_by=self.user,
            cliente=self.perfil,
            original_filename="lista.xlsx",
            source_file=SimpleUploadedFile("lista.xlsx", self._build_single_rack_slot_block_workbook()),
            file_sha256=build_file_sha256(b"abc"),
            mode=IOImportJob.Mode.CREATE_RACK,
            status=IOImportJob.Status.UPLOADED,
            progress_payload={
                "stage": "ai",
                "percent": 62,
                "title": "Correlacionando sinais",
                "message": "Guia REM01 consolidada. 24 linhas uteis acumuladas ate agora.",
                "current_sheet": "REM01",
                "sheets_total": 2,
                "sheets_processed": 1,
                "snapshots": [
                    {
                        "rack_key": "REM01",
                        "rack_name": "REM01",
                        "slots_count": 2,
                        "channels_count": 24,
                        "type_summary": ["DI 16", "DO 08"],
                        "sample_tags": ["DI_001", "DI_002"],
                    }
                ],
            },
        )
        response = self.client_http.get(f"/ios/importacoes/{job.pk}/status/", HTTP_X_REQUESTED_WITH="XMLHttpRequest")
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["ok"])
        self.assertEqual(payload["progress"]["stage"], "ai")
        self.assertEqual(payload["progress"]["percent"], 62)
        self.assertEqual(payload["progress"]["current_sheet"], "REM01")
        self.assertEqual(len(payload["progress"]["snapshots"]), 1)
        self.assertEqual(payload["progress"]["snapshots"][0]["rack_name"], "REM01")
        self.assertEqual(payload["progress"]["steps"]["upload"], "done")
        self.assertEqual(payload["progress"]["steps"]["ai"], "active")

    def test_import_ajax_spawn_exception_returns_json_error(self):
        upload = SimpleUploadedFile(
            "PLANILHA DE IO UBS3 NUTRIEN - REM01 REV03.xlsx",
            self._build_single_rack_slot_block_workbook(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        with patch("core.views._spawn_io_import_job_processor", side_effect=RuntimeError("spawn down")):
            with self.captureOnCommitCallbacks(execute=True):
                response = self.client_http.post(
                    "/ios/importacoes/nova/",
                    {
                        "arquivo": upload,
                    },
                    HTTP_X_REQUESTED_WITH="XMLHttpRequest",
                )

        self.assertEqual(response.status_code, 200)
        self.assertIn("application/json", response["Content-Type"])
        payload = response.json()
        self.assertTrue(payload["ok"])
        job = IOImportJob.objects.latest("id")
        self.assertEqual(job.status, IOImportJob.Status.FAILED)
        self.assertEqual(job.ai_status, IOImportJob.AIStatus.FAILED)
        self.assertIn("spawn down", job.ai_error)
        self.assertTrue((job.progress_payload or {}).get("failed"))

    def test_import_ajax_job_creation_exception_returns_json_error(self):
        upload = SimpleUploadedFile(
            "PLANILHA DE IO UBS3 NUTRIEN - REM01 REV03.xlsx",
            self._build_single_rack_slot_block_workbook(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        with patch("core.views.IOImportJob.objects.create", side_effect=RuntimeError("storage down")):
            response = self.client_http.post(
                "/ios/importacoes/nova/",
                {
                    "arquivo": upload,
                },
                HTTP_X_REQUESTED_WITH="XMLHttpRequest",
            )

        self.assertEqual(response.status_code, 500)
        self.assertIn("application/json", response["Content-Type"])
        payload = response.json()
        self.assertFalse(payload["ok"])
        self.assertIn("Falha interna", payload["message"])

    def test_status_endpoint_hides_technical_failure_from_non_dev(self):
        job = IOImportJob.objects.create(
            created_by=self.user,
            cliente=self.perfil,
            original_filename="lista.xlsx",
            source_file=SimpleUploadedFile("lista.xlsx", self._build_single_rack_slot_block_workbook()),
            file_sha256=build_file_sha256(b"abc"),
            mode=IOImportJob.Mode.CREATE_RACK,
            status=IOImportJob.Status.FAILED,
            warnings=["Falha tecnica: timeout apos 25s na chamada do agente."],
            ai_status=IOImportJob.AIStatus.FAILED,
        )
        with patch("core.views._is_dev_user", return_value=False):
            response = self.client_http.get(f"/ios/importacoes/{job.pk}/status/", HTTP_X_REQUESTED_WITH="XMLHttpRequest")

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["ok"])
        self.assertTrue(payload["failed"])
        self.assertNotIn("timeout", payload["message"].lower())
        self.assertEqual(payload["warnings_count"], 0)
        self.assertEqual(payload["ai_status"], "")

    def test_detail_hides_technical_warning_from_non_dev(self):
        job = IOImportJob.objects.create(
            created_by=self.user,
            cliente=self.perfil,
            original_filename="lista.xlsx",
            source_file=SimpleUploadedFile("lista.xlsx", self._build_single_rack_slot_block_workbook()),
            file_sha256=build_file_sha256(b"abc"),
            mode=IOImportJob.Mode.CREATE_RACK,
            status=IOImportJob.Status.FAILED,
            warnings=["Falha tecnica: timeout apos 25s na chamada do agente."],
            ai_status=IOImportJob.AIStatus.FAILED,
        )
        with patch("core.views._is_dev_user", return_value=False):
            response = self.client_http.get(f"/ios/importacoes/{job.pk}/")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Nao foi possivel concluir a analise da planilha")
        self.assertNotContains(response, "timeout apos 25s")
        self.assertNotContains(response, "Warnings")

    def test_import_admin_renders(self):
        admin_user = User.objects.create_superuser("io-admin", "io-admin@set.local", "123456")
        self.client_http.force_login(admin_user)
        response = self.client_http.get("/ios/importacoes/admin/")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Importacao de planilhas de IO")
        self.assertTrue(IOImportSettings.objects.exists())

    def test_import_xlsx_grouped_by_slot_builds_contextual_modules(self):
        upload = SimpleUploadedFile(
            "PLANILHA DE IO UBS3 NUTRIEN - REM01 REV03.xlsx",
            self._build_single_rack_slot_block_workbook(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response = self.client_http.post(
            "/ios/importacoes/nova/",
            {
                "arquivo": upload,
                "requested_rack_name": "Rack REM01",
                "requested_local": str(self.local.id),
                "requested_grupo": str(self.grupo.id),
                "requested_planta_code": self.planta.codigo,
            },
        )
        self.assertEqual(response.status_code, 302)

        job = IOImportJob.objects.latest("id")
        self.assertEqual(job.status, IOImportJob.Status.REVIEW)
        self.assertEqual(job.file_format, IOImportJob.FileFormat.XLSX)
        self.assertEqual(job.sheet_name, "REM01 6ES7 155-6AU01-0BN0")
        self.assertEqual(job.rows_parsed, 4)

        extracted_rows = job.extracted_payload["rows"]
        self.assertEqual(extracted_rows[0]["slot_index"], 1)
        self.assertEqual(extracted_rows[0]["channel_index"], 1)
        self.assertEqual(extracted_rows[0]["type"], "DI")
        self.assertEqual(extracted_rows[2]["slot_index"], 2)
        self.assertEqual(extracted_rows[2]["channel_index"], 1)
        self.assertEqual(extracted_rows[2]["type"], "DO")

        proposal = job.proposal_payload
        self.assertEqual(proposal["summary"]["modules"], 2)
        self.assertEqual(proposal["summary"]["with_conflicts"], 0)
        self.assertEqual([module["slot_index"] for module in proposal["modules"]], [1, 2])
        self.assertEqual([module["module_type"] for module in proposal["modules"]], ["DI", "DO"])

    def test_import_transport_timeout_do_agente_interrompe_pipeline_ai_first(self):
        settings_obj = IOImportSettings.load()
        settings_obj.enabled = True
        settings_obj.api_key = "test-key"
        settings_obj.save()

        upload = SimpleUploadedFile(
            "PLANILHA DE IO UBS3 NUTRIEN - REM01 REV03.xlsx",
            self._build_single_rack_slot_block_workbook(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        with patch("core.services.io_import.urlrequest.urlopen", side_effect=TimeoutError("timed out")):
            response = self.client_http.post(
                "/ios/importacoes/nova/",
                {
                    "arquivo": upload,
                    "requested_rack_name": "Rack REM01",
                    "requested_local": str(self.local.id),
                    "requested_grupo": str(self.grupo.id),
                    "requested_planta_code": self.planta.codigo,
                },
            )

        self.assertEqual(response.status_code, 302)

        job = IOImportJob.objects.latest("id")
        self.assertEqual(job.status, IOImportJob.Status.FAILED)
        self.assertEqual(job.ai_status, IOImportJob.AIStatus.FAILED)
        self.assertIn("timeout", job.ai_error.lower())
        self.assertTrue(any("timeout" in warning.lower() for warning in (job.warnings or [])))

    def test_call_openai_responses_polls_background_response_until_completed(self):
        settings_obj = IOImportSettings.load()
        settings_obj.enabled = True
        settings_obj.api_key = "test-key"
        settings_obj.save()

        requested = []
        queued_payload = {"id": "resp_123", "status": "queued", "output": []}
        running_payload = {"id": "resp_123", "status": "in_progress", "output": []}
        completed_payload = {
            "id": "resp_123",
            "status": "completed",
            "output": [
                {
                    "content": [
                        {
                            "type": "output_text",
                            "text": json.dumps({"result": "ok"}),
                        }
                    ]
                }
            ],
        }
        responses = [queued_payload, running_payload, completed_payload]

        class _FakeHTTPResponse:
            def __init__(self, payload):
                self.payload = payload

            def read(self):
                return json.dumps(self.payload).encode("utf-8")

            def __enter__(self):
                return self

            def __exit__(self, exc_type, exc, tb):
                return False

        def fake_urlopen(request, timeout=None):
            requested.append(
                {
                    "url": request.full_url,
                    "method": request.get_method(),
                    "timeout": timeout,
                    "data": request.data,
                }
            )
            return _FakeHTTPResponse(responses[len(requested) - 1])

        with patch("core.services.io_import.urlrequest.urlopen", side_effect=fake_urlopen), patch(
            "core.services.io_import.time.sleep"
        ):
            result = _call_openai_responses(
                settings_obj=settings_obj,
                schema_name="io_sheet_semantic_analysis",
                schema={"type": "object", "properties": {"result": {"type": "string"}}, "required": ["result"], "additionalProperties": False},
                system_prompt="system",
                user_prompt="{}",
                request_timeout_seconds=300,
            )

        self.assertEqual(result, {"result": "ok"})
        self.assertEqual([item["method"] for item in requested], ["POST", "GET", "GET"])
        post_payload = json.loads(requested[0]["data"].decode("utf-8"))
        self.assertTrue(post_payload["background"])
        self.assertTrue(post_payload["store"])
        self.assertEqual(post_payload["metadata"]["schema_name"], "io_sheet_semantic_analysis")

    def test_call_openai_responses_does_not_abort_only_because_total_wait_is_long(self):
        settings_obj = IOImportSettings.load()
        settings_obj.enabled = True
        settings_obj.api_key = "test-key"
        settings_obj.save()

        requested = []
        queued_payload = {"id": "resp_123", "status": "queued", "output": []}
        running_payload = {"id": "resp_123", "status": "in_progress", "output": []}
        completed_payload = {
            "id": "resp_123",
            "status": "completed",
            "output": [
                {
                    "content": [
                        {
                            "type": "output_text",
                            "text": json.dumps({"result": "ok"}),
                        }
                    ]
                }
            ],
        }
        responses = [queued_payload, running_payload, completed_payload]

        class _FakeHTTPResponse:
            def __init__(self, payload):
                self.payload = payload

            def read(self):
                return json.dumps(self.payload).encode("utf-8")

            def __enter__(self):
                return self

            def __exit__(self, exc_type, exc, tb):
                return False

        def fake_urlopen(request, timeout=None):
            requested.append(
                {
                    "url": request.full_url,
                    "method": request.get_method(),
                    "timeout": timeout,
                }
            )
            return _FakeHTTPResponse(responses[len(requested) - 1])

        with patch("core.services.io_import.urlrequest.urlopen", side_effect=fake_urlopen), patch(
            "core.services.io_import.time.sleep"
        ), patch("core.services.io_import.time.monotonic", side_effect=[0.0, 600.0, 1800.0]):
            result = _call_openai_responses(
                settings_obj=settings_obj,
                schema_name="io_sheet_semantic_analysis",
                schema={"type": "object", "properties": {"result": {"type": "string"}}, "required": ["result"], "additionalProperties": False},
                system_prompt="system",
                user_prompt="{}",
                request_timeout_seconds=30,
            )

        self.assertEqual(result, {"result": "ok"})
        self.assertEqual([item["method"] for item in requested], ["POST", "GET", "GET"])

    def test_call_openai_responses_retries_polling_same_response_instead_of_resubmitting(self):
        settings_obj = IOImportSettings.load()
        settings_obj.enabled = True
        settings_obj.api_key = "test-key"
        settings_obj.save()

        requested = []

        class _FakeHTTPResponse:
            def __init__(self, payload):
                self.payload = payload

            def read(self):
                return json.dumps(self.payload).encode("utf-8")

            def __enter__(self):
                return self

            def __exit__(self, exc_type, exc, tb):
                return False

        def fake_urlopen(request, timeout=None):
            requested.append({"url": request.full_url, "method": request.get_method()})
            if len(requested) == 1:
                return _FakeHTTPResponse({"id": "resp_retry", "status": "queued", "output": []})
            if len(requested) == 2:
                raise TimeoutError("timed out")
            return _FakeHTTPResponse(
                {
                    "id": "resp_retry",
                    "status": "completed",
                    "output": [{"content": [{"type": "output_text", "text": json.dumps({"result": "ok"})}]}],
                }
            )

        with patch("core.services.io_import.urlrequest.urlopen", side_effect=fake_urlopen), patch(
            "core.services.io_import.time.sleep"
        ):
            result = _call_openai_responses(
                settings_obj=settings_obj,
                schema_name="io_sheet_semantic_analysis",
                schema={"type": "object", "properties": {"result": {"type": "string"}}, "required": ["result"], "additionalProperties": False},
                system_prompt="system",
                user_prompt="{}",
                request_timeout_seconds=300,
            )

        self.assertEqual(result, {"result": "ok"})
        self.assertEqual(sum(1 for item in requested if item["method"] == "POST"), 1)
        self.assertEqual(sum(1 for item in requested if item["method"] == "GET"), 2)
        self.assertTrue(all(item["url"].endswith("/responses/resp_retry") for item in requested[1:]))

    def test_reprocess_import_job_ai_first_can_drive_semantic_parse_from_raw_rows(self):
        settings_obj = IOImportSettings.load()
        settings_obj.enabled = True
        settings_obj.api_key = "test-key"
        settings_obj.save()

        raw_bytes = self._build_ai_first_obscure_workbook()
        job = IOImportJob.objects.create(
            created_by=self.user,
            cliente=self.perfil,
            original_filename="raw-export.xlsx",
            source_file=SimpleUploadedFile(
                "raw-export.xlsx",
                raw_bytes,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
            file_sha256=build_file_sha256(raw_bytes),
            mode=IOImportJob.Mode.CREATE_RACK,
        )
        module_catalog = serialize_module_catalog(ModuloIO.objects.select_related("tipo_base").all())
        parsed = parse_workbook(raw_bytes=raw_bytes, original_filename="raw-export.xlsx")[0]
        with self.assertRaises(Exception):
            normalize_rows(parsed=parsed, module_catalog=module_catalog, ai_result=None)

        sheet_payload = {
            "sheet_role": "data",
            "skip_sheet": False,
            "layout_hint": "tabular",
            "rack_name": "PNL-PRC-01 - Rack 01",
            "column_map": {
                "panel": {"header": "", "confidence": 0, "mode": "fill"},
                "rack": {"header": "", "confidence": 0, "mode": "fill"},
                "slot": {"header": "", "confidence": 0, "mode": "fill"},
                "module_model": {"header": "", "confidence": 0, "mode": "fill"},
                "channel": {"header": "", "confidence": 0, "mode": "fill"},
                "tag": {"header": "", "confidence": 0, "mode": "fill"},
                "description": {"header": "", "confidence": 0, "mode": "fill"},
                "type": {"header": "", "confidence": 0, "mode": "fill"},
            },
            "logical_points": [
                {
                    "source_row": 3,
                    "panel": "PNL-PRC-01",
                    "rack": "Rack 01",
                    "slot": "02",
                    "module_model": "DI-16",
                    "channel": "DI.00",
                    "tag": "P101_RUN_FB",
                    "description": "Feedback bomba 101",
                    "type": "DI",
                    "confidence": 97,
                },
                {
                    "source_row": 4,
                    "panel": "PNL-PRC-01",
                    "rack": "Rack 01",
                    "slot": "03",
                    "module_model": "DO-16",
                    "channel": "DQ.01",
                    "tag": "P101_START_CMD",
                    "description": "Comando partida bomba 101",
                    "type": "DO",
                    "confidence": 97,
                },
            ],
            "row_hints": [
                {
                    "source_row": 1,
                    "row_kind": "noise",
                    "panel": "",
                    "rack": "",
                    "slot": "",
                    "module_model": "",
                    "channel": "",
                    "tag": "",
                    "description": "",
                    "type": "",
                    "confidence": 90,
                },
                {
                    "source_row": 2,
                    "row_kind": "subheader",
                    "panel": "",
                    "rack": "",
                    "slot": "",
                    "module_model": "",
                    "channel": "",
                    "tag": "",
                    "description": "",
                    "type": "",
                    "confidence": 95,
                },
            ],
            "warnings": [],
            "notes": "ok",
        }

        with patch("core.services.io_import._call_openai_responses", side_effect=[sheet_payload]):
            result = reprocess_import_job(job=job, module_catalog=module_catalog, settings_obj=settings_obj)

        self.assertEqual(result["ai_status"], IOImportJob.AIStatus.SUCCESS)
        self.assertEqual(result["rows_parsed"], 2)
        self.assertEqual(result["proposal"]["summary"]["racks"], 1)
        self.assertEqual(result["proposal"]["summary"]["modules"], 2)
        self.assertEqual(result["proposal"]["summary"]["with_conflicts"], 0)
        self.assertEqual(result["proposal"]["rack"]["name"], "PNL-PRC-01 - Rack 01")
        self.assertIn("workbook", result["ai_payload"])
        self.assertEqual(result["normalized_rows"][0]["field_sources"]["tag"], "ai")
        self.assertEqual([module["module_type"] for module in result["proposal"]["modules"]], ["DI", "DO"])

    def test_reprocess_import_job_single_sheet_fast_path_uses_sparse_ai_guidance(self):
        settings_obj = IOImportSettings.load()
        settings_obj.enabled = True
        settings_obj.api_key = "test-key"
        settings_obj.save()

        raw_bytes = self._build_ai_first_obscure_workbook()
        job = IOImportJob.objects.create(
            created_by=self.user,
            cliente=self.perfil,
            original_filename="raw-export.xlsx",
            source_file=SimpleUploadedFile(
                "raw-export.xlsx",
                raw_bytes,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
            file_sha256=build_file_sha256(raw_bytes),
            mode=IOImportJob.Mode.CREATE_RACK,
        )
        module_catalog = serialize_module_catalog(ModuloIO.objects.select_related("tipo_base").all())

        sparse_sheet_payload = {
            "sheet_role": "data",
            "skip_sheet": False,
            "layout_hint": "tabular",
            "rack_name": "PNL-PRC-01 - Rack 01",
            "column_map": {
                "panel": {"header": "", "confidence": 0, "mode": "fill"},
                "rack": {"header": "", "confidence": 0, "mode": "fill"},
                "slot": {"header": "", "confidence": 0, "mode": "fill"},
                "module_model": {"header": "", "confidence": 0, "mode": "fill"},
                "channel": {"header": "", "confidence": 0, "mode": "fill"},
                "location": {"header": "HARDWAREPATH", "confidence": 96, "mode": "fill"},
                "tag": {"header": "REFX", "confidence": 96, "mode": "fill"},
                "description": {"header": "NARRATIVE", "confidence": 96, "mode": "fill"},
                "type": {"header": "SIGCLASS", "confidence": 96, "mode": "fill"},
            },
            "row_hints": [],
            "warnings": [],
            "notes": "sparse-guidance",
        }

        with patch("core.services.io_import._call_openai_responses", side_effect=[sparse_sheet_payload]) as mocked:
            result = reprocess_import_job(job=job, module_catalog=module_catalog, settings_obj=settings_obj)

        self.assertEqual(mocked.call_count, 1)
        self.assertEqual(result["ai_status"], IOImportJob.AIStatus.SUCCESS)
        self.assertEqual(result["rows_parsed"], 2)
        self.assertEqual(result["proposal"]["summary"]["racks"], 1)
        self.assertEqual(result["proposal"]["rack"]["name"], "PNL-PRC-01 - Rack 01")
        self.assertEqual([module["module_type"] for module in result["proposal"]["modules"]], ["DI", "DO"])

    def test_reprocess_import_job_reuses_ai_cache_for_identical_file(self):
        settings_obj = IOImportSettings.load()
        settings_obj.enabled = True
        settings_obj.api_key = "test-key"
        settings_obj.save()

        raw_bytes = self._build_ai_first_obscure_workbook()
        job_one = IOImportJob.objects.create(
            created_by=self.user,
            cliente=self.perfil,
            original_filename="raw-export.xlsx",
            source_file=SimpleUploadedFile(
                "raw-export.xlsx",
                raw_bytes,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
            file_sha256=build_file_sha256(raw_bytes),
            mode=IOImportJob.Mode.CREATE_RACK,
        )
        job_two = IOImportJob.objects.create(
            created_by=self.user,
            cliente=self.perfil,
            original_filename="raw-export.xlsx",
            source_file=SimpleUploadedFile(
                "raw-export.xlsx",
                raw_bytes,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
            file_sha256=build_file_sha256(raw_bytes),
            mode=IOImportJob.Mode.CREATE_RACK,
        )
        module_catalog = serialize_module_catalog(ModuloIO.objects.select_related("tipo_base").all())
        sheet_payload = {
            "sheet_role": "data",
            "skip_sheet": False,
            "layout_hint": "tabular",
            "rack_name": "PNL-PRC-01 - Rack 01",
            "column_map": {
                "panel": {"header": "", "confidence": 0, "mode": "fill"},
                "rack": {"header": "", "confidence": 0, "mode": "fill"},
                "slot": {"header": "", "confidence": 0, "mode": "fill"},
                "module_model": {"header": "", "confidence": 0, "mode": "fill"},
                "channel": {"header": "", "confidence": 0, "mode": "fill"},
                "tag": {"header": "", "confidence": 0, "mode": "fill"},
                "description": {"header": "", "confidence": 0, "mode": "fill"},
                "type": {"header": "", "confidence": 0, "mode": "fill"},
            },
            "logical_points": [
                {
                    "source_row": 3,
                    "panel": "PNL-PRC-01",
                    "rack": "Rack 01",
                    "slot": "02",
                    "module_model": "DI-16",
                    "channel": "DI.00",
                    "tag": "P101_RUN_FB",
                    "description": "Feedback bomba 101",
                    "type": "DI",
                    "confidence": 97,
                },
                {
                    "source_row": 4,
                    "panel": "PNL-PRC-01",
                    "rack": "Rack 01",
                    "slot": "03",
                    "module_model": "DO-16",
                    "channel": "DQ.01",
                    "tag": "P101_START_CMD",
                    "description": "Comando partida bomba 101",
                    "type": "DO",
                    "confidence": 97,
                },
            ],
            "row_hints": [],
            "warnings": [],
            "notes": "ok",
        }

        with patch("core.services.io_import._call_openai_responses", side_effect=[sheet_payload]) as mocked:
            result_one = reprocess_import_job(job=job_one, module_catalog=module_catalog, settings_obj=settings_obj)
            result_two = reprocess_import_job(job=job_two, module_catalog=module_catalog, settings_obj=settings_obj)

        self.assertEqual(mocked.call_count, 1)
        self.assertEqual(result_one["rows_parsed"], 2)
        self.assertEqual(result_two["rows_parsed"], 2)
        self.assertEqual(IOImportAICache.objects.count(), 1)
        self.assertTrue(IOImportAICache.objects.filter(stage=IOImportAICache.Stage.SHEET).exists())

    def test_reprocess_import_job_does_not_reuse_ai_cache_for_updated_file(self):
        settings_obj = IOImportSettings.load()
        settings_obj.enabled = True
        settings_obj.api_key = "test-key"
        settings_obj.save()

        raw_bytes_one = self._build_ai_first_obscure_workbook()
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Raw Export"
        for row in [
            ["Projeto X", "", "", ""],
            ["REFX", "NARRATIVE", "HARDWAREPATH", "SIGCLASS"],
            ["P201_RUN_FB", "Feedback bomba 201", "PNL-PRC-02 / Rack 1 / Slot 02 / DI.00", "DI"],
            ["P201_START_CMD", "Comando partida bomba 201", "PNL-PRC-02 / Rack 1 / Slot 03 / DQ.01", "DO"],
        ]:
            sheet.append(row)
        payload = BytesIO()
        workbook.save(payload)
        raw_bytes_two = payload.getvalue()

        job_one = IOImportJob.objects.create(
            created_by=self.user,
            cliente=self.perfil,
            original_filename="raw-export.xlsx",
            source_file=SimpleUploadedFile("raw-export.xlsx", raw_bytes_one, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
            file_sha256=build_file_sha256(raw_bytes_one),
            mode=IOImportJob.Mode.CREATE_RACK,
        )
        job_two = IOImportJob.objects.create(
            created_by=self.user,
            cliente=self.perfil,
            original_filename="raw-export.xlsx",
            source_file=SimpleUploadedFile("raw-export.xlsx", raw_bytes_two, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
            file_sha256=build_file_sha256(raw_bytes_two),
            mode=IOImportJob.Mode.CREATE_RACK,
        )
        module_catalog = serialize_module_catalog(ModuloIO.objects.select_related("tipo_base").all())
        sheet_payload_one = {
            "sheet_role": "data",
            "skip_sheet": False,
            "layout_hint": "tabular",
            "rack_name": "PNL-PRC-01 - Rack 01",
            "column_map": {key: {"header": "", "confidence": 0, "mode": "fill"} for key in ("panel", "rack", "slot", "module_model", "channel", "tag", "description", "type")},
            "logical_points": [
                {"source_row": 3, "panel": "PNL-PRC-01", "rack": "Rack 01", "slot": "02", "module_model": "DI-16", "channel": "DI.00", "tag": "P101_RUN_FB", "description": "Feedback bomba 101", "type": "DI", "confidence": 97},
                {"source_row": 4, "panel": "PNL-PRC-01", "rack": "Rack 01", "slot": "03", "module_model": "DO-16", "channel": "DQ.01", "tag": "P101_START_CMD", "description": "Comando partida bomba 101", "type": "DO", "confidence": 97},
            ],
            "row_hints": [],
            "warnings": [],
            "notes": "ok",
        }
        sheet_payload_two = {
            "sheet_role": "data",
            "skip_sheet": False,
            "layout_hint": "tabular",
            "rack_name": "PNL-PRC-02 - Rack 01",
            "column_map": {key: {"header": "", "confidence": 0, "mode": "fill"} for key in ("panel", "rack", "slot", "module_model", "channel", "tag", "description", "type")},
            "logical_points": [
                {"source_row": 3, "panel": "PNL-PRC-02", "rack": "Rack 01", "slot": "02", "module_model": "DI-16", "channel": "DI.00", "tag": "P201_RUN_FB", "description": "Feedback bomba 201", "type": "DI", "confidence": 97},
                {"source_row": 4, "panel": "PNL-PRC-02", "rack": "Rack 01", "slot": "03", "module_model": "DO-16", "channel": "DQ.01", "tag": "P201_START_CMD", "description": "Comando partida bomba 201", "type": "DO", "confidence": 97},
            ],
            "row_hints": [],
            "warnings": [],
            "notes": "ok",
        }

        with patch(
            "core.services.io_import._call_openai_responses",
            side_effect=[sheet_payload_one, sheet_payload_two],
        ) as mocked:
            result_one = reprocess_import_job(job=job_one, module_catalog=module_catalog, settings_obj=settings_obj)
            result_two = reprocess_import_job(job=job_two, module_catalog=module_catalog, settings_obj=settings_obj)

        self.assertEqual(mocked.call_count, 2)
        self.assertEqual(result_one["proposal"]["rack"]["name"], "PNL-PRC-01 - Rack 01")
        self.assertEqual(result_two["proposal"]["rack"]["name"], "PNL-PRC-02 - Rack 01")

    def test_run_ai_analysis_sends_compact_contextual_module_catalog(self):
        settings_obj = IOImportSettings.load()
        settings_obj.enabled = True
        settings_obj.api_key = "test-key"
        settings_obj.save()

        parsed = ParsedSpreadsheet(
            file_format="xlsx",
            sheet_name="REM01",
            header_row_index=0,
            rows_total=3,
            headers=["TAG", "DESC", "CH", "POS"],
            raw_rows=[
                ["TAG", "DESC", "CH", "POS"],
                ["P101_RUN_FB", "Feedback bomba 101", "DI.00", "Slot 02"],
                ["P101_START_CMD", "Comando partida bomba 101", "DQ.01", "Slot 03"],
            ],
            column_map={},
            warnings=[],
            layout="tabular",
        )
        module_catalog = []
        for index in range(1, 31):
            module_catalog.append(
                {
                    "id": index,
                    "modelo": f"DI-{index}",
                    "marca": "SIMATIC",
                    "tipo": "DI",
                    "quantidade_canais": 16 if index % 2 else 32,
                }
            )
        for index in range(31, 61):
            module_catalog.append(
                {
                    "id": index,
                    "modelo": f"DO-{index}",
                    "marca": "SIMATIC",
                    "tipo": "DO",
                    "quantidade_canais": 16 if index % 2 else 32,
                }
            )

        captured_payloads = []

        def fake_call(**kwargs):
            captured_payloads.append(json.loads(kwargs["user_prompt"]))
            return {
                "sheet_role": "data",
                "skip_sheet": False,
                "layout_hint": "tabular",
                "rack_name": "REM01",
                "column_map": {key: {"header": "", "confidence": 0, "mode": "fill"} for key in ("panel", "rack", "slot", "module_model", "channel", "tag", "description", "type")},
                "logical_points": [],
                "row_hints": [],
                "warnings": [],
                "notes": "ok",
            }

        with patch("core.services.io_import._call_openai_responses", side_effect=fake_call):
            run_ai_analysis(
                settings_obj=settings_obj,
                parsed=parsed,
                normalized_rows=[],
                module_catalog=module_catalog,
                file_sha256="abc123",
                workbook_plan={"layout_hint": "tabular"},
                request_timeout_seconds=10,
                total_sheets=4,
            )

        self.assertEqual(len(captured_payloads), 1)
        payload = captured_payloads[0]
        self.assertLess(len(payload["module_catalog"]), len(module_catalog))
        self.assertLessEqual(len(payload["module_catalog"]), 18)
        self.assertIn("module_catalog_stats", payload)
        self.assertTrue(all("id" not in item for item in payload["module_catalog"]))

    def test_normalize_rows_ai_can_enrich_row_structure_with_confidence(self):
        parsed = ParsedSpreadsheet(
            file_format="xlsx",
            sheet_name="Painel IA",
            header_row_index=0,
            rows_total=2,
            headers=["SIG", "OBS", "POS", "CH", "CLASS"],
            raw_rows=[
                ["SIG", "OBS", "POS", "CH", "CLASS"],
                ["MTR_FB_01", "Feedback do motor", "Slot 07", "DI.03", "AO"],
            ],
            column_map={
                "tag": {"index": 0, "header": "SIG", "confidence": 80, "source": "heuristic"},
                "description": {"index": 1, "header": "OBS", "confidence": 80, "source": "heuristic"},
            },
            warnings=[],
            layout="tabular",
        )
        module_catalog = serialize_module_catalog(ModuloIO.objects.select_related("tipo_base").all())
        ai_result = {
            "rack_name": "REM-IA-01",
            "column_map": {
                "panel": {"header": "", "confidence": 0, "mode": "fill"},
                "rack": {"header": "", "confidence": 0, "mode": "fill"},
                "slot": {"header": "POS", "confidence": 95, "mode": "fill"},
                "module_model": {"header": "", "confidence": 0, "mode": "fill"},
                "channel": {"header": "CH", "confidence": 95, "mode": "fill"},
                "tag": {"header": "SIG", "confidence": 90, "mode": "fill"},
                "description": {"header": "OBS", "confidence": 90, "mode": "fill"},
                "type": {"header": "CLASS", "confidence": 95, "mode": "override"},
            },
            "row_hints": [
                {
                    "source_row": 2,
                    "row_kind": "data",
                    "panel": "PNL-IA-01",
                    "rack": "Rack 02",
                    "slot": "Slot 07",
                    "module_model": "DI-16",
                    "channel": "DI.03",
                    "tag": "MTR_FB_01",
                    "description": "Feedback do motor",
                    "type": "DI",
                    "confidence": 96,
                }
            ],
            "warnings": [],
            "notes": "ok",
        }

        rows, active_map, warnings = normalize_rows(parsed=parsed, module_catalog=module_catalog, ai_result=ai_result)

        self.assertEqual(len(rows), 1)
        self.assertEqual(active_map["slot"]["header"], "POS")
        self.assertEqual(active_map["channel"]["header"], "CH")
        self.assertEqual(rows[0]["panel_raw"], "PNL-IA-01")
        self.assertEqual(rows[0]["rack_raw"], "Rack 02")
        self.assertEqual(rows[0]["slot_index"], 7)
        self.assertEqual(rows[0]["channel_index"], 4)
        self.assertEqual(rows[0]["module_raw"], "DI-16")
        self.assertEqual(rows[0]["type"], "DI")
        self.assertEqual(rows[0]["field_sources"]["type"], "ai")
        self.assertFalse(warnings)

    def test_normalize_rows_structural_validation_adjusts_type_to_catalog(self):
        parsed = ParsedSpreadsheet(
            file_format="xlsx",
            sheet_name="Painel Catalogo",
            header_row_index=0,
            rows_total=2,
            headers=["TAG", "DESC", "MODULO", "SLOT", "CHANNEL"],
            raw_rows=[
                ["TAG", "DESC", "MODULO", "SLOT", "CHANNEL"],
                ["XIC_001", "Contato auxiliar", "DI-16", "1", "1"],
            ],
            column_map={
                "tag": {"index": 0, "header": "TAG", "confidence": 90, "source": "heuristic"},
                "description": {"index": 1, "header": "DESC", "confidence": 90, "source": "heuristic"},
                "module_model": {"index": 2, "header": "MODULO", "confidence": 90, "source": "heuristic"},
                "slot": {"index": 3, "header": "SLOT", "confidence": 90, "source": "heuristic"},
                "channel": {"index": 4, "header": "CHANNEL", "confidence": 90, "source": "heuristic"},
            },
            warnings=[],
            layout="tabular",
        )
        module_catalog = serialize_module_catalog(ModuloIO.objects.select_related("tipo_base").all())
        ai_result = {
            "rack_name": "REM-IA-02",
            "column_map": {
                "panel": {"header": "", "confidence": 0, "mode": "fill"},
                "rack": {"header": "", "confidence": 0, "mode": "fill"},
                "slot": {"header": "SLOT", "confidence": 95, "mode": "fill"},
                "module_model": {"header": "MODULO", "confidence": 95, "mode": "fill"},
                "channel": {"header": "CHANNEL", "confidence": 95, "mode": "fill"},
                "tag": {"header": "TAG", "confidence": 95, "mode": "fill"},
                "description": {"header": "DESC", "confidence": 95, "mode": "fill"},
                "type": {"header": "", "confidence": 0, "mode": "fill"},
            },
            "row_hints": [
                {
                    "source_row": 2,
                    "row_kind": "data",
                    "panel": "",
                    "rack": "",
                    "slot": "1",
                    "module_model": "DI-16",
                    "channel": "1",
                    "tag": "XIC_001",
                    "description": "Contato auxiliar",
                    "type": "AO",
                    "confidence": 98,
                }
            ],
            "warnings": [],
            "notes": "conflict",
        }

        rows, _, warnings = normalize_rows(parsed=parsed, module_catalog=module_catalog, ai_result=ai_result)

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["module_raw"], "DI-16")
        self.assertEqual(rows[0]["type"], "DI")
        self.assertIn("tipo_ajustado_pelo_catalogo", rows[0]["issues"])
        self.assertTrue(any("conflito com o modulo" in warning.lower() for warning in warnings))

    def test_import_multisheet_builds_multi_rack_proposal(self):
        upload = SimpleUploadedFile(
            "PLANILHA DE IO UBS3 NUTRIEN - REM01 REV03.xlsx",
            self._build_slot_block_workbook(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response = self.client_http.post(
            "/ios/importacoes/nova/",
            {
                "arquivo": upload,
                "requested_local": str(self.local.id),
                "requested_grupo": str(self.grupo.id),
                "requested_planta_code": self.planta.codigo,
            },
        )
        self.assertEqual(response.status_code, 302)

        job = IOImportJob.objects.latest("id")
        self.assertEqual(job.status, IOImportJob.Status.REVIEW)
        self.assertEqual(job.proposal_payload["summary"]["racks"], 2)
        self.assertEqual(job.proposal_payload["summary"]["modules"], 3)
        self.assertEqual(sorted(rack["name"] for rack in job.proposal_payload["racks"]), ["REM01", "UBS3"])
        self.assertEqual(job.rows_parsed, 5)
        self.assertEqual(len(job.extracted_payload["sheets"]), 2)
        detail_response = self.client_http.get(f"/ios/importacoes/{job.pk}/")
        self.assertEqual(detail_response.status_code, 200)
        self.assertContains(detail_response, 'data-import-preview-root', html=False)
        self.assertContains(detail_response, "REM01")
        self.assertContains(detail_response, "UBS3")

    def test_import_single_sheet_can_split_multi_rack_sections(self):
        upload = SimpleUploadedFile(
            "io-consolidado.xlsx",
            self._build_single_sheet_multi_rack_workbook(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response = self.client_http.post(
            "/ios/importacoes/nova/",
            {
                "arquivo": upload,
                "requested_local": str(self.local.id),
                "requested_grupo": str(self.grupo.id),
                "requested_planta_code": self.planta.codigo,
            },
        )
        self.assertEqual(response.status_code, 302)

        job = IOImportJob.objects.latest("id")
        self.assertEqual(job.status, IOImportJob.Status.REVIEW)
        self.assertEqual(job.proposal_payload["summary"]["racks"], 2)
        self.assertEqual(sorted(rack["name"] for rack in job.proposal_payload["racks"]), ["REM01", "UBS3"])
        self.assertEqual(sorted({row["resolved_rack_name"] for row in job.extracted_payload["rows"]}), ["REM01", "UBS3"])

    def test_slot_block_sheet_ignores_internal_point_titles_as_new_racks(self):
        upload = SimpleUploadedFile(
            "aentr.xlsx",
            self._build_slot_block_with_internal_point_titles_workbook(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response = self.client_http.post(
            "/ios/importacoes/nova/",
            {
                "arquivo": upload,
                "requested_local": str(self.local.id),
                "requested_grupo": str(self.grupo.id),
                "requested_planta_code": self.planta.codigo,
            },
        )
        self.assertEqual(response.status_code, 302)

        job = IOImportJob.objects.latest("id")
        self.assertEqual(job.status, IOImportJob.Status.REVIEW)
        self.assertEqual(job.proposal_payload["summary"]["racks"], 1)
        self.assertEqual(job.proposal_payload["racks"][0]["name"].upper(), "AENTR")
        rack_names = [rack["name"] for rack in job.proposal_payload["racks"]]
        self.assertNotIn("CPU CONTROLOGIX 1756-L73 192.100.3.1", rack_names)
        self.assertNotIn("AO2-M21RE5-CAMPO", rack_names)
        self.assertNotIn("AO5-V20MV3-OUT", rack_names)

    def test_build_import_proposal_merges_spurious_interface_alias_into_dominant_rack(self):
        module_catalog = serialize_module_catalog(ModuloIO.objects.select_related("tipo_base").all())
        normalized_rows = []
        for channel_index in range(1, 5):
            normalized_rows.append(
                {
                    "source_sheet": "UBS3-ET200SP",
                    "source_row": channel_index,
                    "panel_raw": "",
                    "rack_raw": "UBS3-ET200SP",
                    "slot_raw": "SLOT 01",
                    "slot_index": 1,
                    "module_raw": "SIMATIC ET 200SP IM 155-6PN - 64 I/O",
                    "channel_raw": f"DI.{channel_index - 1}",
                    "channel_index": channel_index,
                    "location_raw": "",
                    "tag": f"HEAD_{channel_index:02d}",
                    "description": f"Cabeca {channel_index:02d}",
                    "type": "DI",
                    "issues": [],
                }
            )
        for channel_index in range(1, 17):
            normalized_rows.append(
                {
                    "source_sheet": "UBS3-ET200SP",
                    "source_row": 100 + channel_index,
                    "panel_raw": "",
                    "rack_raw": "",
                    "slot_raw": "SLOT 02",
                    "slot_index": 2,
                    "module_raw": "6ES7131-6BH01-0BA0",
                    "channel_raw": f"DI.{channel_index - 1}",
                    "channel_index": channel_index,
                    "location_raw": "",
                    "tag": f"UBS3_FB_{channel_index:02d}",
                    "description": f"Feedback UBS3 {channel_index:02d}",
                    "type": "DI",
                    "issues": [],
                }
            )

        proposal = build_import_proposal(
            original_filename="PLANILHA DE IO UBS3 NUTRIEN 05042025.xlsx",
            normalized_rows=normalized_rows,
            module_catalog=module_catalog,
        )

        self.assertEqual(proposal["summary"]["racks"], 1)
        self.assertEqual(proposal["rack"]["name"], "UBS3")
        self.assertEqual(len(proposal["racks"][0]["modules"]), 2)
        self.assertEqual(
            {row["resolved_rack_name"] for row in normalized_rows},
            {"UBS3"},
        )

    def test_build_import_proposal_merges_tiny_sheet_alias_group_that_duplicates_existing_module(self):
        module_catalog = serialize_module_catalog(ModuloIO.objects.select_related("tipo_base").all())
        normalized_rows = []
        for channel_index in range(1, 3):
            normalized_rows.append(
                {
                    "source_sheet": "UBS3-ET200SP",
                    "source_row": channel_index,
                    "panel_raw": "",
                    "rack_raw": "UBS3-ET200SP",
                    "slot_raw": "SLOT 02",
                    "slot_index": 2,
                    "module_raw": "6ES7131-6BH01-0BA0 (16 DI)",
                    "channel_raw": f"DI.{channel_index - 1}",
                    "channel_index": channel_index,
                    "location_raw": "",
                    "tag": f"FB_ELE{channel_index:02d}",
                    "description": "",
                    "type": "DI",
                    "issues": [],
                }
            )
        for channel_index in range(1, 17):
            normalized_rows.append(
                {
                    "source_sheet": "UBS3-ET200SP",
                    "source_row": 100 + channel_index,
                    "panel_raw": "",
                    "rack_raw": "",
                    "slot_raw": "SLOT 02",
                    "slot_index": 2,
                    "module_raw": "6ES7131-6BH01-0BA0 (16 DI)",
                    "channel_raw": f"DI.{channel_index - 1}",
                    "channel_index": channel_index,
                    "location_raw": "",
                    "tag": f"UBS3_FB_{channel_index:02d}",
                    "description": f"Feedback UBS3 {channel_index:02d}",
                    "type": "DI",
                    "issues": [],
                }
            )
        for channel_index in range(1, 17):
            normalized_rows.append(
                {
                    "source_sheet": "UBS3-ET200SP",
                    "source_row": 200 + channel_index,
                    "panel_raw": "",
                    "rack_raw": "",
                    "slot_raw": "SLOT 03",
                    "slot_index": 3,
                    "module_raw": "6ES7132-6BH01-0BA0 (16 DO)",
                    "channel_raw": f"DO.{channel_index - 1}",
                    "channel_index": channel_index,
                    "location_raw": "",
                    "tag": f"CMD_ELE{channel_index:02d}",
                    "description": f"Comando UBS3 {channel_index:02d}",
                    "type": "DO",
                    "issues": [],
                }
            )

        proposal = build_import_proposal(
            original_filename="PLANILHA DE IO UBS3 NUTRIEN 05042025.xlsx",
            normalized_rows=normalized_rows,
            module_catalog=module_catalog,
        )

        self.assertEqual(proposal["summary"]["racks"], 1)
        self.assertEqual(proposal["rack"]["name"], "UBS3")
        self.assertEqual(
            {row["resolved_rack_name"] for row in normalized_rows},
            {"UBS3"},
        )

    def test_apply_import_multirack_creates_multiple_racks(self):
        upload = SimpleUploadedFile(
            "PLANILHA DE IO UBS3 NUTRIEN - REM01 REV03.xlsx",
            self._build_slot_block_workbook(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        create_response = self.client_http.post(
            "/ios/importacoes/nova/",
            {
                "arquivo": upload,
                "requested_local": str(self.local.id),
                "requested_grupo": str(self.grupo.id),
                "requested_planta_code": self.planta.codigo,
            },
        )
        self.assertEqual(create_response.status_code, 302)

        job = IOImportJob.objects.latest("id")
        apply_response = self.client_http.post(
            f"/ios/importacoes/{job.pk}/",
            {"action": "apply_import"},
        )
        self.assertEqual(apply_response.status_code, 302)

        job.refresh_from_db()
        self.assertEqual(job.status, IOImportJob.Status.APPLIED)
        self.assertEqual(job.apply_log["racks_applied"], 2)
        self.assertEqual(job.apply_log["modules_applied"], 3)
        self.assertEqual(sorted(RackIO.objects.filter(cliente=self.perfil).values_list("nome", flat=True)), ["REM01", "UBS3"])

    def test_apply_import_can_approve_racks_individually_without_duplication(self):
        self.perfil.tipos.clear()
        activate_trial(self.user, DOCUMENTATION_PRODUCT_CODE)
        upload = SimpleUploadedFile(
            "PLANILHA DE IO UBS3 NUTRIEN - REM01 REV03.xlsx",
            self._build_slot_block_workbook(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        create_response = self.client_http.post(
            "/ios/importacoes/nova/",
            {
                "arquivo": upload,
                "requested_local": str(self.local.id),
                "requested_grupo": str(self.grupo.id),
                "requested_planta_code": self.planta.codigo,
            },
        )
        self.assertEqual(create_response.status_code, 302)

        job = IOImportJob.objects.latest("id")
        rack_keys = [rack["rack_key"] for rack in job.proposal_payload["racks"]]

        first_apply = apply_import_job(
            job=job,
            user=self.user,
            rack_model=RackIO,
            rack_slot_model=RackSlotIO,
            rack_module_model=ModuloRackIO,
            channel_model=CanalRackIO,
            module_qs=ModuloIO.objects.select_related("tipo_base").all(),
            plant_model=PlantaIO,
            selected_rack_keys=[rack_keys[0]],
        )
        self.assertEqual(len(first_apply), 1)
        job.refresh_from_db()
        self.assertEqual(job.status, IOImportJob.Status.REVIEW)
        self.assertEqual(job.apply_log["racks_applied"], 1)
        self.assertIsNotNone(job.first_applied_at)
        self.assertEqual(RackIO.objects.filter(cliente=self.perfil).count(), 1)
        self.assertEqual(
            ConsumoImportacaoDiaria.objects.get(
                usuario=self.user,
                produto__codigo=DOCUMENTATION_PRODUCT_CODE,
                modulo=ConsumoImportacaoDiaria.Modulo.IO,
            ).importacoes_bem_sucedidas,
            1,
        )

        duplicated_apply = apply_import_job(
            job=job,
            user=self.user,
            rack_model=RackIO,
            rack_slot_model=RackSlotIO,
            rack_module_model=ModuloRackIO,
            channel_model=CanalRackIO,
            module_qs=ModuloIO.objects.select_related("tipo_base").all(),
            plant_model=PlantaIO,
            selected_rack_keys=[rack_keys[0]],
        )
        self.assertEqual(len(duplicated_apply), 1)
        self.assertEqual(RackIO.objects.filter(cliente=self.perfil).count(), 1)
        self.assertEqual(
            ConsumoImportacaoDiaria.objects.get(
                usuario=self.user,
                produto__codigo=DOCUMENTATION_PRODUCT_CODE,
                modulo=ConsumoImportacaoDiaria.Modulo.IO,
            ).importacoes_bem_sucedidas,
            1,
        )

        second_apply = apply_import_job(
            job=job,
            user=self.user,
            rack_model=RackIO,
            rack_slot_model=RackSlotIO,
            rack_module_model=ModuloRackIO,
            channel_model=CanalRackIO,
            module_qs=ModuloIO.objects.select_related("tipo_base").all(),
            plant_model=PlantaIO,
            selected_rack_keys=[rack_keys[1]],
        )
        self.assertEqual(len(second_apply), 1)
        job.refresh_from_db()
        self.assertEqual(job.status, IOImportJob.Status.APPLIED)
        self.assertEqual(job.apply_log["racks_applied"], 2)
        self.assertEqual(len(job.apply_log["applied_rack_keys"]), 2)
        self.assertEqual(RackIO.objects.filter(cliente=self.perfil).count(), 2)
        self.assertEqual(
            ConsumoImportacaoDiaria.objects.get(
                usuario=self.user,
                produto__codigo=DOCUMENTATION_PRODUCT_CODE,
                modulo=ConsumoImportacaoDiaria.Modulo.IO,
            ).importacoes_bem_sucedidas,
            1,
        )

    def test_generated_fixture_files_all_parse_without_conflicts(self):
        manifest = json.loads((self.FIXTURES_DIR / "manifest.json").read_text(encoding="utf-8"))
        module_catalog = serialize_module_catalog(ModuloIO.objects.select_related("tipo_base").all())

        for item in manifest["files"]:
            file_name = item["file_name"]
            if not file_name.lower().endswith((".xlsx", ".csv", ".tsv")):
                continue
            with self.subTest(file_name=file_name):
                raw_bytes = (self.FIXTURES_DIR / file_name).read_bytes()
                parsed_sheets = parse_workbook(raw_bytes=raw_bytes, original_filename=file_name)
                self.assertGreaterEqual(len(parsed_sheets), 1)

                normalized_rows = []
                for parsed in parsed_sheets:
                    sheet_rows, _, _ = normalize_rows(parsed=parsed, module_catalog=module_catalog, ai_result=None)
                    normalized_rows.extend(sheet_rows)

                proposal = build_import_proposal(
                    original_filename=file_name,
                    normalized_rows=normalized_rows,
                    module_catalog=module_catalog,
                )
                self.assertEqual(proposal["summary"]["rows"], item["expected_logical_points"])
                self.assertEqual(proposal["summary"]["with_conflicts"], 0)

    def test_extended_fixture_files_11_to_60_parse_without_conflicts(self):
        module_catalog = serialize_module_catalog(ModuloIO.objects.select_related("tipo_base").all())
        files = sorted(
            path
            for path in self.FIXTURES_DIR.iterdir()
            if path.is_file()
            and path.suffix.lower() in {".xlsx", ".csv", ".tsv"}
            and path.name[:2].isdigit()
            and 11 <= int(path.name[:2]) <= 60
        )

        for path in files:
            with self.subTest(file_name=path.name):
                parsed_sheets = parse_workbook(raw_bytes=path.read_bytes(), original_filename=path.name)
                normalized_rows = []
                for parsed in parsed_sheets:
                    sheet_rows, _, _ = normalize_rows(parsed=parsed, module_catalog=module_catalog, ai_result=None)
                    normalized_rows.extend(sheet_rows)

                proposal = build_import_proposal(
                    original_filename=path.name,
                    normalized_rows=normalized_rows,
                    module_catalog=module_catalog,
                )
                self.assertEqual(proposal["summary"]["rows"], 120)
                self.assertEqual(proposal["summary"]["with_conflicts"], 0)

    def test_enriched_physical_fixture_files_preserve_expected_topology(self):
        module_catalog = serialize_module_catalog(ModuloIO.objects.select_related("tipo_base").all())
        expected = {
            "29_io_range_first_process.xlsx": {
                "rows": 120,
                "racks": 18,
                "modules": 24,
                "slots": 24,
                "rack_names": {
                    "MCC-TRN-01 - Rack 01",
                    "MCC-TRN-01 - Rack 02",
                    "MCC-TRN-01 - Rack 03",
                    "PNL-BAG-01 - Rack 01",
                    "PNL-BAG-01 - Rack 02",
                    "PNL-BAG-01 - Rack 03",
                    "PNL-MOG-01 - Rack 01",
                    "PNL-MOG-01 - Rack 02",
                    "PNL-MOG-01 - Rack 03",
                    "PNL-PRC-01 - Rack 01",
                    "PNL-PRC-01 - Rack 02",
                    "PNL-PRC-01 - Rack 03",
                    "PNL-RCV-01 - Rack 01",
                    "PNL-RCV-01 - Rack 02",
                    "PNL-RCV-01 - Rack 03",
                    "PNL-UTL-01 - Rack 01",
                    "PNL-UTL-01 - Rack 02",
                    "PNL-UTL-01 - Rack 03",
                },
                "sheet_assertions": {
                    "MCC-TRN-01 - Rack 01": ["ProcessSignals"],
                    "PNL-UTL-01 - Rack 03": ["ProcessSignals"],
                },
            },
            "34_io_cable_core_schedule.xlsx": {
                "rows": 120,
                "racks": 18,
                "modules": 24,
                "slots": 24,
                "rack_names": {
                    "MCC-TRN-01 - Rack 01",
                    "MCC-TRN-01 - Rack 02",
                    "MCC-TRN-01 - Rack 03",
                    "PNL-BAG-01 - Rack 01",
                    "PNL-BAG-01 - Rack 02",
                    "PNL-BAG-01 - Rack 03",
                    "PNL-MOG-01 - Rack 01",
                    "PNL-MOG-01 - Rack 02",
                    "PNL-MOG-01 - Rack 03",
                    "PNL-PRC-01 - Rack 01",
                    "PNL-PRC-01 - Rack 02",
                    "PNL-PRC-01 - Rack 03",
                    "PNL-RCV-01 - Rack 01",
                    "PNL-RCV-01 - Rack 02",
                    "PNL-RCV-01 - Rack 03",
                    "PNL-UTL-01 - Rack 01",
                    "PNL-UTL-01 - Rack 02",
                    "PNL-UTL-01 - Rack 03",
                },
                "sheet_assertions": {
                    "MCC-TRN-01 - Rack 01": ["CableCore"],
                    "PNL-UTL-01 - Rack 03": ["CableCore"],
                },
            },
            "39_io_loop_schedule_multisheet.xlsx": {
                "rows": 120,
                "racks": 18,
                "modules": 24,
                "slots": 24,
                "rack_names": {
                    "MCC-TRN-01 - Rack 01",
                    "MCC-TRN-01 - Rack 02",
                    "MCC-TRN-01 - Rack 03",
                    "PNL-BAG-01 - Rack 01",
                    "PNL-BAG-01 - Rack 02",
                    "PNL-BAG-01 - Rack 03",
                    "PNL-MOG-01 - Rack 01",
                    "PNL-MOG-01 - Rack 02",
                    "PNL-MOG-01 - Rack 03",
                    "PNL-PRC-01 - Rack 01",
                    "PNL-PRC-01 - Rack 02",
                    "PNL-PRC-01 - Rack 03",
                    "PNL-RCV-01 - Rack 01",
                    "PNL-RCV-01 - Rack 02",
                    "PNL-RCV-01 - Rack 03",
                    "PNL-UTL-01 - Rack 01",
                    "PNL-UTL-01 - Rack 02",
                    "PNL-UTL-01 - Rack 03",
                },
                "sheet_assertions": {
                    "MCC-TRN-01 - Rack 01": ["Transporte"],
                    "PNL-BAG-01 - Rack 01": ["Ensacadeira"],
                    "PNL-MOG-01 - Rack 01": ["Moega"],
                    "PNL-PRC-01 - Rack 01": ["Processo"],
                    "PNL-RCV-01 - Rack 01": ["Recebimento"],
                    "PNL-UTL-01 - Rack 01": ["Utilidades"],
                },
            },
            "56_io_field_check_packets.xlsx": {
                "rows": 120,
                "racks": 18,
                "modules": 24,
                "slots": 24,
                "rack_names": {
                    "MCC-TRN-01 - Rack 01",
                    "MCC-TRN-01 - Rack 02",
                    "MCC-TRN-01 - Rack 03",
                    "PNL-BAG-01 - Rack 01",
                    "PNL-BAG-01 - Rack 02",
                    "PNL-BAG-01 - Rack 03",
                    "PNL-MOG-01 - Rack 01",
                    "PNL-MOG-01 - Rack 02",
                    "PNL-MOG-01 - Rack 03",
                    "PNL-PRC-01 - Rack 01",
                    "PNL-PRC-01 - Rack 02",
                    "PNL-PRC-01 - Rack 03",
                    "PNL-RCV-01 - Rack 01",
                    "PNL-RCV-01 - Rack 02",
                    "PNL-RCV-01 - Rack 03",
                    "PNL-UTL-01 - Rack 01",
                    "PNL-UTL-01 - Rack 02",
                    "PNL-UTL-01 - Rack 03",
                },
                "sheet_assertions": {
                    "MCC-TRN-01 - Rack 01": ["Area_A", "Area_B"],
                    "PNL-BAG-01 - Rack 03": ["Area_A"],
                    "PNL-UTL-01 - Rack 03": ["Area_B"],
                },
            },
        }

        for file_name, expected_payload in expected.items():
            with self.subTest(file_name=file_name):
                raw_bytes = (self.FIXTURES_DIR / file_name).read_bytes()
                parsed_sheets = parse_workbook(raw_bytes=raw_bytes, original_filename=file_name)

                normalized_rows = []
                for parsed in parsed_sheets:
                    sheet_rows, _, _ = normalize_rows(parsed=parsed, module_catalog=module_catalog, ai_result=None)
                    normalized_rows.extend(sheet_rows)

                proposal = build_import_proposal(
                    original_filename=file_name,
                    normalized_rows=normalized_rows,
                    module_catalog=module_catalog,
                )
                self.assertEqual(proposal["summary"]["rows"], expected_payload["rows"])
                self.assertEqual(proposal["summary"]["racks"], expected_payload["racks"])
                self.assertEqual(proposal["summary"]["modules"], expected_payload["modules"])
                self.assertEqual(proposal["summary"]["slots"], expected_payload["slots"])
                self.assertEqual(proposal["summary"]["with_conflicts"], 0)

                rack_map = {rack["name"]: rack for rack in proposal["racks"]}
                self.assertSetEqual(set(rack_map.keys()), expected_payload["rack_names"])

                for rack_name, expected_sheets in expected_payload["sheet_assertions"].items():
                    self.assertEqual(rack_map[rack_name]["source_sheets"], expected_sheets)

    def test_all_fixture_files_keep_channel_type_consistent_with_module_type(self):
        module_catalog = serialize_module_catalog(ModuloIO.objects.select_related("tipo_base").all())
        files = sorted(
            path
            for path in self.FIXTURES_DIR.iterdir()
            if path.is_file()
            and path.suffix.lower() in {".xlsx", ".csv", ".tsv"}
            and path.name[:2].isdigit()
        )

        for path in files:
            with self.subTest(file_name=path.name):
                parsed_sheets = parse_workbook(raw_bytes=path.read_bytes(), original_filename=path.name)
                normalized_rows = []
                for parsed in parsed_sheets:
                    sheet_rows, _, _ = normalize_rows(parsed=parsed, module_catalog=module_catalog, ai_result=None)
                    normalized_rows.extend(sheet_rows)

                proposal = build_import_proposal(
                    original_filename=path.name,
                    normalized_rows=normalized_rows,
                    module_catalog=module_catalog,
                )

                for rack in proposal["racks"]:
                    for module in rack["modules"]:
                        module_type = module.get("module_type")
                        self.assertIn(module_type, {"DI", "DO", "AI", "AO"})
                        for channel in module.get("channels") or []:
                            if not channel.get("source_row"):
                                continue
                            self.assertEqual(
                                channel.get("type"),
                                module_type,
                                msg=(
                                    f"{path.name} gerou divergencia no rack {rack['name']} "
                                    f"slot {module['slot_index']} canal {channel['index']}: "
                                    f"modulo={module_type} canal={channel.get('type')}"
                                ),
                            )

    def test_apply_import_creates_custom_modules_when_catalog_has_gap(self):
        raw_bytes = (
            b"panel;rack;slot;channel;module_model;type;tag;description\n"
            b"PNL-CUSTOM-01;1;1;1;AI-CUSTOM-04;AI;AIT_001;Transmissor 01\n"
            b"PNL-CUSTOM-01;1;1;2;AI-CUSTOM-04;AI;AIT_002;Transmissor 02\n"
            b"PNL-CUSTOM-01;1;2;1;AO-CUSTOM-02;AO;FCV_001_CMD;Comando de controle\n"
        )
        job = IOImportJob.objects.create(
            created_by=self.user,
            cliente=self.perfil,
            requested_local=self.local,
            requested_grupo=self.grupo,
            requested_planta_code=self.planta.codigo,
            mode=IOImportJob.Mode.CREATE_RACK,
            status=IOImportJob.Status.UPLOADED,
            file_format=IOImportJob.FileFormat.CSV,
            original_filename="custom-gap.csv",
            file_sha256=build_file_sha256(raw_bytes),
            source_file=SimpleUploadedFile("custom-gap.csv", raw_bytes, content_type="text/csv"),
        )

        restricted_modules = ModuloIO.objects.filter(id__in=[self.modulo_di16.id, self.modulo_do16.id]).select_related("tipo_base")
        result = reprocess_import_job(job=job, module_catalog=serialize_module_catalog(restricted_modules), settings_obj=None)
        job.sheet_name = result["sheet_name"]
        job.header_row_index = result["header_row_index"]
        job.rows_total = result["rows_total"]
        job.rows_parsed = result["rows_parsed"]
        job.column_map = result["column_map"]
        job.extracted_payload = {"rows": result["normalized_rows"], "sheets": result["sheet_summaries"]}
        job.proposal_payload = result["proposal"]
        job.warnings = result["warnings"]
        job.status = IOImportJob.Status.REVIEW
        job.save(
            update_fields=[
                "sheet_name",
                "header_row_index",
                "rows_total",
                "rows_parsed",
                "column_map",
                "extracted_payload",
                "proposal_payload",
                "warnings",
                "status",
            ]
        )

        self.assertEqual(job.proposal_payload["summary"]["with_conflicts"], 0)
        custom_slots = [
            module
            for rack in job.proposal_payload["racks"]
            for module in rack["modules"]
            if module.get("module_model_source") == "custom"
        ]
        self.assertEqual(len(custom_slots), 2)

        applied_racks = apply_import_job(
            job=job,
            user=self.user,
            rack_model=RackIO,
            rack_slot_model=RackSlotIO,
            rack_module_model=ModuloRackIO,
            channel_model=CanalRackIO,
            module_qs=restricted_modules,
            plant_model=PlantaIO,
        )

        self.assertEqual(len(applied_racks), 1)
        self.assertEqual(job.status, IOImportJob.Status.APPLIED)
        ai_custom = ModuloIO.objects.filter(cliente=self.perfil, modelo="AI-CUSTOM-04", tipo_base__nome="AI").first()
        ao_custom = ModuloIO.objects.filter(cliente=self.perfil, modelo="AO-CUSTOM-02", tipo_base__nome="AO").first()
        self.assertIsNotNone(ai_custom)
        self.assertIsNotNone(ao_custom)
        self.assertFalse(ai_custom.is_default)
        self.assertFalse(ao_custom.is_default)


class IPImportPipelineTests(TestCase):
    def setUp(self):
        self.client_http = Client()
        self.tipo_dev = TipoPerfil.objects.get(codigo="DEV")
        self.user = User.objects.create_user(
            username="ip-import@set.local",
            email="ip-import@set.local",
            password="123456",
        )
        self.perfil = PerfilUsuario.objects.create(
            nome="IP Import",
            email="ip-import@set.local",
            usuario=self.user,
        )
        self.perfil.tipos.add(self.tipo_dev)
        self.client_http.force_login(self.user)

    def _post_import_and_process(self, upload, ajax=False):
        headers = {"HTTP_X_REQUESTED_WITH": "XMLHttpRequest"} if ajax else {}
        def run_inline(job_id):
            job = IPImportJob.objects.get(pk=job_id)
            try:
                _reprocess_ip_import_job(job)
            except Exception as exc:
                job.refresh_from_db()
                job.status = IPImportJob.Status.FAILED
                job.ai_status = IPImportJob.AIStatus.FAILED
                job.ai_error = str(exc)
                warnings = list(job.warnings or [])
                warnings.append(str(exc))
                job.warnings = warnings
                job.save(update_fields=["status", "ai_status", "ai_error", "warnings", "updated_at"])

        with patch("core.views._spawn_ip_import_job_processor_safe", side_effect=run_inline):
            with self.captureOnCommitCallbacks(execute=True):
                response = self.client_http.post("/listas-ip/importacoes/nova/", {"arquivo": upload}, **headers)
        return response

    def _build_ip_workbook(self):
        workbook = Workbook()
        sheet1 = workbook.active
        sheet1.title = "Rede Principal"
        for row in [
            ["LISTA", "ID_LISTAIP", "IP", "EQUIPAMENTO", "DESCRICAO", "MAC", "PROTOCOLO"],
            ["PLC PRINCIPAL", "LIP-001", "192.168.10.10", "PLC_MAIN", "Controlador principal", "001122334455", "Modbus TCP"],
            ["PLC PRINCIPAL", "LIP-001", "192.168.10.11", "IHM_MAIN", "Supervisao", "", "HTTP"],
        ]:
            sheet1.append(row)
        sheet2 = workbook.create_sheet("Rede Remota")
        for row in [
            ["LISTA", "ID_LISTAIP", "IP", "EQUIPAMENTO", "DESCRICAO", "MAC", "PROTOCOLO"],
            ["REMOTA", "LIP-002", "10.20.30.40", "SW_REM_01", "Switch remoto", "", "PROFINET"],
            ["REMOTA", "LIP-002", "10.20.30.41", "AP_REM_01", "Access point", "", "TCP/IP"],
        ]:
            sheet2.append(row)
        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def _build_devices_style_ip_workbook(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "DEVICES"
        for row in [
            ["", "", "Milhao Ingredients", "", "", "", "", "", "", ""],
            ["", "", "DEVICES", "", "", "", "", "", "", ""],
            ["", "", "EQP", "NOVO", "IP", "FIM IP", "MAC", "PLC", "DRIVE", "STATUS"],
            ["", "", "MILHAO_MASTER", "", "192.100.3.1", "1", "5C-88-16-F6-E5-7B", "", "CONTROL", "OK"],
            ["", "", "REMCCM1", "", "192.100.3.2", "2", "5C-88-16-B9-A6-13", "CONTROL", "AENTR", "OK"],
        ]:
            sheet.append(row)
        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def _build_multisheet_with_summary_workbook(self):
        workbook = Workbook()
        summary = workbook.active
        summary.title = "Resumo"
        for row in [
            ["AREA", "OBS"],
            ["Rede Principal", "Resumo executivo"],
            ["Rede Remota", "Resumo executivo"],
        ]:
            summary.append(row)
        data_1 = workbook.create_sheet("Rede Principal")
        for row in [
            ["LISTA", "ID_LISTAIP", "IP", "EQUIPAMENTO", "DESCRICAO", "MAC", "PROTOCOLO"],
            ["PLC PRINCIPAL", "LIP-001", "192.168.10.10", "PLC_MAIN", "Controlador principal", "001122334455", "Modbus TCP"],
        ]:
            data_1.append(row)
        data_2 = workbook.create_sheet("Rede Remota")
        for row in [
            ["LISTA", "ID_LISTAIP", "IP", "EQUIPAMENTO", "DESCRICAO", "MAC", "PROTOCOLO"],
            ["REMOTA", "LIP-002", "10.20.30.40", "SW_REM_01", "Switch remoto", "", "PROFINET"],
        ]:
            data_2.append(row)
        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def test_import_csv_creates_review_job_and_preview(self):
        upload = SimpleUploadedFile(
            "lista-ip.csv",
            (
                b"LISTA;ID_LISTAIP;IP;EQUIPAMENTO;DESCRICAO;MAC;PROTOCOLO\n"
                b"PLC PRINCIPAL;LIP-001;192.168.10.10;PLC_MAIN;Controlador principal;001122334455;Modbus TCP\n"
                b"PLC PRINCIPAL;LIP-001;192.168.10.11;IHM_MAIN;Supervisao;;HTTP\n"
            ),
            content_type="text/csv",
        )
        response = self._post_import_and_process(upload)
        self.assertEqual(response.status_code, 302)

        job = IPImportJob.objects.get()
        self.assertEqual(job.status, IPImportJob.Status.REVIEW)
        self.assertEqual(job.file_format, IPImportJob.FileFormat.CSV)
        self.assertEqual(job.rows_parsed, 2)
        self.assertEqual(job.proposal_payload["summary"]["lists"], 1)
        self.assertEqual(job.proposal_payload["summary"]["items"], 2)

        detail_response = self.client_http.get(f"/listas-ip/importacoes/{job.pk}/")
        self.assertEqual(detail_response.status_code, 200)
        self.assertContains(detail_response, "PLC PRINCIPAL")
        self.assertContains(detail_response, "Aprovar e Aplicar")

    def test_import_multisheet_builds_multiple_lists(self):
        upload = SimpleUploadedFile(
            "listas-ip.xlsx",
            self._build_ip_workbook(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response = self._post_import_and_process(upload)
        self.assertEqual(response.status_code, 302)

        job = IPImportJob.objects.latest("id")
        self.assertEqual(job.status, IPImportJob.Status.REVIEW)
        self.assertEqual(job.file_format, IPImportJob.FileFormat.XLSX)
        self.assertEqual(job.proposal_payload["summary"]["lists"], 2)
        self.assertEqual(job.proposal_payload["summary"]["items"], 4)
        self.assertEqual(sorted(item["name"] for item in job.proposal_payload["lists"]), ["PLC PRINCIPAL", "REMOTA"])

    def test_import_devices_layout_maps_eqp_and_composes_description(self):
        upload = SimpleUploadedFile(
            "planilha-ip-devices.xlsx",
            self._build_devices_style_ip_workbook(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response = self._post_import_and_process(upload)
        self.assertEqual(response.status_code, 302)

        job = IPImportJob.objects.latest("id")
        self.assertEqual(job.status, IPImportJob.Status.REVIEW)
        extracted_rows = job.extracted_payload["rows"]
        self.assertEqual(extracted_rows[0]["device_name"], "MILHAO_MASTER")
        self.assertIn("Drive CONTROL", extracted_rows[0]["description"])
        self.assertIn("Status OK", extracted_rows[0]["description"])
        self.assertEqual(extracted_rows[1]["device_name"], "REMCCM1")
        self.assertIn("PLC CONTROL", extracted_rows[1]["description"])
        self.assertIn("Drive AENTR", extracted_rows[1]["description"])

    def test_apply_import_can_approve_lists_individually_without_duplication(self):
        self.perfil.tipos.clear()
        activate_trial(self.user, DOCUMENTATION_PRODUCT_CODE)
        upload = SimpleUploadedFile(
            "listas-ip.xlsx",
            self._build_ip_workbook(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response = self._post_import_and_process(upload)
        self.assertEqual(response.status_code, 302)

        job = IPImportJob.objects.latest("id")
        list_keys = [item["list_key"] for item in job.proposal_payload["lists"]]

        first_apply = apply_ip_import_job(job=job, user=self.user, selected_list_keys=[list_keys[0]])
        self.assertEqual(len(first_apply), 1)
        job.refresh_from_db()
        self.assertEqual(job.status, IPImportJob.Status.REVIEW)
        self.assertIsNotNone(job.first_applied_at)
        self.assertEqual(ListaIP.objects.filter(cliente=self.perfil).count(), 1)
        self.assertEqual(
            ConsumoImportacaoDiaria.objects.get(
                usuario=self.user,
                produto__codigo=DOCUMENTATION_PRODUCT_CODE,
                modulo=ConsumoImportacaoDiaria.Modulo.IP,
            ).importacoes_bem_sucedidas,
            1,
        )

        duplicated_apply = apply_ip_import_job(job=job, user=self.user, selected_list_keys=[list_keys[0]])
        self.assertEqual(len(duplicated_apply), 1)
        self.assertEqual(ListaIP.objects.filter(cliente=self.perfil).count(), 1)
        self.assertEqual(
            ConsumoImportacaoDiaria.objects.get(
                usuario=self.user,
                produto__codigo=DOCUMENTATION_PRODUCT_CODE,
                modulo=ConsumoImportacaoDiaria.Modulo.IP,
            ).importacoes_bem_sucedidas,
            1,
        )

        second_apply = apply_ip_import_job(job=job, user=self.user, selected_list_keys=[list_keys[1]])
        self.assertEqual(len(second_apply), 1)
        job.refresh_from_db()
        self.assertEqual(job.status, IPImportJob.Status.APPLIED)
        self.assertEqual(ListaIP.objects.filter(cliente=self.perfil).count(), 2)
        self.assertEqual(ListaIPItem.objects.filter(lista__cliente=self.perfil).count(), 4)
        self.assertEqual(
            ConsumoImportacaoDiaria.objects.get(
                usuario=self.user,
                produto__codigo=DOCUMENTATION_PRODUCT_CODE,
                modulo=ConsumoImportacaoDiaria.Modulo.IP,
            ).importacoes_bem_sucedidas,
            1,
        )

    def test_import_admin_renders_separate_settings(self):
        admin_user = User.objects.create_superuser("ip-admin", "ip-admin@set.local", "123456")
        self.client_http.force_login(admin_user)
        response = self.client_http.get("/listas-ip/importacoes/admin/")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Importacao de planilhas de IP")
        self.assertTrue(IPImportSettings.objects.exists())

    def test_import_transport_timeout_do_agente_interrompe_pipeline_ai_first(self):
        settings_obj = IPImportSettings.load()
        settings_obj.enabled = True
        settings_obj.api_key = "test-key"
        settings_obj.save()

        upload = SimpleUploadedFile(
            "listas-ip.xlsx",
            self._build_ip_workbook(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        with patch("core.services.ip_import.urlrequest.urlopen", side_effect=TimeoutError("timed out")):
            response = self._post_import_and_process(upload)

        self.assertEqual(response.status_code, 302)
        job = IPImportJob.objects.latest("id")
        self.assertEqual(job.status, IPImportJob.Status.FAILED)
        self.assertEqual(job.ai_status, IPImportJob.AIStatus.FAILED)
        self.assertIn("timeout", job.ai_error.lower())
        self.assertTrue(any("timeout" in warning.lower() for warning in (job.warnings or [])))

    def test_import_ajax_returns_json_redirect_and_status_urls(self):
        upload = SimpleUploadedFile(
            "listas-ip.xlsx",
            self._build_ip_workbook(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response = self._post_import_and_process(upload, ajax=True)
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["ok"])
        self.assertIn("/listas-ip/importacoes/", payload["redirect_url"])
        self.assertIn("/status/", payload["status_url"])

        job = IPImportJob.objects.latest("id")
        self.assertEqual(job.status, IPImportJob.Status.REVIEW)
        self.assertEqual(job.proposal_payload["summary"]["lists"], 2)

    def test_import_ajax_status_endpoint_reports_processing(self):
        upload = SimpleUploadedFile(
            "listas-ip.xlsx",
            self._build_ip_workbook(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        with self.captureOnCommitCallbacks(execute=False):
            response = self.client_http.post(
                "/listas-ip/importacoes/nova/",
                {"arquivo": upload},
                HTTP_X_REQUESTED_WITH="XMLHttpRequest",
            )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        job = IPImportJob.objects.latest("id")
        self.assertEqual(job.status, IPImportJob.Status.UPLOADED)

        status_response = self.client_http.get(
            f"/listas-ip/importacoes/{job.pk}/status/",
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )
        self.assertEqual(status_response.status_code, 200)
        status_payload = status_response.json()
        self.assertTrue(status_payload["ok"])
        self.assertTrue(status_payload["processing"])
        self.assertFalse(status_payload["complete"])
        self.assertEqual(status_payload["progress"]["stage"], "upload")

    def test_ip_call_openai_responses_polls_background_response_until_completed(self):
        settings_obj = IPImportSettings.load()
        settings_obj.enabled = True
        settings_obj.api_key = "test-key"
        settings_obj.save()

        requested = []
        queued_payload = {"id": "resp_ip_123", "status": "queued", "output": []}
        running_payload = {"id": "resp_ip_123", "status": "in_progress", "output": []}
        completed_payload = {
            "id": "resp_ip_123",
            "status": "completed",
            "output": [{"content": [{"type": "output_text", "text": json.dumps({"result": "ok"})}]}],
        }
        responses = [queued_payload, running_payload, completed_payload]

        class _FakeHTTPResponse:
            def __init__(self, payload):
                self.payload = payload

            def read(self):
                return json.dumps(self.payload).encode("utf-8")

            def __enter__(self):
                return self

            def __exit__(self, exc_type, exc, tb):
                return False

        def fake_urlopen(request, timeout=None):
            requested.append(
                {
                    "url": request.full_url,
                    "method": request.get_method(),
                    "timeout": timeout,
                    "data": request.data,
                }
            )
            return _FakeHTTPResponse(responses[len(requested) - 1])

        with patch("core.services.ip_import.urlrequest.urlopen", side_effect=fake_urlopen), patch(
            "core.services.ip_import.time.sleep"
        ):
            result = call_ip_openai_responses(
                settings_obj=settings_obj,
                schema_name="ip_sheet_semantic_analysis",
                schema={"type": "object", "properties": {"result": {"type": "string"}}, "required": ["result"], "additionalProperties": False},
                system_prompt="system",
                user_prompt="{}",
                request_timeout_seconds=300,
            )

        self.assertEqual(result, {"result": "ok"})
        self.assertEqual([item["method"] for item in requested], ["POST", "GET", "GET"])
        post_payload = json.loads(requested[0]["data"].decode("utf-8"))
        self.assertTrue(post_payload["background"])
        self.assertTrue(post_payload["store"])
        self.assertEqual(post_payload["metadata"]["source"], "ip_import")

    def test_ip_call_openai_responses_does_not_abort_only_because_total_wait_is_long(self):
        settings_obj = IPImportSettings.load()
        settings_obj.enabled = True
        settings_obj.api_key = "test-key"
        settings_obj.save()

        requested = []
        queued_payload = {"id": "resp_ip_123", "status": "queued", "output": []}
        running_payload = {"id": "resp_ip_123", "status": "in_progress", "output": []}
        completed_payload = {
            "id": "resp_ip_123",
            "status": "completed",
            "output": [{"content": [{"type": "output_text", "text": json.dumps({"result": "ok"})}]}],
        }
        responses = [queued_payload, running_payload, completed_payload]

        class _FakeHTTPResponse:
            def __init__(self, payload):
                self.payload = payload

            def read(self):
                return json.dumps(self.payload).encode("utf-8")

            def __enter__(self):
                return self

            def __exit__(self, exc_type, exc, tb):
                return False

        def fake_urlopen(request, timeout=None):
            requested.append(
                {
                    "url": request.full_url,
                    "method": request.get_method(),
                    "timeout": timeout,
                }
            )
            return _FakeHTTPResponse(responses[len(requested) - 1])

        with patch("core.services.ip_import.urlrequest.urlopen", side_effect=fake_urlopen), patch(
            "core.services.ip_import.time.sleep"
        ), patch("core.services.ip_import.time.monotonic", side_effect=[0.0, 600.0, 1800.0]):
            result = call_ip_openai_responses(
                settings_obj=settings_obj,
                schema_name="ip_sheet_semantic_analysis",
                schema={"type": "object", "properties": {"result": {"type": "string"}}, "required": ["result"], "additionalProperties": False},
                system_prompt="system",
                user_prompt="{}",
                request_timeout_seconds=30,
            )

        self.assertEqual(result, {"result": "ok"})
        self.assertEqual([item["method"] for item in requested], ["POST", "GET", "GET"])

    def test_reprocess_ip_import_job_ai_can_skip_summary_sheet_and_keep_data_sheets(self):
        settings_obj = IPImportSettings.load()
        settings_obj.enabled = True
        settings_obj.api_key = "test-key"
        settings_obj.save()

        raw_bytes = self._build_multisheet_with_summary_workbook()
        job = IPImportJob.objects.create(
            created_by=self.user,
            cliente=self.perfil,
            original_filename="listas-resumo.xlsx",
            source_file=SimpleUploadedFile(
                "listas-resumo.xlsx",
                raw_bytes,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
            file_sha256=build_file_sha256(raw_bytes),
        )

        workbook_payload = {
            "sheets": [
                {
                    "sheet_name": "Resumo",
                    "use_sheet": False,
                    "sheet_role": "summary",
                    "default_list_name": "",
                    "default_list_code": "",
                    "confidence": 98,
                    "reason": "Guia de resumo sem registros operacionais individuais.",
                },
                {
                    "sheet_name": "Rede Principal",
                    "use_sheet": True,
                    "sheet_role": "data",
                    "default_list_name": "PLC PRINCIPAL",
                    "default_list_code": "LIP-001",
                    "confidence": 94,
                    "reason": "Guia com registros operacionais de IP.",
                },
                {
                    "sheet_name": "Rede Remota",
                    "use_sheet": True,
                    "sheet_role": "data",
                    "default_list_name": "REMOTA",
                    "default_list_code": "LIP-002",
                    "confidence": 94,
                    "reason": "Guia com registros operacionais de IP.",
                },
            ],
            "warnings": [],
            "notes": "ok",
        }
        principal_payload = {
            "skip_sheet": False,
            "sheet_role": "data",
            "default_list_name": "PLC PRINCIPAL",
            "default_list_code": "LIP-001",
            "column_map": {
                "list_name": "LISTA",
                "list_code": "ID_LISTAIP",
                "ip": "IP",
                "device_name": "EQUIPAMENTO",
                "description": "DESCRICAO",
                "mac": "MAC",
                "protocol": "PROTOCOLO",
                "range_start": "",
                "range_end": "",
            },
            "logical_items": [
                {
                    "source_row": 2,
                    "list_name": "PLC PRINCIPAL",
                    "list_code": "LIP-001",
                    "ip": "192.168.10.10",
                    "range_start": "",
                    "range_end": "",
                    "device_name": "PLC_MAIN",
                    "description": "Controlador principal",
                    "mac": "001122334455",
                    "protocol": "Modbus TCP",
                    "confidence": 96,
                }
            ],
            "warnings": [],
            "notes": "ok",
        }
        remota_payload = {
            "skip_sheet": False,
            "sheet_role": "data",
            "default_list_name": "REMOTA",
            "default_list_code": "LIP-002",
            "column_map": {
                "list_name": "LISTA",
                "list_code": "ID_LISTAIP",
                "ip": "IP",
                "device_name": "EQUIPAMENTO",
                "description": "DESCRICAO",
                "mac": "MAC",
                "protocol": "PROTOCOLO",
                "range_start": "",
                "range_end": "",
            },
            "logical_items": [
                {
                    "source_row": 2,
                    "list_name": "REMOTA",
                    "list_code": "LIP-002",
                    "ip": "10.20.30.40",
                    "range_start": "",
                    "range_end": "",
                    "device_name": "SW_REM_01",
                    "description": "Switch remoto",
                    "mac": "",
                    "protocol": "PROFINET",
                    "confidence": 96,
                }
            ],
            "warnings": [],
            "notes": "ok",
        }

        with patch("core.services.ip_import._call_openai_responses", side_effect=[workbook_payload, principal_payload, remota_payload]):
            result = reprocess_ip_import_job(job=job, settings_obj=settings_obj)

        sheet_names = [sheet["sheet_name"] for sheet in result["sheet_summaries"]]
        self.assertEqual(sheet_names, ["Resumo", "Rede Principal", "Rede Remota"])
        resumo = next(item for item in result["sheet_summaries"] if item["sheet_name"] == "Resumo")
        self.assertTrue(resumo["skipped"])
        self.assertEqual(result["proposal"]["summary"]["lists"], 2)
        self.assertEqual(sorted(item["name"] for item in result["proposal"]["lists"]), ["PLC PRINCIPAL", "REMOTA"])


class FinanceiroCompraAnexoFotoTests(TestCase):
    def setUp(self):
        self.tipo_financeiro = TipoPerfil.objects.get(codigo="FINANCEIRO")
        self.financeiro_modulo = ModuloAcesso.objects.get(codigo="FINANCEIRO")
        self.financeiro_modulo.tipos.set([self.tipo_financeiro])

        self.user = User.objects.create_user(
            username="financeiro-anexo@set.local",
            email="financeiro-anexo@set.local",
            password="123456",
        )
        self.perfil = PerfilUsuario.objects.create(
            nome="Financeiro Anexo",
            email="financeiro-anexo@set.local",
            usuario=self.user,
        )
        self.perfil.tipos.add(self.tipo_financeiro)
        self.caderno = Caderno.objects.create(nome="Caderno Teste", criador=self.perfil, ativo=True)
        self.client.force_login(self.user)

    def _build_image_upload(self, name="comprovante.gif"):
        return SimpleUploadedFile(
            name,
            (
                b"GIF89a\x01\x00\x01\x00\x80\x00\x00"
                b"\x00\x00\x00\xff\xff\xff!\xf9\x04\x01"
                b"\x00\x00\x00\x00,\x00\x00\x00\x00\x01"
                b"\x00\x01\x00\x00\x02\x02D\x01\x00;"
            ),
            content_type="image/gif",
        )

    def test_financeiro_nova_cria_compra_com_foto_anexada(self):
        with patch(
            "django.core.files.storage.FileSystemStorage._save",
            side_effect=lambda *args, **kwargs: args[0],
        ):
            response = self.client.post(
                "/financeiro/nova/",
                {
                    "action": "create_compra",
                    "caderno": str(self.caderno.id),
                    "nome": "Compra com foto",
                    "descricao": "Compra teste com comprovante",
                    "data": "2026-04-23",
                    "anexo_foto": self._build_image_upload(),
                    "total_items": "1",
                    "item_nome_0": "",
                    "item_quantidade_0": "",
                    "item_valor_0": "",
                    "item_parcela_0": "",
                    "item_tipo_0": "",
                },
            )

        self.assertEqual(response.status_code, 302)
        compra = Compra.objects.get(nome="Compra com foto")
        self.assertTrue(bool(compra.anexo_foto))
        self.assertIn("financeiro/compras/", compra.anexo_foto.name)

    def test_financeiro_compra_detail_permite_adicionar_e_remover_foto(self):
        compra = Compra.objects.create(caderno=self.caderno, nome="Compra editavel")

        with patch(
            "django.core.files.storage.FileSystemStorage._save",
            side_effect=lambda *args, **kwargs: args[0],
        ):
            response = self.client.post(
                f"/financeiro/compras/{compra.pk}/",
                {
                    "action": "update_compra",
                    "nome": compra.nome,
                    "descricao": "",
                    "data": "",
                    "caderno": str(self.caderno.id),
                    "categoria": "",
                    "centro_custo": "",
                    "anexo_foto": self._build_image_upload("nota.gif"),
                },
            )
        self.assertEqual(response.status_code, 302)

        compra.refresh_from_db()
        self.assertTrue(bool(compra.anexo_foto))

        response = self.client.post(
            f"/financeiro/compras/{compra.pk}/",
            {
                "action": "update_compra",
                "nome": compra.nome,
                "descricao": "",
                "data": "",
                "caderno": str(self.caderno.id),
                "categoria": "",
                "centro_custo": "",
                "remove_anexo_foto": "on",
            },
        )
        self.assertEqual(response.status_code, 302)

        compra.refresh_from_db()
        self.assertFalse(bool(compra.anexo_foto))
