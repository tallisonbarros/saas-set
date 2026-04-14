import csv
import hashlib
import ipaddress
import json
import os
import re
import socket
import time
import unicodedata
import uuid
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import timedelta
from io import BytesIO, StringIO
from pathlib import Path
from urllib import error as urlerror
from urllib import request as urlrequest

from openpyxl import load_workbook
from django.utils import timezone

from core.models import ListaIP, ListaIPID, ListaIPItem
from core.services.billing import register_successful_import_usage


DEFAULT_AI_REQUEST_TIMEOUT_SECONDS = 25.0
DEFAULT_AI_TOTAL_BUDGET_SECONDS = 180.0
DEFAULT_AI_MAX_SHEETS_PER_JOB = 24
DEFAULT_AI_MAX_RETRIES = 2
DEFAULT_AI_CACHE_MAX_ENTRIES = 160
DEFAULT_AI_CACHE_MAX_AGE_DAYS = 14
IP_IMPORT_AI_CACHE_VERSION = "v1"


HEADER_ALIASES = {
    "list_name": [
        "LISTA",
        "LIST NAME",
        "LISTA IP",
        "NOME LISTA",
        "NETWORK",
        "SUBNET",
        "SEGMENT",
        "AREA",
        "PANEL",
    ],
    "list_code": ["ID_LISTAIP", "LIST ID", "LISTA ID", "LIST CODE", "CODIGO", "CODE"],
    "ip": ["IP", "IP ADDRESS", "ENDERECO IP", "ENDERECO", "HOST IP", "HOST"],
    "device_name": [
        "DEVICE",
        "DEVICE NAME",
        "NOME EQUIPAMENTO",
        "EQUIPAMENTO",
        "EQP",
        "EQUIP",
        "EQUIPAMENTO TAG",
        "HOSTNAME",
        "NODE",
        "TAG",
        "TAG NAME",
    ],
    "description": ["DESCRIPTION", "DESCRICAO", "DESC", "SERVICE", "SERVICO"],
    "mac": ["MAC", "MAC ADDRESS", "ENDERECO MAC"],
    "protocol": ["PROTOCOL", "PROTOCOLO", "SERVICE TYPE"],
    "controller": ["PLC", "CONTROLADOR", "CPU", "CLP"],
    "drive": ["DRIVE", "INVERSOR", "ACIONAMENTO"],
    "status_note": ["STATUS", "STATE", "SITUACAO"],
    "novelty": ["NOVO", "NEW"],
    "range_start": ["FAIXA INICIO", "IP INICIAL", "START IP", "RANGE START", "IP START"],
    "range_end": ["FAIXA FIM", "IP FINAL", "END IP", "RANGE END", "IP END"],
}

DEFAULT_HEADER_PROMPT = (
    "Map the spreadsheet headers to the SAAS-SET IP import schema. "
    "Prefer explicit engineering columns such as list_name, list_code, ip, device_name, description, mac, protocol, range_start and range_end. "
    "Only return mappings that are strongly supported by the sheet."
)

DEFAULT_GROUPING_PROMPT = (
    "Given normalized IP rows, suggest the best list grouping and a stable default list name when the spreadsheet omits it. "
    "Do not invent IP addresses or override explicit values."
)


class IPImportError(Exception):
    pass


@dataclass
class ParsedSpreadsheet:
    file_format: str
    sheet_name: str
    header_row_index: int
    rows_total: int
    headers: list
    raw_rows: list
    column_map: dict
    warnings: list
    suggested_list_name: str


def _ascii_upper(value):
    normalized = unicodedata.normalize("NFKD", str(value or ""))
    normalized = normalized.encode("ascii", "ignore").decode("ascii")
    return normalized.strip().upper()


def _cell_to_text(value):
    if value is None:
        return ""
    if isinstance(value, bool):
        return "SIM" if value else "NAO"
    text = str(value).strip()
    if not text:
        return ""
    return re.sub(r"\s+", " ", text)


def _read_positive_float_env(name, default):
    raw_value = str(os.environ.get(name, "")).strip()
    if not raw_value:
        return default
    try:
        parsed = float(raw_value)
    except (TypeError, ValueError):
        return default
    return parsed if parsed > 0 else default


def _read_positive_int_env(name, default):
    raw_value = str(os.environ.get(name, "")).strip()
    if not raw_value:
        return default
    try:
        parsed = int(raw_value)
    except (TypeError, ValueError):
        return default
    return parsed if parsed > 0 else default


def _ai_request_timeout_seconds():
    return _read_positive_float_env("IP_IMPORT_AI_TIMEOUT_SECONDS", DEFAULT_AI_REQUEST_TIMEOUT_SECONDS)


def _ai_total_budget_seconds():
    return _read_positive_float_env("IP_IMPORT_AI_TOTAL_BUDGET_SECONDS", DEFAULT_AI_TOTAL_BUDGET_SECONDS)


def _ai_max_sheets_per_job():
    return _read_positive_int_env("IP_IMPORT_AI_MAX_SHEETS_PER_JOB", DEFAULT_AI_MAX_SHEETS_PER_JOB)


def _ai_max_retries():
    return _read_positive_int_env("IP_IMPORT_AI_MAX_RETRIES", DEFAULT_AI_MAX_RETRIES)


def _ai_cache_max_entries():
    return _read_positive_int_env("IP_IMPORT_AI_CACHE_MAX_ENTRIES", DEFAULT_AI_CACHE_MAX_ENTRIES)


def _ai_cache_max_age_days():
    return _read_positive_int_env("IP_IMPORT_AI_CACHE_MAX_AGE_DAYS", DEFAULT_AI_CACHE_MAX_AGE_DAYS)


def _hash_json_payload(payload):
    encoded = json.dumps(payload, ensure_ascii=True, sort_keys=True, separators=(",", ":")).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def _ai_settings_fingerprint(settings_obj):
    payload = {
        "provider": getattr(settings_obj, "provider", ""),
        "model": getattr(settings_obj, "model", ""),
        "reasoning_effort": getattr(settings_obj, "reasoning_effort", ""),
        "max_rows_for_ai": getattr(settings_obj, "max_rows_for_ai", 0),
        "header_prompt": getattr(settings_obj, "header_prompt", ""),
        "grouping_prompt": getattr(settings_obj, "grouping_prompt", ""),
        "version": IP_IMPORT_AI_CACHE_VERSION,
    }
    return _hash_json_payload(payload)


def _emit_progress(progress_callback, stage, percent, title, message, **extra):
    if not callable(progress_callback):
        return
    payload = {
        "stage": _cell_to_text(stage).lower() or "upload",
        "percent": max(0, min(int(percent or 0), 100)),
        "title": _cell_to_text(title),
        "message": _cell_to_text(message),
    }
    for key, value in extra.items():
        if value is not None:
            payload[key] = value
    progress_callback(payload)


def _compact_token(value):
    return re.sub(r"[^A-Z0-9]+", "", _ascii_upper(value))


def _header_tokens(value):
    return [token for token in re.split(r"[^A-Z0-9]+", _ascii_upper(value)) if token]


def _alias_match_score(header_value, alias):
    header_compact = _compact_token(header_value)
    alias_token = _compact_token(alias)
    if not header_compact or not alias_token:
        return 0
    header_words = set(_header_tokens(header_value))
    if header_compact == alias_token:
        return 100
    if alias_token in header_words:
        return 80
    if len(alias_token) >= 4 and alias_token in header_compact:
        return 60
    return 0


def _non_empty_cells(row):
    return sum(1 for item in row if _cell_to_text(item))


def _score_header_row(row):
    score = 0
    for cell in row:
        text = _cell_to_text(cell)
        if not text:
            continue
        for aliases in HEADER_ALIASES.values():
            score += max(_alias_match_score(text, alias) for alias in aliases)
    return score


def _detect_header_row(rows):
    candidates = []
    for index, row in enumerate(rows[:20]):
        candidates.append((_score_header_row(row), _non_empty_cells(row), -index, index))
    best = max(candidates, default=(0, 0, 0, 0))
    if best[0] <= 0:
        return 0
    return best[3]


def _resolve_file_format(original_filename):
    suffix = Path(original_filename or "").suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return "xlsx"
    if suffix == ".csv":
        return "csv"
    if suffix == ".tsv":
        return "tsv"
    return "unknown"


def _guess_delimiter(sample_text):
    try:
        dialect = csv.Sniffer().sniff(sample_text, delimiters=",;\t|")
        return dialect.delimiter
    except csv.Error:
        if sample_text.count(";") > sample_text.count(","):
            return ";"
        if sample_text.count("\t") > sample_text.count(","):
            return "\t"
        return ","


def _read_csv_rows(raw_bytes, file_format):
    decoded = None
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            decoded = raw_bytes.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if decoded is None:
        raise IPImportError("Nao foi possivel decodificar o arquivo CSV.")
    delimiter = "\t" if file_format == "tsv" else _guess_delimiter(decoded[:4000])
    reader = csv.reader(StringIO(decoded), delimiter=delimiter)
    return [[_cell_to_text(cell) for cell in row] for row in reader]


def _read_xlsx_rows(raw_bytes):
    workbook = load_workbook(filename=BytesIO(raw_bytes), data_only=True, read_only=True)
    sheets = []
    for worksheet in workbook.worksheets:
        rows = []
        for row in worksheet.iter_rows(values_only=True):
            rows.append([_cell_to_text(cell) for cell in row])
        sheets.append({"name": worksheet.title, "rows": rows})
    return sheets


def _clean_list_name(value):
    text = _cell_to_text(value)
    if not text:
        return ""
    text = re.sub(r"\.(xlsx|xlsm|csv|tsv)$", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\b(?:ABA|SHEET)\s*\d+\b", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\s+", " ", text).strip(" -_/")
    return text[:120]


def _suggest_list_name(sheet_name, original_filename):
    sheet_label = _clean_list_name(sheet_name)
    if sheet_label and not re.fullmatch(r"(sheet|planilha|arquivo)\s*\d*", _ascii_upper(sheet_label)):
        return sheet_label
    stem = _clean_list_name(Path(original_filename or "").stem)
    return stem or "Lista importada"


def _detect_column_map(headers, ai_result=None):
    detected = {}
    confidence = {}
    ai_headers = (ai_result or {}).get("column_map") or {}
    for field_name, aliases in HEADER_ALIASES.items():
        best = None
        for index, header in enumerate(headers):
            header_text = _cell_to_text(header)
            if not header_text:
                continue
            score = max(_alias_match_score(header_text, alias) for alias in aliases)
            ai_header = _cell_to_text(ai_headers.get(field_name))
            if ai_header and _compact_token(ai_header) == _compact_token(header_text):
                score = max(score, 115)
            if score <= 0:
                continue
            candidate = {"index": index, "header": header_text}
            if best is None or score > best[0]:
                best = (score, candidate)
        if best:
            detected[field_name] = best[1]
            confidence[field_name] = best[0]
    return detected, confidence


def _build_parsed_sheet(sheet_payload, file_format, original_filename, ai_result=None):
    rows = sheet_payload.get("rows") or []
    if not rows or not any(_non_empty_cells(row) for row in rows):
        raise IPImportError("A planilha nao possui linhas validas para importacao.")
    header_row_index = _detect_header_row(rows)
    header_row_index = min(header_row_index, max(len(rows) - 1, 0))
    headers = rows[header_row_index]
    column_map, confidence = _detect_column_map(headers, ai_result=ai_result)
    warnings = []
    if "ip" not in column_map and not {"range_start", "range_end"} <= set(column_map.keys()):
        warnings.append("Nenhuma coluna clara de IP ou faixa de IP foi encontrada.")
    return ParsedSpreadsheet(
        file_format=file_format,
        sheet_name=sheet_payload.get("name") or "Arquivo",
        header_row_index=header_row_index,
        rows_total=len(rows),
        headers=headers,
        raw_rows=rows,
        column_map={
            key: {
                "index": value["index"],
                "header": value["header"],
                "confidence": confidence.get(key, 0),
            }
            for key, value in column_map.items()
        },
        warnings=warnings,
        suggested_list_name=_suggest_list_name(sheet_payload.get("name"), original_filename),
    )


def parse_workbook(raw_bytes, original_filename):
    file_format = _resolve_file_format(original_filename)
    if file_format == "unknown":
        raise IPImportError("Formato nao suportado. Use arquivos .xlsx, .xlsm, .csv ou .tsv.")
    if file_format in {"csv", "tsv"}:
        return [
            _build_parsed_sheet(
                {"name": "Arquivo", "rows": _read_csv_rows(raw_bytes, file_format)},
                file_format=file_format,
                original_filename=original_filename,
            )
        ]

    sheets = _read_xlsx_rows(raw_bytes)
    parsed = []
    for sheet_payload in sheets:
        rows = sheet_payload.get("rows") or []
        if not rows or not any(_non_empty_cells(row) for row in rows):
            continue
        parsed.append(_build_parsed_sheet(sheet_payload, file_format=file_format, original_filename=original_filename))
    if not parsed:
        raise IPImportError("Nenhuma aba valida foi encontrada no arquivo.")
    return parsed


def _build_raw_rows_payload(parsed, max_rows=18, max_cols=10):
    payload = []
    for row in (parsed.raw_rows or [])[:max_rows]:
        payload.append([_cell_to_text(cell) for cell in row[:max_cols]])
    return payload


def _build_workbook_ai_context(parsed_sheets, original_filename):
    return {
        "original_filename": original_filename,
        "sheets": [
            {
                "sheet_name": parsed.sheet_name,
                "rows_total": parsed.rows_total,
                "header_row_number": parsed.header_row_index + 1,
                "headers": parsed.headers[:18],
                "detected_column_map": parsed.column_map,
                "suggested_list_name": parsed.suggested_list_name,
                "parser_warnings": parsed.warnings,
                "raw_rows": _build_raw_rows_payload(parsed, max_rows=18, max_cols=10),
            }
            for parsed in parsed_sheets
        ],
    }


def _find_workbook_sheet_plan(workbook_ai_payload, sheet_name):
    for item in (workbook_ai_payload or {}).get("sheets") or []:
        if _cell_to_text(item.get("sheet_name")) == _cell_to_text(sheet_name):
            return item
    return {}


def _should_skip_sheet_by_ai(workbook_ai_payload, parsed):
    plan = _find_workbook_sheet_plan(workbook_ai_payload, parsed.sheet_name)
    if not plan:
        return False, ""
    if not plan.get("use_sheet"):
        return True, _cell_to_text(plan.get("reason")) or "Guia classificada como auxiliar."
    return False, ""


def _build_workbook_cache_fingerprint(file_sha256, settings_obj, parsed_sheets):
    payload = {
        "file_sha256": file_sha256,
        "settings": _ai_settings_fingerprint(settings_obj),
        "workbook_context": _build_workbook_ai_context(parsed_sheets=parsed_sheets, original_filename=""),
    }
    return _hash_json_payload(payload)


def _build_sheet_cache_fingerprint(file_sha256, settings_obj, parsed, workbook_plan=None):
    payload = {
        "file_sha256": file_sha256,
        "settings": _ai_settings_fingerprint(settings_obj),
        "sheet_name": parsed.sheet_name,
        "headers": parsed.headers,
        "detected_column_map": parsed.column_map,
        "parser_warnings": parsed.warnings,
        "workbook_plan": workbook_plan or {},
        "raw_rows": _build_raw_rows_payload(parsed, max_rows=max(20, int(getattr(settings_obj, "max_rows_for_ai", 180) or 180)), max_cols=12),
    }
    return _hash_json_payload(payload)


def _load_ai_cache(stage, fingerprint):
    from django.db.models import F
    from django.utils import timezone

    from core.models import IPImportAICache

    item = IPImportAICache.objects.filter(stage=stage, fingerprint=fingerprint).first()
    if not item:
        return None
    IPImportAICache.objects.filter(pk=item.pk).update(hits=F("hits") + 1, last_used_at=timezone.now())
    item.hits += 1
    return item.response_payload or {}


def _prune_ai_cache():
    from django.utils import timezone

    from core.models import IPImportAICache

    max_entries = max(30, _ai_cache_max_entries())
    max_age_days = max(1, _ai_cache_max_age_days())
    cutoff = timezone.now() - timedelta(days=max_age_days)
    IPImportAICache.objects.filter(last_used_at__lt=cutoff).delete()
    total = IPImportAICache.objects.count()
    if total <= max_entries:
        return
    stale_ids = list(
        IPImportAICache.objects.order_by("last_used_at").values_list("id", flat=True)[: max(0, total - max_entries)]
    )
    if stale_ids:
        IPImportAICache.objects.filter(id__in=stale_ids).delete()


def _save_ai_cache(stage, fingerprint, file_sha256, settings_obj, response_payload, sheet_name="", payload_meta=None):
    from core.models import IPImportAICache

    IPImportAICache.objects.update_or_create(
        stage=stage,
        fingerprint=fingerprint,
        defaults={
            "file_sha256": _cell_to_text(file_sha256),
            "sheet_name": _cell_to_text(sheet_name),
            "provider": _cell_to_text(getattr(settings_obj, "provider", "")),
            "model": _cell_to_text(getattr(settings_obj, "model", "")),
            "settings_fingerprint": _ai_settings_fingerprint(settings_obj),
            "response_payload": response_payload or {},
            "payload_meta": payload_meta or {},
        },
    )
    _prune_ai_cache()


def _extract_row_value(row, column_map, field_name):
    mapping = column_map.get(field_name) or {}
    index = mapping.get("index")
    if index is None or index >= len(row):
        return ""
    return _cell_to_text(row[index])


def _normalize_mac(value):
    raw = re.sub(r"[^0-9A-F]", "", _ascii_upper(value))
    if not raw:
        return "", False
    if len(raw) != 12:
        return _cell_to_text(value), False
    return ":".join(raw[index:index + 2] for index in range(0, 12, 2)), True


def _expand_ip_range(start, end, limit=4096):
    try:
        start_ip = ipaddress.ip_address((start or "").strip())
        end_ip = ipaddress.ip_address((end or "").strip())
    except ValueError:
        raise IPImportError("Faixa de IP invalida na planilha.")
    if start_ip.version != end_ip.version:
        raise IPImportError("Faixa de IP mistura versoes diferentes.")
    start_int = int(start_ip)
    end_int = int(end_ip)
    if end_int < start_int:
        raise IPImportError("Faixa de IP invalida: o fim e menor que o inicio.")
    total = end_int - start_int + 1
    if total > limit:
        raise IPImportError(f"Faixa de IP excede o limite de {limit} enderecos.")
    return [str(ipaddress.ip_address(value)) for value in range(start_int, end_int + 1)]


def _validate_ip(value):
    text = _cell_to_text(value)
    if not text:
        return ""
    try:
        return str(ipaddress.ip_address(text))
    except ValueError:
        return ""


def build_file_sha256(raw_bytes):
    digest = hashlib.sha256()
    digest.update(raw_bytes or b"")
    return digest.hexdigest()


def _most_common(values, default=""):
    filtered = [_cell_to_text(value) for value in values if _cell_to_text(value)]
    if not filtered:
        return default
    return Counter(filtered).most_common(1)[0][0]


def _is_sparse_ip_set(ip_values):
    if len(ip_values) <= 1:
        return False
    ordered = sorted(ipaddress.ip_address(ip_value) for ip_value in ip_values)
    return int(ordered[-1]) - int(ordered[0]) + 1 != len(ordered)


def _build_sheet_warning(sheet_name, warning, multi_sheet):
    text = _cell_to_text(warning)
    if not text:
        return ""
    if multi_sheet and sheet_name:
        return f"[{sheet_name}] {text}"
    return text


def _dedupe_ip_items(rows):
    deduped = {}
    duplicates = 0
    for row in rows:
        key = row["ip"]
        if key in deduped:
            duplicates += 1
            existing = deduped[key]
            for field_name in ("device_name", "description", "mac", "protocol", "list_code"):
                if not existing.get(field_name) and row.get(field_name):
                    existing[field_name] = row[field_name]
            continue
        deduped[key] = dict(row)
    return list(deduped.values()), duplicates


def _compose_description(explicit_description, controller, drive, status_note, novelty):
    explicit = _cell_to_text(explicit_description)
    if explicit:
        return explicit[:200]

    parts = []
    controller_text = _cell_to_text(controller)
    drive_text = _cell_to_text(drive)
    status_text = _cell_to_text(status_note)
    novelty_text = _cell_to_text(novelty)

    if controller_text:
        parts.append(f"PLC {controller_text}")
    if drive_text:
        parts.append(f"Drive {drive_text}")
    if status_text:
        parts.append(f"Status {status_text}")
    if novelty_text:
        parts.append(novelty_text)
    return " | ".join(parts)[:200]


def normalize_rows(parsed, ai_result=None):
    column_map, _ = _detect_column_map(parsed.headers, ai_result=ai_result)
    if "ip" not in column_map and not {"range_start", "range_end"} <= set(column_map.keys()):
        raise IPImportError("Nao foi possivel identificar uma coluna de IP ou uma faixa valida.")

    default_list_name = _clean_list_name((ai_result or {}).get("default_list_name")) or parsed.suggested_list_name
    warnings = list(parsed.warnings or [])
    normalized_rows = []

    for row_index, row in enumerate(parsed.raw_rows[parsed.header_row_index + 1 :], start=parsed.header_row_index + 2):
        if not any(_cell_to_text(cell) for cell in row):
            continue

        ip_text = _validate_ip(_extract_row_value(row, column_map, "ip"))
        range_start = _extract_row_value(row, column_map, "range_start")
        range_end = _extract_row_value(row, column_map, "range_end")
        device_name = _extract_row_value(row, column_map, "device_name")
        description = _extract_row_value(row, column_map, "description")
        protocol = _extract_row_value(row, column_map, "protocol")
        controller = _extract_row_value(row, column_map, "controller")
        drive = _extract_row_value(row, column_map, "drive")
        status_note = _extract_row_value(row, column_map, "status_note")
        novelty = _extract_row_value(row, column_map, "novelty")
        list_name = _clean_list_name(_extract_row_value(row, column_map, "list_name")) or default_list_name
        list_code = _cell_to_text(_extract_row_value(row, column_map, "list_code")).upper()
        mac_value, mac_ok = _normalize_mac(_extract_row_value(row, column_map, "mac"))
        description_value = _compose_description(description, controller, drive, status_note, novelty)

        if not ip_text and not (range_start and range_end):
            if any((device_name, description_value, protocol, mac_value)):
                warnings.append(f"Linha {row_index}: ignorada porque nao possui IP ou faixa.")
            continue

        if mac_value and not mac_ok:
            warnings.append(f"Linha {row_index}: MAC mantido como texto bruto porque o formato parece invalido.")

        if ip_text:
            ip_values = [ip_text]
        else:
            try:
                ip_values = _expand_ip_range(range_start, range_end)
            except IPImportError as exc:
                warnings.append(f"Linha {row_index}: {exc}")
                continue

        for ip_value in ip_values:
            normalized_rows.append(
                {
                    "source_sheet": parsed.sheet_name,
                    "source_row": row_index,
                    "list_name": list_name[:120],
                    "list_code": list_code[:60],
                    "ip": ip_value,
                    "device_name": device_name[:120],
                    "description": description_value[:200],
                    "mac": mac_value[:30],
                    "protocol": protocol[:30],
                }
            )

    if not normalized_rows:
        raise IPImportError("Nenhuma linha com IP valido foi encontrada depois do cabecalho.")

    effective_map = {field_name: payload.get("header", "") for field_name, payload in column_map.items()}
    return normalized_rows, effective_map, warnings


def normalize_rows_from_ai_result(parsed, ai_result):
    warnings = list(parsed.warnings or [])
    warnings.extend((ai_result or {}).get("warnings") or [])
    normalized_rows = []

    for item in (ai_result or {}).get("logical_items") or []:
        if not isinstance(item, dict):
            continue
        ip_value = _validate_ip(item.get("ip"))
        if not ip_value:
            range_start = _cell_to_text(item.get("range_start"))
            range_end = _cell_to_text(item.get("range_end"))
            if range_start and range_end:
                try:
                    ip_values = _expand_ip_range(range_start, range_end)
                except IPImportError as exc:
                    source_row = item.get("source_row") or "?"
                    warnings.append(f"Linha {source_row}: {exc}")
                    continue
            else:
                continue
        else:
            ip_values = [ip_value]

        list_name = _clean_list_name(item.get("list_name")) or _clean_list_name(
            (ai_result or {}).get("default_list_name")
        ) or parsed.suggested_list_name
        list_code = _cell_to_text(item.get("list_code")).upper() or _cell_to_text(
            (ai_result or {}).get("default_list_code")
        ).upper()
        device_name = _cell_to_text(item.get("device_name"))[:120]
        description = _cell_to_text(item.get("description"))[:200]
        protocol = _cell_to_text(item.get("protocol"))[:30]
        mac_value, mac_ok = _normalize_mac(item.get("mac"))
        if mac_value and not mac_ok:
            source_row = item.get("source_row") or "?"
            warnings.append(f"Linha {source_row}: MAC mantido como texto bruto porque o formato parece invalido.")
        source_row = int(item.get("source_row") or 0) or parsed.header_row_index + 2

        for expanded_ip in ip_values:
            normalized_rows.append(
                {
                    "source_sheet": parsed.sheet_name,
                    "source_row": source_row,
                    "list_name": list_name[:120],
                    "list_code": list_code[:60],
                    "ip": expanded_ip,
                    "device_name": device_name,
                    "description": description,
                    "mac": mac_value[:30],
                    "protocol": protocol,
                }
            )

    if not normalized_rows:
        raise IPImportError("A IA nao retornou itens operacionais suficientes para montar a lista de IP.")

    effective_map = {field_name: _cell_to_text(value) for field_name, value in ((ai_result or {}).get("column_map") or {}).items()}
    return normalized_rows, effective_map, warnings


def build_import_proposal(original_filename, normalized_rows):
    grouped = defaultdict(list)
    for row in normalized_rows:
        group_key = _compact_token(row.get("list_code") or row.get("list_name") or row.get("source_sheet")) or f"LIST_{len(grouped) + 1}"
        grouped[group_key].append(row)

    proposal_lists = []
    warnings = []
    total_items = 0

    for list_index, (list_token, group_rows) in enumerate(grouped.items(), start=1):
        deduped_rows, duplicate_count = _dedupe_ip_items(group_rows)
        if duplicate_count:
            warnings.append(f"Lista {list_index}: {duplicate_count} IP(s) duplicado(s) foram consolidados na preview.")

        sorted_rows = sorted(deduped_rows, key=lambda item: ipaddress.ip_address(item["ip"]))
        ip_values = [item["ip"] for item in sorted_rows]
        sparse = _is_sparse_ip_set(ip_values)
        list_name = _clean_list_name(_most_common(item.get("list_name") for item in sorted_rows)) or _suggest_list_name(
            sorted_rows[0].get("source_sheet"),
            original_filename,
        )
        list_code = _most_common(item.get("list_code") for item in sorted_rows)
        protocol = _most_common(item.get("protocol") for item in sorted_rows)
        description = _most_common(item.get("description") for item in sorted_rows)
        source_sheets = sorted({item.get("source_sheet") for item in sorted_rows if item.get("source_sheet")})
        total_items += len(sorted_rows)
        if sparse:
            warnings.append(f"Lista {list_name or list_index}: a faixa resume IPs nao contiguos; os itens individuais foram preservados.")
        proposal_lists.append(
            {
                "list_key": f"list_{list_index}_{list_token.lower()}",
                "name": list_name or f"Lista importada {list_index}",
                "id_listaip": list_code,
                "description": description,
                "faixa_inicio": ip_values[0],
                "faixa_fim": ip_values[-1],
                "protocolo_padrao": protocol,
                "total_ips": len(sorted_rows),
                "filled_devices": sum(1 for item in sorted_rows if item.get("device_name")),
                "source_sheets": source_sheets,
                "is_sparse": sparse,
                "items": sorted_rows,
                "preview_items": sorted_rows[:18],
            }
        )

    return {
        "lists": proposal_lists,
        "warnings": warnings,
        "conflicts": [],
        "summary": {
            "lists": len(proposal_lists),
            "items": total_items,
            "rows": len(normalized_rows),
        },
    }


def _build_progress_list_snapshots_from_rows(normalized_rows, limit=4):
    grouped = defaultdict(list)
    for row in normalized_rows or []:
        key = _compact_token(row.get("list_code") or row.get("list_name") or row.get("source_sheet")) or str(len(grouped) + 1)
        grouped[key].append(row)

    snapshots = []
    for _, rows in list(grouped.items())[:limit]:
        sorted_rows = sorted(rows, key=lambda item: (_cell_to_text(item.get("list_name")), item.get("ip")))
        names = [item.get("device_name") for item in sorted_rows if item.get("device_name")][:3]
        ip_values = [item.get("ip") for item in sorted_rows if item.get("ip")]
        snapshots.append(
            {
                "list_name": _clean_list_name(sorted_rows[0].get("list_name")) or _clean_list_name(sorted_rows[0].get("source_sheet")) or "Lista em analise",
                "items": len(sorted_rows),
                "devices": sum(1 for item in sorted_rows if item.get("device_name")),
                "sample_names": names,
                "range_start": ip_values[0] if ip_values else "",
                "range_end": ip_values[-1] if ip_values else "",
            }
        )
    return snapshots


def _build_progress_list_snapshots_from_proposal(proposal, limit=4):
    snapshots = []
    for item in (proposal or {}).get("lists") or []:
        rows = item.get("items") or []
        names = [row.get("device_name") for row in rows if row.get("device_name")][:3]
        snapshots.append(
            {
                "list_name": _cell_to_text(item.get("name")) or "Lista importada",
                "items": int(item.get("total_ips") or len(rows)),
                "devices": int(item.get("filled_devices") or 0),
                "sample_names": names,
                "range_start": _cell_to_text(item.get("faixa_inicio")),
                "range_end": _cell_to_text(item.get("faixa_fim")),
            }
        )
        if len(snapshots) >= limit:
            break
    return snapshots


def _extract_response_text(response_payload):
    if not isinstance(response_payload, dict):
        return ""
    output = response_payload.get("output") or []
    for item in output:
        for content in item.get("content") or []:
            if content.get("type") in {"output_text", "text"} and content.get("text"):
                return content.get("text")
    if response_payload.get("output_text"):
        return response_payload.get("output_text")
    return ""


class _OpenAITransientError(Exception):
    pass


def _extract_response_error_message(payload):
    error = payload.get("error") or {}
    if isinstance(error, dict):
        code = _cell_to_text(error.get("code"))
        message = _cell_to_text(error.get("message"))
        if code and message:
            return f"{code}: {message}"
        if message:
            return message
    incomplete_details = payload.get("incomplete_details") or {}
    if isinstance(incomplete_details, dict):
        reason = _cell_to_text(incomplete_details.get("reason"))
        if reason:
            return reason
    return ""


def _openai_response_is_running(status):
    return _cell_to_text(status).lower() in {"queued", "in_progress", "running"}


def _openai_poll_interval_seconds(elapsed_seconds):
    if elapsed_seconds < 30:
        return 2.0
    if elapsed_seconds < 120:
        return 3.0
    return 5.0


def _openai_request_json(method, url, api_key, timeout_seconds, payload=None):
    headers = {"Authorization": f"Bearer {api_key}"}
    data = None
    if payload is not None:
        headers["Content-Type"] = "application/json"
        data = json.dumps(payload).encode("utf-8")
    http_request = urlrequest.Request(url=url, data=data, method=method, headers=headers)
    transient_http_codes = {408, 409, 425, 429, 500, 502, 503, 504}
    try:
        with urlrequest.urlopen(http_request, timeout=timeout_seconds) as response:
            raw_body = response.read().decode("utf-8")
        return json.loads(raw_body) if raw_body else {}
    except urlerror.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")
        message = f"Falha na chamada do agente: HTTP {exc.code} - {detail[:400]}"
        if exc.code in transient_http_codes:
            raise _OpenAITransientError(message)
        raise IPImportError(message)
    except urlerror.URLError as exc:
        raise _OpenAITransientError(f"Falha na chamada do agente: {exc.reason}")
    except (TimeoutError, socket.timeout):
        raise _OpenAITransientError(f"Falha na chamada do agente: timeout de transporte apos {timeout_seconds:.1f}s.")
    except OSError as exc:
        raise _OpenAITransientError(f"Falha na chamada do agente: {exc}")


def _call_openai_responses(settings_obj, schema_name, schema, system_prompt, user_prompt, request_timeout_seconds=None):
    if not settings_obj.enabled:
        raise IPImportError("Agente de importacao desativado.")
    if settings_obj.provider != settings_obj.Provider.OPENAI:
        raise IPImportError("Provider de agente nao suportado nesta versao.")
    if not settings_obj.api_key:
        raise IPImportError("API key do agente nao configurada.")

    base_url = (settings_obj.api_base_url or "").strip().rstrip("/") or "https://api.openai.com/v1"
    payload = {
        "model": settings_obj.model,
        "input": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
        "text": {
            "format": {
                "type": "json_schema",
                "name": schema_name,
                "strict": True,
                "schema": schema,
            }
        },
    }
    if settings_obj.reasoning_effort and settings_obj.reasoning_effort != "none":
        payload["reasoning"] = {"effort": settings_obj.reasoning_effort}

    payload["background"] = True
    payload["store"] = True
    payload["metadata"] = {
        "schema_name": schema_name[:64],
        "source": "ip_import",
        "client_request_id": uuid.uuid4().hex[:32],
    }

    transport_timeout_seconds = max(
        15.0,
        float(request_timeout_seconds or _ai_request_timeout_seconds() or DEFAULT_AI_REQUEST_TIMEOUT_SECONDS),
    )
    max_attempts = max(1, _ai_max_retries())
    created_payload = None
    last_error = ""
    for attempt in range(1, max_attempts + 1):
        try:
            created_payload = _openai_request_json(
                method="POST",
                url=f"{base_url}/responses",
                api_key=settings_obj.api_key,
                timeout_seconds=transport_timeout_seconds,
                payload=payload,
            )
            break
        except _OpenAITransientError as exc:
            last_error = str(exc)
            if attempt < max_attempts:
                time.sleep(min(4.0, 1.25 * attempt))
                continue
            raise IPImportError(last_error)

    if created_payload is None:
        raise IPImportError(last_error or "Falha na chamada do agente.")

    response_payload = created_payload
    response_id = _cell_to_text(response_payload.get("id"))
    if not response_id:
        raise IPImportError("O agente nao retornou um identificador de resposta utilizavel.")

    polling_started_at = time.monotonic()
    transient_poll_errors = 0
    while True:
        response_text = _extract_response_text(response_payload)
        if response_text:
            break

        status = _cell_to_text(response_payload.get("status")).lower()
        if not _openai_response_is_running(status):
            error_detail = _extract_response_error_message(response_payload)
            raise IPImportError(
                error_detail or f"O agente finalizou sem texto utilizavel (status {status or 'desconhecido'})."
            )

        elapsed_seconds = time.monotonic() - polling_started_at
        time.sleep(_openai_poll_interval_seconds(elapsed_seconds))
        try:
            response_payload = _openai_request_json(
                method="GET",
                url=f"{base_url}/responses/{response_id}",
                api_key=settings_obj.api_key,
                timeout_seconds=transport_timeout_seconds,
            )
            transient_poll_errors = 0
        except _OpenAITransientError as exc:
            transient_poll_errors += 1
            last_error = str(exc)
            if transient_poll_errors >= max_attempts:
                raise IPImportError(last_error)
            continue

    try:
        return json.loads(response_text)
    except json.JSONDecodeError as exc:
        raise IPImportError(f"O agente retornou JSON invalido: {exc}")


def run_ai_workbook_analysis(settings_obj, parsed_sheets, original_filename, file_sha256="", request_timeout_seconds=None):
    schema = {
        "type": "object",
        "properties": {
            "sheets": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "sheet_name": {"type": "string"},
                        "use_sheet": {"type": "boolean"},
                        "sheet_role": {"type": "string", "enum": ["data", "summary", "helper", "noise"]},
                        "default_list_name": {"type": "string"},
                        "default_list_code": {"type": "string"},
                        "confidence": {"type": "integer"},
                        "reason": {"type": "string"},
                    },
                    "required": [
                        "sheet_name",
                        "use_sheet",
                        "sheet_role",
                        "default_list_name",
                        "default_list_code",
                        "confidence",
                        "reason",
                    ],
                    "additionalProperties": False,
                },
            },
            "warnings": {"type": "array", "items": {"type": "string"}},
            "notes": {"type": "string"},
        },
        "required": ["sheets", "warnings", "notes"],
        "additionalProperties": False,
    }
    cache_fingerprint = _build_workbook_cache_fingerprint(file_sha256=file_sha256, settings_obj=settings_obj, parsed_sheets=parsed_sheets)
    cached_payload = _load_ai_cache("WORKBOOK", cache_fingerprint)
    if cached_payload:
        return cached_payload

    system_prompt = (
        "You are the primary interpreter for a heterogeneous industrial IP workbook. "
        "Decide which sheets contain operational IP list data and which sheets are summary, helper or noise. "
        "Prefer skipping summary/index sheets when they only repeat data already represented in dedicated sheets. "
        "Return stable default names/codes when the workbook structure strongly implies them, but do not invent network ranges."
    )
    user_prompt = json.dumps(_build_workbook_ai_context(parsed_sheets=parsed_sheets, original_filename=original_filename), ensure_ascii=True)
    response_payload = _call_openai_responses(
        settings_obj=settings_obj,
        schema_name="ip_workbook_analysis",
        schema=schema,
        system_prompt=system_prompt,
        user_prompt=user_prompt,
        request_timeout_seconds=request_timeout_seconds,
    )
    _save_ai_cache(
        stage="WORKBOOK",
        fingerprint=cache_fingerprint,
        file_sha256=file_sha256,
        settings_obj=settings_obj,
        response_payload=response_payload,
        payload_meta={"sheets_total": len(parsed_sheets or [])},
    )
    return response_payload


def run_ai_analysis(settings_obj, parsed, workbook_plan=None, file_sha256="", request_timeout_seconds=None):
    schema = {
        "type": "object",
        "properties": {
            "skip_sheet": {"type": "boolean"},
            "sheet_role": {"type": "string", "enum": ["data", "summary", "helper", "noise"]},
            "default_list_name": {"type": "string"},
            "default_list_code": {"type": "string"},
            "column_map": {
                "type": "object",
                "properties": {
                    "list_name": {"type": "string"},
                    "list_code": {"type": "string"},
                    "ip": {"type": "string"},
                    "device_name": {"type": "string"},
                    "description": {"type": "string"},
                    "mac": {"type": "string"},
                    "protocol": {"type": "string"},
                    "range_start": {"type": "string"},
                    "range_end": {"type": "string"},
                },
                "required": [
                    "list_name",
                    "list_code",
                    "ip",
                    "device_name",
                    "description",
                    "mac",
                    "protocol",
                    "range_start",
                    "range_end",
                ],
                "additionalProperties": False,
            },
            "logical_items": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "source_row": {"type": "integer"},
                        "list_name": {"type": "string"},
                        "list_code": {"type": "string"},
                        "ip": {"type": "string"},
                        "range_start": {"type": "string"},
                        "range_end": {"type": "string"},
                        "device_name": {"type": "string"},
                        "description": {"type": "string"},
                        "mac": {"type": "string"},
                        "protocol": {"type": "string"},
                        "confidence": {"type": "integer"},
                    },
                    "required": [
                        "source_row",
                        "list_name",
                        "list_code",
                        "ip",
                        "range_start",
                        "range_end",
                        "device_name",
                        "description",
                        "mac",
                        "protocol",
                        "confidence",
                    ],
                    "additionalProperties": False,
                },
            },
            "warnings": {"type": "array", "items": {"type": "string"}},
            "notes": {"type": "string"},
        },
        "required": [
            "skip_sheet",
            "sheet_role",
            "default_list_name",
            "default_list_code",
            "column_map",
            "logical_items",
            "warnings",
            "notes",
        ],
        "additionalProperties": False,
    }
    cache_fingerprint = _build_sheet_cache_fingerprint(
        file_sha256=file_sha256,
        settings_obj=settings_obj,
        parsed=parsed,
        workbook_plan=workbook_plan,
    )
    cached_payload = _load_ai_cache("SHEET", cache_fingerprint)
    if cached_payload:
        return cached_payload

    raw_rows_limit = max(20, min(int(getattr(settings_obj, "max_rows_for_ai", 180) or 180), max(parsed.rows_total, 20)))
    context_payload = {
        "sheet_name": parsed.sheet_name,
        "header_row_number": parsed.header_row_index + 1,
        "headers": parsed.headers,
        "detected_column_map": parsed.column_map,
        "suggested_list_name": parsed.suggested_list_name,
        "parser_warnings": parsed.warnings,
        "workbook_plan": workbook_plan or {},
        "raw_rows": _build_raw_rows_payload(parsed, max_rows=raw_rows_limit, max_cols=12),
    }
    response_payload = _call_openai_responses(
        settings_obj=settings_obj,
        schema_name="ip_sheet_semantic_analysis",
        schema=schema,
        system_prompt=(
            "You are the primary interpreter for an industrial IP list sheet. "
            "Your job is to identify the operational IP entries, determine the correct columns and produce normalized logical items. "
            "Summary/helper rows must be skipped. When there is a clear data sheet, return every operational entry you can recover. "
            f"{settings_obj.header_prompt}\n\n{settings_obj.grouping_prompt}"
        ),
        user_prompt=json.dumps(context_payload, ensure_ascii=True),
        request_timeout_seconds=request_timeout_seconds,
    )
    _save_ai_cache(
        stage="SHEET",
        fingerprint=cache_fingerprint,
        file_sha256=file_sha256,
        settings_obj=settings_obj,
        response_payload=response_payload,
        sheet_name=parsed.sheet_name,
        payload_meta={"rows_total": parsed.rows_total},
    )
    return response_payload


def reprocess_import_job(job, settings_obj=None, progress_callback=None):
    if not job.source_file:
        raise IPImportError("Arquivo fonte da importacao nao encontrado.")
    ai_required = bool(settings_obj and settings_obj.enabled)
    _emit_progress(
        progress_callback,
        stage="upload",
        percent=6,
        title="Arquivo recebido",
        message="O arquivo foi recebido e a leitura inicial da planilha esta comecando.",
        progress_label="Arquivo recebido",
        snapshots=[],
    )
    job.source_file.open("rb")
    raw_bytes = job.source_file.read()
    job.source_file.close()

    parsed_sheets = parse_workbook(raw_bytes=raw_bytes, original_filename=job.original_filename)
    _emit_progress(
        progress_callback,
        stage="parse",
        percent=18,
        title="Estrutura da planilha identificada",
        message=f"{len(parsed_sheets)} guia(s) localizada(s). A estrutura base esta sendo organizada para a analise.",
        progress_label="Estrutura identificada",
        sheets_total=len(parsed_sheets),
        sheets_processed=0,
        snapshots=[],
    )
    multi_sheet = len(parsed_sheets) > 1
    normalized_rows = []
    warnings = []
    effective_map = {}
    sheet_summaries = []
    ai_status = job.AIStatus.SKIPPED
    ai_payload = {"sheets": {}}
    ai_errors = []
    ai_model = settings_obj.model if settings_obj and settings_obj.enabled else ""
    ai_success = 0
    ai_attempts = 0
    processed_sheets = []
    ai_max_sheets = _ai_max_sheets_per_job()
    workbook_ai_payload = {}
    workbook_ai_error = ""

    if settings_obj and settings_obj.enabled:
        _emit_progress(
            progress_callback,
            stage="ai",
            percent=28,
            title="Organizando o contexto da planilha",
            message="A IA esta distinguindo guias operacionais, resumos e agrupamentos antes da leitura detalhada.",
            progress_label="Separando guias uteis",
            sheets_total=len(parsed_sheets),
            sheets_processed=0,
            snapshots=[],
        )
        if len(parsed_sheets) == 1:
            only_sheet = parsed_sheets[0]
            workbook_ai_payload = {
                "sheets": [
                    {
                        "sheet_name": only_sheet.sheet_name,
                        "use_sheet": True,
                        "sheet_role": "data",
                        "default_list_name": only_sheet.suggested_list_name,
                        "default_list_code": "",
                        "confidence": 92,
                        "reason": "Guia unica priorizada para leitura semantica detalhada.",
                    }
                ],
                "warnings": [],
                "notes": "single-sheet-fast-path",
            }
            ai_success += 1
        elif ai_attempts >= ai_max_sheets:
            workbook_ai_error = "A analise com IA excedeu o limite de chamadas previsto para esta importacao."
            ai_errors.append(workbook_ai_error)
            raise IPImportError(workbook_ai_error)
        else:
            try:
                ai_attempts += 1
                workbook_ai_payload = run_ai_workbook_analysis(
                    settings_obj=settings_obj,
                    parsed_sheets=parsed_sheets,
                    original_filename=job.original_filename,
                    file_sha256=job.file_sha256,
                )
                ai_success += 1
            except IPImportError as exc:
                workbook_ai_error = str(exc)
                ai_errors.append(workbook_ai_error)
                raise IPImportError(
                    f"A analise com IA nao conseguiu entender a estrutura geral da planilha. {workbook_ai_error}"
                ) from exc

    total_sheets = max(len(parsed_sheets), 1)
    processed_count = 0
    for sheet_index, parsed in enumerate(parsed_sheets, start=1):
        sheet_ai_payload = {}
        sheet_warnings = []
        skip_by_ai, skip_reason = _should_skip_sheet_by_ai(workbook_ai_payload, parsed)
        current_percent = 34 + int(((sheet_index - 1) / total_sheets) * 46)
        _emit_progress(
            progress_callback,
            stage="ai" if settings_obj and settings_obj.enabled else "parse",
            percent=current_percent,
            title="Correlacionando guias e enderecos",
            message=f"Analisando a guia {sheet_index} de {total_sheets}: {parsed.sheet_name}.",
            progress_label=f"Guia {sheet_index} de {total_sheets}",
            current_sheet=parsed.sheet_name,
            current_sheet_index=sheet_index,
            sheets_total=total_sheets,
            sheets_processed=processed_count,
            snapshots=_build_progress_list_snapshots_from_rows(normalized_rows),
        )

        if skip_by_ai:
            formatted_skip_reason = f"Guia ignorada pela IA: {skip_reason}" if skip_reason else "Guia ignorada pela IA."
            warnings.append(_build_sheet_warning(parsed.sheet_name, formatted_skip_reason, multi_sheet))
            ai_payload["sheets"][parsed.sheet_name] = {
                "sheet_role": "summary",
                "skip_sheet": True,
                "warnings": [formatted_skip_reason],
            }
            sheet_summaries.append(
                {
                    "sheet_name": parsed.sheet_name,
                    "header_row_index": parsed.header_row_index + 1,
                    "rows_total": parsed.rows_total,
                    "rows_parsed": 0,
                    "column_map": {},
                    "skipped": True,
                }
            )
            processed_count += 1
            continue

        try:
            if settings_obj and settings_obj.enabled:
                if ai_attempts >= ai_max_sheets:
                    raise IPImportError(
                        f"A analise com IA atingiu o limite previsto antes de concluir a guia {parsed.sheet_name}."
                    )
                ai_attempts += 1
                sheet_ai_payload = run_ai_analysis(
                    settings_obj=settings_obj,
                    parsed=parsed,
                    workbook_plan=_find_workbook_sheet_plan(workbook_ai_payload, parsed.sheet_name),
                    file_sha256=job.file_sha256,
                )
                if sheet_ai_payload.get("skip_sheet"):
                    skip_reason = _cell_to_text((sheet_ai_payload.get("warnings") or [""])[0]) or "Guia sem dados operacionais."
                    warnings.append(_build_sheet_warning(parsed.sheet_name, f"Guia ignorada pela IA: {skip_reason}", multi_sheet))
                    ai_payload["sheets"][parsed.sheet_name] = sheet_ai_payload
                    sheet_summaries.append(
                        {
                            "sheet_name": parsed.sheet_name,
                            "header_row_index": parsed.header_row_index + 1,
                            "rows_total": parsed.rows_total,
                            "rows_parsed": 0,
                            "column_map": {},
                            "skipped": True,
                        }
                    )
                    ai_success += 1
                    processed_count += 1
                    continue
                if sheet_ai_payload.get("logical_items"):
                    sheet_rows, sheet_map, sheet_warnings = normalize_rows_from_ai_result(parsed=parsed, ai_result=sheet_ai_payload)
                else:
                    sheet_rows, sheet_map, sheet_warnings = normalize_rows(parsed=parsed, ai_result=sheet_ai_payload)
                ai_success += 1
            else:
                sheet_rows, sheet_map, sheet_warnings = normalize_rows(parsed=parsed, ai_result=None)
        except IPImportError as exc:
            ai_errors.append(_build_sheet_warning(parsed.sheet_name, str(exc), multi_sheet))
            if ai_required:
                raise IPImportError(f"A analise com IA nao conseguiu concluir a guia {parsed.sheet_name}. {exc}") from exc
            warnings.append(_build_sheet_warning(parsed.sheet_name, str(exc), multi_sheet))
            sheet_summaries.append(
                {
                    "sheet_name": parsed.sheet_name,
                    "header_row_index": parsed.header_row_index + 1,
                    "rows_total": parsed.rows_total,
                    "rows_parsed": 0,
                    "column_map": {},
                    "skipped": True,
                }
            )
            processed_count += 1
            continue

        normalized_rows.extend(sheet_rows)
        effective_map[parsed.sheet_name] = sheet_map
        ai_payload["sheets"][parsed.sheet_name] = sheet_ai_payload
        sheet_summaries.append(
            {
                "sheet_name": parsed.sheet_name,
                "header_row_index": parsed.header_row_index + 1,
                "rows_total": parsed.rows_total,
                "rows_parsed": len(sheet_rows),
                "column_map": sheet_map,
            }
        )
        processed_sheets.append(parsed)
        processed_count += 1
        for warning in sheet_warnings:
            formatted = _build_sheet_warning(parsed.sheet_name, warning, multi_sheet)
            if formatted:
                warnings.append(formatted)
        _emit_progress(
            progress_callback,
            stage="ai" if settings_obj and settings_obj.enabled else "parse",
            percent=34 + int((processed_count / total_sheets) * 46),
            title="Guia consolidada",
            message=f"{parsed.sheet_name} foi consolidada. {len(normalized_rows)} endereco(s) validado(s) ate agora.",
            progress_label=f"Guia {processed_count} de {total_sheets}",
            current_sheet=parsed.sheet_name,
            current_sheet_index=processed_count,
            sheets_total=total_sheets,
            sheets_processed=processed_count,
            snapshots=_build_progress_list_snapshots_from_rows(normalized_rows),
        )

    if not normalized_rows:
        raise IPImportError("Nenhuma aba util gerou linhas normalizadas para importacao.")

    if settings_obj and settings_obj.enabled:
        expected_successes = len(processed_sheets) + (1 if workbook_ai_payload else 0)
        ai_status = job.AIStatus.SUCCESS if ai_success >= max(expected_successes, 1) else job.AIStatus.FAILED
        ai_error = " | ".join(error for error in ai_errors if error)
        if ai_error:
            warnings.extend(error for error in ai_errors if error)
    else:
        ai_error = ""

    if workbook_ai_payload:
        ai_payload["workbook"] = workbook_ai_payload
    elif workbook_ai_error:
        ai_payload["workbook_error"] = workbook_ai_error

    _emit_progress(
        progress_callback,
        stage="preview",
        percent=90,
        title="Montando a previa para revisao",
        message="As listas sugeridas estao sendo organizadas para abrir a revisao final.",
        progress_label="Montando a previa",
        sheets_total=total_sheets,
        sheets_processed=processed_count,
        snapshots=_build_progress_list_snapshots_from_rows(normalized_rows),
    )
    proposal = build_import_proposal(original_filename=job.original_filename, normalized_rows=normalized_rows)
    for warning in proposal.get("warnings") or []:
        warnings.append(warning)
    final_snapshots = _build_progress_list_snapshots_from_proposal(proposal)
    _emit_progress(
        progress_callback,
        stage="preview",
        percent=100,
        title="Preview pronta",
        message="A estrutura sugerida foi consolidada e esta pronta para revisao.",
        progress_label="Preview pronta",
        sheets_total=total_sheets,
        sheets_processed=processed_count,
        snapshots=final_snapshots,
        summary=proposal.get("summary") or {},
    )

    primary_sheet = processed_sheets[0]
    sheet_label = primary_sheet.sheet_name if len(processed_sheets) == 1 else f"{primary_sheet.sheet_name} +{len(processed_sheets) - 1}"
    return {
        "file_format": primary_sheet.file_format,
        "sheet_name": sheet_label,
        "header_row_index": primary_sheet.header_row_index + 1 if len(processed_sheets) == 1 else None,
        "rows_total": sum(item.rows_total for item in processed_sheets),
        "rows_parsed": len(normalized_rows),
        "column_map": effective_map,
        "warnings": warnings,
        "normalized_rows": normalized_rows,
        "sheet_summaries": sheet_summaries,
        "proposal": proposal,
        "ai_status": ai_status,
        "ai_payload": ai_payload,
        "ai_error": ai_error,
        "ai_model": ai_model,
        "progress_payload": {
            "stage": "preview",
            "percent": 100,
            "title": "Preview pronta",
            "message": "A estrutura sugerida foi consolidada e esta pronta para revisao.",
            "progress_label": "Preview pronta",
            "sheets_total": total_sheets,
            "sheets_processed": processed_count,
            "snapshots": final_snapshots,
            "summary": proposal.get("summary") or {},
        },
    }


def apply_import_job(job, user, selected_list_keys=None):
    from django.db import transaction

    original_job = job
    with transaction.atomic():
        job = type(job).objects.select_for_update().get(pk=original_job.pk)
        proposal = job.proposal_payload or {}
        list_payloads = proposal.get("lists") or []
        if not list_payloads:
            raise IPImportError("A importacao nao possui listas prontas para aplicacao.")
        if not job.cliente:
            raise IPImportError("A importacao nao possui cliente associado.")

        apply_log = dict(job.apply_log or {})
        applied_list_keys = dict(apply_log.get("applied_list_keys") or {})
        applied_list_ids = list(apply_log.get("applied_list_ids") or [])
        requested_keys = {str(item) for item in (selected_list_keys or [payload.get("list_key") for payload in list_payloads]) if str(item)}
        applied = []

        for payload in list_payloads:
            list_key = str(payload.get("list_key") or "")
            if not list_key or list_key not in requested_keys:
                continue

            existing_id = applied_list_keys.get(list_key)
            if existing_id:
                existing_lista = ListaIP.objects.filter(pk=existing_id).first()
                if existing_lista:
                    applied.append(existing_lista)
                    continue

            id_listaip_obj = None
            if payload.get("id_listaip"):
                id_listaip_obj, _ = ListaIPID.objects.get_or_create(codigo=_cell_to_text(payload.get("id_listaip")).upper())

            descricao = _cell_to_text(payload.get("description"))
            if payload.get("is_sparse"):
                sparse_note = "Importado de planilha com IPs nao contiguos."
                descricao = f"{descricao} {sparse_note}".strip() if descricao else sparse_note

            lista = ListaIP.objects.create(
                cliente=job.cliente,
                id_listaip=id_listaip_obj,
                nome=_cell_to_text(payload.get("name"))[:120] or f"Lista importada {len(applied_list_keys) + 1}",
                descricao=descricao[:500],
                faixa_inicio=_cell_to_text(payload.get("faixa_inicio")),
                faixa_fim=_cell_to_text(payload.get("faixa_fim")),
                protocolo_padrao=_cell_to_text(payload.get("protocolo_padrao"))[:30],
            )

            item_payloads, _ = _dedupe_ip_items(payload.get("items") or [])
            ListaIPItem.objects.bulk_create(
                [
                    ListaIPItem(
                        lista=lista,
                        ip=item["ip"],
                        nome_equipamento=_cell_to_text(item.get("device_name"))[:120],
                        descricao=_cell_to_text(item.get("description"))[:200],
                        mac=_cell_to_text(item.get("mac"))[:30],
                        protocolo=_cell_to_text(item.get("protocol"))[:30],
                    )
                    for item in item_payloads
                ]
            )

            applied.append(lista)
            applied_list_keys[list_key] = lista.pk
            if lista.pk not in applied_list_ids:
                applied_list_ids.append(lista.pk)
            job.applied_lista = lista

        if not applied:
            raise IPImportError("Nenhuma lista foi aplicada. Revise a selecao da preview.")

        all_keys = {str(payload.get("list_key") or "") for payload in list_payloads if payload.get("list_key")}
        applied_now = set(applied_list_keys.keys())
        if not job.first_applied_at:
            try:
                register_successful_import_usage(user, "IP")
            except ValueError as exc:
                raise IPImportError(str(exc)) from exc
            job.first_applied_at = timezone.now()
        apply_log["applied_list_keys"] = applied_list_keys
        apply_log["applied_list_ids"] = applied_list_ids
        apply_log["lists_applied"] = len(applied_now)
        apply_log["items_applied"] = ListaIPItem.objects.filter(lista_id__in=applied_list_ids).count()
        job.apply_log = apply_log
        job.status = job.Status.APPLIED if all_keys and all_keys <= applied_now else job.Status.REVIEW
        job.save(update_fields=["applied_lista", "apply_log", "status", "first_applied_at", "updated_at"])
    for field_name in ("applied_lista", "apply_log", "status", "first_applied_at", "updated_at"):
        setattr(original_job, field_name, getattr(job, field_name))
    return applied
