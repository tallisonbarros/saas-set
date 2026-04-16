import csv
import hashlib
import json
import os
import re
import socket
import time
import unicodedata
import uuid
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import timedelta
from io import BytesIO, StringIO
from pathlib import Path
from urllib import error as urlerror
from urllib import request as urlrequest
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from django.utils import timezone

from core.services.billing import register_successful_import_usage


TYPE_ALIASES = {
    "DI": ["DI", "DIGITAL INPUT", "DISCRETE INPUT", "ENTRADA DIGITAL", "DIGITAL_IN"],
    "DO": ["DO", "DQ", "DIGITAL OUTPUT", "DISCRETE OUTPUT", "SAIDA DIGITAL", "DIGITAL_OUT"],
    "AI": ["AI", "ANALOG INPUT", "ENTRADA ANALOGICA", "ANALOG_IN"],
    "AO": ["AO", "AQ", "ANALOG OUTPUT", "SAIDA ANALOGICA", "ANALOG_OUT"],
}

HEADER_ALIASES = {
    "panel": ["PANEL", "PAINEL", "PNL", "PAINEL IO", "PANEL NAME", "DRIVER", "PLC"],
    "rack": ["RACK", "RK", "NOME RACK", "RACK NAME"],
    "slot": ["SLOT", "SL", "POSICAO SLOT", "CARD SLOT", "POSICAO", "CARTAO", "MODULO SLOT"],
    "module_model": [
        "MODULO",
        "MOD",
        "MODELO MODULO",
        "MODEL",
        "MODULE",
        "MODULE MODEL",
        "MODULE TYPE",
        "CARD",
        "CARD MODEL",
        "CARD TYPE",
        "CARD GROUP",
    ],
    "channel": ["CANAL", "CHANNEL", "CHAN", "CH", "CHANNEL NUMBER"],
    "location": [
        "LOCATION",
        "LOCALIZACAO",
        "LOCALIZACAO STRING",
        "LOC",
        "CARD POSITION",
        "CARD POS",
        "PANEL / CARD POS",
        "PANEL / CARD",
        "PANEL/CARD",
        "RACK/SLOT",
        "R/S/C",
        "FIELD WIRING",
        "TERM",
        "TERMINAL",
        "JB/TB",
        "JB TB",
    ],
    "address": ["ADDR", "ADDRESS"],
    "fieldbus": ["FIELDBUS", "BUS", "STATION", "REMOTE NODE", "NODE"],
    "point_ref": ["POINT REF", "LOOP REF", "POINT UID"],
    "tag": [
        "TAG",
        "TAGSET",
        "TAG NAME",
        "TAGNAME",
        "NOME TAG",
        "TAG IO",
        "POINT",
        "PT",
        "OBJECT",
        "OBJECT NAME",
    ],
    "description": [
        "DESCRICAO",
        "DESCRICAO SINAL",
        "DESCRIPTION",
        "SIGNAL DESCRIPTION",
        "SERVICE",
        "SERVICO",
        "SVC",
        "DESC",
        "NOME",
        "VISIBLE TEXT",
        "ALARM TEXT",
    ],
    "signal_hint": ["SIGNAL", "SINAL", "TYPE / SIG", "TYPE / SIGNAL", "SCAN"],
    "type": [
        "TIPO",
        "TYPE",
        "TIPO IO",
        "IO TYPE",
        "I/O",
        "SIGNAL TYPE",
        "TIPO SINAL",
        "TYPE / SIGNAL",
        "CLASS",
        "CODE",
    ],
}

DEFAULT_HEADER_PROMPT = (
    "Map the spreadsheet headers to the SAAS-SET IO import schema. "
    "Prefer stable engineering columns such as rack, slot, module_model, channel, tag, description and type. "
    "Only return explicit mappings that are strongly supported by the sheet content."
)

DEFAULT_GROUPING_PROMPT = (
    "Given the raw worksheet context plus the preliminary normalized IO rows, act as the main interpreter of the sheet. "
    "Suggest reliable column mappings, classify noisy rows, and enrich each logical IO row with rack, slot, module model, "
    "channel, tag, description, and type when the structure supports it. Respect explicit engineering evidence when present, "
    "and only override the preliminary interpretation when confidence is high."
)

DEFAULT_AI_REQUEST_TIMEOUT_SECONDS = 180.0
DEFAULT_AI_TOTAL_BUDGET_SECONDS = 7200.0
DEFAULT_AI_MAX_SHEETS_PER_JOB = 96
DEFAULT_AI_OVERRIDE_CONFIDENCE = 70
DEFAULT_AI_MAX_RAW_ROWS_PER_SHEET = 220
DEFAULT_AI_MAX_RETRIES = 2
DEFAULT_AI_CACHE_MAX_ENTRIES = 240
DEFAULT_AI_CACHE_MAX_AGE_DAYS = 21
IO_IMPORT_AI_CACHE_VERSION = "v1"


class IOImportError(Exception):
    pass


@dataclass
class ParsedSpreadsheet:
    file_format: str
    sheet_name: str
    header_row_index: int
    rows_total: int
    headers: list
    raw_rows: list
    column_map: dict
    warnings: list
    layout: str = "tabular"


def _ascii_upper(value):
    normalized = unicodedata.normalize("NFKD", str(value or ""))
    normalized = normalized.encode("ascii", "ignore").decode("ascii")
    return normalized.strip().upper()


def _compact_token(value):
    return re.sub(r"[^A-Z0-9]+", "", _ascii_upper(value))


def _tokenize_identifier(value):
    return [token for token in re.split(r"[^A-Z0-9]+", _ascii_upper(value)) if token]


def _header_word_tokens(value):
    return [token for token in re.split(r"[^A-Z0-9]+", _ascii_upper(value)) if token]


def _alias_match_score(header_value, alias):
    header_compact = _compact_token(header_value)
    alias_token = _compact_token(alias)
    if not header_compact or not alias_token:
        return 0
    header_words = set(_header_word_tokens(header_value))
    if header_compact == alias_token:
        return 100
    if alias_token in header_words:
        return 80
    if len(alias_token) >= 4 and alias_token in header_compact:
        return 60
    return 0


def _best_alias_match(header_value, aliases):
    best_score = 0
    best_rank = 999
    for rank, alias in enumerate(aliases):
        score = _alias_match_score(header_value, alias)
        if score > best_score or (score and score == best_score and rank < best_rank):
            best_score = score
            best_rank = rank
    return best_score, best_rank


def _sample_column_values(rows, header_row_index, index, limit=12):
    if not rows:
        return []
    values = []
    for row in rows[header_row_index + 1 :]:
        if index >= len(row):
            continue
        value = _cell_to_text(row[index])
        if not value:
            continue
        values.append(value)
        if len(values) >= limit:
            break
    return values


def _looks_like_tag_value(value):
    token = _cell_to_text(value).strip()
    if not token:
        return False
    if " " in token:
        return False
    ascii_token = _ascii_upper(token)
    if len(ascii_token) < 3 or len(ascii_token) > 40:
        return False
    if not any(char.isalpha() for char in ascii_token):
        return False
    if not any(char.isdigit() for char in ascii_token):
        return False
    return bool(re.fullmatch(r"[A-Z0-9_.:/-]+", ascii_token))


def _score_type_values(values):
    if not values:
        return 0
    return sum(1 for value in values if normalize_type(value)) * 15


def _score_tag_values(values):
    if not values:
        return 0
    return sum(1 for value in values if _looks_like_tag_value(value)) * 8


def _cell_to_text(value):
    if value is None:
        return ""
    if isinstance(value, bool):
        return "SIM" if value else "NAO"
    text = str(value).strip()
    if not text:
        return ""
    return re.sub(r"\s+", " ", text)


def _read_positive_float_env(name, default):
    raw_value = str(os.environ.get(name, "")).strip()
    if not raw_value:
        return default
    try:
        parsed = float(raw_value)
    except (TypeError, ValueError):
        return default
    return parsed if parsed > 0 else default


def _read_positive_int_env(name, default):
    raw_value = str(os.environ.get(name, "")).strip()
    if not raw_value:
        return default
    try:
        parsed = int(raw_value)
    except (TypeError, ValueError):
        return default
    return parsed if parsed > 0 else default


def _ai_request_timeout_seconds():
    return _read_positive_float_env("IO_IMPORT_AI_TIMEOUT_SECONDS", DEFAULT_AI_REQUEST_TIMEOUT_SECONDS)


def _ai_total_budget_seconds():
    return _read_positive_float_env("IO_IMPORT_AI_TOTAL_BUDGET_SECONDS", DEFAULT_AI_TOTAL_BUDGET_SECONDS)


def _ai_max_sheets_per_job():
    return _read_positive_int_env("IO_IMPORT_AI_MAX_SHEETS_PER_JOB", DEFAULT_AI_MAX_SHEETS_PER_JOB)


def _ai_override_confidence_threshold():
    threshold = _read_positive_int_env("IO_IMPORT_AI_OVERRIDE_CONFIDENCE", DEFAULT_AI_OVERRIDE_CONFIDENCE)
    return max(1, min(threshold, 100))


def _ai_max_raw_rows_per_sheet():
    return _read_positive_int_env("IO_IMPORT_AI_MAX_RAW_ROWS_PER_SHEET", DEFAULT_AI_MAX_RAW_ROWS_PER_SHEET)


def _ai_max_retries():
    return _read_positive_int_env("IO_IMPORT_AI_MAX_RETRIES", DEFAULT_AI_MAX_RETRIES)


def _ai_cache_max_entries():
    return _read_positive_int_env("IO_IMPORT_AI_CACHE_MAX_ENTRIES", DEFAULT_AI_CACHE_MAX_ENTRIES)


def _ai_cache_max_age_days():
    return _read_positive_int_env("IO_IMPORT_AI_CACHE_MAX_AGE_DAYS", DEFAULT_AI_CACHE_MAX_AGE_DAYS)


def _effective_ai_sheet_row_limit(settings_obj, total_sheets):
    base_limit = max(24, int(getattr(settings_obj, "max_rows_for_ai", 150) or 150))
    if total_sheets >= 16:
        return min(base_limit, 72)
    if total_sheets >= 10:
        return min(base_limit, 96)
    if total_sheets >= 6:
        return min(base_limit, 120)
    return base_limit


def _effective_workbook_context_limit(total_sheets):
    base_limit = _ai_max_raw_rows_per_sheet()
    if total_sheets >= 16:
        return min(base_limit, 24)
    if total_sheets >= 10:
        return min(base_limit, 32)
    if total_sheets >= 6:
        return min(base_limit, 48)
    return min(base_limit, 80)


def normalize_tag(value):
    value = _cell_to_text(value)
    if not value:
        return ""
    normalized = _ascii_upper(value)
    normalized = re.sub(r"[^A-Z0-9]+", "_", normalized)
    return normalized.strip("_")


def normalize_module_alias(value):
    token = _ascii_upper(value)
    token = token.replace("MODULO", "")
    token = token.replace("MODEL", "")
    return re.sub(r"[^A-Z0-9]+", "", token)


def normalize_type(raw_value, fallbacks=None):
    candidates = [_ascii_upper(raw_value)]
    for fallback in fallbacks or []:
        candidates.append(_ascii_upper(fallback))
    compact_candidates = [_compact_token(item) for item in candidates if item]
    combined_text = " | ".join(candidate for candidate in candidates if candidate)
    combined_compact = " ".join(compact_candidates)

    if "SPAREAO" in combined_compact or "RESERVAAO" in combined_compact:
        return "AO"
    if re.search(r"\b(SPARE|RESERVA)[_. -]?(AO)\b", combined_text) or (
        "RESERVA" in combined_text and "SAIDA" in combined_text and "ANALOG" in combined_text
    ):
        return "AO"
    if "SPAREAI" in combined_compact or "RESERVAAI" in combined_compact or "RESERVAANALOG" in combined_compact:
        return "AI"
    if re.search(r"\b(SPARE|RESERVA)[_. -]?(AI|ANALOG)\b", combined_text):
        return "AI"
    if "SPAREDO" in combined_compact or "RESERVADO" in combined_compact:
        return "DO"
    if re.search(r"\b(SPARE|RESERVA)[_. -]?(DO)\b", combined_text) or (
        "RESERVA" in combined_text and "SAIDA" in combined_text and "DIGITAL" in combined_text
    ):
        return "DO"
    if "SPAREDI" in combined_compact or "RESERVADI" in combined_compact:
        return "DI"
    if re.search(r"\b(SPARE|RESERVA)[_. -]?(DI|DIGITAL)\b", combined_text) and "SAIDA" not in combined_text:
        return "DI"

    for candidate in candidates:
        if not candidate:
            continue
        if re.search(r"%IW\d+\.\d+\.\d+\b", candidate):
            return "AI"
        if re.search(r"%QW\d+\.\d+\.\d+\b", candidate):
            return "AO"
        if re.search(r"%I\d+\.\d+\.\d+\b", candidate):
            return "DI"
        if re.search(r"%Q\d+\.\d+\.\d+\b", candidate):
            return "DO"
        compact_candidate = _compact_token(candidate)
        if re.fullmatch(r"DI\d+", compact_candidate):
            return "DI"
        if re.fullmatch(r"(DO|DQ)\d+", compact_candidate):
            return "DO"
        if re.fullmatch(r"AI\d+", compact_candidate):
            return "AI"
        if re.fullmatch(r"(AO|AQ)\d+", compact_candidate):
            return "AO"

    for canonical, aliases in TYPE_ALIASES.items():
        alias_tokens = {_compact_token(alias) for alias in aliases}
        for token in compact_candidates:
            if not token:
                continue
            if token in alias_tokens:
                return canonical

    for canonical, aliases in TYPE_ALIASES.items():
        for alias in aliases:
            alias_text = _ascii_upper(alias)
            alias_compact = _compact_token(alias)
            if len(alias_compact) < 4:
                continue
            if alias_text in combined_text or alias_compact in combined_compact:
                return canonical

    if re.search(r"\b(AIT|FIT|LIT|PIT|TIT|WIT|PHIT)[_.-]?\d+", combined_text):
        return "AI"
    if "_SPD_FBK" in combined_text or "REFERENCIA REAL DE VELOCIDADE" in combined_text:
        return "AI"
    if "_SPD_REF" in combined_compact or "_POS_REF" in combined_compact or "_OUT" in combined_compact or "_REF" in combined_compact:
        return "AO"
    if "SETPOINT" in combined_text:
        return "AO"
    if "TRANSMISSOR" in combined_text or "TRANSMITTER" in combined_text:
        return "AI"
    if re.search(r"\b(ZS|LS|PS|TS|FS)[_.-]?\d+.*_(OPEN|CLOSE)\b", combined_text):
        return "DI"

    analog_markers = ("4-20 MA", "420MA", "0-10 V", "010V", "1-5 V", "15V", "PT100", "RTD", "THERMOCOUPLE")
    analog_hint = any(marker in combined_text or marker.replace(" ", "") in combined_compact for marker in analog_markers)
    analog_output_markers = (
        "SETPOINT",
        "SPD_REF",
        "POS_REF",
        "_REF",
        "_OUT",
        " COMANDO ",
        " COMMAND ",
        "_CMD",
    )
    analog_input_markers = (
        "TRANSMISSOR",
        "TRANSMITTER",
        "INDICADOR",
        "MEASUREMENT",
        "PRESSAO",
        "PRESSURE",
        "TEMPERATURA",
        "TEMPERATURE",
        "VAZAO",
        "FLOW",
        "NIVEL",
        "LEVEL",
        "CONDUTIVIDADE",
        "WEIGHT",
        "PIT_",
        "FIT_",
        "AIT_",
        "TIT_",
        "WIT_",
        "LIT_",
        "PHIT_",
    )
    if analog_hint:
        if any(marker in combined_text or marker in combined_compact for marker in analog_output_markers):
            return "AO"
        if any(marker in combined_text or marker in combined_compact for marker in analog_input_markers):
            return "AI"
        return "AI"

    discrete_output_markers = (
        "COMANDO",
        "COMMAND",
        "_CMD",
        "_START",
        "_STOP",
        "PARTIDA",
        "SOLENOID",
        "BEACON",
        "HORN",
        "STARTER",
        "ENABLE",
    )
    if any(marker in combined_text or marker in combined_compact for marker in discrete_output_markers):
        return "DO"

    discrete_input_markers = (
        "RETORNO",
        "FEEDBACK",
        "HEALTHY",
        "_FB",
        "_FBK",
        "_RUN",
        "_FLT",
        "FAULT",
        "STATUS",
        "AUX CONTACT",
        "DRY CONTACT",
        "LIMIT SWITCH",
        "CHAVE",
        "SENSOR",
        "DETECTOR",
        "FIM DE CURSO",
    )
    if any(marker in combined_text or marker in combined_compact for marker in discrete_input_markers):
        return "DI"

    if "LOCAL:I." in combined_text:
        return "AI" if "ANALOG" in combined_text or "AI" in combined_compact else "DI"
    if "LOCAL:O." in combined_text:
        return "AO" if "ANALOG" in combined_text or "AO" in combined_compact else "DO"
    return ""


def _normalize_int(value):
    text = _cell_to_text(value)
    if not text:
        return None
    match = re.search(r"(\d+)", text)
    if not match:
        return None
    try:
        return int(match.group(1))
    except ValueError:
        return None


def _guess_delimiter(sample_text):
    try:
        dialect = csv.Sniffer().sniff(sample_text, delimiters=",;\t|")
        return dialect.delimiter
    except csv.Error:
        if sample_text.count(";") > sample_text.count(","):
            return ";"
        if sample_text.count("\t") > sample_text.count(","):
            return "\t"
        return ","


def _read_csv_rows(raw_bytes, file_format):
    decoded = None
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            decoded = raw_bytes.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if decoded is None:
        raise IOImportError("Nao foi possivel decodificar o arquivo CSV.")
    delimiter = "\t" if file_format == "tsv" else _guess_delimiter(decoded[:4000])
    reader = csv.reader(StringIO(decoded), delimiter=delimiter)
    rows = []
    for row in reader:
        rows.append([_cell_to_text(cell) for cell in row])
    return rows


def _read_xlsx_rows(raw_bytes):
    try:
        workbook = load_workbook(filename=BytesIO(raw_bytes), data_only=True, read_only=True)
    except (BadZipFile, InvalidFileException, OSError, ValueError) as exc:
        raise IOImportError(
            "Nao foi possivel abrir a planilha Excel. Confirme que o arquivo nao esta corrompido "
            "e que voce selecionou a planilha real, nao um arquivo temporario do Excel."
        ) from exc
    sheets = []
    for worksheet in workbook.worksheets:
        rows = []
        for row in worksheet.iter_rows(values_only=True):
            rows.append([_cell_to_text(cell) for cell in row])
        sheets.append({"name": worksheet.title, "rows": rows})
    return sheets


def _non_empty_cells(row):
    return sum(1 for item in row if _cell_to_text(item))


def _row_text(row):
    return " ".join(_cell_to_text(item) for item in row if _cell_to_text(item))


def _is_mostly_numeric_token(token):
    token = _compact_token(token)
    if not token:
        return False
    letters = sum(1 for char in token if char.isalpha())
    digits = sum(1 for char in token if char.isdigit())
    return digits > letters and digits >= 2


def _is_probable_data_row(row):
    non_empty = [cell for cell in row if _cell_to_text(cell)]
    if len(non_empty) < 4:
        return False
    numeric_tokens = sum(1 for cell in non_empty if _is_mostly_numeric_token(cell))
    type_hits = sum(1 for cell in non_empty if normalize_type(cell))
    tag_hits = sum(1 for cell in non_empty if normalize_tag(cell) and "_" in normalize_tag(cell))
    return numeric_tokens >= 2 or type_hits >= 2 or tag_hits >= 1


def _header_signature(row):
    alias_score = _score_header_row(row)
    non_empty = _non_empty_cells(row)
    data_penalty = 1 if _is_probable_data_row(row) else 0
    return alias_score, non_empty - data_penalty, -data_penalty


def _extract_filename_sheet_hints(original_filename):
    stem = Path(original_filename or "").stem
    tokens = [token for token in _tokenize_identifier(stem) if len(token) >= 3]
    informative_tokens = [
        token
        for token in tokens
        if token not in {"PLANILHA", "LISTA", "NUTRIEN", "REV", "REVISAO", "SHEET", "ABA"}
        and not re.fullmatch(r"REV\d+", token)
    ]
    marker_tokens = [token for token in informative_tokens if re.search(r"[A-Z]+\d+", token)]
    preferred_token = marker_tokens[-1] if marker_tokens else ""
    return informative_tokens, preferred_token


def _parse_slot_block_title(row_or_text):
    if isinstance(row_or_text, str):
        text = _cell_to_text(row_or_text)
    else:
        non_empty = [cell for cell in row_or_text if _cell_to_text(cell)]
        if len(non_empty) == 0 or len(non_empty) > 3:
            return None
        if _score_header_row(row_or_text) >= 6:
            return None
        text = _row_text(row_or_text)
    match = re.search(r"\bSLOT\s*0*(?P<slot>\d+)\b(?P<rest>.*)$", text, flags=re.IGNORECASE)
    if not match:
        return None

    rest = _cell_to_text(match.group("rest"))
    module_text = rest
    declared_channels = None
    declared_type = ""
    trailer = re.search(
        r"(?P<module>.*?)(?:\s*[:\-]\s*(?P<count>\d+)\s*(?P<type>[A-Z]{2}))\s*$",
        rest,
        flags=re.IGNORECASE,
    )
    if trailer:
        module_text = _cell_to_text(trailer.group("module"))
        declared_channels = _normalize_int(trailer.group("count"))
        declared_type = normalize_type(trailer.group("type"))

    return {
        "slot_index": _normalize_int(match.group("slot")),
        "module_raw": module_text.strip(" :-"),
        "declared_channels": declared_channels,
        "declared_type": declared_type,
        "title": text,
    }


def _looks_like_block_subheader(row):
    if _score_header_row(row) >= 6:
        tokens = {_compact_token(cell) for cell in row if _cell_to_text(cell)}
        has_tag = any(token in {"TAG", "TAGS"} or "TAG" in token for token in tokens)
        has_description = any(token in {"DESCRICAO", "DESCRIPTION", "SERVICE", "DESC"} or "DESCR" in token for token in tokens)
        has_structural = any(
            token in {"IO", "PANEL", "CARDPOS", "CARDGROUP", "CHANNEL", "CH", "RACK", "SLOT"}
            for token in tokens
        )
        if has_tag and (has_description or has_structural):
            return True
    tokens = {_compact_token(cell) for cell in row if _cell_to_text(cell)}
    if not tokens:
        return False
    has_io = "IO" in tokens
    has_tag = any(token in {"TAG", "TAGS", "TAGS"} or "TAG" in token for token in tokens)
    has_description = any(token in {"DESCRICAO", "DESCRIPTION"} or "DESCR" in token for token in tokens)
    return has_io and (has_tag or has_description)


def _looks_like_section_title(row):
    if _non_empty_cells(row) != 1:
        return False
    text = _ascii_upper(_row_text(row))
    if not text:
        return False
    explicit_sections = {
        "DISCRETE INPUTS",
        "DISCRETE OUTPUTS",
        "ANALOG INPUTS",
        "ANALOG OUTPUTS",
        "SPARES",
        "SPARES",
        "INPUTS",
        "OUTPUTS",
    }
    if text in explicit_sections:
        return True
    if text.startswith("JUNCTION BOX "):
        return True
    if text.startswith("AREA_"):
        return True
    return False


def _parse_rack_section_title(row):
    if _non_empty_cells(row) == 0:
        return ""
    if _parse_slot_block_title(row) or _looks_like_block_subheader(row):
        return ""
    if _non_empty_cells(row) > 2:
        return ""

    text = _row_text(row)
    upper_text = _ascii_upper(text)
    if not upper_text:
        return ""

    generic_markers = [
        "LISTA DE IO",
        "RESPONSAVEL",
        "RESPONSAVEL TECNICO",
        "NUTRIEN",
        "UNIDADE",
        "REV",
        "REVISAO",
        "FOLHA",
        "SHEET",
    ]
    if any(marker in upper_text for marker in generic_markers):
        return ""

    # Single-cell point labels inside analog blocks must not split the rack.
    if re.match(r"^(AI|AO|DI|DO)\s*[.\-_]?\s*\d+\b", upper_text):
        return ""

    # Hardware/controller banners with explicit IP are metadata, not a new rack.
    if re.search(r"\b\d{1,3}(?:\.\d{1,3}){3}\b", upper_text) and any(
        marker in upper_text for marker in ("CPU", "CONTROLOGIX", "1756-", "1746-", "1734-", "1794-")
    ):
        return ""

    tokens = _tokenize_identifier(upper_text)
    if not tokens:
        return ""
    if not any(re.search(r"[A-Z]+\d+", token) for token in tokens):
        return ""

    rack_name = _infer_rack_name_from_sheet_name(text)
    return rack_name


def _parse_module_section_title(row):
    if _non_empty_cells(row) == 0 or _non_empty_cells(row) > 3:
        return None
    text = _row_text(row)
    upper_text = _ascii_upper(text)
    if not upper_text or "PANEL" not in upper_text or "SLOT" not in upper_text:
        return None
    if "|" not in text and " / " not in text:
        return None

    panel = _parse_panel_token(upper_text)
    rack_match = re.search(r"\bRACK\s*0*(\d+)\b", upper_text)
    slot_match = re.search(r"\bSLOT\s*0*(\d+)\b", upper_text)
    if not rack_match or not slot_match:
        return None

    module_text = ""
    if "|" in text:
        module_text = _cell_to_text(text.split("|")[-1])
    module_text = re.sub(r"^\s*SLOT\s*\d+\s*", "", module_text, flags=re.IGNORECASE).strip()
    try:
        rack_index = int(rack_match.group(1))
        slot_index = int(slot_match.group(1))
    except ValueError:
        return None
    return {
        "panel": panel,
        "rack": rack_index,
        "slot": slot_index,
        "module_model": module_text,
    }


def _count_slot_block_rows(rows):
    return sum(1 for row in rows if _parse_slot_block_title(row))


def _detect_sheet_header_row(rows):
    best_signature = None
    best_row_index = 0
    for index, row in enumerate((rows or [])[:20]):
        signature = _header_signature(row)
        if best_signature is None or signature > best_signature:
            best_signature = signature
            best_row_index = index
    return best_row_index, best_signature[0] if best_signature else 0


def _score_header_row(row):
    score = 0
    matched_fields = set()
    for cell in row:
        if not _compact_token(cell):
            continue
        for field_name, aliases in HEADER_ALIASES.items():
            if field_name in matched_fields:
                continue
            best_score = max((_alias_match_score(cell, alias) for alias in aliases), default=0)
            if best_score >= 100:
                score += 3
                matched_fields.add(field_name)
                break
            if best_score >= 60:
                score += 1
                matched_fields.add(field_name)
                break
    return score


def _choose_best_sheet(sheet_payloads, original_filename=""):
    best_sheet = None
    best_signature = None
    filename_tokens, preferred_token = _extract_filename_sheet_hints(original_filename)
    for sheet in sheet_payloads:
        rows = sheet.get("rows") or []
        best_row_index, best_row_score = _detect_sheet_header_row(rows)
        sheet_name_token = _compact_token(sheet.get("name") or "")
        preferred_bonus = 1 if preferred_token and preferred_token in sheet_name_token else 0
        overlap_score = sum(1 for token in filename_tokens if token and token in sheet_name_token)
        block_score = _count_slot_block_rows(rows[:200])
        signature = (preferred_bonus, overlap_score, block_score, best_row_score, len(rows), -best_row_index)
        if best_signature is None or signature > best_signature:
            best_signature = signature
            best_sheet = {
                "name": sheet.get("name") or "Sheet1",
                "rows": rows,
                "header_row_index": best_row_index,
            }
    return best_sheet


def _detect_column_map(headers, rows=None, header_row_index=0):
    used_indexes = set()
    used_fields = set()
    mapping = {}
    confidence = {}
    normalized_headers = [(_compact_token(header), header) for header in headers]
    candidates = []
    for field_name, aliases in HEADER_ALIASES.items():
        for index, (token, raw_header) in enumerate(normalized_headers):
            if not token:
                continue
            score, alias_rank = _best_alias_match(raw_header, aliases)
            if field_name == "description" and token in {"DESCRICAO", "DESCRIPTION", "NOME", "SIGNAL"}:
                score = 50
            elif field_name == "module_model" and token in {"MODULO", "MODEL", "MODELO"}:
                score = 50
            elif field_name == "channel" and token == "IO":
                score = max(score, 45)
            sample_values = _sample_column_values(rows or [], header_row_index, index) if rows else []
            semantic_bonus = 0
            if field_name == "type":
                semantic_bonus = _score_type_values(sample_values)
            elif field_name == "tag":
                semantic_bonus = _score_tag_values(sample_values)
            if score > 0:
                candidates.append(
                    (
                        score + semantic_bonus,
                        score,
                        -alias_rank,
                        -len(token),
                        -index,
                        field_name,
                        index,
                        raw_header,
                    )
                )
    for score, base_score, _, _, _, field_name, index, raw_header in sorted(candidates, reverse=True):
        if field_name in used_fields or index in used_indexes:
            continue
        mapping[field_name] = {"index": index, "header": raw_header}
        confidence[field_name] = max(score, base_score)
        used_fields.add(field_name)
        used_indexes.add(index)
    return mapping, confidence


def _looks_like_summary_headers(headers):
    tokens = {_compact_token(header) for header in headers if _cell_to_text(header)}
    aggregate_hits = {"POINTS", "DI", "DO", "AI", "AO", "SPARE"} & tokens
    if "POINTS" in tokens and len(aggregate_hits) >= 4:
        return True
    if "PRIMARYSHEET" in tokens or "CONTEUDO" in tokens:
        return True
    if tokens and tokens.issubset({"FIELD", "VALUE", "REVISION", "DATE", "AUTHOR", "SCOPE"}):
        return True
    return False


def _count_non_empty_data_rows(rows, header_row_index):
    start_index = min(header_row_index + 1, len(rows))
    count = 0
    for row in rows[start_index:]:
        if _non_empty_cells(row) == 0:
            continue
        if _parse_module_section_title(row) or _parse_slot_block_title(row) or _looks_like_block_subheader(row):
            continue
        count += 1
    return count


def _should_skip_parsed_sheet(parsed):
    sheet_token = _compact_token(parsed.sheet_name)
    helper_markers = (
        "INDEX",
        "INDICE",
        "LEGEND",
        "LEGENDA",
        "SUMMARY",
        "RESUMO",
        "RACKSUMMARY",
        "RACKMAP",
        "RACKXREF",
        "REVISION",
        "REVLOG",
        "CAPA",
        "COVER",
        "CHKINDEX",
        "STARTHERE",
        "START",
        "ALARMS",
        "SLOTSUMMARY",
        "JBSUMMARY",
        "PANELSUMMARY",
        "AREASUMMARY",
    )
    if _looks_like_summary_headers(parsed.headers):
        return True

    core_fields = {"tag", "description"} & set(parsed.column_map.keys())
    structural_fields = {"slot", "channel", "module_model", "type", "location", "point_ref"} & set(parsed.column_map.keys())
    data_rows = _count_non_empty_data_rows(parsed.raw_rows, parsed.header_row_index)
    block_layout = parsed.layout == "slot_blocks" or _has_slot_block_context(parsed.raw_rows)

    if any(marker in sheet_token for marker in helper_markers):
        if sheet_token in {"ALARMS", "START", "STARTHERE", "INDEX", "INDICE"}:
            return True
        if not block_layout and (not core_fields or not structural_fields):
            return True
        if data_rows < 2:
            return True

    if not block_layout and not core_fields:
        return True
    return False


def _extract_sheet_tag_set(parsed):
    mapping = parsed.column_map.get("tag")
    if not mapping:
        return set()
    tag_index = mapping.get("index")
    if tag_index is None:
        return set()
    tags = set()
    for row in parsed.raw_rows[parsed.header_row_index + 1 :]:
        if tag_index >= len(row):
            continue
        if _looks_like_repeated_tabular_header(row) or _looks_like_section_title(row):
            continue
        tag = normalize_tag(row[tag_index])
        if tag:
            tags.add(tag)
    return tags


def _sheet_quality_signature(parsed):
    structural = len(
        {"panel", "rack", "slot", "channel", "module_model", "location", "address", "fieldbus", "type"}
        & set(parsed.column_map.keys())
    )
    semantic = len({"tag", "description", "point_ref"} & set(parsed.column_map.keys()))
    return (
        structural,
        semantic,
        _count_non_empty_data_rows(parsed.raw_rows, parsed.header_row_index),
        len(parsed.column_map),
    )


def _dedupe_parsed_sheets(parsed_sheets):
    if len(parsed_sheets) <= 1:
        return parsed_sheets
    kept = []
    kept_meta = []
    for position, parsed in enumerate(parsed_sheets):
        tag_set = _extract_sheet_tag_set(parsed)
        duplicate_of = None
        for kept_index, meta in enumerate(kept_meta):
            kept_tags = meta["tags"]
            if not tag_set or not kept_tags:
                continue
            overlap = len(tag_set & kept_tags)
            min_size = min(len(tag_set), len(kept_tags))
            if min_size < 20:
                continue
            if overlap >= int(min_size * 0.9):
                current_quality = _sheet_quality_signature(parsed)
                kept_quality = meta["quality"]
                if current_quality > kept_quality:
                    kept[kept_index] = parsed
                    kept_meta[kept_index] = {
                        "tags": tag_set,
                        "quality": current_quality,
                        "position": position,
                    }
                duplicate_of = kept_index
                break
        if duplicate_of is None:
            kept.append(parsed)
            kept_meta.append(
                {"tags": tag_set, "quality": _sheet_quality_signature(parsed), "position": position}
            )
    return [item for _, item in sorted(zip((meta["position"] for meta in kept_meta), kept), key=lambda pair: pair[0])]


def _resolve_file_format(filename):
    suffix = Path(filename or "").suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return "xlsx"
    if suffix == ".csv":
        return "csv"
    if suffix == ".tsv":
        return "tsv"
    return "unknown"


def build_file_sha256(raw_bytes):
    digest = hashlib.sha256()
    digest.update(raw_bytes or b"")
    return digest.hexdigest()


def _build_parsed_sheet(sheet_payload, file_format):
    rows = sheet_payload.get("rows") or []
    if not rows:
        raise IOImportError("A planilha nao possui linhas para importacao.")

    header_row_index, _ = _detect_sheet_header_row(rows)
    header_row_index = min(header_row_index, max(len(rows) - 1, 0))
    headers = rows[header_row_index]
    column_map, confidence = _detect_column_map(headers, rows=rows, header_row_index=header_row_index)
    explicit_tabular_columns = {"tag"} <= set(column_map.keys()) and {"rack", "slot", "channel"} & set(column_map.keys())
    layout = "slot_blocks" if _count_slot_block_rows(rows) and not explicit_tabular_columns else "tabular"
    warnings = []
    if "tag" not in column_map and "description" not in column_map:
        warnings.append("Nao foi possivel mapear claramente as colunas principais de TAG/descricao.")
    if "type" not in column_map:
        warnings.append("Tipo de IO nao encontrado de forma explicita; sera inferido quando possivel.")
    if layout == "slot_blocks":
        warnings.append("Layout agrupado por SLOT detectado; slot, modulo e tipo serao inferidos a partir dos titulos dos blocos.")

    return ParsedSpreadsheet(
        file_format=file_format,
        sheet_name=sheet_payload.get("name") or "Sheet1",
        header_row_index=header_row_index,
        rows_total=len(rows),
        headers=headers,
        raw_rows=rows,
        column_map={
            key: {
                "index": value["index"],
                "header": value["header"],
                "confidence": confidence.get(key, 0),
            }
            for key, value in column_map.items()
        },
        warnings=warnings,
        layout=layout,
    )


def parse_workbook(raw_bytes, original_filename):
    file_format = _resolve_file_format(original_filename)
    if file_format == "unknown":
        raise IOImportError("Formato nao suportado. Use arquivos .xlsx, .xlsm, .csv ou .tsv.")

    if file_format in {"csv", "tsv"}:
        sheet_payload = {"name": "Arquivo", "rows": _read_csv_rows(raw_bytes, file_format)}
        return [_build_parsed_sheet(sheet_payload, file_format=file_format)]

    sheets = _read_xlsx_rows(raw_bytes)
    if not sheets:
        raise IOImportError("Nenhuma planilha foi encontrada no arquivo.")

    parsed_sheets = []
    for sheet_payload in sheets:
        rows = sheet_payload.get("rows") or []
        if not rows or not any(_non_empty_cells(row) for row in rows):
            continue
        parsed_sheets.append(_build_parsed_sheet(sheet_payload, file_format=file_format))
    if not parsed_sheets:
        raise IOImportError("Nao foi possivel detectar abas validas para importacao.")
    filtered_sheets = [parsed for parsed in parsed_sheets if not _should_skip_parsed_sheet(parsed)]
    return _dedupe_parsed_sheets(filtered_sheets or parsed_sheets)


def parse_spreadsheet(raw_bytes, original_filename):
    file_format = _resolve_file_format(original_filename)
    if file_format == "unknown":
        raise IOImportError("Formato nao suportado. Use arquivos .xlsx, .xlsm, .csv ou .tsv.")

    if file_format in {"csv", "tsv"}:
        sheet_payload = {"name": "Arquivo", "rows": _read_csv_rows(raw_bytes, file_format)}
    else:
        sheets = _read_xlsx_rows(raw_bytes)
        if not sheets:
            raise IOImportError("Nenhuma planilha foi encontrada no arquivo.")
        sheet_payload = _choose_best_sheet(sheets, original_filename=original_filename)
        if not sheet_payload:
            raise IOImportError("Nao foi possivel detectar uma aba valida para importacao.")
    return _build_parsed_sheet(sheet_payload, file_format=file_format)


def _extract_row_value(row, column_map, field_name):
    mapping = column_map.get(field_name)
    if not mapping:
        return ""
    index = mapping.get("index")
    if index is None or index >= len(row):
        return ""
    return _cell_to_text(row[index])


def _find_first_value(*values):
    for value in values:
        text = _cell_to_text(value)
        if text:
            return text
    return ""


def _expand_panel_name(token):
    token = _clean_rack_name(token.replace("_", "-"))
    if not token:
        return ""
    upper_token = _ascii_upper(token)
    upper_token = re.sub(r"-(?:R|RACK)\d{1,2}$", "", upper_token, flags=re.IGNORECASE)
    if upper_token.startswith(("MCC-", "PNL-", "REM-", "UBS-")):
        return _clean_rack_name(upper_token)
    upper_token = re.sub(r"^(?:ET200|FB|PLC)-", "", upper_token, flags=re.IGNORECASE)
    if re.fullmatch(r"[A-Z]{3}-\d{2}", upper_token):
        if upper_token.startswith("TRN-"):
            return f"MCC-{upper_token}"
        return f"PNL-{upper_token}"
    return _clean_rack_name(upper_token)


def _parse_panel_token(text):
    upper_text = _ascii_upper(text)
    if not upper_text:
        return ""
    match = re.search(r"\b(?:MCC|PNL|PLC|ET200|REM|UBS|FB)[-_][A-Z0-9]+(?:[-_][A-Z0-9]+)*\b", upper_text)
    if match:
        return _expand_panel_name(match.group(0))
    short_match = re.fullmatch(r"[A-Z]{3}-\d{2}", upper_text.strip())
    if short_match:
        return _expand_panel_name(short_match.group(0))
    return ""


def _parse_hardware_context(*values):
    text = " ".join(_cell_to_text(value) for value in values if _cell_to_text(value))
    upper_text = _ascii_upper(text)
    if not upper_text:
        return {}

    panel = _parse_panel_token(upper_text)
    rack = None
    slot = None
    channel = None

    rack_match = re.search(r"\bRACK\s*0*(\d+)\b", upper_text)
    if not rack_match:
        rack_match = re.search(r"(?:^|[\s/.\-])R\s*0*(\d+)\b", upper_text)
    if not rack_match:
        rack_match = re.search(r"[-_/]R0*(\d+)\b", upper_text)
    if rack_match:
        try:
            rack = int(rack_match.group(1))
        except ValueError:
            rack = None

    slot_match = re.search(r"\bSLOT\s*0*(\d+)\b", upper_text)
    if not slot_match:
        slot_match = re.search(r"(?:^|[\s/.\-])S\s*0*(\d+)\b", upper_text)
    if not slot_match:
        slot_match = re.search(r"[-_/]S0*(\d+)\b", upper_text)
    if slot_match:
        try:
            slot = int(slot_match.group(1))
        except ValueError:
            slot = None

    channel_match = re.search(r"\bCH(?:ANNEL)?\s*0*(\d+)\b", upper_text)
    if not channel_match:
        channel_match = re.search(r"(?:^|[\s/.\-])C\s*0*(\d+)\b", upper_text)
    if not channel_match:
        channel_match = re.search(r"[-_/]CH?0*(\d+)\b", upper_text)
    if channel_match:
        try:
            channel = int(channel_match.group(1))
        except ValueError:
            channel = None

    combined_match = re.search(r"\bR0*(\d+)\s*[/\-]\s*S0*(\d+)(?:\s*[/\-]\s*CH?0*(\d+))?\b", upper_text)
    if combined_match:
        rack = rack or int(combined_match.group(1))
        slot = slot or int(combined_match.group(2))
        if combined_match.group(3):
            channel = channel or int(combined_match.group(3))

    std_addr_match = re.search(r"\b(\d+)\.(\d+)\.(\d+)\b", upper_text)
    if std_addr_match:
        rack = rack or int(std_addr_match.group(1))
        slot = slot or int(std_addr_match.group(2))
        channel = channel or int(std_addr_match.group(3))

    unity_match = re.search(r"%[IQ](?:W)?(\d+)\.(\d+)\.(\d+)\b", upper_text)
    if unity_match:
        rack = rack or int(unity_match.group(1))
        slot = slot or int(unity_match.group(2))
        channel = channel or (int(unity_match.group(3)) + 1)

    logix_match = re.search(r"\bSLOT\s*0*(\d+)\.DATA\.(\d+)\b", upper_text)
    if logix_match:
        slot = slot or int(logix_match.group(1))
        channel = channel or (int(logix_match.group(2)) + 1)

    bit_match = re.search(r"\bBIT\s*0*(\d+)\b", upper_text)
    if bit_match:
        channel = channel or (int(bit_match.group(1)) + 1)

    return {
        "panel": panel,
        "rack": rack,
        "slot": slot,
        "channel": channel,
    }


def _parse_symbolic_channel(value):
    text = _ascii_upper(value)
    if not text:
        return None
    match = re.search(r"\b(?P<type>[A-Z]{1,3})\s*[\.\-_/ ]\s*(?P<index>\d+)\b", text)
    if not match:
        return None
    channel_type = normalize_type(match.group("type"))
    if not channel_type:
        return None
    try:
        channel_index = int(match.group("index")) + 1
    except ValueError:
        return None
    return {"type": channel_type, "index": channel_index}


def _normalize_channel_index(value):
    symbolic = _parse_symbolic_channel(value)
    if symbolic:
        return symbolic["index"]
    return _normalize_int(value)


def _normalize_ai_confidence(value, default=0):
    try:
        parsed = int(float(value))
    except (TypeError, ValueError):
        return default
    return max(0, min(parsed, 100))


def _normalize_ai_mode(value):
    return "override" if _ascii_upper(value) == "OVERRIDE" else "fill"


def _normalize_ai_row_kind(value):
    normalized = _cell_to_text(value).strip().lower()
    if normalized in {"section", "subheader", "noise"}:
        return normalized
    return "data"


def _sample_raw_rows_for_ai(parsed, limit=18):
    if not parsed.raw_rows:
        return []
    selected_indexes = []
    header_start = max(0, parsed.header_row_index - 3)
    header_end = min(len(parsed.raw_rows), parsed.header_row_index + 5)
    selected_indexes.extend(range(header_start, header_end))

    body_indexes = [
        index
        for index in range(parsed.header_row_index + 1, len(parsed.raw_rows))
        if _non_empty_cells(parsed.raw_rows[index]) > 0
    ]
    if body_indexes:
        sample_positions = []
        head_count = min(6, len(body_indexes))
        sample_positions.extend(body_indexes[:head_count])
        if len(body_indexes) > head_count:
            middle_index = body_indexes[len(body_indexes) // 2]
            sample_positions.append(middle_index)
        tail_indexes = body_indexes[max(len(body_indexes) - 4, head_count) :]
        sample_positions.extend(tail_indexes)
        selected_indexes.extend(sample_positions)

    sampled_rows = []
    for index in sorted(set(selected_indexes))[:limit]:
        sampled_rows.append(_raw_row_payload(index + 1, parsed.raw_rows[index], headers=parsed.headers, max_cells=12))
    return sampled_rows


def _compact_header_key(header, fallback_index):
    token = _compact_token(header)
    return token[:24] if token else f"C{fallback_index + 1}"


def _compact_cell_value_for_ai(value, max_len=84):
    text = _cell_to_text(value)
    if len(text) <= max_len:
        return text
    return text[: max_len - 3].rstrip() + "..."


def _raw_row_payload(row_number, row, headers=None, max_cells=12):
    entries = []
    for index, cell in enumerate(list(row)):
        text = _compact_cell_value_for_ai(cell)
        if not text:
            continue
        label = _compact_header_key(headers[index] if headers and index < len(headers) else "", index)
        entries.append(f"{label}={text}")
        if len(entries) >= max_cells:
            break
    return {
        "row_number": row_number,
        "entries": entries,
    }


def _effective_ai_max_cells_for_sheet(parsed):
    layout = _cell_to_text(getattr(parsed, "layout", "")).lower()
    if layout == "slot_blocks":
        return 10
    if layout == "tabular":
        return 12
    return 14


def _select_raw_row_indexes_for_ai(parsed, max_rows=None):
    max_rows = max_rows or _ai_max_raw_rows_per_sheet()
    rows = parsed.raw_rows or []
    if len(rows) <= max_rows:
        return list(range(len(rows)))

    selected = set()
    header_start = max(0, parsed.header_row_index - 4)
    header_end = min(len(rows), parsed.header_row_index + 6)
    selected.update(range(header_start, header_end))

    if parsed.layout == "slot_blocks":
        for index, row in enumerate(rows):
            if _parse_rack_section_title(row) or _parse_slot_block_title(row) or _parse_module_section_title(row):
                start = max(0, index - 1)
                end = min(len(rows), index + 5)
                selected.update(range(start, end))
        non_empty_indexes = [index for index, row in enumerate(rows) if _non_empty_cells(row) > 0]
        if non_empty_indexes:
            selected.update(non_empty_indexes[: min(18, len(non_empty_indexes))])
            selected.update(non_empty_indexes[max(0, len(non_empty_indexes) - 12) :])
            if len(non_empty_indexes) > 30:
                chunk = max(1, len(non_empty_indexes) // 6)
                for position in range(chunk, len(non_empty_indexes), chunk):
                    selected.add(non_empty_indexes[min(position, len(non_empty_indexes) - 1)])
        ordered = sorted(selected)
        if len(ordered) > max_rows:
            ordered = ordered[:max_rows]
        return ordered

    non_empty_indexes = [index for index, row in enumerate(rows) if _non_empty_cells(row) > 0]
    if non_empty_indexes:
        selected.update(non_empty_indexes[: min(30, len(non_empty_indexes))])
        selected.update(non_empty_indexes[max(0, len(non_empty_indexes) - 20) :])

        if len(non_empty_indexes) > 40:
            chunk = max(1, len(non_empty_indexes) // 8)
            for position in range(chunk, len(non_empty_indexes), chunk):
                selected.add(non_empty_indexes[min(position, len(non_empty_indexes) - 1)])

    for index, row in enumerate(rows):
        if _parse_rack_section_title(row) or _parse_slot_block_title(row) or _parse_module_section_title(row):
            start = max(0, index - 2)
            end = min(len(rows), index + 6)
            selected.update(range(start, end))

    ordered = sorted(selected)
    if len(ordered) > max_rows:
        trimmed = ordered[: max_rows // 2]
        remaining = ordered[max_rows // 2 :]
        step = max(1, len(remaining) // max(1, max_rows - len(trimmed)))
        trimmed.extend(remaining[::step][: max_rows - len(trimmed)])
        ordered = sorted(set(trimmed))
    return ordered[:max_rows]


def _build_raw_rows_payload(parsed, max_rows=None, max_cells=24):
    max_cells = min(max_cells, _effective_ai_max_cells_for_sheet(parsed))
    indexes = _select_raw_row_indexes_for_ai(parsed, max_rows=max_rows)
    return [
        _raw_row_payload(index + 1, parsed.raw_rows[index], headers=parsed.headers, max_cells=max_cells)
        for index in indexes
    ]


def _sheet_signal_bundle(parsed):
    sample_text = " ".join(
        filter(
            None,
            [
                parsed.sheet_name,
                " ".join(parsed.headers or []),
                " ".join(" ".join(item.get("entries") or []) for item in _sample_raw_rows_for_ai(parsed, limit=10)),
            ],
        )
    )
    compact = _ascii_upper(sample_text)
    type_hints = []
    for candidate in ("DI", "DO", "AI", "AO"):
        if candidate in compact:
            type_hints.append(candidate)
    channel_counts = sorted(set(int(value) for value in re.findall(r"\b(4|8|16|32|64)\b", compact)))
    return {
        "text": compact,
        "types": type_hints,
        "channel_counts": channel_counts,
    }


def _module_family_key(module_name):
    compact = _compact_token(module_name)
    compact = re.sub(r"(DI|DO|AI|AO)\d+$", "", compact)
    compact = re.sub(r"\d+$", "", compact)
    return compact[:24] or _compact_token(module_name)[:24]


def _build_compact_module_catalog(module_catalog, parsed=None, workbook_plan=None, limit=18):
    if not module_catalog:
        return []

    signal_bundle = _sheet_signal_bundle(parsed) if parsed else {"text": "", "types": [], "channel_counts": []}
    evidence_text = signal_bundle["text"]
    desired_types = set(signal_bundle["types"])
    desired_counts = set(signal_bundle["channel_counts"])
    plan_layout = _cell_to_text((workbook_plan or {}).get("layout_hint"))

    scored = []
    for item in module_catalog:
        modelo = _cell_to_text(item.get("modelo"))
        marca = _cell_to_text(item.get("marca"))
        tipo = _cell_to_text(item.get("tipo"))
        quantidade_canais = int(item.get("quantidade_canais") or 0)
        score = 0
        compact_model = _compact_token(modelo)
        if compact_model and compact_model in _compact_token(evidence_text):
            score += 120
        if tipo and tipo in desired_types:
            score += 65
        if desired_counts and quantidade_canais in desired_counts:
            score += 25
        if marca and _compact_token(marca) and _compact_token(marca) in _compact_token(evidence_text):
            score += 18
        if plan_layout == "slot_blocks":
            score += 5
        scored.append(
            {
                "score": score,
                "family": _module_family_key(modelo),
                "tipo": tipo,
                "quantidade_canais": quantidade_canais,
                "modelo": modelo,
                "marca": marca,
            }
        )

    if desired_types:
        scored.sort(key=lambda item: (item["tipo"] not in desired_types, -item["score"], item["quantidade_canais"], item["modelo"]))
    else:
        scored.sort(key=lambda item: (-item["score"], item["tipo"], item["quantidade_canais"], item["modelo"]))

    compact_items = []
    seen_keys = set()
    for item in scored:
        dedupe_key = (item["tipo"], item["quantidade_canais"], item["family"])
        if dedupe_key in seen_keys:
            continue
        seen_keys.add(dedupe_key)
        compact_items.append(
            {
                "modelo": item["modelo"],
                "marca": item["marca"],
                "tipo": item["tipo"],
                "quantidade_canais": item["quantidade_canais"],
            }
        )
        if len(compact_items) >= limit:
            break

    if not compact_items:
        return []
    return compact_items


def _module_catalog_stats(module_catalog):
    by_type = Counter()
    for item in module_catalog or []:
        if item.get("tipo"):
            by_type[item["tipo"]] += 1
    return {
        "total": len(module_catalog or []),
        "by_type": dict(by_type),
    }


def _sheet_layout_markers(parsed, limit=8):
    markers = {
        "slot_titles": [],
        "module_sections": [],
        "rack_sections": [],
    }
    for row in parsed.raw_rows or []:
        if len(markers["slot_titles"]) < limit:
            slot_block = _parse_slot_block_title(row)
            if slot_block and slot_block.get("title") and slot_block["title"] not in markers["slot_titles"]:
                markers["slot_titles"].append(slot_block["title"])
        if len(markers["module_sections"]) < limit:
            module_section = _parse_module_section_title(row)
            if module_section:
                candidate = _cell_to_text(module_section.get("title")) or _cell_to_text(module_section.get("module_model"))
                if candidate and candidate not in markers["module_sections"]:
                    markers["module_sections"].append(candidate)
        if len(markers["rack_sections"]) < limit:
            rack_section = _parse_rack_section_title(row)
            if rack_section and rack_section not in markers["rack_sections"]:
                markers["rack_sections"].append(rack_section)
    return markers


def _build_sheet_semantic_summary(parsed):
    signal_bundle = _sheet_signal_bundle(parsed)
    markers = _sheet_layout_markers(parsed)
    return {
        "sheet_name": parsed.sheet_name,
        "layout": parsed.layout,
        "rows_total": parsed.rows_total,
        "headers": parsed.headers[:12],
        "non_empty_data_rows": _count_non_empty_data_rows(parsed.raw_rows, parsed.header_row_index),
        "slot_block_rows": _count_slot_block_rows(parsed.raw_rows),
        "type_hints": signal_bundle.get("types") or [],
        "channel_count_hints": signal_bundle.get("channel_counts") or [],
        "slot_title_examples": markers["slot_titles"][:6],
        "module_section_examples": markers["module_sections"][:6],
        "rack_section_examples": markers["rack_sections"][:6],
    }


def _effective_reasoning_effort(settings_obj, schema_name, parsed=None, total_sheets=1):
    configured = _cell_to_text(getattr(settings_obj, "reasoning_effort", "")).lower() or "medium"
    if configured == "none":
        return ""
    if schema_name == "io_sheet_semantic_analysis" and parsed is not None:
        layout = _cell_to_text(getattr(parsed, "layout", "")).lower()
        rows_total = int(getattr(parsed, "rows_total", 0) or 0)
        if total_sheets == 1 and layout in {"slot_blocks", "tabular"} and rows_total <= 1200:
            if configured in {"high", "medium"}:
                return "low"
    if schema_name == "io_workbook_analysis" and total_sheets <= 3 and configured in {"high", "medium"}:
        return "low"
    return configured


def _hash_json_payload(payload):
    encoded = json.dumps(payload, ensure_ascii=True, sort_keys=True, separators=(",", ":")).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def _ai_settings_fingerprint(settings_obj):
    payload = {
        "provider": getattr(settings_obj, "provider", ""),
        "model": getattr(settings_obj, "model", ""),
        "reasoning_effort": getattr(settings_obj, "reasoning_effort", ""),
        "header_prompt": getattr(settings_obj, "header_prompt", ""),
        "grouping_prompt": getattr(settings_obj, "grouping_prompt", ""),
        "version": IO_IMPORT_AI_CACHE_VERSION,
    }
    return _hash_json_payload(payload)


def _build_workbook_cache_fingerprint(file_sha256, settings_obj, parsed_sheets):
    payload = {
        "file_sha256": file_sha256,
        "settings": _ai_settings_fingerprint(settings_obj),
        "workbook_context": _build_workbook_ai_context(parsed_sheets=parsed_sheets, original_filename=""),
    }
    return _hash_json_payload(payload)


def _build_sheet_cache_fingerprint(
    file_sha256,
    settings_obj,
    parsed,
    workbook_plan=None,
    total_sheets=1,
    compact_catalog=None,
):
    sheet_row_limit = _effective_ai_sheet_row_limit(settings_obj, total_sheets)
    payload = {
        "file_sha256": file_sha256,
        "settings": _ai_settings_fingerprint(settings_obj),
        "sheet_name": parsed.sheet_name,
        "layout": parsed.layout,
        "headers": parsed.headers,
        "warnings": parsed.warnings,
        "workbook_plan": workbook_plan or {},
        "raw_rows": _build_raw_rows_payload(parsed, max_rows=sheet_row_limit),
        "module_catalog": compact_catalog or [],
    }
    return _hash_json_payload(payload)


def _load_ai_cache(stage, fingerprint):
    from django.db.models import F
    from django.utils import timezone

    from core.models import IOImportAICache

    item = IOImportAICache.objects.filter(stage=stage, fingerprint=fingerprint).first()
    if not item:
        return None
    IOImportAICache.objects.filter(pk=item.pk).update(hits=F("hits") + 1, last_used_at=timezone.now())
    item.hits += 1
    return item.response_payload or {}


def _prune_ai_cache():
    from django.utils import timezone

    from core.models import IOImportAICache

    max_entries = max(40, _ai_cache_max_entries())
    max_age_days = max(1, _ai_cache_max_age_days())
    cutoff = timezone.now() - timedelta(days=max_age_days)
    IOImportAICache.objects.filter(last_used_at__lt=cutoff).delete()
    total = IOImportAICache.objects.count()
    if total <= max_entries:
        return
    stale_ids = list(
        IOImportAICache.objects.order_by("last_used_at").values_list("id", flat=True)[: max(0, total - max_entries)]
    )
    if stale_ids:
        IOImportAICache.objects.filter(id__in=stale_ids).delete()


def _save_ai_cache(stage, fingerprint, file_sha256, settings_obj, response_payload, sheet_name="", payload_meta=None):
    from core.models import IOImportAICache

    IOImportAICache.objects.update_or_create(
        stage=stage,
        fingerprint=fingerprint,
        defaults={
            "file_sha256": _cell_to_text(file_sha256),
            "sheet_name": _cell_to_text(sheet_name),
            "provider": _cell_to_text(getattr(settings_obj, "provider", "")),
            "model": _cell_to_text(getattr(settings_obj, "model", "")),
            "settings_fingerprint": _ai_settings_fingerprint(settings_obj),
            "response_payload": response_payload or {},
            "payload_meta": payload_meta or {},
        },
    )
    _prune_ai_cache()


def _build_workbook_ai_context(parsed_sheets, original_filename):
    workbook_row_limit = _effective_workbook_context_limit(len(parsed_sheets or []))
    return {
        "original_filename": original_filename,
        "sheets": [
            {
                "sheet_name": parsed.sheet_name,
                "rows_total": parsed.rows_total,
                "header_row_number": parsed.header_row_index + 1,
                "layout": parsed.layout,
                "headers": parsed.headers,
                "parser_warnings": parsed.warnings,
                "column_map": parsed.column_map,
                "data_rows": _count_non_empty_data_rows(parsed.raw_rows, parsed.header_row_index),
                "slot_block_rows": _count_slot_block_rows(parsed.raw_rows),
                "raw_rows": _build_raw_rows_payload(parsed, max_rows=workbook_row_limit),
            }
            for parsed in parsed_sheets
        ],
    }


def _find_workbook_sheet_plan(workbook_ai_payload, sheet_name):
    for item in (workbook_ai_payload or {}).get("sheets") or []:
        if _cell_to_text(item.get("sheet_name")) == _cell_to_text(sheet_name):
            return item
    return {}


def _build_single_sheet_fast_workbook_plan(parsed_sheets):
    if len(parsed_sheets or []) != 1:
        return {}
    parsed = parsed_sheets[0]
    return {
        "sheets": [
            {
                "sheet_name": parsed.sheet_name,
                "use_sheet": True,
                "layout_hint": parsed.layout or "unknown",
                "confidence": 88,
                "reason": "Guia unica com leitura estrutural local suficiente para seguir direto para a analise semantica.",
            }
        ],
        "warnings": [],
        "notes": "single-sheet-fast-path",
    }


def _should_skip_sheet_by_ai(workbook_ai_payload, parsed):
    plan = _find_workbook_sheet_plan(workbook_ai_payload, parsed.sheet_name)
    if not plan:
        return False, ""
    if not plan.get("use_sheet"):
        confidence = _normalize_ai_confidence(plan.get("confidence"), default=0)
        if confidence >= _ai_override_confidence_threshold():
            return True, _cell_to_text(plan.get("reason")) or "IA classificou a guia como nao operacional."
    return False, ""


def _merge_ai_column_map(headers, detected_map, ai_result):
    active_map = dict(detected_map)
    ai_mapping = (ai_result or {}).get("column_map") or {}
    override_threshold = _ai_override_confidence_threshold()

    for field_name, suggestion in ai_mapping.items():
        header_name = ""
        confidence = 80
        mode = "fill"
        if isinstance(suggestion, str):
            header_name = suggestion
        elif isinstance(suggestion, dict):
            header_name = suggestion.get("header") or ""
            confidence = _normalize_ai_confidence(suggestion.get("confidence"), default=0)
            mode = _normalize_ai_mode(suggestion.get("mode"))
        header_name = _cell_to_text(header_name)
        if not header_name:
            continue

        matching_index = None
        for index, header in enumerate(headers):
            if _compact_token(header) == _compact_token(header_name):
                matching_index = index
                header_name = header
                break
        if matching_index is None:
            continue

        existing_mapping = active_map.get(field_name)
        if existing_mapping and mode != "override":
            continue
        if existing_mapping and mode == "override" and confidence < override_threshold:
            continue

        active_map[field_name] = {
            "index": matching_index,
            "header": header_name,
            "confidence": max(confidence, 1),
            "source": "ai",
        }

    return active_map


def _build_ai_row_hints(ai_result):
    hints_by_row = {}
    for item in (ai_result or {}).get("row_hints") or []:
        if not isinstance(item, dict):
            continue
        source_row = _normalize_int(item.get("source_row"))
        if not source_row:
            continue
        normalized_hint = {
            "source_row": source_row,
            "row_kind": _normalize_ai_row_kind(item.get("row_kind")),
            "panel": _cell_to_text(item.get("panel")),
            "rack": _cell_to_text(item.get("rack")),
            "slot": _cell_to_text(item.get("slot")),
            "module_model": _cell_to_text(item.get("module_model")),
            "channel": _cell_to_text(item.get("channel")),
            "tag": _cell_to_text(item.get("tag")),
            "description": _cell_to_text(item.get("description")),
            "type": _cell_to_text(item.get("type")),
            "confidence": _normalize_ai_confidence(item.get("confidence"), default=0),
        }
        current = hints_by_row.get(source_row)
        if current is None or normalized_hint["confidence"] >= current["confidence"]:
            hints_by_row[source_row] = normalized_hint
    return hints_by_row


def _ensure_row_metadata(row):
    row.setdefault("field_sources", {})
    row.setdefault("field_confidence", {})
    row.setdefault("row_kind", "data")
    for field_name in ("panel_raw", "rack_raw", "slot_raw", "module_raw", "channel_raw", "tag", "description", "type"):
        if _cell_to_text(row.get(field_name)) and field_name not in row["field_sources"]:
            row["field_sources"][field_name] = "heuristic"
            row["field_confidence"][field_name] = 55
    return row


def _apply_ai_value_to_row(row, target_key, value, confidence):
    value = _cell_to_text(value)
    if not value:
        return
    override_threshold = _ai_override_confidence_threshold()
    existing_value = _cell_to_text(row.get(target_key))
    if existing_value and confidence < override_threshold:
        return
    row[target_key] = value
    row["field_sources"][target_key] = "ai"
    row["field_confidence"][target_key] = confidence


def _merge_ai_row_hints_into_rows(normalized_rows, ai_result):
    row_hints = _build_ai_row_hints(ai_result)
    if not row_hints:
        return normalized_rows, []

    merged_rows = []
    warnings = []
    override_threshold = _ai_override_confidence_threshold()
    for row in normalized_rows:
        row = _ensure_row_metadata(dict(row))
        hint = row_hints.get(row.get("source_row"))
        if not hint:
            merged_rows.append(row)
            continue

        row["row_kind"] = hint["row_kind"]
        row["ai_hint_confidence"] = hint["confidence"]
        if hint["row_kind"] in {"section", "subheader", "noise"} and hint["confidence"] >= override_threshold:
            warnings.append(
                f"Linha {row.get('source_row')} da guia {row.get('source_sheet') or '-'} foi descartada pela IA como {hint['row_kind']}."
            )
            continue

        _apply_ai_value_to_row(row, "panel_raw", hint.get("panel"), hint["confidence"])
        _apply_ai_value_to_row(row, "rack_raw", hint.get("rack"), hint["confidence"])
        _apply_ai_value_to_row(row, "slot_raw", hint.get("slot"), hint["confidence"])
        _apply_ai_value_to_row(row, "module_raw", hint.get("module_model"), hint["confidence"])
        _apply_ai_value_to_row(row, "channel_raw", hint.get("channel"), hint["confidence"])
        if hint.get("tag"):
            tag_value = normalize_tag(hint["tag"])
            if tag_value:
                _apply_ai_value_to_row(row, "tag", tag_value, hint["confidence"])
        _apply_ai_value_to_row(row, "description", hint.get("description"), hint["confidence"])
        if hint.get("type"):
            inferred_type = normalize_type(hint["type"]) or normalize_type(
                "",
                fallbacks=[row.get("channel_raw"), row.get("module_raw"), row.get("description"), row.get("tag")],
            )
            if inferred_type:
                _apply_ai_value_to_row(row, "type", inferred_type, hint["confidence"])

        if _cell_to_text(row.get("slot_raw")):
            row["slot_index"] = _normalize_int(row.get("slot_raw"))
        if _cell_to_text(row.get("channel_raw")):
            row["channel_index"] = _normalize_channel_index(row.get("channel_raw"))
        merged_rows.append(row)

    return merged_rows, warnings


def _validate_normalized_rows(normalized_rows, module_catalog):
    warnings = []
    known_models = {normalize_module_alias(item["modelo"]): item for item in module_catalog}
    validated_rows = []

    for row in normalized_rows:
        row = _ensure_row_metadata(dict(row))
        row["panel_raw"] = _cell_to_text(row.get("panel_raw"))
        row["location_raw"] = _cell_to_text(row.get("location_raw"))
        row["rack_raw"] = _cell_to_text(row.get("rack_raw"))
        row["slot_raw"] = _cell_to_text(row.get("slot_raw"))
        row["module_raw"] = _cell_to_text(row.get("module_raw"))
        row["channel_raw"] = _cell_to_text(row.get("channel_raw"))
        row["tag"] = normalize_tag(row.get("tag"))
        row["description"] = _cell_to_text(row.get("description"))
        explicit_type = normalize_type(row.get("type"))
        row["type"] = explicit_type or normalize_type(
            "",
            fallbacks=[row.get("channel_raw"), row.get("module_raw"), row.get("description"), row.get("tag"), row.get("location_raw")],
        )

        slot_index = _normalize_int(row.get("slot_raw")) or _normalize_int(row.get("slot_index"))
        channel_index = _normalize_channel_index(row.get("channel_raw")) or _normalize_int(row.get("channel_index"))
        row["slot_index"] = slot_index if slot_index and slot_index > 0 else None
        row["channel_index"] = channel_index if channel_index and channel_index > 0 else None

        module_alias = normalize_module_alias(row.get("module_raw"))
        matched_module = known_models.get(module_alias) if module_alias else None
        if matched_module and row["type"] and matched_module["tipo"] != row["type"]:
            warnings.append(
                f"Linha {row.get('source_row')} da guia {row.get('source_sheet') or '-'} tinha tipo {row['type']} em conflito com o modulo {matched_module['modelo']}; o tipo foi ajustado para {matched_module['tipo']}."
            )
            row["type"] = matched_module["tipo"]
            row["issues"].append("tipo_ajustado_pelo_catalogo")
            row["field_sources"]["type"] = "catalog"
            row["field_confidence"]["type"] = 100
        elif matched_module and not row["type"]:
            row["type"] = matched_module["tipo"]
            row["field_sources"]["type"] = "catalog"
            row["field_confidence"]["type"] = 100

        hardware_context = _parse_hardware_context(
            row.get("panel_raw"),
            row.get("location_raw"),
            row.get("rack_raw"),
            row.get("slot_raw"),
            row.get("channel_raw"),
            row.get("module_raw"),
        )
        if not row["panel_raw"] and hardware_context.get("panel"):
            row["panel_raw"] = hardware_context["panel"]
        if not row["rack_raw"] and hardware_context.get("rack"):
            row["rack_raw"] = str(hardware_context["rack"])
        if not row["slot_index"] and hardware_context.get("slot"):
            row["slot_index"] = hardware_context["slot"]
            row["slot_raw"] = str(hardware_context["slot"])
        if not row["channel_index"] and hardware_context.get("channel"):
            row["channel_index"] = hardware_context["channel"]
            row["channel_raw"] = str(hardware_context["channel"])

        validated_rows.append(row)

    return validated_rows, warnings


def normalize_rows_from_ai_result(parsed, module_catalog, ai_result):
    warnings = list(parsed.warnings or [])
    warnings.extend((ai_result or {}).get("warnings") or [])
    active_map = _merge_ai_column_map(parsed.headers, parsed.column_map, ai_result)

    normalized_rows = []
    for point in (ai_result or {}).get("logical_points") or []:
        if not isinstance(point, dict):
            continue
        source_row = _normalize_int(point.get("source_row"))
        if not source_row:
            continue
        row = {
            "source_row": source_row,
            "source_sheet": parsed.sheet_name,
            "panel_raw": _cell_to_text(point.get("panel")),
            "location_raw": "",
            "rack_raw": _cell_to_text(point.get("rack")),
            "slot_raw": _cell_to_text(point.get("slot")),
            "module_raw": _cell_to_text(point.get("module_model")),
            "channel_raw": _cell_to_text(point.get("channel")),
            "tag": normalize_tag(point.get("tag")),
            "description": _cell_to_text(point.get("description")),
            "type": normalize_type(point.get("type")),
            "slot_index": _normalize_int(point.get("slot")),
            "channel_index": _normalize_channel_index(point.get("channel")),
            "issues": [],
            "field_sources": {},
            "field_confidence": {},
            "row_kind": "data",
            "ai_hint_confidence": _normalize_ai_confidence(point.get("confidence"), default=0),
        }
        row = _ensure_row_metadata(row)
        for field_name in ("panel_raw", "rack_raw", "slot_raw", "module_raw", "channel_raw", "tag", "description", "type"):
            if _cell_to_text(row.get(field_name)):
                row["field_sources"][field_name] = "ai"
                row["field_confidence"][field_name] = row["ai_hint_confidence"] or 90
        if not any([row["tag"], row["description"], row["channel_raw"], row["slot_raw"], row["module_raw"], row["rack_raw"], row["panel_raw"]]):
            continue
        normalized_rows.append(row)

    normalized_rows, ai_row_warnings = _merge_ai_row_hints_into_rows(normalized_rows, ai_result)
    warnings.extend(ai_row_warnings)

    if not normalized_rows:
        raise IOImportError("A IA nao retornou pontos operacionais suficientes para montar a importacao.")

    normalized_rows, validation_warnings = _validate_normalized_rows(normalized_rows, module_catalog)
    warnings.extend(validation_warnings)

    known_models = {normalize_module_alias(item["modelo"]): item for item in module_catalog}
    for row in normalized_rows:
        module_alias = normalize_module_alias(row.get("module_raw"))
        row["matched_module"] = known_models.get(module_alias) if module_alias else None
        if not row.get("type"):
            row["issues"].append("tipo_nao_inferido")
        if not row.get("tag") and not row.get("description"):
            row["issues"].append("linha_sem_tag_e_descricao")

    return normalized_rows, active_map, warnings


def serialize_module_catalog(modules_qs):
    return [
        {
            "id": module.id,
            "modelo": (module.modelo or module.nome or f"Modulo {module.id}").strip(),
            "marca": (module.marca or "").strip(),
            "tipo": module.tipo_base.nome if module.tipo_base_id else "",
            "quantidade_canais": int(module.quantidade_canais),
        }
        for module in modules_qs
    ]


def _has_slot_block_context(rows):
    titles = 0
    headers = 0
    for row in rows:
        if _parse_slot_block_title(row):
            titles += 1
        elif _looks_like_block_subheader(row):
            headers += 1
    return titles >= 1 and headers >= 1


def _looks_like_repeated_tabular_header(row):
    return _score_header_row(row) >= 6 and not _is_probable_data_row(row)


def _row_repeats_headers(row, headers):
    if not headers:
        return False
    comparisons = []
    for index, header in enumerate(headers):
        header_token = _compact_token(header)
        if not header_token or index >= len(row):
            continue
        row_token = _compact_token(row[index])
        if not row_token:
            continue
        comparisons.append((row_token, header_token))
    if len(comparisons) < 3:
        return False
    matches = sum(1 for row_token, header_token in comparisons if row_token == header_token)
    return matches >= 3 and matches >= int(len(comparisons) * 0.6)


def _normalize_slot_block_rows(parsed, active_map):
    normalized_rows = []
    current_block = None
    current_rack_context = ""
    for row_offset, row in enumerate(parsed.raw_rows, start=1):
        if _non_empty_cells(row) == 0:
            continue
        if _looks_like_section_title(row):
            continue
        rack_section_title = _parse_rack_section_title(row)
        if rack_section_title:
            current_rack_context = rack_section_title
            current_block = None
            continue
        slot_block = _parse_slot_block_title(row)
        if slot_block:
            current_block = slot_block
            continue
        if _looks_like_block_subheader(row):
            continue
        if _row_repeats_headers(row, parsed.headers):
            continue
        if current_block is None:
            continue

        channel_raw = _extract_row_value(row, active_map, "channel")
        address_raw = _extract_row_value(row, active_map, "address")
        fieldbus_raw = _extract_row_value(row, active_map, "fieldbus")
        tag_raw = _extract_row_value(row, active_map, "tag")
        description_raw = _extract_row_value(row, active_map, "description")
        type_raw = _extract_row_value(row, active_map, "type")
        if not channel_raw and _parse_symbolic_channel(type_raw):
            channel_raw = type_raw
            type_raw = ""
        panel_raw = _extract_row_value(row, active_map, "panel")
        location_raw = _extract_row_value(row, active_map, "location")
        rack_raw = _extract_row_value(row, active_map, "rack") or current_rack_context
        hardware_context = _parse_hardware_context(
            panel_raw,
            location_raw,
            address_raw,
            fieldbus_raw,
            rack_raw,
            channel_raw,
            current_rack_context,
        )
        if not panel_raw:
            panel_raw = hardware_context.get("panel") or ""
        if not rack_raw and hardware_context.get("rack"):
            rack_raw = str(hardware_context["rack"])
        symbolic_channel = _parse_symbolic_channel(channel_raw)
        inferred_type = normalize_type(
            type_raw,
            fallbacks=[
                current_block.get("declared_type"),
                channel_raw,
                address_raw,
                tag_raw,
                description_raw,
                current_block.get("module_raw"),
            ],
        )

        if not any([channel_raw, tag_raw, description_raw]):
            continue

        normalized_rows.append(
            {
                "source_row": row_offset,
                "source_sheet": parsed.sheet_name,
                "panel_raw": panel_raw,
                "location_raw": location_raw,
                "rack_raw": rack_raw,
                "slot_raw": str(current_block.get("slot_index") or hardware_context.get("slot") or ""),
                "module_raw": current_block.get("module_raw") or "",
                "channel_raw": channel_raw,
                "tag": normalize_tag(tag_raw),
                "description": description_raw,
                "type": inferred_type,
                "slot_index": current_block.get("slot_index") or hardware_context.get("slot"),
                "channel_index": (symbolic_channel or {}).get("index")
                or hardware_context.get("channel")
                or _normalize_channel_index(channel_raw),
                "issues": [],
            }
        )
    return normalized_rows


def _normalize_tabular_rows(parsed, active_map, start_index=None):
    normalized_rows = []
    row_start = parsed.header_row_index + 1 if start_index is None else max(0, start_index)
    current_section_context = {}
    if row_start > 0:
        for previous_row in reversed(parsed.raw_rows[:row_start]):
            previous_section_context = _parse_module_section_title(previous_row)
            if previous_section_context:
                current_section_context = previous_section_context
                break
    for row_offset, row in enumerate(parsed.raw_rows[row_start:], start=row_start + 1):
        if _non_empty_cells(row) == 0:
            continue
        if _looks_like_section_title(row):
            continue

        section_context = _parse_module_section_title(row)
        if section_context:
            current_section_context = section_context
            continue
        if _looks_like_repeated_tabular_header(row):
            continue
        if _row_repeats_headers(row, parsed.headers):
            continue

        panel_raw = _extract_row_value(row, active_map, "panel") or current_section_context.get("panel") or ""
        rack_raw = _extract_row_value(row, active_map, "rack")
        slot_raw = _extract_row_value(row, active_map, "slot")
        module_raw = _extract_row_value(row, active_map, "module_model") or current_section_context.get("module_model") or ""
        channel_raw = _extract_row_value(row, active_map, "channel")
        address_raw = _extract_row_value(row, active_map, "address")
        fieldbus_raw = _extract_row_value(row, active_map, "fieldbus")
        signal_raw = _extract_row_value(row, active_map, "signal_hint")
        tag_raw = _extract_row_value(row, active_map, "tag")
        description_raw = _extract_row_value(row, active_map, "description")
        type_raw = _extract_row_value(row, active_map, "type")
        location_raw = _find_first_value(
            _extract_row_value(row, active_map, "location"),
            fieldbus_raw,
            address_raw,
            _extract_row_value(row, active_map, "point_ref"),
        )
        hardware_context = _parse_hardware_context(
            panel_raw,
            location_raw,
            address_raw,
            fieldbus_raw,
            rack_raw,
            slot_raw,
            channel_raw,
            _row_text(row),
        )
        if not panel_raw:
            panel_raw = hardware_context.get("panel") or ""
        if not rack_raw and current_section_context.get("rack"):
            rack_raw = str(current_section_context["rack"])
        if not rack_raw and hardware_context.get("rack"):
            rack_raw = str(hardware_context["rack"])
        if not slot_raw and current_section_context.get("slot"):
            slot_raw = str(current_section_context["slot"])
        if not slot_raw and hardware_context.get("slot"):
            slot_raw = str(hardware_context["slot"])
        if not channel_raw and hardware_context.get("channel"):
            channel_raw = str(hardware_context["channel"])

        if not any([rack_raw, slot_raw, channel_raw, tag_raw, description_raw, location_raw, panel_raw]):
            continue
        if not tag_raw and not description_raw:
            continue

        inferred_type = normalize_type(
            type_raw,
            fallbacks=[signal_raw, channel_raw, address_raw, tag_raw, description_raw, module_raw, location_raw],
        )
        normalized_rows.append(
            {
                "source_row": row_offset,
                "source_sheet": parsed.sheet_name,
                "panel_raw": panel_raw,
                "location_raw": location_raw,
                "rack_raw": rack_raw,
                "slot_raw": slot_raw,
                "module_raw": module_raw,
                "channel_raw": channel_raw,
                "tag": normalize_tag(tag_raw),
                "description": description_raw,
                "type": inferred_type,
                "slot_index": _normalize_int(slot_raw) or hardware_context.get("slot"),
                "channel_index": _normalize_channel_index(channel_raw) or hardware_context.get("channel"),
                "issues": [],
            }
        )
    return normalized_rows


def normalize_rows(parsed, module_catalog, ai_result=None):
    normalized_rows = []
    warnings = list(parsed.warnings)
    active_map = _merge_ai_column_map(parsed.headers, parsed.column_map, ai_result)

    block_layout_active = parsed.layout == "slot_blocks" or (
        "slot" not in active_map and "module_model" not in active_map and _has_slot_block_context(parsed.raw_rows)
    )
    if block_layout_active:
        normalized_rows = _normalize_slot_block_rows(parsed, active_map)
    else:
        normalized_rows = _normalize_tabular_rows(parsed, active_map)
        if not normalized_rows:
            normalized_rows = _normalize_tabular_rows(parsed, active_map, start_index=0)
            if normalized_rows:
                warnings.append(
                    "O cabecalho detectado nao foi confiavel para esta guia; a leitura foi refeita a partir das linhas brutas para preservar a interpretacao."
                )

    normalized_rows, ai_row_warnings = _merge_ai_row_hints_into_rows(normalized_rows, ai_result)
    warnings.extend(ai_row_warnings)

    if not normalized_rows:
        sheet_role = _cell_to_text((ai_result or {}).get("sheet_role")).lower()
        if (ai_result or {}).get("skip_sheet") or sheet_role in {"helper", "noise"}:
            warnings.append("A guia foi classificada como nao operacional durante a leitura semantica.")
            return [], active_map, warnings
        raise IOImportError("A IA nao conseguiu materializar linhas operacionais suficientes a partir desta guia.")

    normalized_rows, validation_warnings = _validate_normalized_rows(normalized_rows, module_catalog)
    warnings.extend(validation_warnings)

    known_models = {normalize_module_alias(item["modelo"]): item for item in module_catalog}
    for row in normalized_rows:
        module_alias = normalize_module_alias(row.get("module_raw"))
        row["matched_module"] = known_models.get(module_alias) if module_alias else None
        if not row["type"]:
            row["issues"].append("tipo_nao_inferido")
        if not row["tag"] and not row["description"]:
            row["issues"].append("linha_sem_tag_e_descricao")

    warnings.extend((ai_result or {}).get("warnings") or [])
    return normalized_rows, active_map, warnings


def _best_module_for_type(module_catalog, channel_type, needed_channels, explicit_model=None):
    type_modules = [item for item in module_catalog if item["tipo"] == channel_type]
    if explicit_model:
        alias = normalize_module_alias(explicit_model)
        for item in type_modules:
            if normalize_module_alias(item["modelo"]) == alias:
                return item
    exact = [item for item in type_modules if item["quantidade_canais"] == needed_channels]
    if exact:
        return exact[0]
    larger = sorted(
        [item for item in type_modules if item["quantidade_canais"] >= needed_channels],
        key=lambda item: (item["quantidade_canais"], item["modelo"]),
    )
    if larger:
        return larger[0]
    smaller = sorted(
        [item for item in type_modules if item["quantidade_canais"] < needed_channels],
        key=lambda item: (-item["quantidade_canais"], item["modelo"]),
    )
    if smaller:
        return smaller[0]
    if channel_type:
        capacity = max(int(needed_channels or 0), 1)
        model_name = _cell_to_text(explicit_model)[:80] or f"{channel_type}-{capacity:02d} CUSTOM"
        display_name = _cell_to_text(explicit_model)[:120] or f"Modulo {channel_type} {capacity} canais"
        return {
            "id": None,
            "modelo": model_name,
            "nome": display_name,
            "marca": "IMPORTADO",
            "tipo": channel_type,
            "quantidade_canais": capacity,
            "source": "custom",
        }
    return None


def _pick_default_rack_name(original_filename, requested_rack_name, ai_result, normalized_rows, target_rack=None):
    if target_rack:
        return target_rack.nome
    if requested_rack_name:
        return requested_rack_name.strip()
    ai_name = (ai_result or {}).get("rack_name") or ""
    if ai_name:
        return ai_name.strip()
    for row in normalized_rows:
        rack_raw = (row.get("rack_raw") or "").strip()
        if rack_raw:
            return rack_raw[:120]
    stem = Path(original_filename or "rack_importado").stem
    stem = re.sub(r"[_-]+", " ", stem).strip()
    return (stem or "Rack importado")[:120]


def _clean_rack_name(value):
    text = _cell_to_text(value)
    if not text:
        return ""
    text = re.sub(r"\s+", " ", text).strip(" -_/")
    return text[:120]


def _infer_rack_name_from_sheet_name(sheet_name, original_filename=""):
    sheet_name = _clean_rack_name(sheet_name)
    if not sheet_name:
        return _pick_default_rack_name(original_filename, "", None, [])

    tokens = _tokenize_identifier(sheet_name)
    generic_tokens = {"SHEET", "PLANILHA", "LISTA", "ABA", "PAINEL", "RACK"}
    technical_prefixes = ("6ES", "ET200", "CPU", "IM", "DI", "DQ", "AI", "AQ")
    meaningful = []
    for token in tokens:
        if token in generic_tokens:
            continue
        if any(token.startswith(prefix) for prefix in technical_prefixes):
            break
        meaningful.append(token)
    if meaningful:
        if len(meaningful) >= 2 and meaningful[0].isalpha() and meaningful[1].isdigit():
            meaningful = [f"{meaningful[0]}{meaningful[1]}"] + meaningful[2:]
        return _clean_rack_name("-".join(meaningful))
    return sheet_name[:120]


def _compose_rack_name(panel_raw, rack_raw, source_sheet, original_filename):
    panel_name = _parse_panel_token(panel_raw) or _clean_rack_name(panel_raw)
    rack_name = _clean_rack_name(rack_raw)
    sheet_name = _infer_rack_name_from_sheet_name(source_sheet, original_filename=original_filename)

    if panel_name and rack_name:
        if re.fullmatch(r"\d+", rack_name):
            return f"{panel_name} - Rack {int(rack_name):02d}"
        if normalize_tag(panel_name) == normalize_tag(rack_name):
            return panel_name
        return f"{panel_name} - {rack_name}"
    if panel_name:
        return panel_name
    if rack_name:
        if re.fullmatch(r"\d+", rack_name):
            base_name = sheet_name or _pick_default_rack_name(original_filename, "", None, [])
            return f"{base_name} - Rack {int(rack_name):02d}"
        return rack_name
    return sheet_name


def _resolve_rack_partition(row, original_filename):
    panel_raw = row.get("panel_raw") or ""
    rack_raw = row.get("rack_raw") or ""
    rack_name = _compose_rack_name(
        panel_raw=panel_raw,
        rack_raw=rack_raw,
        source_sheet=row.get("source_sheet"),
        original_filename=original_filename,
    )
    rack_name = rack_name or _pick_default_rack_name(original_filename, "", None, [row])
    rack_key = normalize_tag(rack_name) or "RACK_IMPORTADO"
    return rack_key, rack_name


def _looks_like_interface_head_module(module_raw):
    text = _ascii_upper(module_raw)
    alias = normalize_module_alias(module_raw)
    if not text and not alias:
        return False
    if "SIMATIC ET 200SP IM" in text or "IM 155" in text:
        return True
    if "CONTROLOGIX" in text or "CPU" in text:
        return True
    if "INTERFACE" in text or "HEAD" in text or "COUPLER" in text:
        return True
    if "AENTR" in alias or alias.startswith("CPU"):
        return True
    return bool(re.search(r"(^|[^A-Z])IM\d|(^|[^A-Z])IM($|[^A-Z0-9])", text))


def _normalize_module_signature(module_raw):
    text = _ascii_upper(module_raw)
    if not text:
        return ""
    text = re.sub(r"\([^)]*\)", "", text)
    text = re.sub(r"\b\d+\s*(DI|DO|AI|AO)\b", "", text)
    text = re.sub(r"\s+", " ", text).strip(" :-_/")
    return normalize_module_alias(text)


def _rack_group_slot_module_signatures(rows):
    signatures = set()
    for row in rows:
        slot_index = row.get("slot_index") or 0
        module_alias = _normalize_module_signature(row.get("module_raw"))
        channel_type = _ascii_upper(row.get("type"))
        if not slot_index and not module_alias and not channel_type:
            continue
        signatures.add((slot_index, module_alias, channel_type))
    return signatures


def _should_merge_spurious_interface_rack_group(
    rack_rows,
    dominant_rows,
    source_sheet,
    dominant_name,
    original_filename,
):
    if not rack_rows or not dominant_rows or not source_sheet:
        return False
    if len(rack_rows) > max(4, int(len(dominant_rows) * 0.08)):
        return False
    if len(dominant_rows) < max(8, len(rack_rows) * 4):
        return False

    rack_raw_values = [normalize_tag(row.get("rack_raw")) for row in rack_rows if row.get("rack_raw")]
    if not rack_raw_values:
        return False

    sheet_aliases = {
        normalize_tag(source_sheet),
        normalize_tag(_clean_rack_name(source_sheet)),
        normalize_tag(_infer_rack_name_from_sheet_name(source_sheet, original_filename=original_filename)),
    }
    if not all(value in sheet_aliases for value in rack_raw_values):
        return False

    module_values = [row.get("module_raw") for row in rack_rows if row.get("module_raw")]
    if not module_values:
        return False

    slot_indexes = {row.get("slot_index") for row in rack_rows if row.get("slot_index")}
    interface_only_group = all(_looks_like_interface_head_module(value) for value in module_values)
    if interface_only_group and slot_indexes and not slot_indexes.issubset({1}):
        return False

    duplicate_slot_module_group = False
    if not interface_only_group:
        rack_signatures = _rack_group_slot_module_signatures(rack_rows)
        dominant_signatures = _rack_group_slot_module_signatures(dominant_rows)
        duplicate_slot_module_group = bool(rack_signatures and rack_signatures.issubset(dominant_signatures))

    if not interface_only_group and not duplicate_slot_module_group:
        return False

    dominant_modules = [row.get("module_raw") for row in dominant_rows if row.get("module_raw")]
    if dominant_modules and all(_looks_like_interface_head_module(value) for value in dominant_modules):
        return False

    return True


def _collapse_spurious_interface_rack_aliases(normalized_rows, original_filename):
    if not normalized_rows:
        return

    rows_by_sheet = defaultdict(list)
    for row in normalized_rows:
        source_sheet = _cell_to_text(row.get("source_sheet"))
        if source_sheet:
            rows_by_sheet[source_sheet].append(row)

    for source_sheet, sheet_rows in rows_by_sheet.items():
        groups = defaultdict(list)
        for row in sheet_rows:
            groups[row.get("resolved_rack_key") or "RACK_IMPORTADO"].append(row)
        if len(groups) <= 1:
            continue

        dominant_key, dominant_rows = max(groups.items(), key=lambda item: len(item[1]))
        dominant_name = dominant_rows[0].get("resolved_rack_name") or dominant_key
        dominant_raw_name = next(
            (_clean_rack_name(row.get("rack_raw")) for row in dominant_rows if _clean_rack_name(row.get("rack_raw"))),
            "",
        ) or dominant_name

        for rack_key, rack_rows in list(groups.items()):
            if rack_key == dominant_key:
                continue
            if not _should_merge_spurious_interface_rack_group(
                rack_rows=rack_rows,
                dominant_rows=dominant_rows,
                source_sheet=source_sheet,
                dominant_name=dominant_name,
                original_filename=original_filename,
            ):
                continue
            for row in rack_rows:
                row["rack_raw"] = dominant_raw_name
                row["resolved_rack_key"] = dominant_key
                row["resolved_rack_name"] = dominant_name


def _group_rows_by_slot(normalized_rows):
    by_slot = defaultdict(list)
    without_slot = []
    for row in normalized_rows:
        if row.get("slot_index"):
            by_slot[row["slot_index"]].append(row)
        else:
            without_slot.append(row)
    return by_slot, without_slot


def _pack_rows_without_slot(rows, module_catalog, starting_slot):
    slot_cursor = starting_slot
    module_groups = []
    by_type = defaultdict(list)
    for row in rows:
        key = row.get("type") or "SEM_TIPO"
        by_type[key].append(row)

    for channel_type in sorted(by_type.keys()):
        grouped_rows = sorted(by_type[channel_type], key=lambda item: item["source_row"])
        while grouped_rows:
            desired = len(grouped_rows)
            module_model = _best_module_for_type(module_catalog, channel_type, desired)
            if not module_model:
                take = min(16, len(grouped_rows))
                selected_rows = grouped_rows[:take]
                grouped_rows = grouped_rows[take:]
                module_groups.append(
                    {"slot_index": slot_cursor, "module_model": None, "rows": selected_rows, "source": "packed"}
                )
                slot_cursor += 1
                continue
            capacity = max(1, module_model["quantidade_canais"])
            if len(grouped_rows) <= capacity:
                take = len(grouped_rows)
            else:
                smaller_sizes = [
                    item["quantidade_canais"]
                    for item in module_catalog
                    if item["tipo"] == channel_type and item["quantidade_canais"] < len(grouped_rows)
                ]
                take = max(smaller_sizes) if smaller_sizes else capacity
            selected_rows = grouped_rows[:take]
            grouped_rows = grouped_rows[take:]
            module_groups.append(
                {
                    "slot_index": slot_cursor,
                    "module_model": _best_module_for_type(module_catalog, channel_type, len(selected_rows)),
                    "rows": selected_rows,
                    "source": "packed",
                }
            )
            slot_cursor += 1
    return module_groups


def _build_channel_payloads(group_rows, module_model):
    capacity = module_model["quantidade_canais"] if module_model else len(group_rows)
    rows_by_channel = {}
    next_channel = 1
    for row in group_rows:
        channel_index = row.get("channel_index")
        if channel_index and channel_index not in rows_by_channel:
            rows_by_channel[channel_index] = row
            continue
        while next_channel in rows_by_channel:
            next_channel += 1
        rows_by_channel[next_channel] = row
        next_channel += 1

    channels = []
    for index in range(1, capacity + 1):
        row = rows_by_channel.get(index)
        if row:
            channels.append(
                {
                    "index": index,
                    "tag": row.get("tag") or "",
                    "description": row.get("description") or "",
                    "type": row.get("type") or (module_model or {}).get("tipo") or "",
                    "source_row": row.get("source_row"),
                    "issues": list(row.get("issues") or []),
                }
            )
        else:
            channels.append(
                {
                    "index": index,
                    "tag": "",
                    "description": "",
                    "type": (module_model or {}).get("tipo") or "",
                    "source_row": None,
                    "issues": ["canal_sem_linha_de_origem"],
                }
            )
    return channels


def _build_single_rack_proposal(rack_name, normalized_rows, module_catalog, target_rack=None):
    proposal_warnings = []
    by_slot, rows_without_slot = _group_rows_by_slot(normalized_rows)
    module_groups = []
    for slot_index in sorted(by_slot.keys()):
        slot_rows = sorted(by_slot[slot_index], key=lambda item: (item.get("channel_index") or 999999, item["source_row"]))
        explicit_model = next((item["module_raw"] for item in slot_rows if item.get("module_raw")), "")
        channel_type = next((item["type"] for item in slot_rows if item.get("type")), "")
        module_groups.append(
            {
                "slot_index": slot_index,
                "module_model": _best_module_for_type(module_catalog, channel_type, len(slot_rows), explicit_model=explicit_model),
                "rows": slot_rows,
                "source": "slot",
            }
        )

    next_slot = (max((group["slot_index"] for group in module_groups), default=0) or 0) + 1
    if rows_without_slot:
        module_groups.extend(_pack_rows_without_slot(rows_without_slot, module_catalog, next_slot))
        proposal_warnings.append(
            f"{len(rows_without_slot)} linhas nao traziam slot explicito e foram agrupadas automaticamente."
        )

    modules = []
    max_slot = 0
    conflicts = []
    for group in sorted(module_groups, key=lambda item: item["slot_index"]):
        slot_index = group["slot_index"]
        max_slot = max(max_slot, slot_index)
        module_model = group.get("module_model")
        custom_module = None
        if module_model and module_model.get("id") is None and module_model.get("tipo"):
            custom_module = {
                "modelo": module_model.get("modelo") or "",
                "nome": module_model.get("nome") or module_model.get("modelo") or "",
                "marca": module_model.get("marca") or "IMPORTADO",
                "tipo": module_model.get("tipo") or "",
                "quantidade_canais": int(module_model.get("quantidade_canais") or len(group["rows"]) or 1),
            }
        if not module_model:
            conflicts.append(
                {
                    "slot_index": slot_index,
                    "message": "Nao foi possivel resolver um modelo de modulo para este agrupamento.",
                }
            )
        modules.append(
            {
                "slot_index": slot_index,
                "module_model_id": module_model["id"] if module_model else None,
                "module_model_name": module_model["modelo"] if module_model else "",
                "module_type": module_model["tipo"] if module_model else "",
                "channel_capacity": module_model["quantidade_canais"] if module_model else len(group["rows"]),
                "module_model_source": "custom" if custom_module else "catalog",
                "custom_module": custom_module,
                "channels": _build_channel_payloads(group["rows"], module_model),
                "source": group["source"],
            }
        )

    return {
        "rack_key": normalize_tag(rack_name) or "RACK_IMPORTADO",
        "name": rack_name,
        "slots_total": max(max_slot, len(modules), 1),
        "target_rack_id": target_rack.id if target_rack else None,
        "source_sheets": sorted({row.get("source_sheet") for row in normalized_rows if row.get("source_sheet")}),
        "modules": modules,
        "warnings": proposal_warnings,
        "conflicts": conflicts,
        "summary": {
            "rows": len(normalized_rows),
            "modules": len(modules),
            "slots": max(max_slot, len(modules), 1),
            "with_conflicts": len(conflicts),
        },
    }


def build_import_proposal(
    original_filename,
    normalized_rows,
    module_catalog,
    requested_rack_name="",
    target_rack=None,
    ai_result=None,
):
    if not normalized_rows:
        raise IOImportError("Nao ha linhas normalizadas para montar a proposta.")

    racks_by_key = defaultdict(list)
    for row in normalized_rows:
        rack_key, rack_name = _resolve_rack_partition(row, original_filename=original_filename)
        row["resolved_rack_key"] = rack_key
        row["resolved_rack_name"] = rack_name
        racks_by_key[rack_key].append(row)

    _collapse_spurious_interface_rack_aliases(normalized_rows, original_filename=original_filename)
    racks_by_key = defaultdict(list)
    for row in normalized_rows:
        racks_by_key[row.get("resolved_rack_key") or "RACK_IMPORTADO"].append(row)

    rack_items = sorted(
        racks_by_key.items(),
        key=lambda item: (item[1][0].get("resolved_rack_name") or item[0], item[0]),
    )
    is_single_rack = len(rack_items) == 1

    racks = []
    global_warnings = []
    global_conflicts = []
    total_modules = 0
    total_slots = 0
    for rack_key, rack_rows in rack_items:
        rack_name = rack_rows[0].get("resolved_rack_name") or rack_key
        if is_single_rack:
            rack_name = _pick_default_rack_name(
                original_filename=original_filename,
                requested_rack_name=requested_rack_name,
                ai_result=ai_result,
                normalized_rows=rack_rows,
                target_rack=target_rack,
            )
        rack_payload = _build_single_rack_proposal(
            rack_name=rack_name,
            normalized_rows=rack_rows,
            module_catalog=module_catalog,
            target_rack=target_rack if is_single_rack else None,
        )
        racks.append(rack_payload)
        total_modules += rack_payload["summary"]["modules"]
        total_slots += rack_payload["summary"]["slots"]
        for warning in rack_payload.get("warnings") or []:
            global_warnings.append(f"[{rack_name}] {warning}")
        for conflict in rack_payload.get("conflicts") or []:
            global_conflicts.append(
                {
                    "rack_key": rack_payload["rack_key"],
                    "rack_name": rack_name,
                    "slot_index": conflict.get("slot_index"),
                    "message": conflict.get("message"),
                }
            )

    if target_rack and len(racks) > 1:
        global_conflicts.append(
            {
                "rack_key": normalize_tag(target_rack.nome) or "RACK_DESTINO",
                "rack_name": target_rack.nome,
                "slot_index": None,
                "message": "A proposta gerou multiplos racks; nao e possivel aplicar tudo em um unico rack destino.",
            }
        )

    proposal = {
        "racks": racks,
        "warnings": global_warnings,
        "conflicts": global_conflicts,
        "summary": {
            "rows": len(normalized_rows),
            "modules": total_modules,
            "slots": total_slots,
            "racks": len(racks),
            "with_conflicts": len(global_conflicts),
        },
    }
    if len(racks) == 1:
        proposal["rack"] = {
            "name": racks[0]["name"],
            "slots_total": racks[0]["slots_total"],
            "target_rack_id": racks[0]["target_rack_id"],
        }
        proposal["modules"] = racks[0]["modules"]
    else:
        proposal["rack"] = {}
        proposal["modules"] = []
    return proposal


def _build_progress_rack_snapshots_from_rows(original_filename, normalized_rows, limit=3):
    racks = {}
    for row in normalized_rows or []:
        rack_key, rack_name = _resolve_rack_partition(row, original_filename=original_filename)
        snapshot = racks.setdefault(
            rack_key,
            {
                "rack_key": rack_key,
                "rack_name": rack_name,
                "slots": set(),
                "channels": 0,
                "types": Counter(),
                "sample_tags": [],
            },
        )
        if row.get("slot_index"):
            snapshot["slots"].add(int(row["slot_index"]))
        if row.get("channel_index"):
            snapshot["channels"] += 1
        if row.get("type"):
            snapshot["types"][row["type"]] += 1
        tag_value = _cell_to_text(row.get("tag"))
        if tag_value and len(snapshot["sample_tags"]) < 3 and tag_value not in snapshot["sample_tags"]:
            snapshot["sample_tags"].append(tag_value)

    ordered = sorted(
        racks.values(),
        key=lambda item: (-item["channels"], -len(item["slots"]), item["rack_name"] or item["rack_key"]),
    )
    snapshots = []
    for item in ordered[:limit]:
        top_types = [f"{channel_type} {total}" for channel_type, total in item["types"].most_common(3)]
        snapshots.append(
            {
                "rack_key": item["rack_key"],
                "rack_name": item["rack_name"],
                "slots_count": len(item["slots"]),
                "channels_count": int(item["channels"]),
                "type_summary": top_types,
                "sample_tags": item["sample_tags"],
            }
        )
    return snapshots


def _build_progress_rack_snapshots_from_proposal(proposal, limit=3):
    snapshots = []
    for rack_payload in (proposal or {}).get("racks") or []:
        type_counter = Counter()
        sample_tags = []
        channels_total = 0
        for module_payload in rack_payload.get("modules") or []:
            for channel in module_payload.get("channels") or []:
                if channel.get("type"):
                    type_counter[channel["type"]] += 1
                if channel.get("tag"):
                    channels_total += 1
                    if len(sample_tags) < 3 and channel["tag"] not in sample_tags:
                        sample_tags.append(channel["tag"])
        snapshots.append(
            {
                "rack_key": rack_payload.get("rack_key") or normalize_tag(rack_payload.get("name")) or "RACK_IMPORTADO",
                "rack_name": rack_payload.get("name") or "Rack importado",
                "slots_count": int(rack_payload.get("slots_total") or len(rack_payload.get("modules") or []) or 0),
                "channels_count": channels_total,
                "type_summary": [f"{channel_type} {total}" for channel_type, total in type_counter.most_common(3)],
                "sample_tags": sample_tags,
            }
        )
    return snapshots[:limit]


def _emit_progress(progress_callback, stage, percent, title, message, **extra):
    if not callable(progress_callback):
        return
    payload = {
        "stage": _cell_to_text(stage).lower() or "upload",
        "percent": max(0, min(int(percent or 0), 100)),
        "title": _cell_to_text(title),
        "message": _cell_to_text(message),
    }
    for key, value in extra.items():
        if value is not None:
            payload[key] = value
    progress_callback(payload)


def _extract_response_text(payload):
    output = payload.get("output") or []
    for item in output:
        content = item.get("content") or []
        for content_item in content:
            if content_item.get("type") == "output_text":
                return content_item.get("text") or ""
    return payload.get("output_text") or ""


class _OpenAITransientError(Exception):
    pass


def _extract_response_error_message(payload):
    error = payload.get("error") or {}
    if isinstance(error, dict):
        code = _cell_to_text(error.get("code"))
        message = _cell_to_text(error.get("message"))
        if code and message:
            return f"{code}: {message}"
        if message:
            return message
    incomplete_details = payload.get("incomplete_details") or {}
    if isinstance(incomplete_details, dict):
        reason = _cell_to_text(incomplete_details.get("reason"))
        if reason:
            return reason
    return ""


def _openai_response_is_running(status):
    return _cell_to_text(status).lower() in {"queued", "in_progress", "running"}


def _openai_poll_interval_seconds(elapsed_seconds):
    if elapsed_seconds < 30:
        return 2.0
    if elapsed_seconds < 120:
        return 3.0
    return 5.0


def _openai_request_json(method, url, api_key, timeout_seconds, payload=None):
    headers = {"Authorization": f"Bearer {api_key}"}
    data = None
    if payload is not None:
        headers["Content-Type"] = "application/json"
        data = json.dumps(payload).encode("utf-8")
    http_request = urlrequest.Request(url=url, data=data, method=method, headers=headers)
    transient_http_codes = {408, 409, 425, 429, 500, 502, 503, 504}
    try:
        with urlrequest.urlopen(http_request, timeout=timeout_seconds) as response:
            raw_body = response.read().decode("utf-8")
        return json.loads(raw_body) if raw_body else {}
    except urlerror.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")
        message = f"Falha na chamada do agente: HTTP {exc.code} - {detail[:400]}"
        if exc.code in transient_http_codes:
            raise _OpenAITransientError(message)
        raise IOImportError(message)
    except urlerror.URLError as exc:
        raise _OpenAITransientError(f"Falha na chamada do agente: {exc.reason}")
    except (TimeoutError, socket.timeout):
        raise _OpenAITransientError(f"Falha na chamada do agente: timeout de transporte apos {timeout_seconds:.1f}s.")
    except OSError as exc:
        raise _OpenAITransientError(f"Falha na chamada do agente: {exc}")


def _call_openai_responses(
    settings_obj,
    schema_name,
    schema,
    system_prompt,
    user_prompt,
    request_timeout_seconds=None,
    reasoning_effort_override=None,
):
    if not settings_obj.enabled:
        raise IOImportError("Agente de importacao desativado.")
    if settings_obj.provider != settings_obj.Provider.OPENAI:
        raise IOImportError("Provider de agente nao suportado nesta versao.")
    if not settings_obj.api_key:
        raise IOImportError("API key do agente nao configurada.")

    base_url = (settings_obj.api_base_url or "").strip().rstrip("/")
    if not base_url:
        base_url = "https://api.openai.com/v1"
    payload = {
        "model": settings_obj.model,
        "input": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
        "text": {
            "format": {
                "type": "json_schema",
                "name": schema_name,
                "strict": True,
                "schema": schema,
            }
        },
        "background": True,
        "store": True,
        "metadata": {
            "schema_name": schema_name[:64],
            "source": "io_import",
            "client_request_id": uuid.uuid4().hex[:32],
        },
    }
    effective_reasoning = _cell_to_text(reasoning_effort_override or settings_obj.reasoning_effort).lower()
    if effective_reasoning and effective_reasoning != "none":
        payload["reasoning"] = {"effort": effective_reasoning}

    transport_timeout_seconds = max(15.0, float(request_timeout_seconds or _ai_request_timeout_seconds()))
    max_attempts = max(1, _ai_max_retries())
    created_payload = None
    last_error = ""
    for attempt in range(1, max_attempts + 1):
        try:
            created_payload = _openai_request_json(
                method="POST",
                url=f"{base_url}/responses",
                api_key=settings_obj.api_key,
                timeout_seconds=transport_timeout_seconds,
                payload=payload,
            )
            break
        except _OpenAITransientError as exc:
            last_error = str(exc)
            if attempt < max_attempts:
                time.sleep(min(4.0, 1.25 * attempt))
                continue
            raise IOImportError(last_error)

    if created_payload is None:
        raise IOImportError(last_error or "Falha na chamada do agente.")

    response_payload = created_payload
    response_id = _cell_to_text(response_payload.get("id"))
    if not response_id:
        raise IOImportError("O agente nao retornou um identificador de resposta utilizavel.")

    polling_started_at = time.monotonic()
    transient_poll_errors = 0
    while True:
        response_text = _extract_response_text(response_payload)
        if response_text:
            break

        status = _cell_to_text(response_payload.get("status")).lower()
        if not _openai_response_is_running(status):
            error_detail = _extract_response_error_message(response_payload)
            raise IOImportError(
                error_detail or f"O agente finalizou sem texto utilizavel (status {status or 'desconhecido'})."
            )

        elapsed_seconds = time.monotonic() - polling_started_at
        time.sleep(_openai_poll_interval_seconds(elapsed_seconds))
        try:
            response_payload = _openai_request_json(
                method="GET",
                url=f"{base_url}/responses/{response_id}",
                api_key=settings_obj.api_key,
                timeout_seconds=transport_timeout_seconds,
            )
            transient_poll_errors = 0
        except _OpenAITransientError as exc:
            transient_poll_errors += 1
            last_error = str(exc)
            if transient_poll_errors >= max_attempts:
                raise IOImportError(last_error)
            continue

    try:
        return json.loads(response_text)
    except json.JSONDecodeError as exc:
        raise IOImportError(f"O agente retornou JSON invalido: {exc}")


def run_ai_workbook_analysis(settings_obj, parsed_sheets, original_filename, file_sha256="", request_timeout_seconds=None):
    schema = {
        "type": "object",
        "properties": {
            "sheets": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "sheet_name": {"type": "string"},
                        "use_sheet": {"type": "boolean"},
                        "layout_hint": {"type": "string", "enum": ["tabular", "slot_blocks", "hybrid", "unknown"]},
                        "confidence": {"type": "integer"},
                        "reason": {"type": "string"},
                    },
                    "required": ["sheet_name", "use_sheet", "layout_hint", "confidence", "reason"],
                    "additionalProperties": False,
                },
            },
            "warnings": {"type": "array", "items": {"type": "string"}},
            "notes": {"type": "string"},
        },
        "required": ["sheets", "warnings", "notes"],
        "additionalProperties": False,
    }
    cache_fingerprint = _build_workbook_cache_fingerprint(file_sha256=file_sha256, settings_obj=settings_obj, parsed_sheets=parsed_sheets)
    cached_payload = _load_ai_cache("WORKBOOK", cache_fingerprint)
    if cached_payload:
        return cached_payload

    system_prompt = (
        "You are the primary interpreter for a heterogeneous industrial IO workbook. "
        "Decide which sheets carry operational IO data, which are helper/index/noise sheets, "
        "and what layout each relevant sheet uses. Prefer broad inclusion when uncertain."
    )
    workbook_context = None
    response_payload = None
    last_exc = None
    workbook_row_profiles = sorted(
        {
            _effective_workbook_context_limit(len(parsed_sheets or [])),
            max(16, _effective_workbook_context_limit(len(parsed_sheets or [])) // 2),
        },
        reverse=True,
    )
    for row_limit in workbook_row_profiles:
        try:
            workbook_context = {
                "original_filename": original_filename,
                "sheets": [
                    {
                        "sheet_name": parsed.sheet_name,
                        "rows_total": parsed.rows_total,
                        "header_row_number": parsed.header_row_index + 1,
                        "layout": parsed.layout,
                        "headers": parsed.headers,
                        "parser_warnings": parsed.warnings,
                        "column_map": parsed.column_map,
                        "data_rows": _count_non_empty_data_rows(parsed.raw_rows, parsed.header_row_index),
                        "slot_block_rows": _count_slot_block_rows(parsed.raw_rows),
                        "raw_rows": _build_raw_rows_payload(parsed, max_rows=row_limit),
                    }
                    for parsed in parsed_sheets
                ],
            }
            user_prompt = json.dumps(workbook_context, ensure_ascii=True)
            response_payload = _call_openai_responses(
                settings_obj=settings_obj,
                schema_name="io_workbook_analysis",
                schema=schema,
                system_prompt=system_prompt,
                user_prompt=user_prompt,
                request_timeout_seconds=request_timeout_seconds,
                reasoning_effort_override=_effective_reasoning_effort(
                    settings_obj=settings_obj,
                    schema_name="io_workbook_analysis",
                    parsed=None,
                    total_sheets=len(parsed_sheets or []),
                ),
            )
            break
        except IOImportError as exc:
            last_exc = exc
            continue
    if response_payload is None:
        raise last_exc or IOImportError("Falha na analise estrutural com IA.")
    _save_ai_cache(
        stage="WORKBOOK",
        fingerprint=cache_fingerprint,
        file_sha256=file_sha256,
        settings_obj=settings_obj,
        response_payload=response_payload,
        payload_meta={
            "sheets_total": len(parsed_sheets or []),
            "context_rows_total": sum(len((item or {}).get("raw_rows") or []) for item in workbook_context.get("sheets") or []),
        },
    )
    return response_payload


def run_ai_analysis(
    settings_obj,
    parsed,
    normalized_rows,
    module_catalog,
    file_sha256="",
    workbook_plan=None,
    request_timeout_seconds=None,
    total_sheets=1,
):
    sheet_row_limit = _effective_ai_sheet_row_limit(settings_obj, total_sheets)
    if parsed.layout == "slot_blocks":
        sheet_row_limit = min(sheet_row_limit, 72 if total_sheets <= 2 else 60)
    elif parsed.layout == "tabular":
        sheet_row_limit = min(sheet_row_limit, 96 if total_sheets == 1 else 120)
    bootstrap_rows = []
    for row in (normalized_rows or [])[: min(len(normalized_rows or []), sheet_row_limit)]:
        bootstrap_rows.append(
            {
                "source_row": row["source_row"],
                "rack_raw": row["rack_raw"],
                "slot_raw": row["slot_raw"],
                "module_raw": row["module_raw"],
                "channel_raw": row["channel_raw"],
                "tag": row["tag"],
                "description": row["description"],
                "type": row["type"],
            }
        )

    schema = {
        "type": "object",
        "properties": {
            "sheet_role": {"type": "string", "enum": ["data", "helper", "noise"]},
            "skip_sheet": {"type": "boolean"},
            "layout_hint": {"type": "string", "enum": ["tabular", "slot_blocks", "hybrid", "unknown"]},
            "rack_name": {"type": "string"},
            "column_map": {
                "type": "object",
                "properties": {
                    "panel": {
                        "type": "object",
                        "properties": {
                            "header": {"type": "string"},
                            "confidence": {"type": "integer"},
                            "mode": {"type": "string", "enum": ["fill", "override"]},
                        },
                        "required": ["header", "confidence", "mode"],
                        "additionalProperties": False,
                    },
                    "rack": {
                        "type": "object",
                        "properties": {
                            "header": {"type": "string"},
                            "confidence": {"type": "integer"},
                            "mode": {"type": "string", "enum": ["fill", "override"]},
                        },
                        "required": ["header", "confidence", "mode"],
                        "additionalProperties": False,
                    },
                    "slot": {
                        "type": "object",
                        "properties": {
                            "header": {"type": "string"},
                            "confidence": {"type": "integer"},
                            "mode": {"type": "string", "enum": ["fill", "override"]},
                        },
                        "required": ["header", "confidence", "mode"],
                        "additionalProperties": False,
                    },
                    "module_model": {
                        "type": "object",
                        "properties": {
                            "header": {"type": "string"},
                            "confidence": {"type": "integer"},
                            "mode": {"type": "string", "enum": ["fill", "override"]},
                        },
                        "required": ["header", "confidence", "mode"],
                        "additionalProperties": False,
                    },
                    "channel": {
                        "type": "object",
                        "properties": {
                            "header": {"type": "string"},
                            "confidence": {"type": "integer"},
                            "mode": {"type": "string", "enum": ["fill", "override"]},
                        },
                        "required": ["header", "confidence", "mode"],
                        "additionalProperties": False,
                    },
                    "location": {
                        "type": "object",
                        "properties": {
                            "header": {"type": "string"},
                            "confidence": {"type": "integer"},
                            "mode": {"type": "string", "enum": ["fill", "override"]},
                        },
                        "required": ["header", "confidence", "mode"],
                        "additionalProperties": False,
                    },
                    "tag": {
                        "type": "object",
                        "properties": {
                            "header": {"type": "string"},
                            "confidence": {"type": "integer"},
                            "mode": {"type": "string", "enum": ["fill", "override"]},
                        },
                        "required": ["header", "confidence", "mode"],
                        "additionalProperties": False,
                    },
                    "description": {
                        "type": "object",
                        "properties": {
                            "header": {"type": "string"},
                            "confidence": {"type": "integer"},
                            "mode": {"type": "string", "enum": ["fill", "override"]},
                        },
                        "required": ["header", "confidence", "mode"],
                        "additionalProperties": False,
                    },
                    "type": {
                        "type": "object",
                        "properties": {
                            "header": {"type": "string"},
                            "confidence": {"type": "integer"},
                            "mode": {"type": "string", "enum": ["fill", "override"]},
                        },
                        "required": ["header", "confidence", "mode"],
                        "additionalProperties": False,
                    },
                },
                "required": ["panel", "rack", "slot", "module_model", "channel", "location", "tag", "description", "type"],
                "additionalProperties": False,
            },
            "logical_points": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "source_row": {"type": "integer"},
                        "panel": {"type": "string"},
                        "rack": {"type": "string"},
                        "slot": {"type": "string"},
                        "module_model": {"type": "string"},
                        "channel": {"type": "string"},
                        "tag": {"type": "string"},
                        "description": {"type": "string"},
                        "type": {"type": "string"},
                        "confidence": {"type": "integer"},
                    },
                    "required": [
                        "source_row",
                        "panel",
                        "rack",
                        "slot",
                        "module_model",
                        "channel",
                        "tag",
                        "description",
                        "type",
                        "confidence",
                    ],
                    "additionalProperties": False,
                },
            },
            "row_hints": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "source_row": {"type": "integer"},
                        "row_kind": {"type": "string", "enum": ["data", "section", "subheader", "noise"]},
                        "panel": {"type": "string"},
                        "rack": {"type": "string"},
                        "slot": {"type": "string"},
                        "module_model": {"type": "string"},
                        "channel": {"type": "string"},
                        "tag": {"type": "string"},
                        "description": {"type": "string"},
                        "type": {"type": "string"},
                        "confidence": {"type": "integer"},
                    },
                    "required": [
                        "source_row",
                        "row_kind",
                        "panel",
                        "rack",
                        "slot",
                        "module_model",
                        "channel",
                        "tag",
                        "description",
                        "type",
                        "confidence",
                    ],
                    "additionalProperties": False,
                },
            },
            "warnings": {"type": "array", "items": {"type": "string"}},
            "notes": {"type": "string"},
        },
        "required": ["sheet_role", "skip_sheet", "layout_hint", "rack_name", "column_map", "logical_points", "row_hints", "warnings", "notes"],
        "additionalProperties": False,
    }
    system_prompt = (
        f"{settings_obj.header_prompt}\n\n{settings_obj.grouping_prompt}\n\n"
        "You are the primary semantic parser for this IO sheet. Use the raw rows as the source of truth. "
        "Prefer a sparse guidance strategy: return a strong column_map, rack_name and only the row_hints that "
        "correct, discard or contextualize ambiguous rows. Only return logical_points when the sheet cannot be "
        "materialized safely from column_map plus selective row_hints alone. For regular tabular or slot-block "
        "sheets, avoid enumerating every point. The local parser will materialize the full dataset from your "
        "semantic guidance."
    )
    row_profiles = sorted({sheet_row_limit, max(24, sheet_row_limit // 2)}, reverse=True)
    base_catalog_limit = 14 if total_sheets >= 12 else 18
    compact_catalog_seed = _build_compact_module_catalog(
        module_catalog=module_catalog,
        parsed=parsed,
        workbook_plan=workbook_plan,
        limit=base_catalog_limit,
    )
    cache_fingerprint = _build_sheet_cache_fingerprint(
        file_sha256=file_sha256,
        settings_obj=settings_obj,
        parsed=parsed,
        workbook_plan=workbook_plan,
        total_sheets=total_sheets,
        compact_catalog=compact_catalog_seed,
    )
    cached_payload = _load_ai_cache("SHEET", cache_fingerprint)
    if cached_payload:
        return cached_payload

    catalog_profiles = sorted({base_catalog_limit, 10}, reverse=True)
    user_payload = None
    compact_catalog = []
    response_payload = None
    last_exc = None
    for row_limit in row_profiles:
        for catalog_limit in catalog_profiles:
            compact_catalog = _build_compact_module_catalog(
                module_catalog=module_catalog,
                parsed=parsed,
                workbook_plan=workbook_plan,
                limit=catalog_limit,
            )
            user_payload = {
                "sheet_name": parsed.sheet_name,
                "layout": parsed.layout,
                "header_row_number": parsed.header_row_index + 1,
                "headers": parsed.headers,
                "sheet_summary": _build_sheet_semantic_summary(parsed),
                "parser_warnings": parsed.warnings,
                "detected_column_map": parsed.column_map,
                "workbook_plan": workbook_plan or {},
                "raw_rows": _build_raw_rows_payload(parsed, max_rows=row_limit),
                "module_catalog": compact_catalog,
                "module_catalog_stats": _module_catalog_stats(compact_catalog),
            }
            if bootstrap_rows:
                user_payload["bootstrap_rows"] = bootstrap_rows
            try:
                user_prompt = json.dumps(user_payload, ensure_ascii=True)
                response_payload = _call_openai_responses(
                    settings_obj=settings_obj,
                    schema_name="io_sheet_semantic_analysis",
                    schema=schema,
                    system_prompt=system_prompt,
                    user_prompt=user_prompt,
                    request_timeout_seconds=request_timeout_seconds,
                    reasoning_effort_override=_effective_reasoning_effort(
                        settings_obj=settings_obj,
                        schema_name="io_sheet_semantic_analysis",
                        parsed=parsed,
                        total_sheets=total_sheets,
                    ),
                )
                break
            except IOImportError as exc:
                last_exc = exc
                continue
        if response_payload is not None:
            break
    if response_payload is None:
        raise last_exc or IOImportError(f"Falha na analise com IA da guia {parsed.sheet_name}.")
    _save_ai_cache(
        stage="SHEET",
        fingerprint=cache_fingerprint,
        file_sha256=file_sha256,
        sheet_name=parsed.sheet_name,
        settings_obj=settings_obj,
        response_payload=response_payload,
        payload_meta={
            "raw_rows_sent": len(user_payload.get("raw_rows") or []),
            "catalog_items_sent": len(compact_catalog),
            "layout": parsed.layout,
        },
    )
    return response_payload


def _format_sheet_warning(sheet_name, warning, multi_sheet):
    warning = _cell_to_text(warning)
    if not warning:
        return ""
    if multi_sheet and sheet_name:
        return f"[{sheet_name}] {warning}"
    return warning


def reprocess_import_job(job, module_catalog, settings_obj=None, progress_callback=None):
    if not job.source_file:
        raise IOImportError("Arquivo fonte da importacao nao encontrado.")
    ai_required = bool(settings_obj and settings_obj.enabled)
    _emit_progress(
        progress_callback,
        stage="upload",
        percent=6,
        title="Arquivo recebido",
        message="O arquivo foi recebido e a leitura inicial da planilha esta começando.",
        progress_label="Arquivo recebido",
        snapshots=[],
    )
    job.source_file.open("rb")
    raw_bytes = job.source_file.read()
    job.source_file.close()

    parsed_sheets = parse_workbook(raw_bytes=raw_bytes, original_filename=job.original_filename)
    _emit_progress(
        progress_callback,
        stage="parse",
        percent=18,
        title="Estrutura da planilha identificada",
        message=(
            f"{len(parsed_sheets)} guia(s) localizada(s). A estrutura base esta sendo organizada para a analise."
        ),
        progress_label="Estrutura identificada",
        sheets_total=len(parsed_sheets),
        sheets_processed=0,
        snapshots=[],
    )
    multi_sheet = len(parsed_sheets) > 1
    normalized_rows = []
    warnings = []
    effective_map = {}
    sheet_summaries = []
    ai_status = job.AIStatus.SKIPPED
    ai_payload = {"sheets": {}}
    ai_errors = []
    ai_model = settings_obj.model if settings_obj and settings_obj.enabled else ""
    ai_success = 0
    ai_attempts = 0
    processed_sheets = []
    ai_max_sheets = _ai_max_sheets_per_job()
    ai_rack_names = []
    workbook_ai_payload = {}
    workbook_ai_error = ""
    single_sheet_fast_path = len(parsed_sheets) == 1 and _cell_to_text(parsed_sheets[0].layout).lower() in {"tabular", "slot_blocks", "hybrid"}

    if settings_obj and settings_obj.enabled:
        _emit_progress(
            progress_callback,
            stage="ai",
            percent=28,
            title="Entendendo a estrutura da importacao",
            message="A analise esta separando guias uteis, resumos e contextos de rack antes da leitura detalhada.",
            progress_label="Preparando leitura guiada",
            sheets_total=len(parsed_sheets),
            sheets_processed=0,
            snapshots=[],
        )
        if single_sheet_fast_path:
            workbook_ai_payload = _build_single_sheet_fast_workbook_plan(parsed_sheets)
            ai_success += 1
            _emit_progress(
                progress_callback,
                stage="ai",
                percent=31,
                title="Estrutura principal identificada",
                message="A guia unica foi validada e a analise semantica detalhada vai comecar.",
                progress_label="Guia principal validada",
                sheets_total=len(parsed_sheets),
                sheets_processed=0,
                snapshots=[],
            )
        elif ai_attempts < ai_max_sheets:
            try:
                ai_attempts += 1
                workbook_ai_payload = run_ai_workbook_analysis(
                    settings_obj=settings_obj,
                    parsed_sheets=parsed_sheets,
                    original_filename=job.original_filename,
                    file_sha256=job.file_sha256,
                )
                ai_success += 1
            except IOImportError as exc:
                workbook_ai_error = str(exc)
                ai_errors.append(workbook_ai_error)
                if ai_required:
                    raise IOImportError(
                        f"A analise com IA nao conseguiu entender a estrutura geral da planilha. {workbook_ai_error}"
                    ) from exc
        else:
            workbook_ai_error = "A analise com IA excedeu o limite de chamadas previsto para esta importacao."
            ai_errors.append(workbook_ai_error)
            if ai_required:
                raise IOImportError(workbook_ai_error)

    total_sheets = max(len(parsed_sheets), 1)
    processed_count = 0
    for sheet_index, parsed in enumerate(parsed_sheets, start=1):
        sheet_ai_payload = {}
        sheet_warnings = []
        skip_by_ai, skip_reason = _should_skip_sheet_by_ai(workbook_ai_payload, parsed)
        current_percent = 34 + int(((sheet_index - 1) / total_sheets) * 46)
        _emit_progress(
            progress_callback,
            stage="ai" if settings_obj and settings_obj.enabled else "parse",
            percent=current_percent,
            title="Correlacionando guias e sinais",
            message=f"Analisando a guia {sheet_index} de {total_sheets}: {parsed.sheet_name}.",
            progress_label=f"Guia {sheet_index} de {total_sheets}",
            current_sheet=parsed.sheet_name,
            current_sheet_index=sheet_index,
            sheets_total=total_sheets,
            sheets_processed=processed_count,
            snapshots=_build_progress_rack_snapshots_from_rows(job.original_filename, normalized_rows),
        )
        if skip_by_ai:
            formatted_skip_reason = (
                f"Guia ignorada pela IA: {skip_reason}" if skip_reason else "Guia ignorada pela IA."
            )
            warnings.append(_format_sheet_warning(parsed.sheet_name, formatted_skip_reason, multi_sheet))
            ai_payload["sheets"][parsed.sheet_name] = {
                "sheet_role": "helper",
                "skip_sheet": True,
                "warnings": [formatted_skip_reason],
            }
            sheet_summaries.append(
                {
                    "sheet_name": parsed.sheet_name,
                    "header_row_index": parsed.header_row_index + 1,
                    "rows_total": parsed.rows_total,
                    "rows_parsed": 0,
                    "layout": parsed.layout,
                    "column_map": {},
                      "skipped": True,
                  }
              )
            processed_count += 1
            continue

        try:
            if settings_obj and settings_obj.enabled:
                if ai_attempts >= ai_max_sheets:
                    message = (
                        f"A analise com IA atingiu o limite previsto antes de concluir a guia {parsed.sheet_name}."
                    )
                    ai_errors.append(_format_sheet_warning(parsed.sheet_name, message, multi_sheet))
                    if ai_required:
                        raise IOImportError(message)
                    sheet_warnings.append(message)
                    sheet_rows, sheet_map, fallback_warnings = normalize_rows(parsed=parsed, module_catalog=module_catalog, ai_result=None)
                    sheet_warnings.extend(fallback_warnings)
                else:
                    try:
                        ai_attempts += 1
                        sheet_ai_payload = run_ai_analysis(
                            settings_obj=settings_obj,
                            parsed=parsed,
                            normalized_rows=[],
                            module_catalog=module_catalog,
                            file_sha256=job.file_sha256,
                            workbook_plan=_find_workbook_sheet_plan(workbook_ai_payload, parsed.sheet_name),
                            total_sheets=total_sheets,
                        )
                        if sheet_ai_payload.get("skip_sheet"):
                            skip_reason = _cell_to_text((sheet_ai_payload.get("warnings") or [""])[0]) or "Guia sem dados operacionais."
                            warnings.append(_format_sheet_warning(parsed.sheet_name, f"Guia ignorada pela IA: {skip_reason}", multi_sheet))
                            ai_payload["sheets"][parsed.sheet_name] = sheet_ai_payload
                            sheet_summaries.append(
                                {
                                    "sheet_name": parsed.sheet_name,
                                    "header_row_index": parsed.header_row_index + 1,
                                    "rows_total": parsed.rows_total,
                                    "rows_parsed": 0,
                                    "layout": parsed.layout,
                                    "column_map": {},
                                    "skipped": True,
                                }
                            )
                            ai_success += 1
                            continue
                        if (sheet_ai_payload.get("logical_points") or []):
                            sheet_rows, sheet_map, sheet_warnings = normalize_rows_from_ai_result(
                                parsed=parsed,
                                module_catalog=module_catalog,
                                ai_result=sheet_ai_payload,
                            )
                        else:
                            sheet_rows, sheet_map, sheet_warnings = normalize_rows(
                                parsed=parsed,
                                module_catalog=module_catalog,
                                ai_result=sheet_ai_payload,
                            )
                        if _cell_to_text(sheet_ai_payload.get("rack_name")):
                            ai_rack_names.append(_cell_to_text(sheet_ai_payload.get("rack_name")))
                        ai_success += 1
                    except IOImportError as exc:
                        ai_errors.append(_format_sheet_warning(parsed.sheet_name, str(exc), multi_sheet))
                        if ai_required:
                            raise IOImportError(
                                f"A analise com IA nao conseguiu concluir a guia {parsed.sheet_name}. {exc}"
                            ) from exc
                        sheet_warnings.append(str(exc))
                        sheet_rows, sheet_map, fallback_warnings = normalize_rows(
                            parsed=parsed,
                            module_catalog=module_catalog,
                            ai_result=None,
                        )
                        sheet_warnings.extend(fallback_warnings)
            else:
                sheet_rows, sheet_map, sheet_warnings = normalize_rows(
                    parsed=parsed,
                    module_catalog=module_catalog,
                    ai_result=None,
                )
        except IOImportError as exc:
            if ai_required:
                raise
            warnings.append(_format_sheet_warning(parsed.sheet_name, str(exc), multi_sheet))
            sheet_summaries.append(
                {
                    "sheet_name": parsed.sheet_name,
                    "header_row_index": parsed.header_row_index + 1,
                    "rows_total": parsed.rows_total,
                    "rows_parsed": 0,
                    "layout": parsed.layout,
                    "column_map": {},
                    "skipped": True,
                }
            )
            processed_count += 1
            continue
        if not sheet_rows:
            skip_reason = _cell_to_text((sheet_ai_payload or {}).get("notes")) or "Guia sem linhas operacionais materializadas."
            warnings.append(_format_sheet_warning(parsed.sheet_name, f"Guia ignorada apos a leitura semantica: {skip_reason}", multi_sheet))
            ai_payload["sheets"][parsed.sheet_name] = sheet_ai_payload
            sheet_summaries.append(
                {
                    "sheet_name": parsed.sheet_name,
                    "header_row_index": parsed.header_row_index + 1,
                    "rows_total": parsed.rows_total,
                    "rows_parsed": 0,
                    "layout": parsed.layout,
                    "column_map": {},
                    "skipped": True,
                }
            )
            processed_count += 1
            continue
        normalized_rows.extend(sheet_rows)
        effective_map[parsed.sheet_name] = sheet_map
        ai_payload["sheets"][parsed.sheet_name] = sheet_ai_payload
        sheet_summaries.append(
            {
                "sheet_name": parsed.sheet_name,
                "header_row_index": parsed.header_row_index + 1,
                "rows_total": parsed.rows_total,
                "rows_parsed": len(sheet_rows),
                "layout": parsed.layout,
                "column_map": sheet_map,
            }
        )
        processed_sheets.append(parsed)
        processed_count += 1
        for warning in sheet_warnings:
            formatted = _format_sheet_warning(parsed.sheet_name, warning, multi_sheet)
            if formatted:
                warnings.append(formatted)
        _emit_progress(
            progress_callback,
            stage="ai" if settings_obj and settings_obj.enabled else "parse",
            percent=34 + int((processed_count / total_sheets) * 46),
            title="Guia consolidada",
            message=(
                f"{parsed.sheet_name} foi consolidada. "
                f"{len(normalized_rows)} linha(s) util(is) acumulada(s) ate agora."
            ),
            progress_label=f"Guia {processed_count} de {total_sheets}",
            current_sheet=parsed.sheet_name,
            current_sheet_index=processed_count,
            sheets_total=total_sheets,
            sheets_processed=processed_count,
            snapshots=_build_progress_rack_snapshots_from_rows(job.original_filename, normalized_rows),
        )

    if not normalized_rows:
        raise IOImportError("Nenhuma aba util gerou linhas normalizadas para importacao.")

    if settings_obj and settings_obj.enabled:
        expected_successes = len(processed_sheets) + (1 if workbook_ai_payload else 0)
        if ai_success >= max(expected_successes, 1):
            ai_status = job.AIStatus.SUCCESS
        elif ai_success == 0:
            ai_status = job.AIStatus.FAILED
        else:
            ai_status = job.AIStatus.FAILED
        ai_error = " | ".join(error for error in ai_errors if error)
        if ai_error:
            warnings.extend(error for error in ai_errors if error)
    else:
        ai_error = ""

    if len(set(name for name in ai_rack_names if name)) == 1:
        ai_payload["rack_name"] = ai_rack_names[0]
    if workbook_ai_payload:
      ai_payload["workbook"] = workbook_ai_payload
    elif workbook_ai_error:
      ai_payload["workbook_error"] = workbook_ai_error

    _emit_progress(
        progress_callback,
        stage="preview",
        percent=90,
        title="Montando sugestoes para revisao",
        message="Os racks sugeridos estao sendo organizados para abrir a revisao final.",
        progress_label="Montando sugestoes",
        sheets_total=total_sheets,
        sheets_processed=processed_count,
        snapshots=_build_progress_rack_snapshots_from_rows(job.original_filename, normalized_rows),
    )
    proposal = build_import_proposal(
        original_filename=job.original_filename,
        normalized_rows=normalized_rows,
        module_catalog=module_catalog,
        requested_rack_name=job.requested_rack_name,
        target_rack=job.target_rack,
        ai_result=ai_payload,
    )
    final_snapshots = _build_progress_rack_snapshots_from_proposal(proposal)
    _emit_progress(
        progress_callback,
        stage="preview",
        percent=100,
        title="Sugestoes prontas",
        message="A estrutura sugerida foi consolidada e esta pronta para revisao.",
        progress_label="Sugestoes prontas",
        sheets_total=total_sheets,
        sheets_processed=processed_count,
        snapshots=final_snapshots,
        summary=proposal.get("summary") or {},
    )
    primary_sheet = processed_sheets[0]
    sheet_label = primary_sheet.sheet_name if len(processed_sheets) == 1 else f"{primary_sheet.sheet_name} +{len(processed_sheets) - 1}"
    return {
        "file_format": primary_sheet.file_format,
        "sheet_name": sheet_label,
        "header_row_index": primary_sheet.header_row_index + 1 if len(processed_sheets) == 1 else None,
        "rows_total": sum(item.rows_total for item in processed_sheets),
        "rows_parsed": len(normalized_rows),
        "column_map": effective_map,
        "warnings": warnings,
        "normalized_rows": normalized_rows,
        "sheet_summaries": sheet_summaries,
        "proposal": proposal,
        "ai_status": ai_status,
        "ai_payload": ai_payload,
        "ai_error": ai_error,
        "ai_model": ai_model,
        "progress_payload": {
            "stage": "preview",
            "percent": 100,
            "title": "Sugestoes prontas",
            "message": "A estrutura sugerida foi consolidada e esta pronta para revisao.",
            "progress_label": "Sugestoes prontas",
            "sheets_total": total_sheets,
            "sheets_processed": processed_count,
            "snapshots": final_snapshots,
            "summary": proposal.get("summary") or {},
        },
    }


def _apply_single_rack_payload(
    job,
    rack_payload,
    target_rack,
    module_map,
    module_qs,
    rack_model,
    rack_slot_model,
    rack_module_model,
    channel_model,
    plant_model,
):
    type_model = module_qs.model._meta.get_field("tipo_base").remote_field.model
    modules_payload = rack_payload.get("modules") or []
    if not modules_payload:
        raise IOImportError("Um dos racks propostos nao possui modulos para aplicar.")

    rack = target_rack
    if rack is None:
        rack = rack_model.objects.create(
            cliente=job.cliente,
            inventario=job.requested_inventario,
            local=job.requested_local,
            grupo=job.requested_grupo,
            id_planta=(
                plant_model.objects.get_or_create(codigo=job.requested_planta_code.upper())[0]
                if job.requested_planta_code
                else None
            ),
            nome=(rack_payload.get("name") or job.requested_rack_name or "Rack importado")[:120],
            descricao=f"Importado de planilha: {job.original_filename}"[:500],
            slots_total=max(int(rack_payload.get("slots_total") or len(modules_payload) or 1), 1),
        )
        rack_slot_model.objects.bulk_create(
            [rack_slot_model(rack=rack, posicao=index) for index in range(1, rack.slots_total + 1)]
        )
    else:
        max_slot = max(int(module["slot_index"]) for module in modules_payload)
        if max_slot > rack.slots_total:
            rack_slot_model.objects.bulk_create(
                [rack_slot_model(rack=rack, posicao=index) for index in range(rack.slots_total + 1, max_slot + 1)]
            )
            rack.slots_total = max_slot
            rack.save(update_fields=["slots_total"])

    slot_map = {
        slot.posicao: slot
        for slot in rack.slots.select_related("modulo", "modulo__modulo_modelo").all()
    }

    for module_payload in modules_payload:
        slot_index = int(module_payload["slot_index"])
        slot = slot_map.get(slot_index)
        if not slot:
            raise IOImportError(f"Slot {slot_index} nao encontrado no rack destino.")
        module_model_id = module_payload.get("module_model_id")
        module_model = module_map.get(str(module_model_id)) if module_model_id else None
        if not module_model:
            custom_module = module_payload.get("custom_module") or {}
            module_type = normalize_type(custom_module.get("tipo") or module_payload.get("module_type"))
            capacity = int(
                custom_module.get("quantidade_canais")
                or module_payload.get("channel_capacity")
                or len(module_payload.get("channels") or [])
                or 1
            )
            if not module_type:
                raise IOImportError(f"Modulo do slot {slot_index} nao esta catalogado e o tipo nao pode ser inferido.")
            tipo_base = type_model.objects.filter(nome=module_type).first()
            if not tipo_base:
                raise IOImportError(f"Tipo base {module_type} nao encontrado para criar modulo do slot {slot_index}.")
            modelo = (
                _cell_to_text(custom_module.get("modelo"))
                or _cell_to_text(module_payload.get("module_model_name"))
                or f"{module_type}-{capacity:02d} CUSTOM"
            )[:80]
            nome = (_cell_to_text(custom_module.get("nome")) or modelo or f"Modulo {module_type}")[:120]
            marca = (_cell_to_text(custom_module.get("marca")) or "IMPORTADO")[:80]
            module_model = module_qs.model.objects.filter(
                cliente=job.cliente,
                modelo=modelo,
                tipo_base=tipo_base,
                quantidade_canais=capacity,
            ).first()
            if not module_model:
                module_model = module_qs.model.objects.create(
                    cliente=job.cliente,
                    nome=nome,
                    modelo=modelo,
                    marca=marca,
                    quantidade_canais=capacity,
                    tipo_base=tipo_base,
                    is_default=False,
                )
            module_map[str(module_model.id)] = module_model

        rack_module = slot.modulo
        if rack_module:
            if rack_module.modulo_modelo_id != module_model.id:
                raise IOImportError(
                    f"Conflito no slot {slot_index}: ja existe um modulo diferente no rack destino."
                )
        else:
            rack_module = rack_module_model.objects.create(rack=rack, modulo_modelo=module_model)
            channel_model.objects.bulk_create(
                [
                    channel_model(modulo=rack_module, indice=index, descricao="", tipo=module_model.tipo_base)
                    for index in range(1, module_model.quantidade_canais + 1)
                ]
            )
            slot.modulo = rack_module
            slot.save(update_fields=["modulo"])
            slot_map[slot_index] = slot

        channel_lookup = {channel.indice: channel for channel in rack_module.canais.select_related("tipo").all()}
        to_update = []
        for channel_payload in module_payload.get("channels") or []:
            channel = channel_lookup.get(int(channel_payload["index"]))
            if not channel:
                continue
            channel.tag = normalize_tag(channel_payload.get("tag"))
            channel.descricao = (channel_payload.get("description") or "").strip()
            channel_type = normalize_type(channel_payload.get("type"))
            if channel_type:
                matching_type = next(
                    (module.tipo_base for module in module_qs if module.tipo_base.nome == channel_type),
                    None,
                )
                if matching_type:
                    channel.tipo = matching_type
            to_update.append(channel)
        if to_update:
            channel_model.objects.bulk_update(to_update, ["tag", "descricao", "tipo"])
    return rack


def apply_import_job(
    job,
    user,
    rack_model,
    rack_slot_model,
    rack_module_model,
    channel_model,
    module_qs,
    plant_model,
    selected_rack_keys=None,
):
    from django.db import transaction

    original_job = job
    with transaction.atomic():
        job = type(job).objects.select_for_update().get(pk=original_job.pk)
        proposal = job.proposal_payload or {}
        proposal_racks = proposal.get("racks") or []
        if not proposal_racks and proposal.get("modules"):
            proposal_racks = [
                {
                    "name": (proposal.get("rack") or {}).get("name") or job.requested_rack_name or "Rack importado",
                    "slots_total": (proposal.get("rack") or {}).get("slots_total") or len(proposal.get("modules") or []),
                    "modules": proposal.get("modules") or [],
                    "summary": proposal.get("summary") or {},
                }
            ]

        conflicts = proposal.get("conflicts") or []
        if conflicts:
            raise IOImportError("A proposta possui conflitos e nao pode ser aplicada.")
        if not proposal_racks:
            raise IOImportError("A proposta nao possui racks para aplicar.")
        if job.target_rack and len(proposal_racks) > 1:
            raise IOImportError("Esta importacao gerou multiplos racks e nao pode ser aplicada em um rack destino unico.")
        all_source_racks = list(proposal_racks)

        module_map = {str(module.id): module for module in module_qs}
        existing_apply_log = dict(job.apply_log or {})
        applied_rack_key_map = dict(existing_apply_log.get("applied_rack_keys") or {})
        if selected_rack_keys:
            selected_set = {str(item) for item in selected_rack_keys if str(item).strip()}
            proposal_racks = [
                rack_payload
                for rack_payload in proposal_racks
                if str(rack_payload.get("rack_key") or "").strip() in selected_set
            ]
            if not proposal_racks:
                raise IOImportError("Nenhum rack selecionado para aplicacao.")
        proposal_racks = [
            rack_payload
            for rack_payload in proposal_racks
            if str(rack_payload.get("rack_key") or "").strip() not in applied_rack_key_map
        ]
        if not proposal_racks:
            selected_ids = [
                rack_id
                for rack_key, rack_id in applied_rack_key_map.items()
                if not selected_rack_keys or rack_key in {str(item) for item in selected_rack_keys}
            ]
            return list(rack_model.objects.filter(id__in=selected_ids))

        applied_racks = []
        for index, rack_payload in enumerate(proposal_racks):
            rack = _apply_single_rack_payload(
                job=job,
                rack_payload=rack_payload,
                target_rack=job.target_rack if index == 0 else None,
                module_map=module_map,
                module_qs=module_qs,
                rack_model=rack_model,
                rack_slot_model=rack_slot_model,
                rack_module_model=rack_module_model,
                channel_model=channel_model,
                plant_model=plant_model,
            )
            applied_racks.append(rack)
            rack_key = str(rack_payload.get("rack_key") or "").strip()
            if rack_key:
                applied_rack_key_map[rack_key] = rack.id

        all_proposal_keys = {
            str(rack_payload.get("rack_key") or "").strip()
            for rack_payload in all_source_racks
            if str(rack_payload.get("rack_key") or "").strip()
        }
        all_applied_ids = list(dict.fromkeys(list(applied_rack_key_map.values())))
        if not job.first_applied_at:
            try:
                register_successful_import_usage(user, "IO")
            except ValueError as exc:
                raise IOImportError(str(exc)) from exc
            job.first_applied_at = timezone.now()
        job.applied_rack = rack_model.objects.filter(id__in=all_applied_ids).order_by("id").first()
        applied_all_racks = False
        if all_proposal_keys:
            applied_all_racks = all_proposal_keys.issubset(applied_rack_key_map.keys())
        else:
            applied_all_racks = len(all_applied_ids) >= len(all_source_racks)
        job.status = job.Status.APPLIED if applied_all_racks else job.Status.REVIEW
        job.apply_log = {
            "applied_by": getattr(user, "username", "") or "",
            "racks_applied": len(all_applied_ids),
            "modules_applied": sum(
                len(rack_payload.get("modules") or [])
                for rack_payload in all_source_racks
                if not all_proposal_keys or str(rack_payload.get("rack_key") or "").strip() in applied_rack_key_map
            ),
            "applied_rack_ids": all_applied_ids,
            "applied_rack_keys": applied_rack_key_map,
        }
        job.save(update_fields=["applied_rack", "status", "apply_log", "first_applied_at", "updated_at"])
    for field_name in ("applied_rack", "status", "apply_log", "first_applied_at", "updated_at"):
        setattr(original_job, field_name, getattr(job, field_name))
    return applied_racks
