from __future__ import annotations

import csv
import json
import random
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SEED = 20260412
DOC_DATE = date(2026, 4, 12)
OUTPUT_DIR = Path(__file__).resolve().parent
EXPECTED_COUNTS = {"DI": 40, "DO": 32, "AI": 28, "AO": 12, "SPARE": 8}
EXPECTED_TOTAL = sum(EXPECTED_COUNTS.values())
AUTHOR = "Codex IO Dataset Generator"


CANONICAL_COLUMNS = [
    "point_uid",
    "area",
    "panel",
    "plc",
    "remote_node",
    "rack",
    "slot",
    "module_model",
    "module_type",
    "channel",
    "io_type",
    "tag",
    "description",
    "signal_type",
    "eng_unit",
    "range_min",
    "range_max",
    "power_or_signal",
    "device",
    "p_and_id",
    "junction_box",
    "terminal_block",
    "terminal",
    "cable",
    "location_string",
    "comment",
]


SUMMARY_BY_FILE = {
    "01_io_flat_ptbr.xlsx": "Planilha limpa em PT-BR com colunas explícitas de hardware, sinal e processo.",
    "02_io_flat_english.xlsx": "Lista linear em inglês com ordem de colunas diferente e abreviações mistas.",
    "03_io_compact_location.xlsx": "Formato compacto com localização combinada e guias auxiliares de índice e racks.",
    "04_io_grouped_by_module.xlsx": "Layout agrupado por módulo com múltiplos racks na mesma guia de dados e abas auxiliares.",
    "05_io_multisheet_by_panel.xlsx": "Workbook com capa, legenda e uma aba de dados para cada painel.",
    "06_io_two_header_levels.xlsx": "Planilha com título, revisão, cabeçalhos em dois níveis e guias auxiliares de revisão/legenda.",
    "07_io_noisy_export.xlsx": "Exportação bagunçada com linhas de observação, colunas extras, espaços supérfluos e guia auxiliar.",
    "08_io_decimal_comma_ptbr.xlsx": "Formato PT-BR com vírgula decimal e ranges tratados como texto regional.",
    "09_io_minimal_headers.xlsx": "Exportação de cabeçalhos curtos e pouco amigáveis, porém consistente.",
    "10_io_mixed_real_world.xlsx": "Planilha de campo com bloco de revisão e dados distribuídos em guias não óbvias, cada uma com múltiplos racks.",
}


MODULE_LIBRARY = {
    "DI16": {
        "module_model": "ET200SP DI 16x24VDC HF",
        "module_type": "DI16",
        "capacity": 16,
    },
    "DO16": {
        "module_model": "ET200SP DQ 16x24VDC/0.5A HF",
        "module_type": "DO16",
        "capacity": 16,
    },
    "AI8": {
        "module_model": "ET200SP AI 8xI 4-20mA ST",
        "module_type": "AI8",
        "capacity": 8,
    },
    "AO8": {
        "module_model": "ET200SP AQ 8xI 4-20mA ST",
        "module_type": "AO8",
        "capacity": 8,
    },
}


@dataclass(frozen=True)
class AreaContext:
    key: str
    area_pt: str
    area_en: str
    code: str
    panel: str
    plc: str
    remote_node: str
    pid: str
    default_location: str


AREA_CONTEXTS = {
    "rcv": AreaContext(
        key="rcv",
        area_pt="Recebimento",
        area_en="Receiving",
        code="RCV",
        panel="PNL-RCV-01",
        plc="PLC-RCV-01",
        remote_node="ET200-RCV-01",
        pid="PID-RCV-100",
        default_location="Recebimento norte / moega rodoviaria",
    ),
    "mog": AreaContext(
        key="mog",
        area_pt="Moega",
        area_en="Hopper",
        code="MOG",
        panel="PNL-MOG-01",
        plc="PLC-MOG-01",
        remote_node="ET200-MOG-01",
        pid="PID-MOG-200",
        default_location="Predio da moega / nivel operacional",
    ),
    "trn": AreaContext(
        key="trn",
        area_pt="Transporte",
        area_en="Conveying",
        code="TRN",
        panel="MCC-TRN-01",
        plc="PLC-TRN-01",
        remote_node="ET200-TRN-01",
        pid="PID-TRN-300",
        default_location="Galeria de transporte / trecho principal",
    ),
    "prc": AreaContext(
        key="prc",
        area_pt="Processo",
        area_en="Process",
        code="PRC",
        panel="PNL-PRC-01",
        plc="PLC-PRC-01",
        remote_node="ET200-PRC-01",
        pid="PID-PRC-400",
        default_location="Sala de processo / plataforma principal",
    ),
    "bag": AreaContext(
        key="bag",
        area_pt="Ensacadeira",
        area_en="Bagging",
        code="BAG",
        panel="PNL-BAG-01",
        plc="PLC-BAG-01",
        remote_node="ET200-BAG-01",
        pid="PID-BAG-500",
        default_location="Linha de ensaque / corredor de maquinas",
    ),
    "utl": AreaContext(
        key="utl",
        area_pt="Utilidades",
        area_en="Utilities",
        code="UTL",
        panel="PNL-UTL-01",
        plc="PLC-UTL-01",
        remote_node="ET200-UTL-01",
        pid="PID-UTL-600",
        default_location="Casa de utilidades / skid de servicos",
    ),
}


def default_family(io_type: str) -> str:
    return {
        "DI": "DI16",
        "DO": "DO16",
        "AI": "AI8",
        "AO": "AO8",
        "SPARE": "DI16",
    }[io_type]


def build_canonical_dataset() -> list[dict]:
    rows: list[dict] = []
    area_sequence = Counter()

    def add_point(
        area_key: str,
        io_type: str,
        tag: str,
        description_pt: str,
        description_en: str,
        signal_type: str,
        eng_unit: str = "",
        range_min: float | None = None,
        range_max: float | None = None,
        power_or_signal: str = "",
        device: str | None = None,
        p_and_id: str | None = None,
        junction_box: str | None = None,
        terminal_block: str | None = None,
        terminal: str | None = None,
        cable: str | None = None,
        field_location: str | None = None,
        comment_pt: str = "",
        comment_en: str = "",
        module_family: str | None = None,
    ) -> None:
        ctx = AREA_CONTEXTS[area_key]
        area_sequence[area_key] += 1
        seq = area_sequence[area_key]
        rows.append(
            {
                "point_uid": f"IO-{len(rows) + 1:04d}",
                "area": ctx.area_pt,
                "panel": ctx.panel,
                "plc": ctx.plc,
                "remote_node": ctx.remote_node,
                "rack": None,
                "slot": None,
                "module_model": "",
                "module_type": "",
                "channel": None,
                "io_type": io_type,
                "tag": tag,
                "description": description_pt,
                "signal_type": signal_type,
                "eng_unit": eng_unit,
                "range_min": range_min,
                "range_max": range_max,
                "power_or_signal": power_or_signal,
                "device": device or tag,
                "p_and_id": p_and_id or ctx.pid,
                "junction_box": junction_box or f"JB-{ctx.code}-{((seq - 1) % 4) + 1:02d}",
                "terminal_block": terminal_block or f"X{((seq - 1) % 3) + 1}",
                "terminal": terminal or f"{((seq - 1) % 16) + 1:02d}",
                "cable": cable or f"CBL-{ctx.code}-{seq:03d}",
                "location_string": field_location or ctx.default_location,
                "comment": comment_pt,
                "_area_key": area_key,
                "_area_en": ctx.area_en,
                "_description_en": description_en,
                "_comment_en": comment_en,
                "_module_family": module_family or default_family(io_type),
            }
        )

    def add_ai(
        area_key: str,
        tag: str,
        description_pt: str,
        description_en: str,
        eng_unit: str,
        range_min: float,
        range_max: float,
        device: str,
        field_location: str,
        signal_type: str = "4-20 mA",
        comment_pt: str = "",
        comment_en: str = "",
        module_family: str = "AI8",
    ) -> None:
        add_point(
            area_key=area_key,
            io_type="AI",
            tag=tag,
            description_pt=description_pt,
            description_en=description_en,
            signal_type=signal_type,
            eng_unit=eng_unit,
            range_min=range_min,
            range_max=range_max,
            power_or_signal=signal_type,
            device=device,
            field_location=field_location,
            comment_pt=comment_pt,
            comment_en=comment_en,
            module_family=module_family,
        )

    def add_ao(
        area_key: str,
        tag: str,
        description_pt: str,
        description_en: str,
        eng_unit: str,
        range_min: float,
        range_max: float,
        device: str,
        field_location: str,
        signal_type: str = "4-20 mA",
        comment_pt: str = "",
        comment_en: str = "",
        module_family: str = "AO8",
    ) -> None:
        add_point(
            area_key=area_key,
            io_type="AO",
            tag=tag,
            description_pt=description_pt,
            description_en=description_en,
            signal_type=signal_type,
            eng_unit=eng_unit,
            range_min=range_min,
            range_max=range_max,
            power_or_signal=signal_type,
            device=device,
            field_location=field_location,
            comment_pt=comment_pt,
            comment_en=comment_en,
            module_family=module_family,
        )

    def add_di(
        area_key: str,
        tag: str,
        description_pt: str,
        description_en: str,
        device: str,
        field_location: str,
        signal_type: str = "Dry Contact",
        power_or_signal: str = "24 Vdc / contato seco",
        comment_pt: str = "",
        comment_en: str = "",
        module_family: str = "DI16",
    ) -> None:
        add_point(
            area_key=area_key,
            io_type="DI",
            tag=tag,
            description_pt=description_pt,
            description_en=description_en,
            signal_type=signal_type,
            power_or_signal=power_or_signal,
            device=device,
            field_location=field_location,
            comment_pt=comment_pt,
            comment_en=comment_en,
            module_family=module_family,
        )

    def add_do(
        area_key: str,
        tag: str,
        description_pt: str,
        description_en: str,
        device: str,
        field_location: str,
        signal_type: str = "24 Vdc Command",
        power_or_signal: str = "24 Vdc",
        comment_pt: str = "",
        comment_en: str = "",
        module_family: str = "DO16",
    ) -> None:
        add_point(
            area_key=area_key,
            io_type="DO",
            tag=tag,
            description_pt=description_pt,
            description_en=description_en,
            signal_type=signal_type,
            power_or_signal=power_or_signal,
            device=device,
            field_location=field_location,
            comment_pt=comment_pt,
            comment_en=comment_en,
            module_family=module_family,
        )

    def add_spare(
        area_key: str,
        tag: str,
        description_pt: str,
        description_en: str,
        module_family: str,
        comment_pt: str = "",
        comment_en: str = "",
    ) -> None:
        add_point(
            area_key=area_key,
            io_type="SPARE",
            tag=tag,
            description_pt=description_pt,
            description_en=description_en,
            signal_type=f"Reserved {module_family}",
            power_or_signal=f"Reserva em {module_family}",
            device=tag,
            field_location=AREA_CONTEXTS[area_key].default_location,
            comment_pt=comment_pt,
            comment_en=comment_en,
            module_family=module_family,
        )

    def add_motor_vfd(
        area_key: str,
        base_tag: str,
        service_pt: str,
        service_en: str,
        field_location: str,
        speed_unit: str = "%",
        speed_max: float = 100.0,
        speed_feedback_signal: str = "4-20 mA",
    ) -> None:
        add_di(
            area_key,
            f"{base_tag}_RUN",
            f"Retorno em operacao do {service_pt}",
            f"{service_en} running feedback",
            device=base_tag,
            field_location=field_location,
            signal_type="Aux Contact",
            power_or_signal="24 Vdc / contato auxiliar",
        )
        add_di(
            area_key,
            f"{base_tag}_FLT",
            f"Falha do {service_pt}",
            f"{service_en} fault feedback",
            device=base_tag,
            field_location=field_location,
            signal_type="Aux Contact",
            power_or_signal="24 Vdc / contato auxiliar",
            comment_pt="Sinal proveniente do inversor.",
            comment_en="Signal sourced from the drive.",
        )
        add_do(
            area_key,
            f"{base_tag}_START",
            f"Comando de partida do {service_pt}",
            f"{service_en} start command",
            device=base_tag,
            field_location=field_location,
            signal_type="24 Vdc Run Command",
            comment_pt="Comando habilitado por intertravamento de processo.",
            comment_en="Command released by process interlock.",
        )
        add_ai(
            area_key,
            f"{base_tag}_SPD_FBK",
            f"Referencia real de velocidade do {service_pt}",
            f"{service_en} actual speed feedback",
            eng_unit=speed_unit,
            range_min=0.0,
            range_max=speed_max,
            device=base_tag,
            field_location=field_location,
            signal_type=speed_feedback_signal,
        )
        add_ao(
            area_key,
            f"{base_tag}_SPD_REF",
            f"Setpoint de velocidade do {service_pt}",
            f"{service_en} speed setpoint",
            eng_unit=speed_unit,
            range_min=0.0,
            range_max=speed_max,
            device=base_tag,
            field_location=field_location,
        )

    def add_motor_dol(
        area_key: str,
        base_tag: str,
        service_pt: str,
        service_en: str,
        field_location: str,
        include_fault: bool = False,
    ) -> None:
        add_di(
            area_key,
            f"{base_tag}_RUN",
            f"Retorno em operacao do {service_pt}",
            f"{service_en} running feedback",
            device=base_tag,
            field_location=field_location,
            signal_type="Aux Contact",
            power_or_signal="24 Vdc / contato auxiliar",
        )
        if include_fault:
            add_di(
                area_key,
                f"{base_tag}_FLT",
                f"Falha do {service_pt}",
                f"{service_en} fault feedback",
                device=base_tag,
                field_location=field_location,
                signal_type="Aux Contact",
                power_or_signal="24 Vdc / contato auxiliar",
            )
        add_do(
            area_key,
            f"{base_tag}_START",
            f"Comando de partida do {service_pt}",
            f"{service_en} start command",
            device=base_tag,
            field_location=field_location,
            signal_type="24 Vdc Run Command",
        )

    def add_onoff_valve(
        area_key: str,
        base_tag: str,
        service_pt: str,
        service_en: str,
        field_location: str,
        with_feedback: bool = True,
    ) -> None:
        add_do(
            area_key,
            f"{base_tag}_CMD",
            f"Comando da {service_pt}",
            f"{service_en} command",
            device=base_tag,
            field_location=field_location,
            signal_type="24 Vdc Solenoid",
            power_or_signal="24 Vdc solenóide",
        )
        if with_feedback:
            add_di(
                area_key,
                f"{base_tag}_FB",
                f"Retorno de posicao da {service_pt}",
                f"{service_en} position feedback",
                device=base_tag,
                field_location=field_location,
                signal_type="Dry Contact",
                power_or_signal="24 Vdc / fim de curso",
            )

    # Recebimento
    add_onoff_valve("rcv", "XV_101", "comporta de descarga do fosso", "truck pit discharge gate", "Fosso rodoviario / extremidade norte")
    add_di("rcv", "LSH_101", "Chave de nivel alto da moega rodoviaria", "Truck hopper high level switch", "LSH_101", "Moega rodoviaria", comment_pt="Ajuste de campo em 85%.", comment_en="Field set point adjusted to 85%.")
    add_di("rcv", "LSL_101", "Chave de nivel baixo da moega rodoviaria", "Truck hopper low level switch", "LSL_101", "Moega rodoviaria")
    add_ai("rcv", "LIT_101", "Transmissor de nivel da moega rodoviaria", "Truck hopper level transmitter", "%", 0.0, 100.0, "LIT_101", "Moega rodoviaria", comment_pt="Escala ajustada para operacao entre 10% e 90%.", comment_en="Scale adjusted for normal operation between 10% and 90%.")
    add_motor_vfd("rcv", "CV_111", "rosca de descarga do recebimento", "receiving discharge screw conveyor", "Base da moega rodoviaria")
    add_motor_dol("rcv", "MTR_112", "exaustor do filtro do recebimento", "receiving dust collector fan", "Cobertura do filtro de mangas")
    add_ai("rcv", "PIT_113", "Transmissor de pressao da linha de fluidizacao", "Fluidization line pressure transmitter", "bar", 0.0, 6.0, "PIT_113", "Linha de ar de fluidizacao")
    add_ai("rcv", "FIT_114", "Transmissor de vazao de descarga do recebimento", "Receiving discharge flow transmitter", "t/h", 0.0, 60.0, "FIT_114", "Calha de descarga do recebimento")
    add_onoff_valve("rcv", "YV_115", "valvula desviadora de amostragem", "sampling diverter valve", "Bica de amostragem do recebimento")
    add_ao("rcv", "FCV_116_OUT", "Setpoint da valvula de dosagem de agua", "Water dosing valve setpoint", "%", 0.0, 100.0, "FCV_116", "Linha de umectacao do recebimento")
    add_do("rcv", "PMP_117_START", "Comando de partida da bomba de lavagem", "Wash water pump start command", "PMP_117", "Skid de lavagem do piso", signal_type="24 Vdc Run Command")
    add_spare("rcv", "SPARE_DI_RCV_01", "Reserva de entrada digital no painel de recebimento", "Reserved digital input on receiving panel", "DI16", comment_pt="Reservado para futura chave de nivel.", comment_en="Reserved for future level switch.")

    # Moega
    add_di("mog", "LSH_201", "Chave de nivel alto do silo pulmão da moega", "Hopper buffer bin high level switch", "LSH_201", "Silo pulmão da moega")
    add_di("mog", "LSL_201", "Chave de nivel baixo do silo pulmão da moega", "Hopper buffer bin low level switch", "LSL_201", "Silo pulmão da moega")
    add_ai("mog", "LIT_202", "Transmissor de nivel do silo pulmão da moega", "Hopper buffer bin level transmitter", "%", 0.0, 100.0, "LIT_202", "Silo pulmão da moega")
    add_ai("mog", "FIT_203", "Transmissor de vazao da calha vibratoria da moega", "Hopper feeder flow transmitter", "t/h", 0.0, 80.0, "FIT_203", "Calha vibratoria da moega")
    add_motor_vfd("mog", "CV_211", "transportador principal da moega", "hopper main conveyor", "Galeria de transferencia da moega", speed_unit="Hz", speed_max=60.0)
    add_motor_dol("mog", "CV_212", "transportador auxiliar da moega", "hopper auxiliary conveyor", "Galeria de transferencia da moega")
    add_do("mog", "YV_213_CMD", "Comando da comporta desviadora da moega", "Hopper diverter gate command", "YV_213", "Bifurcacao de descarga da moega", signal_type="24 Vdc Solenoid", power_or_signal="24 Vdc solenóide")
    add_di("mog", "ZS_213_OPEN", "Fim de curso aberto da comporta desviadora da moega", "Hopper diverter gate open limit switch", "YV_213", "Bifurcacao de descarga da moega", power_or_signal="24 Vdc / fim de curso")
    add_onoff_valve("mog", "XV_214", "valvula de alivio do silo de transicao", "transition bin vent valve", "Silo de transicao da moega")
    add_motor_dol("mog", "MTR_215", "vibrador do silo de transicao", "transition bin vibrator", "Cone inferior do silo de transicao")
    add_ai("mog", "PIT_216", "Transmissor de pressao da linha de despoeiramento", "Dust collection line pressure transmitter", "mbar", -50.0, 50.0, "PIT_216", "Coletor de despoeiramento da moega")
    add_ai("mog", "TIT_217", "Transmissor de temperatura do mancal do transportador principal", "Main conveyor bearing temperature transmitter", "°C", 0.0, 120.0, "TIT_217", "Mancal motriz do CV_211", signal_type="4-20 mA / temp")
    add_do("mog", "ALM_220_HORN", "Comando da buzina de alarme da moega", "Hopper alarm horn command", "ALM_220", "Passarela da moega", signal_type="24 Vdc Alarm")
    add_ao("mog", "FCV_219_OUT", "Setpoint da valvula de recirculacao de po", "Dust recirculation control valve setpoint", "%", 0.0, 100.0, "FCV_219", "Linha de recirculacao de po")
    add_spare("mog", "SPARE_AI_MOG_01", "Reserva analogica no rack de moega", "Reserved analog input on hopper rack", "AI8", comment_pt="Previsto para transmissor de vibracao.", comment_en="Provisioned for future vibration transmitter.")

    # Transporte
    add_motor_vfd("trn", "CV_301", "transportador de correia principal", "main belt conveyor", "Galeria principal de transporte", speed_unit="Hz", speed_max=60.0)
    add_motor_dol("trn", "CV_302", "transportador de correia secundario", "secondary belt conveyor", "Galeria secundaria de transporte", include_fault=True)
    add_motor_dol("trn", "BE_303", "elevador de canecas", "bucket elevator", "Torre do elevador", include_fault=True)
    add_onoff_valve("trn", "XV_304", "valvula guilhotina do chute A", "chute A knife gate valve", "Chute de transferencia A")
    add_onoff_valve("trn", "YV_305", "valvula desviadora do chute B", "chute B diverter valve", "Chute de transferencia B")
    add_ai("trn", "FIT_307", "Transmissor de vazao da linha de transporte principal", "Main conveying line flow transmitter", "t/h", 0.0, 120.0, "FIT_307", "Chute de transferencia principal")
    add_ai("trn", "WIT_308", "Transmissor de carga da balanca integradora", "Belt scale load transmitter", "t/h", 0.0, 150.0, "WIT_308", "Balanca integradora do CV_301")
    add_ai("trn", "PIT_309", "Transmissor de pressao da linha de transporte pneumático", "Pneumatic conveying line pressure transmitter", "bar", 0.0, 2.5, "PIT_309", "Coletor da linha pneumática")
    add_ai("trn", "TIT_310", "Transmissor de temperatura do mancal do elevador", "Bucket elevator bearing temperature transmitter", "°C", 0.0, 120.0, "TIT_310", "Mancal superior do elevador", signal_type="4-20 mA / temp")
    add_do("trn", "ALM_314_BEACON", "Comando do sinaleiro do transporte", "Conveying beacon command", "ALM_314", "Passarela principal de transporte", signal_type="24 Vdc Alarm")
    add_do("trn", "SV_315_CMD", "Comando da valvula de spray de mitigacao de poeira", "Dust suppression spray valve command", "SV_315", "Chute de transferencia principal", signal_type="24 Vdc Solenoid", power_or_signal="24 Vdc solenóide")
    add_ao("trn", "FCV_312_OUT", "Setpoint da valvula de controle de vazao do bypass", "Bypass flow control valve setpoint", "%", 0.0, 100.0, "FCV_312", "Linha de bypass do transporte")
    add_ao("trn", "DMPR_313_POS_REF", "Setpoint do damper da exaustao do transporte", "Conveying exhaust damper position setpoint", "%", 0.0, 100.0, "DMPR_313", "Duto de exaustao do transporte")
    add_spare("trn", "SPARE_DO_TRN_01", "Reserva de saida digital do transporte", "Reserved digital output on conveying panel", "DO16", comment_pt="Disponivel para sirene adicional.", comment_en="Available for future extra siren.")
    add_spare("trn", "SPARE_AI_TRN_02", "Reserva analogica do transporte", "Reserved analog input on conveying panel", "AI8", comment_pt="Canal reservado para monitor de vibracao.", comment_en="Channel reserved for vibration monitor.")

    # Processo
    add_di("prc", "LSH_401", "Chave de nivel alto do tanque de mistura", "Mix tank high level switch", "LSH_401", "Tanque de mistura TQ_401")
    add_di("prc", "LSL_401", "Chave de nivel baixo do tanque de mistura", "Mix tank low level switch", "LSL_401", "Tanque de mistura TQ_401")
    add_ai("prc", "LIT_402", "Transmissor de nivel do tanque de mistura", "Mix tank level transmitter", "%", 0.0, 100.0, "LIT_402", "Tanque de mistura TQ_401")
    add_ai("prc", "PIT_403", "Transmissor de pressao do manifold de processo", "Process manifold pressure transmitter", "bar", 0.0, 10.0, "PIT_403", "Manifold de processo")
    add_ai("prc", "TIT_404", "Transmissor de temperatura do tanque de mistura", "Mix tank temperature transmitter", "°C", 0.0, 120.0, "TIT_404", "Tanque de mistura TQ_401", signal_type="4-20 mA / temp")
    add_ai("prc", "PHIT_405", "Transmissor de pH da linha de preparo", "Preparation line pH transmitter", "pH", 0.0, 14.0, "PHIT_405", "Linha de preparo do processo")
    add_ai("prc", "FIT_406", "Transmissor de vazao da linha de dosagem", "Dosing line flow transmitter", "m3/h", 0.0, 25.0, "FIT_406", "Linha de dosagem do processo")
    add_ai("prc", "AIT_407", "Transmissor de condutividade da linha de lavagem", "Wash line conductivity transmitter", "mS/cm", 0.0, 20.0, "AIT_407", "Linha CIP do processo")
    add_motor_vfd("prc", "MTR_411", "agitador principal do tanque de mistura", "mix tank main agitator", "Topo do tanque de mistura")
    add_motor_dol("prc", "PMP_412", "bomba de recirculacao de processo", "process recirculation pump", "Skid de recirculacao do processo")
    add_onoff_valve("prc", "XV_413", "valvula de entrada de agua de processo", "process water inlet valve", "Linha de agua de processo")
    add_onoff_valve("prc", "XV_414", "valvula de descarga do tanque de mistura", "mix tank outlet valve", "Bocal inferior do tanque de mistura")
    add_di("prc", "UV_415_HEALTHY", "Saude do modulo UV da linha de sanitizacao", "Sanitizing line UV module healthy status", "UV_415", "Skid de sanitizacao", signal_type="Dry Contact", power_or_signal="24 Vdc / contato seco")
    add_do("prc", "UV_415_ENABLE", "Comando de habilitacao do modulo UV", "UV module enable command", "UV_415", "Skid de sanitizacao", signal_type="24 Vdc Enable")
    add_ao("prc", "FCV_416_OUT", "Setpoint da valvula de controle de vazao de processo", "Process flow control valve setpoint", "%", 0.0, 100.0, "FCV_416", "Linha principal de processo")
    add_ao("prc", "PCV_417_OUT", "Setpoint da valvula de controle de pressao de processo", "Process pressure control valve setpoint", "%", 0.0, 100.0, "PCV_417", "Linha principal de processo")
    add_do("prc", "ALM_418_HORN", "Comando da buzina de alarme do processo", "Process area alarm horn command", "ALM_418", "Sala de processo", signal_type="24 Vdc Alarm")
    add_spare("prc", "SPARE_DI_PRC_01", "Reserva de entrada digital do processo", "Reserved digital input on process panel", "DI16", comment_pt="Reservado para pressostato futuro.", comment_en="Reserved for future pressure switch.")
    add_spare("prc", "SPARE_AO_PRC_02", "Reserva analogica de saida do processo", "Reserved analog output on process panel", "AO8", comment_pt="Previsto para malha de dosagem adicional.", comment_en="Provisioned for future extra dosing loop.")

    # Ensacadeira
    add_motor_dol("bag", "CV_501", "transportador de alimentacao da ensacadeira", "bagger feed conveyor", "Entrada da ensacadeira")
    add_motor_dol("bag", "MTR_502", "motor da maquina de costura", "bag sewing machine motor", "Cabecote da costura")
    add_ai("bag", "WIT_503", "Transmissor de peso da balanca ensacadeira", "Bagging scale weight transmitter", "kg", 0.0, 50.0, "WIT_503", "Balanca da ensacadeira")
    add_di("bag", "LSH_504", "Chave de nivel alto do funil da ensacadeira", "Bagger hopper high level switch", "LSH_504", "Funil da ensacadeira")
    add_di("bag", "LSL_504", "Chave de nivel baixo do funil da ensacadeira", "Bagger hopper low level switch", "LSL_504", "Funil da ensacadeira")
    add_ai("bag", "LIT_505", "Transmissor de nivel do funil da ensacadeira", "Bagger hopper level transmitter", "%", 0.0, 100.0, "LIT_505", "Funil da ensacadeira")
    add_motor_dol("bag", "MTR_506", "exaustor do filtro da ensacadeira", "bagger dust collector fan", "Filtro de mangas da ensacadeira")
    add_do("bag", "XV_507_CMD", "Comando da valvula de corte da ensacadeira", "Bagger cutoff valve command", "XV_507", "Linha de enchimento", signal_type="24 Vdc Solenoid", power_or_signal="24 Vdc solenóide")
    add_do("bag", "ALM_508_BEACON", "Comando do sinaleiro da ensacadeira", "Bagging line beacon command", "ALM_508", "Coluna de sinaleiro da ensacadeira", signal_type="24 Vdc Alarm")
    add_ai("bag", "PIT_509", "Transmissor de pressao do ar de servico da ensacadeira", "Bagging service air pressure transmitter", "bar", 0.0, 8.0, "PIT_509", "Linha de ar comprimido da ensacadeira")
    add_ai("bag", "TIT_510", "Transmissor de temperatura da seladora", "Sealer temperature transmitter", "°C", 0.0, 250.0, "TIT_510", "Cabecote da seladora", signal_type="4-20 mA / temp")
    add_ao("bag", "CV_511_SPD_REF", "Setpoint de velocidade do dosador de rosca da ensacadeira", "Bagger screw feeder speed setpoint", "%", 0.0, 100.0, "CV_511", "Dosador de rosca da ensacadeira")
    add_spare("bag", "SPARE_DO_BAG_01", "Reserva de saida digital da ensacadeira", "Reserved digital output on bagging panel", "DO16", comment_pt="Canal livre para luz de stack adicional.", comment_en="Free channel for future stack light.")

    # Utilidades
    add_di("utl", "COMP_601_RUN", "Retorno em operacao do compressor principal", "Main compressor running feedback", "COMP_601", "Casa de utilidades / skid do compressor", signal_type="Aux Contact", power_or_signal="24 Vdc / contato auxiliar")
    add_di("utl", "COMP_601_FLT", "Falha do compressor principal", "Main compressor fault feedback", "COMP_601", "Casa de utilidades / skid do compressor", signal_type="Aux Contact", power_or_signal="24 Vdc / contato auxiliar")
    add_do("utl", "COMP_601_START", "Comando de partida do compressor principal", "Main compressor start command", "COMP_601", "Casa de utilidades / skid do compressor", signal_type="24 Vdc Run Command")
    add_ai("utl", "PIT_602", "Transmissor de pressao do header de ar comprimido", "Compressed air header pressure transmitter", "bar", 0.0, 10.0, "PIT_602", "Header principal de ar comprimido")
    add_ai("utl", "FIT_603", "Transmissor de vazao de agua industrial", "Industrial water flow transmitter", "m3/h", 0.0, 40.0, "FIT_603", "Linha de agua industrial")
    add_ai("utl", "TIT_604", "Transmissor de temperatura da linha de agua gelada", "Chilled water line temperature transmitter", "°C", 0.0, 30.0, "TIT_604", "Linha de agua gelada", signal_type="4-20 mA / temp")
    add_onoff_valve("utl", "XV_605", "valvula de bloqueio do header de ar", "air header isolation valve", "Header de ar comprimido")
    add_motor_dol("utl", "PMP_606", "bomba de recirculacao da torre de resfriamento", "cooling tower recirculation pump", "Skid da torre de resfriamento")
    add_ao("utl", "PCV_607_OUT", "Setpoint da valvula de controle de pressao do ar", "Air pressure control valve setpoint", "%", 0.0, 100.0, "PCV_607", "Header de ar comprimido")
    add_spare("utl", "SPARE_AI_UTL_01", "Reserva analogica de utilidades", "Reserved analog input on utilities panel", "AI8", comment_pt="Reservado para transmissor de ponto de orvalho.", comment_en="Reserved for future dew point transmitter.")

    allocate_hardware(rows)
    validate_canonical_counts(rows)
    return rows


def allocate_hardware(rows: list[dict]) -> None:
    module_pool = {
        "rcv": [("DI16", 1, 1), ("DO16", 1, 2), ("AI8", 2, 1), ("AO8", 3, 1)],
        "mog": [("DI16", 1, 1), ("DO16", 1, 2), ("AI8", 2, 1), ("AO8", 3, 1)],
        "trn": [("DI16", 1, 1), ("DO16", 1, 2), ("AI8", 2, 1), ("AO8", 3, 1)],
        "prc": [("DI16", 1, 1), ("DO16", 1, 2), ("AI8", 2, 1), ("AO8", 3, 1)],
        "bag": [("DI16", 1, 1), ("DO16", 1, 2), ("AI8", 2, 1), ("AO8", 3, 1)],
        "utl": [("DI16", 1, 1), ("DO16", 1, 2), ("AI8", 2, 1), ("AO8", 3, 1)],
    }

    grouped: dict[tuple[str, str], list[dict]] = defaultdict(list)
    for row in rows:
        grouped[(row["_area_key"], row["_module_family"])].append(row)

    for (area_key, family), family_rows in grouped.items():
        family_rows.sort(key=lambda item: item["tag"])
        modules = [entry for entry in module_pool[area_key] if entry[0] == family]
        if not modules:
            raise ValueError(f"No module available for area={area_key} family={family}")
        module_index = 0
        channel_index = 0
        for row in family_rows:
            module_family, rack, slot = modules[module_index]
            capacity = MODULE_LIBRARY[module_family]["capacity"]
            if channel_index >= capacity:
                module_index += 1
                channel_index = 0
                if module_index >= len(modules):
                    raise ValueError(f"Insufficient module capacity for area={area_key} family={family}")
                module_family, rack, slot = modules[module_index]
            channel_index += 1
            row["rack"] = rack
            row["slot"] = slot
            row["channel"] = channel_index
            row["module_model"] = MODULE_LIBRARY[module_family]["module_model"]
            row["module_type"] = MODULE_LIBRARY[module_family]["module_type"]
            row["location_string"] = (
                f"{row['panel']} / Rack {rack} / Slot {slot} / Ch {channel_index:02d} / {row['location_string']}"
            )

    rows.sort(key=lambda item: (item["panel"], item["rack"], item["slot"], item["channel"], item["tag"]))


def validate_canonical_counts(rows: Iterable[dict]) -> None:
    counts = Counter(row["io_type"] for row in rows)
    if counts != Counter(EXPECTED_COUNTS):
        raise ValueError(f"Unexpected I/O distribution: {counts}")
    if sum(counts.values()) != EXPECTED_TOTAL:
        raise ValueError(f"Unexpected total rows: {sum(counts.values())}")


def sanitize_rows_for_export(rows: Iterable[dict]) -> list[dict]:
    return [{key: value for key, value in row.items() if not key.startswith("_")} for row in rows]


def format_number(value: float | int | None, decimal_comma: bool = False) -> str:
    if value in (None, ""):
        return ""
    if float(value).is_integer():
        text = f"{value:.1f}" if decimal_comma else f"{int(value)}"
    else:
        text = f"{value:.1f}"
    return text.replace(".", ",") if decimal_comma else text


def format_range(row: dict, decimal_comma: bool = False) -> str:
    if row["range_min"] in (None, "") and row["range_max"] in (None, ""):
        return ""
    left = format_number(row["range_min"], decimal_comma=decimal_comma)
    right = format_number(row["range_max"], decimal_comma=decimal_comma)
    return f"{left} .. {right}"


def card_position(row: dict, with_leading_zero: bool = True) -> str:
    rack = f"{row['rack']:02d}" if with_leading_zero else str(row["rack"])
    slot = f"{row['slot']:02d}" if with_leading_zero else str(row["slot"])
    channel = f"{row['channel']:02d}" if with_leading_zero else str(row["channel"])
    return f"R{rack}-S{slot}-CH{channel}"


def compact_location_variants(row: dict) -> tuple[str, str]:
    selectors = [
        f"R{row['rack']:02d}-S{row['slot']:02d}-CH{row['channel']:02d}",
        f"Rack {row['rack']} / Slot {row['slot']} / Ch {row['channel']:02d}",
        f"{row['panel']}.R{row['rack']}.S{row['slot']}.C{row['channel']}",
    ]
    position = selectors[(row["channel"] - 1) % len(selectors)]
    return f"{row['panel']}::{row['tag']}", position


def auto_size_worksheet(ws, minimum: int = 10) -> None:
    for column_cells in ws.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[column_letter].width = max(minimum, min(max_length + 2, 42))


def apply_table_header(ws, row_index: int) -> None:
    fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[row_index]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def freeze_header(ws, row_index: int) -> None:
    ws.freeze_panes = f"A{row_index + 1}"


def write_rows(ws, headers: list[str], rows: Iterable[list]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append(row)
    apply_table_header(ws, 1)
    auto_size_worksheet(ws)
    freeze_header(ws, 1)


def append_rack_summary_sheet(wb: Workbook, rows: list[dict], title: str = "Rack Summary") -> None:
    ws = wb.create_sheet(title)
    headers = ["Panel", "Rack", "Points", "DI", "DO", "AI", "AO", "SPARE"]
    grouped = defaultdict(list)
    for row in rows:
        grouped[(row["panel"], row["rack"])].append(row)
    data_rows = []
    for (panel, rack), items in sorted(grouped.items()):
        counts = Counter(item["io_type"] for item in items)
        data_rows.append(
            [
                panel,
                rack,
                len(items),
                counts.get("DI", 0),
                counts.get("DO", 0),
                counts.get("AI", 0),
                counts.get("AO", 0),
                counts.get("SPARE", 0),
            ]
        )
    write_rows(ws, headers, data_rows)


def append_panel_index_sheet(wb: Workbook, rows: list[dict], title: str = "Index") -> None:
    ws = wb.create_sheet(title)
    headers = ["Panel", "Area", "Racks", "Points", "Primary Sheet"]
    grouped = defaultdict(list)
    for row in rows:
        grouped[row["panel"]].append(row)
    data_rows = []
    for panel, items in sorted(grouped.items()):
        racks = ",".join(str(rack) for rack in sorted({item["rack"] for item in items}))
        areas = ", ".join(sorted({item["area"] for item in items}))
        data_rows.append([panel, areas, racks, len(items), "refer to data tabs"])
    write_rows(ws, headers, data_rows)


def append_simple_kv_sheet(wb: Workbook, title: str, pairs: list[tuple[str, str]]) -> None:
    ws = wb.create_sheet(title)
    write_rows(ws, ["Field", "Value"], pairs)


def write_canonical_csv(rows: list[dict]) -> None:
    path = OUTPUT_DIR / "00_canonical_io.csv"
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=CANONICAL_COLUMNS)
        writer.writeheader()
        writer.writerows(sanitize_rows_for_export(rows))


def write_canonical_xlsx(rows: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Canonical IO"
    write_rows(ws, CANONICAL_COLUMNS, ([row[col] for col in CANONICAL_COLUMNS] for row in sanitize_rows_for_export(rows)))
    wb.save(OUTPUT_DIR / "00_canonical_io.xlsx")


def write_01_flat_ptbr(rows: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Lista IO"
    headers = [
        "Rack",
        "Slot",
        "Modulo",
        "Canal",
        "Tag",
        "Descricao",
        "Tipo de I/O",
        "Sinal",
        "Unidade",
        "Faixa",
        "Painel",
        "Area",
        "Comentario",
    ]
    data_rows = []
    for row in rows:
        data_rows.append(
            [
                f"{row['rack']:02d}",
                f"{row['slot']:02d}",
                row["module_model"],
                f"{row['channel']:02d}",
                row["tag"],
                row["description"],
                row["io_type"],
                row["signal_type"],
                row["eng_unit"],
                format_range(row),
                row["panel"],
                row["area"],
                row["comment"],
            ]
        )
    write_rows(ws, headers, data_rows)
    wb.save(OUTPUT_DIR / "01_io_flat_ptbr.xlsx")


def write_02_flat_english(rows: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "IO List"
    headers = [
        "Panel",
        "Area",
        "Card",
        "Rack",
        "Slot",
        "Chan",
        "TAG",
        "Desc",
        "Type",
        "Signal",
        "Unit",
        "Range",
        "Field Notes",
    ]
    data_rows = []
    for row in rows:
        data_rows.append(
            [
                row["panel"],
                row["_area_en"],
                row["module_type"],
                row["rack"],
                row["slot"],
                row["channel"],
                row["tag"],
                row["_description_en"],
                row["io_type"],
                row["signal_type"],
                row["eng_unit"],
                format_range(row),
                row["_comment_en"],
            ]
        )
    write_rows(ws, headers, data_rows)
    wb.save(OUTPUT_DIR / "02_io_flat_english.xlsx")


def write_03_compact_location(rows: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Compact IO"
    headers = [
        "Point Ref",
        "Location",
        "Card Position",
        "Type",
        "Signal",
        "Tag",
        "Service",
        "Panel",
        "Field Wiring",
        "Range / Unit",
        "Remarks",
    ]
    data_rows = []
    for row in rows:
        point_ref, position = compact_location_variants(row)
        range_unit = format_range(row)
        if row["eng_unit"]:
            range_unit = f"{range_unit} {row['eng_unit']}".strip()
        data_rows.append(
            [
                point_ref,
                row["location_string"],
                position,
                row["io_type"],
                row["signal_type"],
                row["tag"],
                row["_description_en"],
                row["panel"],
                f"{row['junction_box']} / {row['terminal_block']}-{row['terminal']}",
                range_unit,
                row["_comment_en"] or row["comment"],
            ]
        )
    write_rows(ws, headers, data_rows)
    append_panel_index_sheet(wb, rows, "Index")
    append_rack_summary_sheet(wb, rows, "Rack Map")
    wb.save(OUTPUT_DIR / "03_io_compact_location.xlsx")


def write_04_grouped_by_module(rows: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Indice"
    write_rows(
        ws,
        ["Guia", "Conteudo"],
        [
            ["Modules_All", "Todos os módulos e racks na mesma guia principal"],
            ["Rack Summary", "Resumo por painel e rack"],
        ],
    )

    data_ws = wb.create_sheet("Modules_All")
    data_ws.append(["I/O List Grouped by Module"])
    data_ws["A1"].font = Font(size=14, bold=True)
    data_ws.merge_cells("A1:J1")
    data_ws.append([])

    grouped = defaultdict(list)
    for row in rows:
        grouped[(row["panel"], row["rack"], row["slot"], row["module_model"])].append(row)

    for key in sorted(grouped):
        panel, rack, slot, module_model = key
        module_rows = sorted(grouped[key], key=lambda item: item["channel"])
        data_ws.append([f"Panel {panel} | Rack {rack:02d} | Slot {slot:02d} | {module_model}"])
        section_row = data_ws.max_row
        data_ws.merge_cells(start_row=section_row, start_column=1, end_row=section_row, end_column=10)
        data_ws.cell(section_row, 1).font = Font(bold=True)
        data_ws.cell(section_row, 1).fill = PatternFill(fill_type="solid", fgColor="D9E2F3")
        data_ws.append(["Rack", "Slot", "Ch", "I/O", "Tag", "Descricao", "Sinal", "Faixa", "Area", "Comentario"])
        apply_table_header(data_ws, data_ws.max_row)
        for row in module_rows:
            data_ws.append(
                [
                    f"{row['rack']:02d}",
                    f"{row['slot']:02d}",
                    f"{row['channel']:02d}",
                    row["io_type"],
                    row["tag"],
                    row["description"],
                    row["signal_type"],
                    format_range(row),
                    row["area"],
                    row["comment"],
                ]
            )
        data_ws.append([])
    auto_size_worksheet(data_ws)
    data_ws.freeze_panes = "A4"
    append_rack_summary_sheet(wb, rows, "Rack Summary")
    wb.save(OUTPUT_DIR / "04_io_grouped_by_module.xlsx")


def write_05_multisheet_by_panel(rows: list[dict]) -> None:
    wb = Workbook()
    cover = wb.active
    cover.title = "Capa"
    cover["A1"] = "Industrial I/O Import Test Pack"
    cover["A1"].font = Font(size=16, bold=True)
    cover["A3"] = "Documento: conjunto de planilhas para validar importacao multi-formato"
    cover["A4"] = f"Total de pontos logicos: {len(rows)}"
    cover["A5"] = f"Data de geracao: {DOC_DATE.isoformat()}"
    cover["A6"] = f"Seed fixa: {SEED}"
    cover["A7"] = "Cada painel contém múltiplos racks e os dados estão separados por guia."

    legend = wb.create_sheet("Legenda")
    legend_rows = [
        ["Abreviacao", "Significado"],
        ["DI", "Digital Input"],
        ["DO", "Digital Output"],
        ["AI", "Analog Input"],
        ["AO", "Analog Output"],
        ["SPARE", "Canal de reserva"],
    ]
    write_rows(legend, legend_rows[0], legend_rows[1:])

    panel_groups = defaultdict(list)
    for row in rows:
        panel_groups[row["panel"]].append(row)

    for panel in sorted(panel_groups):
        ws = wb.create_sheet(panel.replace("PNL-", "").replace("MCC-", ""))
        headers = [
            "Rack",
            "Slot",
            "Canal",
            "Tag",
            "Descricao",
            "Tipo",
            "Sinal",
            "Unidade",
            "Faixa",
            "JB",
            "TB",
            "Terminal",
            "Area",
        ]
        data_rows = []
        for row in sorted(panel_groups[panel], key=lambda item: (item["rack"], item["slot"], item["channel"], item["tag"])):
            data_rows.append(
                [
                    row["rack"],
                    row["slot"],
                    row["channel"],
                    row["tag"],
                    row["description"],
                    row["io_type"],
                    row["signal_type"],
                    row["eng_unit"],
                    format_range(row),
                    row["junction_box"],
                    row["terminal_block"],
                    row["terminal"],
                    row["area"],
                ]
            )
        write_rows(ws, headers, data_rows)
    append_rack_summary_sheet(wb, rows, "Resumo_Racks")
    wb.save(OUTPUT_DIR / "05_io_multisheet_by_panel.xlsx")


def write_06_two_header_levels(rows: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "IO RevB"
    ws["A1"] = "Lista Geral de Entradas e Saidas"
    ws["A1"].font = Font(size=16, bold=True)
    ws.merge_cells("A1:P1")
    ws["A2"] = "Revisao"
    ws["B2"] = "B"
    ws["D2"] = "Data"
    ws["E2"] = DOC_DATE.isoformat()
    ws["G2"] = "Autor"
    ws["H2"] = AUTHOR

    group_headers = [
        ("A5:D5", "Hardware"),
        ("E5:H5", "Signal"),
        ("I5:K5", "Process"),
        ("L5:P5", "Installation"),
    ]
    for cell_range, label in group_headers:
        ws.merge_cells(cell_range)
        cell = ws[cell_range.split(":")[0]]
        cell.value = label
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="4F81BD")
        cell.alignment = Alignment(horizontal="center")

    subheaders = [
        "Rack",
        "Slot",
        "Modulo",
        "Canal",
        "Tag",
        "Descricao",
        "Tipo I/O",
        "Sinal",
        "Unidade",
        "Faixa Min",
        "Faixa Max",
        "Painel",
        "Area",
        "JB",
        "TB / Tm",
        "Comentario",
    ]
    ws.append([])
    ws.append(subheaders)
    apply_table_header(ws, 6)

    for row in rows:
        ws.append(
            [
                row["rack"],
                row["slot"],
                row["module_type"],
                row["channel"],
                row["tag"],
                row["description"],
                row["io_type"],
                row["signal_type"],
                row["eng_unit"],
                row["range_min"],
                row["range_max"],
                row["panel"],
                row["area"],
                row["junction_box"],
                f"{row['terminal_block']} / {row['terminal']}",
                row["comment"],
            ]
        )
    auto_size_worksheet(ws)
    ws.freeze_panes = "A7"
    append_simple_kv_sheet(
        wb,
        "Revision Log",
        [
            ("Revision", "B"),
            ("Date", DOC_DATE.isoformat()),
            ("Author", AUTHOR),
            ("Scope", "Main sheet plus auxiliary guides for rack and legend validation"),
        ],
    )
    append_simple_kv_sheet(
        wb,
        "Legend",
        [
            ("DI", "Digital input"),
            ("DO", "Digital output"),
            ("AI", "Analog input"),
            ("AO", "Analog output"),
            ("SPARE", "Reserved logical point"),
        ],
    )
    append_rack_summary_sheet(wb, rows, "Racks")
    wb.save(OUTPUT_DIR / "06_io_two_header_levels.xlsx")


def write_07_noisy_export(rows: list[dict], rng: random.Random) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "EXPORT_01"
    ws["A1"] = "Legacy export generated from site database"
    ws["A2"] = "NOTE: blank lines and comments below should be ignored by import."
    ws["A4"] = "Revision C / Drawing package IO-EXP-07"

    headers = [
        "Panel",
        "Rack",
        "Slot",
        "Chan",
        "Tag",
        "Desc",
        "Type",
        "Signal",
        "Unit",
        "Range Min",
        "Range Max",
        "Drawing",
        "Revision",
        "Loop No.",
        "Skid",
        "Remark 1",
        "Remark 2",
    ]
    start_row = 6
    for col_index, header in enumerate(headers, start=1):
        ws.cell(start_row, col_index, header)
    apply_table_header(ws, start_row)

    current_row = start_row + 1
    for index, row in enumerate(rows, start=1):
        padded_tag = f" {row['tag']} " if index % 11 == 0 else row["tag"]
        padded_desc = f"  {row['description']}  " if index % 13 == 0 else row["description"]
        remark_1 = row["comment"] or ("check marshalling" if index % 9 == 0 else "")
        remark_2 = "legacy export" if index % 7 == 0 else ""
        ws.append(
            [
                f" {row['panel']} " if index % 10 == 0 else row["panel"],
                row["rack"],
                row["slot"],
                row["channel"],
                padded_tag,
                padded_desc,
                row["io_type"],
                row["signal_type"],
                row["eng_unit"],
                row["range_min"],
                row["range_max"],
                f"DWG-{row['_area_key'].upper()}-{100 + (index % 9)}",
                rng.choice(["A", "B", "C"]),
                f"LP-{index:04d}",
                row["device"],
                remark_1,
                remark_2,
            ]
        )
        current_row += 1
        if index in {24, 61, 97}:
            ws.append([])
            current_row += 1
    ws[f"A{current_row + 1}"] = f"TOTAL LOGICAL POINTS SHOWN = {len(rows)}"
    ws[f"A{current_row + 2}"] = "Legend: rows with extra spaces are intentional."
    auto_size_worksheet(ws)
    ws.freeze_panes = "A7"
    append_simple_kv_sheet(
        wb,
        "IGNORE_ME",
        [
            ("Note", "This auxiliary tab simulates a workbook with extra guides."),
            ("Hint", "Importer should focus on the export guide containing the actual rows."),
            ("Racks Present", ",".join(str(r) for r in sorted({row['rack'] for row in rows}))),
        ],
    )
    wb.save(OUTPUT_DIR / "07_io_noisy_export.xlsx")


def write_08_decimal_comma_ptbr(rows: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Lista Regional"
    headers = [
        "Painel",
        "Rack",
        "Slot",
        "Canal",
        "Tag",
        "Descricao",
        "Tipo de I/O",
        "Sinal",
        "Unidade",
        "Faixa Min",
        "Faixa Max",
        "Observacao",
    ]
    data_rows = []
    for row in rows:
        data_rows.append(
            [
                row["panel"],
                f"{row['rack']:02d}",
                f"{row['slot']:02d}",
                f"{row['channel']:02d}",
                row["tag"],
                row["description"],
                row["io_type"],
                row["signal_type"],
                row["eng_unit"],
                format_number(row["range_min"], decimal_comma=True),
                format_number(row["range_max"], decimal_comma=True),
                row["comment"],
            ]
        )
    write_rows(ws, headers, data_rows)
    wb.save(OUTPUT_DIR / "08_io_decimal_comma_ptbr.xlsx")


def write_09_minimal_headers(rows: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "EXP_MIN"
    headers = ["Pnl", "Loc", "Rk", "Sl", "Mod", "Ch", "Pt", "Svc", "I/O", "Sig", "JB", "TB", "Tm", "Cbl"]
    data_rows = []
    for row in rows:
        data_rows.append(
            [
                row["panel"],
                row["location_string"],
                row["rack"],
                row["slot"],
                row["module_type"],
                row["channel"],
                row["tag"],
                row["description"],
                row["io_type"],
                row["signal_type"],
                row["junction_box"],
                row["terminal_block"],
                row["terminal"],
                row["cable"],
            ]
        )
    write_rows(ws, headers, data_rows)
    wb.save(OUTPUT_DIR / "09_io_minimal_headers.xlsx")


def write_10_mixed_real_world(rows: list[dict], rng: random.Random) -> None:
    wb = Workbook()
    cover = wb.active
    cover.title = "ChkIndex"
    cover["A1"] = "PROJECT: INDUSTRIAL I/O CHECK LIST"
    cover["A2"] = "DOC: MIXED REAL WORLD EXPORT"
    cover["A3"] = "REV: C"
    cover["C3"] = "DATE"
    cover["D3"] = DOC_DATE.strftime("%d/%m/%Y")
    cover["F3"] = "AUTHOR"
    cover["G3"] = AUTHOR
    cover["A5"] = "Dados distribuídos nas guias 'AreaBatch_A' e 'zz_field_dump'."

    headers = [
        "Loop Ref",
        "Panel / Card Pos",
        "Point",
        "Type / Signal",
        "Service",
        "Area",
        "Wiring",
        "Range",
        "Remark",
        "Asset",
        "P&ID",
    ]
    spare_labels = ["SPARE", "RES", "RESERVA"]
    tab_rows = [
        ("AreaBatch_A", rows[::2]),
        ("zz_field_dump", rows[1::2]),
    ]
    row_counter = 0
    for sheet_name, sheet_rows in tab_rows:
        ws = wb.create_sheet(sheet_name)
        ws["A1"] = "PROJECT: INDUSTRIAL I/O CHECK LIST"
        ws["A2"] = f"SOURCE TAB: {sheet_name}"
        ws["A3"] = "REV: C"
        ws["C3"] = "DATE"
        ws["D3"] = DOC_DATE.strftime("%d/%m/%Y")
        ws["A5"] = "Data below starts after revision block."
        header_row = 7
        for col_index, header in enumerate(headers, start=1):
            ws.cell(header_row, col_index, header)
        apply_table_header(ws, header_row)
        for row in sheet_rows:
            row_counter += 1
            service_value = row["description"]
            if row["io_type"] == "SPARE":
                service_value = spare_labels[(row_counter - 1) % len(spare_labels)]
            type_signal = {
                "DI": "DI Dry Contact",
                "DO": "DO 24Vdc",
                "AI": "AI 4-20mA",
                "AO": "AO 4-20mA",
                "SPARE": f"{row['module_type']} RESERVED",
            }[row["io_type"]]
            panel_card = f"{row['panel']} / {card_position(row, with_leading_zero=row_counter % 2 == 0)}"
            range_text = format_range(row)
            if row["eng_unit"]:
                range_text = f"{range_text} {row['eng_unit']}".strip()
            remark = row["comment"] or rng.choice(["field check", "as built", "confirmed", ""])
            ws.append(
                [
                    f"{row['device']}-{row_counter:03d}",
                    panel_card,
                    row["tag"],
                    type_signal,
                    service_value,
                    row["area"],
                    f"{row['junction_box']} / {row['terminal_block']}-{row['terminal']} / {row['cable']}",
                    range_text,
                    remark,
                    row["device"],
                    row["p_and_id"],
                ]
            )
        auto_size_worksheet(ws)
        ws.freeze_panes = "A8"
    append_rack_summary_sheet(wb, rows, "Rack_Xref")
    wb.save(OUTPUT_DIR / "10_io_mixed_real_world.xlsx")


def build_manifest_entries() -> list[dict]:
    entries = [
        {
            "file_name": "00_canonical_io.csv",
            "description": "Base canônica em CSV com todos os campos de normalização.",
            "expected_logical_points": EXPECTED_TOTAL,
            "header_languages": ["pt-BR", "technical"],
            "multi_sheet": False,
            "multi_rack": True,
            "multi_racks_in_single_sheet": True,
            "multi_level_header": False,
            "combined_rack_slot_channel": False,
            "contains_noise_rows": False,
            "contains_decimal_comma": False,
        },
        {
            "file_name": "00_canonical_io.xlsx",
            "description": "Base canônica em Excel com todos os campos de validação.",
            "expected_logical_points": EXPECTED_TOTAL,
            "header_languages": ["pt-BR", "technical"],
            "multi_sheet": False,
            "multi_rack": True,
            "multi_racks_in_single_sheet": True,
            "multi_level_header": False,
            "combined_rack_slot_channel": False,
            "contains_noise_rows": False,
            "contains_decimal_comma": False,
        },
    ]
    for file_name, description in SUMMARY_BY_FILE.items():
        entries.append(
            {
                "file_name": file_name,
                "description": description,
                "expected_logical_points": EXPECTED_TOTAL,
                "header_languages": {
                    "01_io_flat_ptbr.xlsx": ["pt-BR"],
                    "02_io_flat_english.xlsx": ["en"],
                    "03_io_compact_location.xlsx": ["en", "technical"],
                    "04_io_grouped_by_module.xlsx": ["pt-BR"],
                    "05_io_multisheet_by_panel.xlsx": ["pt-BR"],
                    "06_io_two_header_levels.xlsx": ["pt-BR"],
                    "07_io_noisy_export.xlsx": ["en", "technical"],
                    "08_io_decimal_comma_ptbr.xlsx": ["pt-BR"],
                    "09_io_minimal_headers.xlsx": ["technical"],
                    "10_io_mixed_real_world.xlsx": ["en", "pt-BR", "technical"],
                }[file_name],
                "multi_sheet": file_name in {
                    "03_io_compact_location.xlsx",
                    "04_io_grouped_by_module.xlsx",
                    "05_io_multisheet_by_panel.xlsx",
                    "06_io_two_header_levels.xlsx",
                    "07_io_noisy_export.xlsx",
                    "10_io_mixed_real_world.xlsx",
                },
                "multi_rack": True,
                "multi_racks_in_single_sheet": file_name in {
                    "01_io_flat_ptbr.xlsx",
                    "02_io_flat_english.xlsx",
                    "03_io_compact_location.xlsx",
                    "04_io_grouped_by_module.xlsx",
                    "05_io_multisheet_by_panel.xlsx",
                    "06_io_two_header_levels.xlsx",
                    "07_io_noisy_export.xlsx",
                    "08_io_decimal_comma_ptbr.xlsx",
                    "09_io_minimal_headers.xlsx",
                    "10_io_mixed_real_world.xlsx",
                },
                "multi_level_header": file_name == "06_io_two_header_levels.xlsx",
                "combined_rack_slot_channel": file_name in {
                    "03_io_compact_location.xlsx",
                    "04_io_grouped_by_module.xlsx",
                    "10_io_mixed_real_world.xlsx",
                },
                "contains_noise_rows": file_name in {
                    "04_io_grouped_by_module.xlsx",
                    "06_io_two_header_levels.xlsx",
                    "07_io_noisy_export.xlsx",
                    "10_io_mixed_real_world.xlsx",
                },
                "contains_decimal_comma": file_name == "08_io_decimal_comma_ptbr.xlsx",
            }
        )
    entries.extend(
        [
            {
                "file_name": "README.md",
                "description": "Documentação do pacote gerado.",
                "expected_logical_points": EXPECTED_TOTAL,
                "header_languages": ["pt-BR"],
                "multi_sheet": False,
                "multi_rack": True,
                "multi_racks_in_single_sheet": False,
                "multi_level_header": False,
                "combined_rack_slot_channel": False,
                "contains_noise_rows": False,
                "contains_decimal_comma": False,
            },
            {
                "file_name": "generate_io_test_files.py",
                "description": "Script Python reprodutível que gera todos os artefatos localmente.",
                "expected_logical_points": EXPECTED_TOTAL,
                "header_languages": ["technical"],
                "multi_sheet": False,
                "multi_rack": True,
                "multi_racks_in_single_sheet": False,
                "multi_level_header": False,
                "combined_rack_slot_channel": False,
                "contains_noise_rows": False,
                "contains_decimal_comma": False,
            },
        ]
    )
    return entries


def write_manifest() -> None:
    manifest = {
        "package_name": "industrial_io_import_test_files",
        "generated_on": DOC_DATE.isoformat(),
        "seed": SEED,
        "canonical_logical_points": EXPECTED_TOTAL,
        "io_distribution": EXPECTED_COUNTS,
        "files": build_manifest_entries(),
    }
    with (OUTPUT_DIR / "manifest.json").open("w", encoding="utf-8") as handle:
        json.dump(manifest, handle, indent=2, ensure_ascii=False)


def write_readme(rows: list[dict]) -> None:
    counts = Counter(row["io_type"] for row in rows)
    lines = [
        "# Industrial I/O Import Test Files",
        "",
        "## Visão geral",
        "",
        "Este pacote contém uma base canônica e 10 planilhas Excel estruturalmente diferentes para testar importação de listas de I/O industriais.",
        "Todas as 10 planilhas de teste representam exatamente o mesmo conjunto lógico de pontos da base canônica.",
        "O conjunto atualizado contempla explicitamente cenários com múltiplos racks na mesma guia e também múltiplas guias em diferentes workbooks.",
        "",
        "## Quantidade total de pontos lógicos",
        "",
        f"- Total: {len(rows)}",
        f"- DI: {counts['DI']}",
        f"- DO: {counts['DO']}",
        f"- AI: {counts['AI']}",
        f"- AO: {counts['AO']}",
        f"- SPARE: {counts['SPARE']}",
        "",
        "## Arquivos gerados",
        "",
        "- `00_canonical_io.csv`: base canônica em CSV com todos os campos de normalização.",
        "- `00_canonical_io.xlsx`: base canônica em Excel com todos os campos de validação.",
        "- `01_io_flat_ptbr.xlsx`: planilha limpa em português com colunas explícitas.",
        "- `02_io_flat_english.xlsx`: planilha linear em inglês com ordem e nomenclatura alteradas.",
        "- `03_io_compact_location.xlsx`: layout compacto com localização combinada, índice e resumo por rack.",
        "- `04_io_grouped_by_module.xlsx`: layout agrupado por módulo com múltiplos racks na mesma guia principal e abas auxiliares.",
        "- `05_io_multisheet_by_panel.xlsx`: workbook com capa, legenda, resumo de racks e abas por painel.",
        "- `06_io_two_header_levels.xlsx`: documento com título, revisão, cabeçalho em dois níveis e guias auxiliares.",
        "- `07_io_noisy_export.xlsx`: exportação bagunçada com colunas irrelevantes, ruído e guia auxiliar.",
        "- `08_io_decimal_comma_ptbr.xlsx`: planilha PT-BR com vírgula decimal.",
        "- `09_io_minimal_headers.xlsx`: exportação com cabeçalhos curtos e pouco amigáveis.",
        "- `10_io_mixed_real_world.xlsx`: planilha de campo com dados distribuídos em guias não óbvias, cada uma contendo múltiplos racks.",
        "- `manifest.json`: metadados do pacote e características de cada arquivo.",
        "- `generate_io_test_files.py`: script reprodutível para gerar todos os artefatos.",
        "",
        "## Observações",
        "",
        "- Os arquivos `.xlsx` foram gerados com `openpyxl` e não utilizam macros.",
        "- Todos os formatos derivam da mesma base canônica; a variação está apenas na organização e apresentação.",
        "- Todos os datasets gerados permanecem multi-rack na representação física dos canais, inclusive em guias únicas de vários formatos.",
        "- O campo `point_uid` existe apenas na base canônica para rastreabilidade e validação.",
    ]
    (OUTPUT_DIR / "README.md").write_text("\n".join(lines) + "\n", encoding="utf-8")


def validate_generated_workbooks() -> None:
    workbook_files = [
        "00_canonical_io.xlsx",
        "01_io_flat_ptbr.xlsx",
        "02_io_flat_english.xlsx",
        "03_io_compact_location.xlsx",
        "04_io_grouped_by_module.xlsx",
        "05_io_multisheet_by_panel.xlsx",
        "06_io_two_header_levels.xlsx",
        "07_io_noisy_export.xlsx",
        "08_io_decimal_comma_ptbr.xlsx",
        "09_io_minimal_headers.xlsx",
        "10_io_mixed_real_world.xlsx",
    ]
    for file_name in workbook_files:
        load_workbook(OUTPUT_DIR / file_name)


def build_tree_lines() -> list[str]:
    tree_lines = [OUTPUT_DIR.name]
    for path in sorted(OUTPUT_DIR.iterdir()):
        tree_lines.append(f"+-- {path.name}")
    return tree_lines


def generate_all() -> dict:
    rng = random.Random(SEED)
    rows = build_canonical_dataset()
    write_canonical_csv(rows)
    write_canonical_xlsx(rows)
    write_01_flat_ptbr(rows)
    write_02_flat_english(rows)
    write_03_compact_location(rows)
    write_04_grouped_by_module(rows)
    write_05_multisheet_by_panel(rows)
    write_06_two_header_levels(rows)
    write_07_noisy_export(rows, rng)
    write_08_decimal_comma_ptbr(rows)
    write_09_minimal_headers(rows)
    write_10_mixed_real_world(rows, rng)
    write_readme(rows)
    write_manifest()
    validate_generated_workbooks()
    return {
        "rows": rows,
        "tree_lines": build_tree_lines(),
        "counts": Counter(row["io_type"] for row in rows),
    }


def main() -> None:
    summary = generate_all()
    print("Generated files:")
    for line in summary["tree_lines"]:
        print(line)
    print()
    print(f"Canonical logical points: {len(summary['rows'])}")
    print(f"I/O distribution: {dict(summary['counts'])}")


if __name__ == "__main__":
    main()
