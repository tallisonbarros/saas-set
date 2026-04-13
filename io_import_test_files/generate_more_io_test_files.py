from __future__ import annotations

import csv
import random
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SEED = 20260412
TODAY = date(2026, 4, 12)
BASE_DIR = Path(__file__).resolve().parent
CANONICAL_CSV = BASE_DIR / "00_canonical_io.csv"
EXPECTED_ROWS = 120

AREA_EN = {
    "Recebimento": "Receiving",
    "Moega": "Hopper",
    "Transporte": "Conveying",
    "Processo": "Process",
    "Ensacadeira": "Bagging",
    "Utilidades": "Utilities",
}

IO_CLASS = {
    "DI": "Discrete Input",
    "DO": "Discrete Output",
    "AI": "Analog Input",
    "AO": "Analog Output",
    "SPARE": "Spare",
}

FAMILY_SUMMARIES = []


def load_rows() -> list[dict]:
    with CANONICAL_CSV.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))
    for row in rows:
        row["rack"] = int(row["rack"])
        row["slot"] = int(row["slot"])
        row["channel"] = int(row["channel"])
        row["range_min"] = float(row["range_min"]) if row["range_min"] else None
        row["range_max"] = float(row["range_max"]) if row["range_max"] else None
        row["area_en"] = AREA_EN[row["area"]]
        row["panel_short"] = row["panel"].replace("PNL-", "").replace("MCC-", "")
        row["io_class"] = IO_CLASS[row["io_type"]]
        row["range_text"] = format_range(row)
        row["range_text_pt"] = format_range(row, decimal_comma=True)
        row["card_pos"] = f"R{row['rack']:02d}-S{row['slot']:02d}-CH{row['channel']:02d}"
        row["card_id"] = f"R{row['rack']:02d}-S{row['slot']:02d}-{row['module_type']}"
        row["rack_slot"] = f"R{row['rack']:02d}/S{row['slot']:02d}"
        row["tb_term"] = f"{row['terminal_block']}-{row['terminal']}"
        row["loop_no"] = f"LP-{row['point_uid'].split('-')[-1]}"
        row["signal_group"] = "DISCRETE" if row["io_type"] in {"DI", "DO"} else "ANALOG" if row["io_type"] in {"AI", "AO"} else "SPARE"
        row["fieldbus"] = f"FB-{row['panel_short']}-R{row['rack']:02d}"
        row["remote_station"] = f"{row['remote_node']}-R{row['rack']:02d}"
        row["scada_object"] = row["tag"].replace("_", ".")
        row["alarm_text"] = build_alarm_text(row)
        row["plc_addr_std"] = f"{row['rack']}.{row['slot']}.{row['channel']:02d}"
        row["plc_addr_logix"] = build_logix_address(row)
        row["plc_addr_unity"] = build_unity_address(row)
        row["plc_addr_wordbit"] = f"Rack {row['rack']} Slot {row['slot']} Bit {row['channel'] - 1:02d}"
        row["terminal_full"] = f"{row['junction_box']} / {row['terminal_block']}-{row['terminal']}"
        row["cable_core"] = f"{row['cable']}-C{((row['channel'] - 1) % 4) + 1}"
        row["module_group"] = f"{row['module_type']} / Rack {row['rack']} / Slot {row['slot']}"
    if len(rows) != EXPECTED_ROWS:
        raise ValueError(f"Unexpected canonical row count: {len(rows)}")
    return rows


def build_alarm_text(row: dict) -> str:
    if row["io_type"] == "DI":
        return f"{row['tag']} status change"
    if row["io_type"] == "DO":
        return f"{row['tag']} command mismatch"
    if row["io_type"] == "AI":
        return f"{row['tag']} out of range"
    if row["io_type"] == "AO":
        return f"{row['tag']} output deviation"
    return f"{row['tag']} reserved"


def build_logix_address(row: dict) -> str:
    prefix = {"DI": "Local:I", "DO": "Local:O", "AI": "Local:I", "AO": "Local:O", "SPARE": "Local:I"}[row["io_type"]]
    return f"{prefix}.Slot{row['slot']}.Data.{row['channel'] - 1}"


def build_unity_address(row: dict) -> str:
    prefix = {"DI": "%I", "DO": "%Q", "AI": "%IW", "AO": "%QW", "SPARE": "%I"}[row["io_type"]]
    return f"{prefix}{row['rack']}.{row['slot']}.{row['channel'] - 1}"


def format_number(value: float | None, decimal_comma: bool = False) -> str:
    if value is None:
        return ""
    if float(value).is_integer():
        text = f"{value:.1f}" if decimal_comma else f"{int(value)}"
    else:
        text = f"{value:.1f}"
    return text.replace(".", ",") if decimal_comma else text


def format_range(row: dict, decimal_comma: bool = False) -> str:
    if row["range_min"] is None and row["range_max"] is None:
        return ""
    return f"{format_number(row['range_min'], decimal_comma)} .. {format_number(row['range_max'], decimal_comma)}"


def safe_sheet_name(value: str) -> str:
    clean = value.replace("/", "-").replace("\\", "-").replace("[", "(").replace("]", ")").replace("*", "").replace("?", "")
    return clean[:31]


def add_title_block(ws, title: str, subtitle: str | None = None, end_col: int = 12) -> int:
    ws.cell(1, 1, title)
    ws.cell(1, 1).font = Font(size=14, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    if subtitle:
        ws.cell(2, 1, subtitle)
        ws.cell(2, 1).font = Font(italic=True)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_col)
        return 3
    return 2


def style_header_row(ws, row_index: int, fill_color: str = "1F4E78") -> None:
    fill = PatternFill(fill_type="solid", fgColor=fill_color)
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[row_index]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def auto_size(ws, minimum: int = 9) -> None:
    for column_cells in ws.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[column_letter].width = max(minimum, min(max_length + 2, 40))


def write_table(ws, start_row: int, headers: list[str], rows: list[list], freeze: bool = True, fill_color: str = "1F4E78") -> None:
    for col_index, header in enumerate(headers, start=1):
        ws.cell(start_row, col_index, header)
    style_header_row(ws, start_row, fill_color=fill_color)
    row_ptr = start_row + 1
    for row in rows:
        for col_index, value in enumerate(row, start=1):
            ws.cell(row_ptr, col_index, value)
        row_ptr += 1
    auto_size(ws)
    if freeze:
        ws.freeze_panes = f"A{start_row + 1}"


def add_key_value_sheet(wb: Workbook, title: str, items: list[tuple[str, str]]) -> None:
    ws = wb.create_sheet(safe_sheet_name(title))
    write_table(ws, 1, ["Field", "Value"], [[k, v] for k, v in items])


def add_summary_sheet(wb: Workbook, title: str, rows: list[dict], group_by: str) -> None:
    ws = wb.create_sheet(safe_sheet_name(title))
    grouped = defaultdict(list)
    for row in rows:
        grouped[row[group_by]].append(row)
    out = []
    for key, items in sorted(grouped.items(), key=lambda item: str(item[0])):
        counts = Counter(item["io_type"] for item in items)
        out.append([key, len(items), counts.get("DI", 0), counts.get("DO", 0), counts.get("AI", 0), counts.get("AO", 0), counts.get("SPARE", 0)])
    write_table(ws, 1, [group_by, "Points", "DI", "DO", "AI", "AO", "SPARE"], out)


def sort_rows(rows: list[dict], mode: str) -> list[dict]:
    if mode == "panel":
        return sorted(rows, key=lambda r: (r["panel"], r["rack"], r["slot"], r["channel"], r["tag"]))
    if mode == "area":
        return sorted(rows, key=lambda r: (r["area"], r["panel"], r["rack"], r["slot"], r["channel"]))
    if mode == "type":
        return sorted(rows, key=lambda r: (r["io_type"], r["panel"], r["rack"], r["slot"], r["channel"]))
    if mode == "tag":
        return sorted(rows, key=lambda r: r["tag"])
    return sorted(rows, key=lambda r: (r["rack"], r["slot"], r["channel"], r["tag"]))


def write_flat_variant(path: Path, rows: list[dict], cfg: dict) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = safe_sheet_name(cfg["sheet"])
    start_row = 1
    if cfg.get("title"):
        start_row = add_title_block(ws, cfg["title"], cfg.get("subtitle"), end_col=len(cfg["columns"]))
        if cfg.get("meta_line"):
            ws.cell(start_row, 1, cfg["meta_line"])
            ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(cfg["columns"]))
            start_row += 2
    ordered = sort_rows(rows, cfg.get("sort", "panel"))
    data_rows = []
    for row in ordered:
        data_rows.append([value_from_key(row, key, cfg) for _, key in cfg["columns"]])
        if cfg.get("blank_every") and len(data_rows) % cfg["blank_every"] == 0:
            data_rows.append(["" for _ in cfg["columns"]])
    write_table(ws, start_row, [header for header, _ in cfg["columns"]], data_rows, fill_color=cfg.get("fill", "1F4E78"))
    if cfg.get("extra_index"):
        add_summary_sheet(wb, "Index", rows, cfg["extra_index"])
    wb.save(path)


def write_plc_import_variant(path: Path, rows: list[dict], cfg: dict) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = safe_sheet_name(cfg["sheet"])
    start_row = add_title_block(ws, cfg["title"], cfg.get("subtitle"), end_col=11)
    if cfg.get("preface"):
        ws.cell(start_row, 1, cfg["preface"])
        start_row += 2
    headers = cfg["headers"]
    out_rows: list[list] = []
    current_group = None
    for row in sort_rows(rows, cfg.get("sort", "panel")):
        group_value = cfg.get("group_by") and row[cfg["group_by"]]
        if cfg.get("insert_breaks") and current_group is not None and group_value != current_group:
            out_rows.append(["BREAK"] + [""] * (len(headers) - 1))
        current_group = group_value
        out_rows.append([
            cfg.get("code", "PLC"),
            row["tag"],
            row["plc_addr_std"] if cfg.get("addr_style") == "std" else row["plc_addr_logix"] if cfg.get("addr_style") == "logix" else row["plc_addr_unity"],
            row["rack"],
            row["slot"],
            row["channel"],
            row["module_type"],
            row["terminal_full"],
            row["description"],
            row["panel"],
            row["location_string"],
        ])
    write_table(ws, start_row, headers, out_rows, fill_color="5B9BD5")
    if cfg.get("legend"):
        add_key_value_sheet(wb, "Legend", [("BREAK", "Section separator"), ("ADDR", cfg["addr_style"]), ("Rows", str(len(rows)))])
    wb.save(path)


def write_grouped_signal_variant(path: Path, rows: list[dict], cfg: dict) -> None:
    wb = Workbook()
    if cfg.get("tabs"):
        first = True
        groups = ["DI", "DO", "AI", "AO", "SPARE"]
        for io_type in groups:
            ws = wb.active if first else wb.create_sheet()
            first = False
            ws.title = safe_sheet_name(io_type if cfg.get("short_tabs") else IO_CLASS[io_type])
            subset = [row for row in sort_rows(rows, "panel") if row["io_type"] == io_type]
            start_row = add_title_block(ws, f"{cfg['title']} - {ws.title}", cfg.get("subtitle"), end_col=10)
            headers = cfg["headers"]
            out_rows = [[value_from_key(row, key, cfg) for _, key in cfg["columns"]] for row in subset]
            write_table(ws, start_row, headers, out_rows, fill_color=cfg.get("fill", "4472C4"))
    else:
        ws = wb.active
        ws.title = safe_sheet_name(cfg["sheet"])
        start_row = add_title_block(ws, cfg["title"], cfg.get("subtitle"), end_col=len(cfg["columns"]))
        row_ptr = start_row
        for io_type in ["DI", "DO", "AI", "AO", "SPARE"]:
            ws.cell(row_ptr, 1, cfg.get("section_names", {}).get(io_type, IO_CLASS[io_type]))
            ws.cell(row_ptr, 1).font = Font(bold=True)
            ws.merge_cells(start_row=row_ptr, start_column=1, end_row=row_ptr, end_column=len(cfg["columns"]))
            row_ptr += 1
            headers = [header for header, _ in cfg["columns"]]
            for col_index, header in enumerate(headers, start=1):
                ws.cell(row_ptr, col_index, header)
            style_header_row(ws, row_ptr, fill_color=cfg.get("fill", "4472C4"))
            row_ptr += 1
            for row in [r for r in sort_rows(rows, cfg.get("sort", "panel")) if r["io_type"] == io_type]:
                for col_index, (_, key) in enumerate(cfg["columns"], start=1):
                    ws.cell(row_ptr, col_index, value_from_key(row, key, cfg))
                row_ptr += 1
            row_ptr += 1
        auto_size(ws)
        ws.freeze_panes = f"A{start_row + 1}"
    if cfg.get("summary"):
        add_summary_sheet(wb, "Type Summary", rows, "io_type")
    wb.save(path)


def write_multisheet_variant(path: Path, rows: list[dict], cfg: dict) -> None:
    wb = Workbook()
    cover = wb.active
    cover.title = safe_sheet_name(cfg.get("cover", "Index"))
    write_table(cover, 1, ["Sheet", "Purpose"], cfg["cover_rows"])
    grouped = defaultdict(list)
    for row in rows:
        grouped[row[cfg["split_by"]]].append(row)
    for key, items in sorted(grouped.items(), key=lambda item: str(item[0])):
        ws = wb.create_sheet(safe_sheet_name(str(key)))
        start_row = 1
        if cfg.get("title"):
            start_row = add_title_block(ws, cfg["title"], f"{cfg['split_by']}: {key}", end_col=len(cfg["columns"]))
        out_rows = [[value_from_key(row, field, cfg) for _, field in cfg["columns"]] for row in sort_rows(items, cfg.get("sort", "panel"))]
        write_table(ws, start_row, [header for header, _ in cfg["columns"]], out_rows, fill_color=cfg.get("fill", "1F4E78"))
    if cfg.get("summary_by"):
        add_summary_sheet(wb, f"{cfg['summary_by']} Summary", rows, cfg["summary_by"])
    wb.save(path)


def write_marshalling_variant(path: Path, rows: list[dict], cfg: dict) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = safe_sheet_name(cfg["sheet"])
    start_row = add_title_block(ws, cfg["title"], cfg.get("subtitle"), end_col=len(cfg["columns"]))
    ordered = sort_rows(rows, cfg.get("sort", "panel"))
    if cfg.get("group_by_jb"):
        row_ptr = start_row
        for jb, items in sorted(defaultdict(list, {k: [r for r in ordered if r["junction_box"] == k] for k in sorted({r["junction_box"] for r in ordered})}).items()):
            ws.cell(row_ptr, 1, f"Junction Box {jb}")
            ws.cell(row_ptr, 1).font = Font(bold=True)
            ws.merge_cells(start_row=row_ptr, start_column=1, end_row=row_ptr, end_column=len(cfg["columns"]))
            row_ptr += 1
            for col_index, (header, _) in enumerate(cfg["columns"], start=1):
                ws.cell(row_ptr, col_index, header)
            style_header_row(ws, row_ptr, fill_color="8064A2")
            row_ptr += 1
            for row in items:
                for col_index, (_, key) in enumerate(cfg["columns"], start=1):
                    ws.cell(row_ptr, col_index, value_from_key(row, key, cfg))
                row_ptr += 1
            row_ptr += 1
        auto_size(ws)
        ws.freeze_panes = f"A{start_row + 1}"
    else:
        data_rows = [[value_from_key(row, key, cfg) for _, key in cfg["columns"]] for row in ordered]
        write_table(ws, start_row, [header for header, _ in cfg["columns"]], data_rows, fill_color="8064A2")
    if cfg.get("extras"):
        add_summary_sheet(wb, "JB Summary", rows, "junction_box")
        add_summary_sheet(wb, "Rack Summary", rows, "rack")
    wb.save(path)


def write_scada_variant(path: Path, rows: list[dict], cfg: dict) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = safe_sheet_name(cfg["sheet"])
    start_row = add_title_block(ws, cfg["title"], cfg.get("subtitle"), end_col=len(cfg["columns"]))
    ordered = sort_rows(rows, cfg.get("sort", "tag"))
    if cfg.get("noise_lines"):
        ws.cell(start_row, 1, "Generated from control database export")
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(cfg["columns"]))
        start_row += 2
    out_rows = []
    for row in ordered:
        out_rows.append([value_from_key(row, key, cfg) for _, key in cfg["columns"]])
        if cfg.get("blank_every") and len(out_rows) % cfg["blank_every"] == 0:
            out_rows.append(["" for _ in cfg["columns"]])
    write_table(ws, start_row, [header for header, _ in cfg["columns"]], out_rows, fill_color="70AD47")
    if cfg.get("alarm_sheet"):
        alarm_rows = [[row["tag"], row["alarm_text"], row["io_type"], row["panel"]] for row in ordered]
        alarm_ws = wb.create_sheet("Alarms")
        write_table(alarm_ws, 1, ["Tag", "Alarm Text", "Type", "Panel"], alarm_rows, fill_color="C00000")
    wb.save(path)


def write_card_variant(path: Path, rows: list[dict], cfg: dict) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = safe_sheet_name(cfg["sheet"])
    start_row = add_title_block(ws, cfg["title"], cfg.get("subtitle"), end_col=10)
    grouped = defaultdict(list)
    for row in sort_rows(rows, "panel"):
        grouped[(row["panel"], row["rack"], row["slot"], row["module_type"])].append(row)
    row_ptr = start_row
    for key in sorted(grouped):
        panel, rack, slot, module_type = key
        ws.cell(row_ptr, 1, f"{panel} / Rack {rack:02d} / Slot {slot:02d} / {module_type}")
        ws.cell(row_ptr, 1).font = Font(bold=True)
        ws.merge_cells(start_row=row_ptr, start_column=1, end_row=row_ptr, end_column=10)
        row_ptr += 1
        headers = [header for header, _ in cfg["columns"]]
        for col_index, header in enumerate(headers, start=1):
            ws.cell(row_ptr, col_index, header)
        style_header_row(ws, row_ptr, fill_color="9E480E")
        row_ptr += 1
        for row in grouped[key]:
            for col_index, (_, field) in enumerate(cfg["columns"], start=1):
                ws.cell(row_ptr, col_index, value_from_key(row, field, cfg))
            row_ptr += 1
        row_ptr += 1
    auto_size(ws)
    ws.freeze_panes = f"A{start_row + 1}"
    if cfg.get("summary"):
        add_summary_sheet(wb, "Slot Summary", rows, "rack_slot")
    wb.save(path)


def write_extreme_variant(path: Path, rows: list[dict], cfg: dict) -> None:
    wb = Workbook()
    cover = wb.active
    cover.title = safe_sheet_name(cfg.get("cover", "START_HERE"))
    write_table(cover, 1, ["Block", "Value"], [["Title", cfg["title"]], ["Date", TODAY.isoformat()], ["Hint", cfg["hint"]]])
    if cfg.get("legend"):
        add_key_value_sheet(wb, "LEGEND", cfg["legend"])
    for sheet_name, subset in cfg["splitter"](rows):
        ws = wb.create_sheet(safe_sheet_name(sheet_name))
        top = add_title_block(ws, cfg["title"], sheet_name, end_col=len(cfg["columns"]))
        if cfg.get("double_header"):
            groups = cfg["double_header"]
            col_ptr = 1
            for label, span in groups:
                ws.merge_cells(start_row=top, start_column=col_ptr, end_row=top, end_column=col_ptr + span - 1)
                ws.cell(top, col_ptr, label)
                ws.cell(top, col_ptr).font = Font(bold=True, color="FFFFFF")
                ws.cell(top, col_ptr).fill = PatternFill(fill_type="solid", fgColor="5B9BD5")
                col_ptr += span
            top += 1
        data_rows = [[value_from_key(row, field, cfg) for _, field in cfg["columns"]] for row in subset]
        write_table(ws, top, [header for header, _ in cfg["columns"]], data_rows, fill_color=cfg.get("fill", "1F4E78"))
    wb.save(path)


def value_from_key(row: dict, key: str, cfg: dict | None = None):
    decimal_comma = bool(cfg and cfg.get("decimal_comma"))
    if key == "panel":
        return row["panel"]
    if key == "panel_short":
        return row["panel_short"]
    if key == "area":
        return row["area"]
    if key == "area_en":
        return row["area_en"]
    if key == "rack":
        return row["rack"]
    if key == "rack02":
        return f"{row['rack']:02d}"
    if key == "slot":
        return row["slot"]
    if key == "slot02":
        return f"{row['slot']:02d}"
    if key == "channel":
        return row["channel"]
    if key == "channel02":
        return f"{row['channel']:02d}"
    if key == "card_pos":
        return row["card_pos"]
    if key == "rack_slot":
        return row["rack_slot"]
    if key == "module_model":
        return row["module_model"]
    if key == "module_type":
        return row["module_type"]
    if key == "tag":
        return row["tag"]
    if key == "tag_spaced":
        return f" {row['tag']} "
    if key == "description":
        return row["description"]
    if key == "description_upper":
        return row["description"].upper()
    if key == "io_type":
        return row["io_type"]
    if key == "io_class":
        return row["io_class"]
    if key == "signal_type":
        return row["signal_type"]
    if key == "eng_unit":
        return row["eng_unit"]
    if key == "range":
        return row["range_text_pt"] if decimal_comma else row["range_text"]
    if key == "range_min":
        return format_number(row["range_min"], decimal_comma=decimal_comma)
    if key == "range_max":
        return format_number(row["range_max"], decimal_comma=decimal_comma)
    if key == "power_or_signal":
        return row["power_or_signal"]
    if key == "device":
        return row["device"]
    if key == "p_and_id":
        return row["p_and_id"]
    if key == "junction_box":
        return row["junction_box"]
    if key == "tb_term":
        return row["tb_term"]
    if key == "terminal":
        return row["terminal"]
    if key == "terminal_full":
        return row["terminal_full"]
    if key == "cable":
        return row["cable"]
    if key == "cable_core":
        return row["cable_core"]
    if key == "location":
        return row["location_string"]
    if key == "loop_no":
        return row["loop_no"]
    if key == "fieldbus":
        return row["fieldbus"]
    if key == "remote_station":
        return row["remote_station"]
    if key == "scada_object":
        return row["scada_object"]
    if key == "alarm_text":
        return row["alarm_text"]
    if key == "plc":
        return row["plc"]
    if key == "remote_node":
        return row["remote_node"]
    if key == "plc_addr_std":
        return row["plc_addr_std"]
    if key == "plc_addr_logix":
        return row["plc_addr_logix"]
    if key == "plc_addr_unity":
        return row["plc_addr_unity"]
    if key == "plc_addr_wordbit":
        return row["plc_addr_wordbit"]
    if key == "module_group":
        return row["module_group"]
    if key == "signal_group":
        return row["signal_group"]
    if key == "comment":
        return row["comment"]
    return row.get(key, "")


def register_summary(file_name: str, summary: str) -> None:
    FAMILY_SUMMARIES.append((file_name, summary))


def family_flat(rows: list[dict]) -> None:
    configs = [
        ("11_io_autocad_like_register.xlsx", "AutoCAD-like register with technical but still flat columns.", {
            "sheet": "PLC_Register",
            "title": "PLC Register",
            "subtitle": "AutoCAD Electrical inspired flat import table",
            "columns": [("CODE", "io_type"), ("TAGNAME", "tag"), ("DESC", "description"), ("RACK", "rack02"), ("SLOT", "slot02"), ("CH", "channel02"), ("ADDR", "plc_addr_std"), ("CARD", "module_type"), ("LOC", "location"), ("PANEL", "panel")],
            "sort": "panel",
            "fill": "2F75B5",
            "extra_index": "panel",
        }),
        ("12_io_flat_db_register.xlsx", "DB export style with metadata rows before the header and mixed business columns.", {
            "sheet": "DB_Export",
            "title": "Site I/O DB Export",
            "subtitle": "Generated from engineering register",
            "meta_line": "Revision D / Export source: ENG_IO_REGISTER",
            "columns": [("Panel", "panel"), ("Area", "area"), ("Rack", "rack"), ("Slot", "slot"), ("Chan", "channel"), ("Point", "tag"), ("Description", "description"), ("Type", "io_type"), ("Signal", "signal_type"), ("Note", "comment")],
            "sort": "area",
            "fill": "5B9BD5",
        }),
        ("13_io_flat_hybrid_headers.xlsx", "Hybrid PT/EN headers with combined technical naming and blank spacer rows.", {
            "sheet": "Lista",
            "title": "Lista IO Híbrida",
            "columns": [("Painel", "panel"), ("Área", "area"), ("R/S/C", "card_pos"), ("Tag", "tag"), ("Descrição", "description"), ("I/O Type", "io_type"), ("Signal", "signal_type"), ("Range", "range"), ("Device", "device")],
            "sort": "panel",
            "blank_every": 25,
            "fill": "4472C4",
        }),
        ("14_io_flat_decimal_text_mix.xlsx", "Regional flat file with decimal comma and technical columns kept as text.", {
            "sheet": "Regional",
            "title": "Regional Flat Export",
            "columns": [("Painel", "panel"), ("Rack", "rack02"), ("Slot", "slot02"), ("Canal", "channel02"), ("Tag", "tag"), ("Descricao", "description"), ("Tipo", "io_type"), ("Faixa Min", "range_min"), ("Faixa Max", "range_max"), ("Unid", "eng_unit")],
            "sort": "panel",
            "decimal_comma": True,
            "fill": "70AD47",
        }),
        ("15_io_flat_operator_view.xlsx", "Operator-oriented view with object-like naming and minimal engineering columns.", {
            "sheet": "OperatorPts",
            "title": "Operator Point Register",
            "columns": [("Object", "scada_object"), ("Alarm", "alarm_text"), ("Type", "io_class"), ("Tag", "tag"), ("Panel", "panel_short"), ("Card Pos", "card_pos"), ("Unit", "eng_unit"), ("Range", "range"), ("Service", "description")],
            "sort": "tag",
            "fill": "A5A5A5",
        }),
    ]
    for file_name, summary, cfg in configs:
        write_flat_variant(BASE_DIR / file_name, rows, cfg)
        register_summary(file_name, summary)


def family_plc(rows: list[dict]) -> None:
    configs = [
        ("16_io_plc_import_code_addr.xlsx", "PLC import sheet with CODE, TAG, ADDR, rack, slot and location.", {"sheet": "PLC_Import", "title": "Spreadsheet to PLC I/O", "headers": ["CODE", "TAGNAME", "ADDR", "RACK", "SLOT", "CH", "CARD", "TERM", "DESC", "PANEL", "LOC"], "addr_style": "std", "code": "PLC", "group_by": "panel", "legend": True}),
        ("17_io_plc_import_breaks.xlsx", "Auto-generated PLC import with BREAK rows between panels.", {"sheet": "PLC_Input", "title": "PLC Import with Breaks", "headers": ["CODE", "TAGNAME", "ADDR", "RACK", "SLOT", "CH", "CARD", "TERM", "DESC", "PANEL", "LOC"], "addr_style": "std", "code": "BREAKABLE", "group_by": "panel", "insert_breaks": True, "legend": True}),
        ("18_io_rslogix_prepared.xlsx", "RSLogix-prepared export using Local:I/O style addresses.", {"sheet": "RSLogixPrep", "title": "RSLogix Export to Spreadsheet", "headers": ["CODE", "TAGNAME", "ADDR", "RACK", "SLOT", "CH", "CARD", "TERM", "DESC", "PANEL", "LOC"], "addr_style": "logix", "code": "RSLOGIX", "group_by": "rack"}),
        ("19_io_unitypro_prepared.xlsx", "Unity Pro inspired export using percent addresses.", {"sheet": "UnityPrep", "title": "Unity Pro Export to Spreadsheet", "headers": ["CODE", "TAGNAME", "ADDR", "RACK", "SLOT", "CH", "CARD", "TERM", "DESC", "PANEL", "LOC"], "addr_style": "unity", "code": "UNITY", "group_by": "rack", "legend": True}),
        ("20_io_plc_min_columns.xlsx", "Minimal PLC import with compact import-facing columns only.", {"sheet": "PLC_Min", "title": "Compact PLC Import", "headers": ["CODE", "TAGNAME", "ADDR", "RACK", "SLOT", "CH", "CARD", "TERM", "DESC", "PANEL", "LOC"], "addr_style": "logix", "code": "PLCMIN", "preface": "Columns kept intentionally terse for import mapping."}),
    ]
    for file_name, summary, cfg in configs:
        write_plc_import_variant(BASE_DIR / file_name, rows, cfg)
        register_summary(file_name, summary)


def family_fieldbus(rows: list[dict]) -> None:
    configs = [
        ("21_io_device_list_single_rack_style.xlsx", "Device I/O list styled as single-rack template while still mixing racks in one sheet.", {
            "sheet": "DeviceIO",
            "title": "Device IO List - Single Rack Style",
            "columns": [("Fieldbus", "fieldbus"), ("Remote Station", "remote_station"), ("Card Pos", "card_pos"), ("Device", "device"), ("Point", "tag"), ("Description", "description"), ("I/O", "io_type"), ("Signal", "signal_type"), ("Panel", "panel")],
            "sort": "panel",
            "fill": "2E75B6",
            "extra_index": "rack",
        }),
        ("22_io_device_list_single_fieldbus.xlsx", "Single-fieldbus style list with all points under one device-oriented structure.", {
            "sheet": "Fieldbus_A",
            "title": "Device IO List - Single Fieldbus",
            "columns": [("Bus", "fieldbus"), ("Node", "remote_station"), ("Card", "module_type"), ("Channel", "channel"), ("Tag", "tag"), ("Service", "description"), ("Type", "io_type"), ("Addr", "plc_addr_std"), ("JB/TB", "terminal_full")],
            "sort": "rack",
            "fill": "5B9BD5",
        }),
        ("23_io_device_list_multi_fieldbus.xlsx", "Multiple-fieldbus workbook split by remote station.", {
            "cover": "Index",
            "cover_rows": [["Index", "Summary and per remote station tabs"], ["Summary", "By fieldbus segment"]],
            "split_by": "remote_station",
            "title": "Multiple Fieldbus Device IO",
            "columns": [("Fieldbus", "fieldbus"), ("Panel", "panel"), ("Card Pos", "card_pos"), ("Tag", "tag"), ("Description", "description"), ("I/O", "io_type"), ("Signal", "signal_type"), ("Range", "range")],
            "summary_by": "fieldbus",
            "fill": "4472C4",
        }),
        ("24_io_remote_io_by_node.xlsx", "Remote I/O by node with workbook tabs per panel short name.", {
            "cover": "Start",
            "cover_rows": [["Guide", "Remote IO by panel"], ["Coverage", "All points preserved"]],
            "split_by": "panel_short",
            "title": "Remote IO by Node",
            "columns": [("Remote Node", "remote_node"), ("Fieldbus", "fieldbus"), ("Rack/Slot", "rack_slot"), ("Ch", "channel02"), ("Tag", "tag"), ("Service", "description"), ("Type", "io_type"), ("Cable", "cable")],
            "summary_by": "remote_node",
            "fill": "70AD47",
        }),
        ("25_io_remote_panel_device.xlsx", "Panel-device list with emphasis on remote station and device associations.", {
            "sheet": "PanelDevice",
            "title": "Panel / Device IO Register",
            "columns": [("Panel", "panel"), ("Station", "remote_station"), ("Device", "device"), ("Point", "tag"), ("Class", "io_class"), ("Card Group", "module_group"), ("Term", "terminal_full"), ("Comment", "comment")],
            "sort": "panel",
            "fill": "A5A5A5",
            "extra_index": "remote_station",
        }),
    ]
    for file_name, summary, cfg in configs:
        if "split_by" in cfg:
            write_multisheet_variant(BASE_DIR / file_name, rows, cfg)
        else:
            write_flat_variant(BASE_DIR / file_name, rows, cfg)
        register_summary(file_name, summary)


def family_signal(rows: list[dict]) -> None:
    shared_columns = [("Panel", "panel"), ("Area", "area"), ("Card Pos", "card_pos"), ("Tag", "tag"), ("Description", "description"), ("Type", "io_type"), ("Signal", "signal_type"), ("Unit", "eng_unit"), ("Range", "range"), ("P&ID", "p_and_id")]
    configs = [
        ("26_io_signal_grouped_winnipeg.xlsx", "Signal-grouped schedule inspired by municipal standards with EU range and P&ID.", {"sheet": "BySignal", "title": "I/O List Grouped by Signal Type", "columns": shared_columns, "summary": True, "fill": "5B9BD5"}),
        ("27_io_signal_tabs_by_type.xlsx", "Workbook split into DI, DO, AI, AO and SPARE tabs.", {"title": "Signal Group Tabs", "subtitle": "One tab per signal class", "columns": shared_columns, "headers": [h for h, _ in shared_columns], "tabs": True, "short_tabs": True, "fill": "4472C4"}),
        ("28_io_discrete_then_analog.xlsx", "Single-sheet grouped export with discrete sections first and analog later.", {"sheet": "Grouped", "title": "Discrete and Analog Grouped", "columns": shared_columns, "section_names": {"DI": "DISCRETE INPUTS", "DO": "DISCRETE OUTPUTS", "AI": "ANALOG INPUTS", "AO": "ANALOG OUTPUTS", "SPARE": "SPARES"}, "fill": "2F5597"}),
        ("29_io_range_first_process.xlsx", "Process-first list emphasizing units and engineering ranges, but now carrying full physical routing.", {"sheet": "ProcessSignals", "title": "Process Signal Schedule", "columns": [("Area", "area"), ("P&ID", "p_and_id"), ("Panel", "panel"), ("Rack", "rack02"), ("Slot", "slot02"), ("Ch", "channel02"), ("Loop", "loop_no"), ("Tag", "tag"), ("Desc", "description"), ("Unit", "eng_unit"), ("Low EU", "range_min"), ("High EU", "range_max"), ("Signal", "signal_type"), ("I/O", "io_type")], "sort": "area", "fill": "70AD47"}),
        ("30_io_type_grouped_decimal_comma.xlsx", "Signal-grouped PT-BR workbook using decimal comma for ranges.", {"sheet": "Tipos", "title": "Lista Agrupada por Tipo de Sinal", "columns": [("Área", "area"), ("Painel", "panel"), ("R/S/C", "card_pos"), ("Tag", "tag"), ("Descrição", "description"), ("Tipo", "io_type"), ("Unidade", "eng_unit"), ("Faixa", "range"), ("Sinal", "signal_type")], "decimal_comma": True, "sort": "panel", "fill": "ED7D31", "summary": True}),
    ]
    for file_name, summary, cfg in configs:
        if cfg.get("tabs") or cfg.get("section_names"):
            write_grouped_signal_variant(BASE_DIR / file_name, rows, cfg)
        else:
            write_flat_variant(BASE_DIR / file_name, rows, cfg)
        register_summary(file_name, summary)


def family_marshalling(rows: list[dict]) -> None:
    configs = [
        ("31_io_marshalling_schedule.xlsx", "Marshalling schedule focused on JB, TB, terminal and cable continuity.", {"sheet": "Marshall", "title": "Marshalling Schedule", "columns": [("Panel", "panel"), ("JB", "junction_box"), ("TB-Term", "tb_term"), ("Terminal", "terminal_full"), ("Cable", "cable"), ("Core", "cable_core"), ("Tag", "tag"), ("Description", "description"), ("I/O", "io_type"), ("Addr", "plc_addr_std")], "sort": "panel", "extras": True}),
        ("32_io_marshalling_by_junction_box.xlsx", "Marshalling workbook grouped by junction box sections in a single sheet.", {"sheet": "ByJB", "title": "Marshalling by Junction Box", "columns": [("JB", "junction_box"), ("Cable", "cable"), ("Core", "cable_core"), ("TB-Term", "tb_term"), ("Tag", "tag"), ("Service", "description"), ("Signal", "signal_type"), ("Panel", "panel"), ("Card Pos", "card_pos")], "group_by_jb": True}),
        ("33_io_terminal_plan.xlsx", "Terminal plan with terminal-centric wording and full installation references.", {"sheet": "TerminalPlan", "title": "Terminal Plan", "columns": [("Panel", "panel"), ("Rack/Slot", "rack_slot"), ("Terminal", "terminal_full"), ("Cable", "cable"), ("Core", "cable_core"), ("Point", "tag"), ("Svc", "description"), ("Type", "io_type"), ("P&ID", "p_and_id")], "sort": "rack", "extras": True}),
        ("34_io_cable_core_schedule.xlsx", "Cable/core schedule crossing point tags to marshalling and PLC card positions with explicit panel and physical channel fields.", {"sheet": "CableCore", "title": "Cable Core Schedule", "columns": [("Panel", "panel"), ("Rack", "rack02"), ("Slot", "slot02"), ("Ch", "channel02"), ("Cable", "cable"), ("Core", "cable_core"), ("JB", "junction_box"), ("TB", "tb_term"), ("Card Pos", "card_pos"), ("Tag", "tag"), ("Type", "io_type"), ("Signal", "signal_type"), ("Comment", "comment")], "sort": "panel"}),
        ("35_io_wiring_cross_reference.xlsx", "Cross-reference style wiring list with cable, terminal and PLC address in one line.", {"sheet": "XREF", "title": "Wiring Cross Reference", "columns": [("Tag", "tag"), ("Description", "description"), ("Cable", "cable"), ("Core", "cable_core"), ("JB/TB", "terminal_full"), ("Addr", "plc_addr_wordbit"), ("Panel", "panel"), ("Area", "area"), ("Range", "range")], "sort": "tag", "extras": True}),
    ]
    for file_name, summary, cfg in configs:
        write_marshalling_variant(BASE_DIR / file_name, rows, cfg)
        register_summary(file_name, summary)


def family_loop(rows: list[dict]) -> None:
    configs = [
        ("36_io_instrument_index_compact.xlsx", "Instrument-index-like compact register with loop, P&ID and service columns.", {"sheet": "InstIndex", "title": "Instrument Index Derived IO List", "columns": [("Loop No", "loop_no"), ("Tag", "tag"), ("Service", "description"), ("P&ID", "p_and_id"), ("Signal", "signal_type"), ("Unit", "eng_unit"), ("Range", "range"), ("Panel", "panel"), ("Addr", "plc_addr_std")], "sort": "tag"}),
        ("37_io_instrument_panel_first.xlsx", "Instrument-derived list ordered by panel and card position first.", {"sheet": "InstPanel", "title": "Instrument Index by Panel", "columns": [("Panel", "panel"), ("Card Pos", "card_pos"), ("Loop", "loop_no"), ("Tag", "tag"), ("Description", "description"), ("P&ID", "p_and_id"), ("Signal", "signal_type"), ("I/O", "io_type"), ("Unit", "eng_unit")], "sort": "panel"}),
        ("38_io_loop_schedule.xlsx", "Loop schedule style with explicit low/high engineering units.", {"sheet": "LoopSched", "title": "Loop Schedule", "columns": [("Loop", "loop_no"), ("Area", "area"), ("Tag", "tag"), ("Service", "description"), ("Unit", "eng_unit"), ("Low", "range_min"), ("High", "range_max"), ("Signal", "signal_type"), ("Location", "location")], "sort": "area"}),
        ("39_io_loop_schedule_multisheet.xlsx", "Loop schedule workbook split by area, but each sheet now carries full panel/rack/slot/channel topology.", {"cover": "Index", "cover_rows": [["Index", "Loop schedules by area"], ["Scope", "All logical points included"], ["Topology", "Panel + rack + slot + channel on every data row"]], "split_by": "area", "title": "Loop Schedule by Area", "columns": [("Loop", "loop_no"), ("Tag", "tag"), ("Service", "description"), ("P&ID", "p_and_id"), ("Panel", "panel"), ("Rack", "rack02"), ("Slot", "slot02"), ("Ch", "channel02"), ("Unit", "eng_unit"), ("Range", "range"), ("Signal", "signal_type")], "summary_by": "area"}),
        ("40_io_pid_reference_list.xlsx", "P&ID-reference-driven list with panel/card routing appended.", {"sheet": "PIDRef", "title": "P&ID Reference IO List", "columns": [("P&ID", "p_and_id"), ("Loop", "loop_no"), ("Tag", "tag"), ("Description", "description"), ("Panel", "panel"), ("Card Pos", "card_pos"), ("JB/TB", "terminal_full"), ("Comment", "comment")], "sort": "area"}),
    ]
    for file_name, summary, cfg in configs:
        if "split_by" in cfg:
            write_multisheet_variant(BASE_DIR / file_name, rows, cfg)
        else:
            write_flat_variant(BASE_DIR / file_name, rows, cfg)
        register_summary(file_name, summary)


def family_scada(rows: list[dict]) -> None:
    configs = [
        ("41_io_scada_object_list.xlsx", "SCADA object list with object naming, alarm text and standard PLC address.", {"sheet": "Objects", "title": "SCADA Object List", "columns": [("Object", "scada_object"), ("Tag", "tag"), ("Description", "description"), ("Address", "plc_addr_std"), ("Panel", "panel"), ("Type", "io_class"), ("Alarm", "alarm_text"), ("Unit", "eng_unit"), ("Range", "range")], "sort": "tag", "alarm_sheet": True}),
        ("42_io_scada_address_table.xlsx", "SCADA address table emphasizing address, driver station and scan groups.", {"sheet": "AddressTbl", "title": "SCADA Address Table", "columns": [("Object", "scada_object"), ("Driver", "remote_node"), ("Address", "plc_addr_logix"), ("Scan", "signal_group"), ("Alarm", "alarm_text"), ("Service", "description"), ("Area", "area_en"), ("Fieldbus", "fieldbus")], "sort": "tag"}),
        ("43_io_igss_import_style.xlsx", "IGSS-like import style using object-centric rows and alarm wording.", {"sheet": "IGSS", "title": "IGSS Import Style", "columns": [("Object Name", "scada_object"), ("Description", "description"), ("PLC", "plc"), ("Address", "plc_addr_unity"), ("Area", "area_en"), ("Alarm Text", "alarm_text"), ("Unit", "eng_unit"), ("Range", "range")], "sort": "tag", "noise_lines": True}),
        ("44_io_alarm_tag_register.xlsx", "Alarm register split from process points but preserving the same point set.", {"sheet": "AlarmTags", "title": "Alarm / Tag Register", "columns": [("Tag", "tag"), ("Alarm Text", "alarm_text"), ("Class", "io_class"), ("Area", "area"), ("Panel", "panel"), ("Address", "plc_addr_std"), ("Comment", "comment")], "sort": "tag", "alarm_sheet": True}),
        ("45_io_operator_point_list.xlsx", "Operator point list with object, visible text and process range focus.", {"sheet": "OperatorList", "title": "Operator Point List", "columns": [("Object", "scada_object"), ("Visible Text", "description"), ("Type", "io_type"), ("Unit", "eng_unit"), ("Range", "range"), ("Panel", "panel_short"), ("Location", "location"), ("Alarm", "alarm_text")], "sort": "tag", "blank_every": 18}),
    ]
    for file_name, summary, cfg in configs:
        write_scada_variant(BASE_DIR / file_name, rows, cfg)
        register_summary(file_name, summary)


def family_legacy(rows: list[dict]) -> None:
    configs = [
        ("46_io_legacy_db_dump.xlsx", "Legacy DB dump with preface rows and wide technical columns.", {"sheet": "LEGACY_DB", "title": "Legacy DB Dump", "subtitle": "Migrated engineering register", "meta_line": "Data starts below / source = IO_MASTER", "columns": [("RecID", "loop_no"), ("Panel", "panel"), ("Rack", "rack"), ("Slot", "slot"), ("Chan", "channel"), ("Tag", "tag_spaced"), ("Desc", "description"), ("Type", "io_type"), ("Signal", "signal_type"), ("Term", "terminal_full"), ("Loc", "location")], "sort": "panel", "fill": "7F6000"}),
        ("47_io_legacy_csv_wrapped.xlsx", "CSV-wrapped export with compact names and reordered columns.", {"sheet": "CSV_WRAP", "title": "Wrapped CSV Export", "columns": [("Pt", "tag"), ("Svc", "description_upper"), ("Pnl", "panel_short"), ("Loc", "card_pos"), ("I/O", "io_type"), ("Sig", "signal_type"), ("JB", "junction_box"), ("TB", "tb_term"), ("Cbl", "cable"), ("Rem", "comment")], "sort": "tag", "fill": "A5A5A5"}),
        ("48_io_noisy_site_export_v2.xlsx", "Noisy site export with wide rows, spacing and mixed business columns.", {"sheet": "SITE_EXPORT", "title": "Noisy Site Export", "subtitle": "As-built extract", "meta_line": "Ignore blank lines and repeated tag spaces.", "columns": [("Revision", "p_and_id"), ("Skid", "device"), ("Panel", "panel"), ("R/S/C", "card_pos"), ("Tag", "tag_spaced"), ("Description", "description"), ("I/O", "io_type"), ("Address", "plc_addr_wordbit"), ("Remark", "comment"), ("Area", "area")], "sort": "panel", "blank_every": 17, "fill": "C55A11"}),
        ("49_io_revision_history_export.xlsx", "Workbook with revision sheet plus a legacy-style export table.", {"cover": "RevHist", "cover_rows": [["Revision", "A"], ["Date", TODAY.isoformat()], ["Comment", "Legacy export reissued"]], "split_by": "signal_group", "title": "Revision History Export", "columns": [("Group", "signal_group"), ("Tag", "tag"), ("Desc", "description"), ("Panel", "panel"), ("Addr", "plc_addr_std"), ("Term", "terminal_full"), ("Comment", "comment")], "summary_by": "panel"}),
        ("50_io_asbuilt_field_notes.xlsx", "As-built field note workbook with mixed engineering and field comment emphasis.", {"sheet": "AsBuilt", "title": "As-Built Field Notes", "columns": [("Area", "area"), ("Panel", "panel"), ("Card", "card_id"), ("Tag", "tag"), ("Description", "description"), ("Signal", "signal_type"), ("Field Note", "comment"), ("Cable", "cable"), ("Terminal", "terminal_full"), ("Location", "location")], "sort": "area", "fill": "548235"}),
    ]
    for file_name, summary, cfg in configs:
        if "split_by" in cfg:
            write_multisheet_variant(BASE_DIR / file_name, rows, cfg)
        else:
            write_flat_variant(BASE_DIR / file_name, rows, cfg)
        register_summary(file_name, summary)


def family_cards(rows: list[dict]) -> None:
    configs = [
        ("51_io_card_face_schedule.xlsx", "Card-face schedule grouped by panel, rack and slot.", {"sheet": "CardFace", "title": "Card Face Schedule", "columns": [("Rack", "rack02"), ("Slot", "slot02"), ("Ch", "channel02"), ("Type", "io_type"), ("Tag", "tag"), ("Description", "description"), ("Signal", "signal_type"), ("Range", "range"), ("TB", "tb_term"), ("Panel", "panel")], "summary": True}),
        ("52_io_card_face_multirack.xlsx", "Card schedule explicitly mixing multiple racks in each section ordering.", {"sheet": "CardFaceMR", "title": "Multi-Rack Card Schedule", "columns": [("Panel", "panel"), ("Card Pos", "card_pos"), ("Card", "module_type"), ("Tag", "tag"), ("Svc", "description"), ("I/O", "io_type"), ("Cable", "cable"), ("Term", "terminal_full")], "summary": True}),
        ("53_io_slot_channel_register.xlsx", "Slot/channel register with hardware-first ordering and terse technical naming.", {"sheet": "SlotCh", "title": "Slot Channel Register", "columns": [("Rack", "rack"), ("Slot", "slot"), ("Channel", "channel"), ("Module", "module_model"), ("Point", "tag"), ("Type", "io_type"), ("Addr", "plc_addr_std"), ("Location", "location")], "summary": True}),
        ("54_io_module_capacity_view.xlsx", "Module capacity view grouped by physical cards and installation references.", {"sheet": "Capacity", "title": "Module Capacity View", "columns": [("Card Group", "module_group"), ("Card Pos", "card_pos"), ("Tag", "tag"), ("Description", "description"), ("Signal", "signal_type"), ("Panel", "panel"), ("JB/TB", "terminal_full")], "summary": True}),
        ("55_io_module_channel_matrix.xlsx", "Module-channel style schedule with channel-first rows and terminal routing.", {"sheet": "MatrixView", "title": "Module Channel Matrix", "columns": [("Ch", "channel02"), ("Tag", "tag"), ("Desc", "description"), ("I/O", "io_type"), ("Signal", "signal_type"), ("Range", "range"), ("TB", "tb_term"), ("Cable", "cable"), ("Rack/Slot", "rack_slot"), ("Panel", "panel")], "summary": True}),
    ]
    for file_name, summary, cfg in configs:
        write_card_variant(BASE_DIR / file_name, rows, cfg)
        register_summary(file_name, summary)


def family_extreme(rows: list[dict]) -> None:
    configs = [
        ("56_io_field_check_packets.xlsx", "Field-check workbook split by packets with title blocks and auxiliary legend, now preserving full physical topology on each row.", {
            "title": "Field Check Packet",
            "hint": "Main data is split across Area_A and Area_B sheets.",
            "legend": [("DI", "Digital input"), ("AI", "Analog input"), ("Addr", "Embedded in text")],
            "columns": [("Loop", "loop_no"), ("Panel", "panel"), ("Rack", "rack02"), ("Slot", "slot02"), ("Ch", "channel02"), ("Panel / Card", "card_pos"), ("Point", "tag"), ("Type / Sig", "signal_type"), ("Service", "description"), ("Area", "area"), ("Term", "terminal_full"), ("Range", "range"), ("Remark", "comment")],
            "splitter": lambda src: [("Area_A", src[::2]), ("Area_B", src[1::2])],
            "double_header": [("Topology", 5), ("Identity", 3), ("Process", 2), ("Installation", 3)],
            "fill": "5B9BD5",
        }),
        ("57_io_hybrid_pten_headers.xlsx", "Hybrid PT/EN workbook with mixed tabs and bilingual headers.", {
            "title": "Hybrid PT-EN Workbook",
            "hint": "Headers intentionally mix Portuguese and English naming.",
            "legend": [("Painel", "Panel"), ("Descricao", "Description"), ("Faixa", "Range")],
            "columns": [("Painel", "panel"), ("Área", "area"), ("Card Pos", "card_pos"), ("Tag", "tag"), ("Descricao", "description"), ("Type", "io_type"), ("Unid", "eng_unit"), ("Faixa", "range"), ("Coment", "comment")],
            "splitter": lambda src: [("MainData", sort_rows(src, "panel")), ("AltOrder", sort_rows(src, "tag"))],
            "fill": "70AD47",
        }),
        ("58_io_obscure_headers.xlsx", "Obscure-header workbook with strongly abbreviated columns and helper legend tab.", {
            "title": "Obscure Header Export",
            "hint": "Importer should infer columns from short names.",
            "legend": [("Pnl", "Panel"), ("Loc", "Location"), ("Svc", "Service"), ("Sig", "Signal"), ("Tm", "Terminal")],
            "columns": [("Pnl", "panel_short"), ("Loc", "card_pos"), ("Pt", "tag"), ("Svc", "description"), ("I/O", "io_type"), ("Sig", "signal_type"), ("JB", "junction_box"), ("Tm", "tb_term"), ("Cbl", "cable"), ("Rm", "comment")],
            "splitter": lambda src: [("EXP_A", sort_rows(src, "panel"))],
            "fill": "A5A5A5",
        }),
        ("59_io_staggered_sections.xlsx", "Workbook with staggered sections and two-level header blocks.", {
            "title": "Staggered Section Export",
            "hint": "Sections are grouped by panel but remain in one workbook.",
            "legend": [("Section", "Panel grouping"), ("Double Header", "Present"), ("Data", "Starts below title block")],
            "columns": [("Panel", "panel"), ("Rack", "rack02"), ("Slot", "slot02"), ("Ch", "channel02"), ("Tag", "tag"), ("Desc", "description"), ("Type", "io_type"), ("Signal", "signal_type"), ("TB", "tb_term"), ("Loc", "location")],
            "splitter": lambda src: [(panel, items) for panel, items in sorted(defaultdict(list, {p: [r for r in src if r['panel'] == p] for p in sorted({r['panel'] for r in src})}).items())],
            "double_header": [("Hardware", 4), ("Point", 3), ("Wiring", 3)],
            "fill": "ED7D31",
        }),
        ("60_io_extreme_mixed_workbook.xlsx", "Extreme mixed workbook with multilingual cover, split data tabs and helper references.", {
            "title": "Extreme Mixed Workbook",
            "hint": "Data exists in non-obvious tabs and uses varied column semantics.",
            "legend": [("Guide", "Use all data tabs"), ("Racks", "1..3 mixed"), ("Fieldbus", "Embedded in remote station")],
            "columns": [("Object", "scada_object"), ("Panel/Card", "card_pos"), ("Tag", "tag"), ("Type", "io_type"), ("Service", "description"), ("Addr", "plc_addr_logix"), ("Fieldbus", "fieldbus"), ("JB/TB", "terminal_full"), ("Range", "range"), ("Area", "area_en")],
            "splitter": lambda src: [("zz_dump_A", sort_rows(src[::3], "tag")), ("Data_Core", sort_rows(src[1::3], "panel")), ("Ref_Mixed", sort_rows(src[2::3], "rack"))],
            "double_header": [("Object", 3), ("PLC Mapping", 3), ("Installation", 4)],
            "fill": "C00000",
        }),
    ]
    for file_name, summary, cfg in configs:
        write_extreme_variant(BASE_DIR / file_name, rows, cfg)
        register_summary(file_name, summary)


def generate_all() -> None:
    rows = load_rows()
    random.seed(SEED)
    family_flat(rows)
    family_plc(rows)
    family_fieldbus(rows)
    family_signal(rows)
    family_marshalling(rows)
    family_loop(rows)
    family_scada(rows)
    family_legacy(rows)
    family_cards(rows)
    family_extreme(rows)


def validate_outputs() -> None:
    for number in range(11, 61):
        matches = list(BASE_DIR.glob(f"{number:02d}_*.xlsx"))
        if len(matches) != 1:
            raise ValueError(f"Expected exactly one file for prefix {number:02d}, found {len(matches)}")
        load_workbook(matches[0])


def print_summary() -> None:
    for file_name, summary in FAMILY_SUMMARIES:
        print(f"{file_name}: {summary}")


def main() -> None:
    generate_all()
    validate_outputs()
    print(f"Generated {len(FAMILY_SUMMARIES)} additional workbooks from canonical dataset.")
    print_summary()


if __name__ == "__main__":
    main()
